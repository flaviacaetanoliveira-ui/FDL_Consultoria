from __future__ import annotations

import html
from datetime import date, datetime, timedelta, timezone
from email.utils import parsedate_to_datetime
from io import BytesIO
import base64
import json
import hashlib
import math
import numbers
import re
import os
from pathlib import Path
import shutil
import time
import unicodedata
from typing import Any, Callable, Literal
from textwrap import dedent
from urllib.parse import parse_qsl, quote, urlencode, urlparse, urlunparse, urljoin
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen
import zipfile
from zoneinfo import ZoneInfo

import pandas as pd

# Bases grandes (ex.: repasse Mega Fácil) excedem o limite default do Styler (262144 células) e o
# Streamlit rebenta com StreamlitAPIException ao renderizar ``st.dataframe(..., column_config=…)``.
try:
    _fdl_styler_max = int(os.environ.get("FDL_STYLER_MAX_ELEMENTS", "2097152").strip())
except ValueError:
    _fdl_styler_max = 2_097_152
try:
    pd.set_option("styler.render.max_elements", _fdl_styler_max)
except Exception:
    pass

# Paleta semântica — status na UI Resultado Gerencial (tabela NF e referência para badges)
CORES_STATUS: dict[str, str] = {
    "Lucro": "#22c55e",
    "Prejuízo": "#ef4444",
    "Neutro": "#6b7280",
    "Atenção": "#f59e0b",
    "Saudável": "#16a34a",
}

import streamlit as st
from streamlit.column_config import DatetimeColumn, NumberColumn, SelectboxColumn, TextColumn
from openpyxl.styles import numbers as oxl_number_formats
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

import comercial_pedidos_analise as cpa

from carregamento_bases import PIPELINE_DATA_REVISION
from etapa4b_integracao_contas_receber import BASE_DIR, carregar_tabela_final_operacional
from faturamento_dre_recorte import (
    _fdl_fr_mask_nf_emissao_no_periodo,
    apply_recorte_modulo,
    faturamento_recorte_state_from_session,
)
from processing.faturamento.fiscal_commercial_nf_merge import merge_fiscal_base_with_commercial_nf_dataframe
from processing.faturamento.fiscal_materializado import fiscal_contract_dataframe_valid
from processing.faturamento.nf_materializado import nf_first_contract_dataframe_valid
from processing.faturamento.nf_panel_materializado import nf_panel_materializado_dataframe_valid
from processing.faturamento.normalize import (
    normalize_empresa_fiscal_commercial_join_key,
    normalize_nf_fiscal_commercial_join_key,
)
from processing.faturamento.nf_table_display_filters import nf_table_filter_mask as _fdl_nf_table_filter_mask
from processing.repasse_contract import REPASSE_ARTIFACT_FILENAME
from processing.repasse_load import (
    postprocess_repasse_parquet_dataframe,
    read_repasse_parquet,
    repasse_use_parquet_flag,
)
from processing.repasse_ui_grid import (
    REPASSE_UI_GRID_ROW_CAP,
    repasse_ui_apply_grid_styler,
    repasse_ui_grid_display_slice,
)
from processing.repasse_ui_session import (
    COL_ACAO_LEGACY_UI,
    COL_DATA_PERIODO_REPASSE,
    repasse_ui_acao_column,
    repasse_ui_apply_filtro_somente_linhas_com_data_pagamento,
    repasse_ui_apply_pipeline_exclusao_na_ui,
    repasse_ui_periodo_series_parquet,
)
from faturamento_dre_recorte_minimo import (
    CommercialCoverageStats,
    FaturamentoFiscalBaseStats,
    FaturamentoRecorteMinState,
    _min_cal_limits,
    build_faturamento_fiscal_base_slice,
    build_nf_panel_aligned_to_fiscal_base,
    compute_commercial_coverage_stats,
    compute_nf_panel_kpis,
    dre_imposto_para_linha_dre_gerencial,
    faturamento_min_series_nf_emissao_bounds_dates,
    faturamento_recorte_min_state_from_session,
    nf_grain_plataforma_label_for_ui,
    nf_grain_plataforma_match_key,
    nf_grain_plataforma_ui_options,
)

try:
    from app.components.faturamento_dre_ui import (
        build_dre_gerencial_premium_html,
        build_kpi_nf_premium_shell_html,
        fat_dre_premium_css,
        faturamento_section_rule_html,
    )

    _FAT_DRE_UI_V2 = True
except ImportError:
    _FAT_DRE_UI_V2 = False
    build_dre_gerencial_premium_html = None  # type: ignore[misc,assignment]
    build_kpi_nf_premium_shell_html = None  # type: ignore[misc,assignment]
    fat_dre_premium_css = lambda: ""  # type: ignore[misc,assignment]
    faturamento_section_rule_html = lambda _lbl: ""  # type: ignore[misc,assignment]
from app.components.rg_layout_helpers import (
    rg_pick_empresa_maior_receita_mes_fechado as _rg_pick_empresa_maior_receita_mes_fechado,
)
from fdl_paths import resolve_pasta_vendas_ml
from operacional_app_context import (
    SESSION_ACTIVE_ORG_KEY,
    get_active_organization,
    logout_operacional_user,
    nomes_permitidos_com_registro,
    organizacao_por_nome_cadastrado,
    require_app_user,
)

# MVP Faturamento & DRE: área de produto separada de Financeiro (Repasse/Frete).
SESSION_FDL_PRODUCT_AREA_KEY = "fdl_product_area"
FDL_PRODUCT_AREA_FINANCEIRO = "financeiro"
FDL_PRODUCT_AREA_FATURAMENTO_DRE = "faturamento_dre"
FDL_PRODUCT_AREA_COMERCIAL_PEDIDOS = "comercial_pedidos"
FDL_PRODUCT_AREA_APURACAO_FISCAL = "apuracao_fiscal"

# Filtros globais / escopo de carga (MVP)
FAT_DRE_ESCOPO_EMPRESA = "empresa_ativa"
FAT_DRE_ESCOPO_CONSOLIDADO = "consolidado"
# Carga Faturamento & DRE: consolidado não usa org da sidebar; chave de cache estável.
FAT_DRE_CACHE_ACTIVE_ORG_PLACEHOLDER = "_fdl_fat_dre_consolidado_load"

# Select "Visão" no painel Faturamento: rótulo explícito vs. legado em sessão.
_FAT_PAINEL_VISAO_COM_NF = "Com NF (pedido)"
_FAT_PAINEL_VISAO_COM_NF_LEGACY = "Com NF"

from operacional_data_config import DATASET_EMPRESA
from operacional_frete import (
    FRETE_ML_COL,
    FRETE_UI_ANUNCIO,
    FRETE_UI_ANALISADO_COBRADO_MAIOR,
    FRETE_UI_ANALISADO_COBRADO_MENOR,
    FRETE_UI_ANALISADO_REPASSE_FRETE,
    FRETE_UI_CLASSIFICACAO,
    FRETE_UI_DIFERENCA,
    FRETE_UI_N_VENDA,
    FRETE_UI_RECEBIDO,
    FRETE_UI_SITUACAO_FRETE,
    FRETE_UI_STATUS_CONC,
    FRETE_UI_VAL_DIVERGENCIA,
    FRETE_SITUACAO_FRETE_VALORES_FILTRO,
    FRETE_VAL_RECEBIDO_NAO,
    FRETE_VAL_RECEBIDO_SIM,
    FontesFrete,
    carregar_tabela_final_frete_operacional,
    compute_frete_situacao_frete_column,
    dataframe_frete_conciliacao_principal,
    descobrir_fontes_frete,
    frete_vendas_loader_args,
    frete_kpis_executivos,
    frete_series_for_date_filter,
    frete_series_normalize_sale_dt,
    frete_tabela_anuncios_cobrado_maior,
    frete_tabela_anuncios_repasse_frete,
    normalize_frete_status_conc_display,
    stable_mtime_ns_for_frete_url,
    validate_frete_operacional_dataframe,
)
from operacional_frete_ui import (
    _column_config_frete,
    _dataframe_frete_grid,
    _frete_conciliacao_grid_com_icones,
    frete_executivo_display_styled,
)

_REPO_APP_ROOT = Path(__file__).resolve().parent
BUILD_TAG = "build-20260329-repasse-ui-saas"


def _sidebar_version_display() -> str:
    """Rótulo curto para a sidebar (ex.: v20260329)."""
    for tok in BUILD_TAG.replace("build-", "").split("-"):
        if len(tok) == 8 and tok.isdigit():
            return f"v{tok}"
    return "v—"



try:
    st.set_page_config(page_title="FDL Analytics — Financeiro", layout="wide")
except Exception:
    pass  # já definido por app.py no primeiro arranque


def _fdl_global_trace(msg: str) -> None:
    """Marca etapa do rerun (FDL_DEBUG_BOOTSTRAP). Deve existir antes do restante do módulo."""
    try:
        st.session_state["_fdl_bootstrap_stage"] = msg
        log = st.session_state.setdefault("_fdl_bootstrap_log", [])
        if not isinstance(log, list):
            log = []
            st.session_state["_fdl_bootstrap_log"] = log
        log.append(msg)
        if len(log) > 48:
            del log[:24]
    except Exception:
        pass


def _materialized_path_mode() -> str:
    """
    fixed — FDL_REPASSE_MATERIALIZED_PATH / frete / faturamento nos secrets (legado).
    dynamic — repasse/frete/faturamento derivados de data_products/<cliente>/<org_id>/... pela org ativa.
    """
    raw = os.environ.get("FDL_MATERIALIZED_PATH_MODE", "").strip().lower()
    if raw in {"dynamic", "fixed"}:
        return raw
    try:
        s = str(st.secrets.get("FDL_MATERIALIZED_PATH_MODE", "")).strip().lower()
        if s in {"dynamic", "fixed"}:
            return s
    except Exception:
        pass
    return "fixed"


def _materialized_cliente_slug() -> str:
    """Segmento de pasta do cliente (ex.: cliente_2), alinhado a materialize_financeiro --cliente."""
    raw = os.environ.get("FDL_MATERIALIZED_CLIENTE_SLUG", "").strip()
    if raw:
        return raw
    try:
        return str(st.secrets.get("FDL_MATERIALIZED_CLIENTE_SLUG", "")).strip()
    except Exception:
        return ""


def _materialized_data_products_root() -> str:
    raw = os.environ.get("FDL_DATA_PRODUCTS_ROOT", "").strip()
    if raw:
        return raw
    try:
        v = str(st.secrets.get("FDL_DATA_PRODUCTS_ROOT", "data_products")).strip()
        return v or "data_products"
    except Exception:
        return "data_products"


def _dynamic_materialized_repasse_rel_path(org_id: str) -> str:
    cliente = _materialized_cliente_slug()
    if not cliente or not (org_id or "").strip():
        return ""
    root = _materialized_data_products_root().strip().strip("/\\")
    oid = org_id.strip()
    if _repasse_use_parquet():
        return f"{root}/{cliente}/{oid}/repasse/current/{REPASSE_ARTIFACT_FILENAME}"
    return f"{root}/{cliente}/{oid}/repasse/current/dataset_repasse_app.csv"


def _dynamic_materialized_frete_rel_path(org_id: str) -> str:
    """Mesmo layout que materialize_financeiro; não depender do CSV de repasse existir para derivar o path."""
    cliente = _materialized_cliente_slug()
    if not cliente or not (org_id or "").strip():
        return ""
    root = _materialized_data_products_root().strip().strip("/\\")
    oid = org_id.strip()
    return f"{root}/{cliente}/{oid}/frete/current/dataset_frete_app.csv"


def _dynamic_materialized_devolucoes_rel_path(org_id: str) -> str:
    """``data_products/<slug>/<org_id>/devolucoes/current/`` — o app prefere ``dataset.parquet`` se existir ao lado do CSV."""
    cliente = _materialized_cliente_slug()
    if not cliente or not (org_id or "").strip():
        return ""
    root = _materialized_data_products_root().strip().strip("/\\")
    oid = org_id.strip()
    return f"{root}/{cliente}/{oid}/devolucoes/current/dataset_devolucoes_app.csv"


_fdl_global_trace("01: início app_operacional (módulo reexecutado)")
_app_ctx = require_app_user()
_fdl_global_trace("02: após autenticação (require_app_user)")
_active_org = get_active_organization(_app_ctx)


def _apply_antomoveis_single_tenant_env_defaults() -> None:
    """
    Inquilino só Antomóveis (ex.: Everton): alinha leitura da base final a
    ``data_products/default/antomoveis/...`` mesmo quando o Main file do Streamlit é ``app.py``
    em vez de ``app_cliente_everton.py``.

    Usa ``os.environ.setdefault`` para não sobrepor variáveis de ambiente já definidas no deploy.
    """
    orgs = getattr(_app_ctx, "organizations", None) or ()
    if len(orgs) != 1 or orgs[0].org_id != "antomoveis":
        return
    os.environ.setdefault("FDL_MATERIALIZED_CLIENTE_SLUG", "default")
    os.environ.setdefault("FDL_MATERIALIZED_PATH_MODE", "dynamic")
    os.environ.setdefault("FDL_REPASSE_CONSUME_MODE", "materialized")
    os.environ.setdefault("FDL_FRETE_CONSUME_MODE", "materialized")
    os.environ.setdefault("FDL_ENABLED_FINANCE_MODULES", "repasse,frete,faturamento")


_apply_antomoveis_single_tenant_env_defaults()


def _dataset_empresa_label() -> str:
    """Rótulo da coluna `empresa` quando o dataset não a traz; em dynamic alinha à org ativa."""
    if _materialized_path_mode() == "dynamic":
        return _active_org.display_name
    return DATASET_EMPRESA


def _enabled_finance_modules() -> set[str]:
    """
    Módulos visíveis na sidebar (default: repasse, frete, faturamento).
    ``faturamento`` controla o módulo **Faturamento & DRE** (fora do grupo Financeiro).
    Ex.: FDL_ENABLED_FINANCE_MODULES=repasse,frete para clientes sem faturamento.
    """
    raw = os.environ.get("FDL_ENABLED_FINANCE_MODULES", "").strip()
    if not raw:
        try:
            raw = str(st.secrets.get("FDL_ENABLED_FINANCE_MODULES", "")).strip()
        except Exception:
            raw = ""
    if not raw:
        return {"repasse", "frete", "faturamento"}
    out = {x.strip().lower() for x in raw.split(",") if x.strip()}
    valid = {"repasse", "frete", "devolucoes", "faturamento"}
    return out & valid or {"repasse", "frete", "faturamento"}


def _user_perfil_acesso_operacional_only() -> bool:
    """Cadastro ``perfil_acesso=operacional``: só módulos operacionais (Repasse/Frete/Devoluções), sem Faturamento nem Comercial na sidebar."""
    if not st.session_state.get("logged_in"):
        return False
    p = str(st.session_state.get("fdl_perfil_acesso", "completo")).strip().lower()
    return p == "operacional"


def _repasse_sem_bling() -> bool:
    """Cliente sem Bling: ações operacionais usam «Baixado» em vez de «Baixar no Bling»."""
    raw = os.environ.get("FDL_REPASSE_SEM_BLING", "").strip().lower()
    if raw in {"1", "true", "yes", "on"}:
        return True
    if raw in {"0", "false", "no", "off"}:
        return False
    try:
        sec = st.secrets.get("FDL_REPASSE_SEM_BLING", False)
        if isinstance(sec, bool):
            return sec
        return str(sec).strip().lower() in {"1", "true", "yes", "on"}
    except Exception:
        return False


def _repasse_vendas_liberacoes_only() -> bool:
    """
    Cenário específico (ex.: Thiago): repasse sem notas/contas, só vendas x liberações.
    """
    raw = os.environ.get("FDL_REPASSE_VENDAS_LIBERACOES_ONLY", "").strip().lower()
    if raw in {"1", "true", "yes", "on"}:
        return True
    if raw in {"0", "false", "no", "off"}:
        return False
    try:
        sec = st.secrets.get("FDL_REPASSE_VENDAS_LIBERACOES_ONLY", False)
        if isinstance(sec, bool):
            return sec
        return str(sec).strip().lower() in {"1", "true", "yes", "on"}
    except Exception:
        return False


def _filtrar_df_col_empresa_por_contexto(
    df: pd.DataFrame,
    *,
    repasse_org_scoped_fallback: bool = False,
) -> pd.DataFrame:
    """
    Em modo fixed/live, restringe às empresas permitidas ou, em cenários multi-empresa no mesmo ficheiro,
    alinha ao contexto.

    Em **dynamic**, cada artefacto já vive em ``data_products/<cliente>/<org_id>/...`` — não filtrar pela
    coluna ``empresa``: muitas materializações usam ``FDL_DATASET_EMPRESA`` global (ex.: «Antomóveis») e o
    filtro por ``display_name`` esvaziava o repasse/faturamento para todas as orgs.

    Com ``repasse_org_scoped_fallback=True``, se o filtro esvaziar o quadro mas o ficheiro materializado
    de repasse for claramente da org ativa (path/URL contém ``/<org_id>/``), devolve o quadro sem filtrar
    pela coluna ``empresa`` (rótulo desatualizado na base final).
    """
    if df.empty or "empresa" not in df.columns:
        return df
    if _materialized_path_mode() == "dynamic":
        return df
    empresas = st.session_state["empresas_permitidas"]
    out = df[df["empresa"].isin(empresas)].copy()
    if not out.empty or not repasse_org_scoped_fallback:
        return out
    oid = (_active_org.org_id or "").strip()
    if not oid or _repasse_consume_mode() != "materialized":
        return out
    path = (_repasse_materialized_path_str() or "").replace("\\", "/").lower()
    url = (_repasse_materialized_url_str() or "").replace("\\", "/").lower()
    needle = f"/{oid.lower()}/"
    if needle in path or needle in url:
        return df
    return out


if "op_financeiro_view" not in st.session_state:
    st.session_state["op_financeiro_view"] = "repasse"
elif st.session_state["op_financeiro_view"] not in ("repasse", "frete", "devolucoes", "faturamento"):
    st.session_state["op_financeiro_view"] = "repasse"

_enabled_modules = _enabled_finance_modules()

# Área de produto: Financeiro (repasse/frete) vs Faturamento & DRE (módulo próprio).
if SESSION_FDL_PRODUCT_AREA_KEY not in st.session_state:
    if st.session_state.get("op_financeiro_view") == "faturamento":
        st.session_state[SESSION_FDL_PRODUCT_AREA_KEY] = FDL_PRODUCT_AREA_FATURAMENTO_DRE
        st.session_state["op_financeiro_view"] = "repasse"
    else:
        st.session_state[SESSION_FDL_PRODUCT_AREA_KEY] = FDL_PRODUCT_AREA_FINANCEIRO
# Migração de sessões antigas que ainda tinham só op_financeiro_view == faturamento
if st.session_state.get("op_financeiro_view") == "faturamento":
    st.session_state[SESSION_FDL_PRODUCT_AREA_KEY] = FDL_PRODUCT_AREA_FATURAMENTO_DRE
    st.session_state["op_financeiro_view"] = "repasse"
if st.session_state["op_financeiro_view"] not in ("repasse", "frete", "devolucoes"):
    st.session_state["op_financeiro_view"] = "repasse"

if st.session_state["op_financeiro_view"] not in _enabled_modules:
    for _cand in ("repasse", "frete", "devolucoes"):
        if _cand in _enabled_modules:
            st.session_state["op_financeiro_view"] = _cand
            break
    else:
        st.session_state["op_financeiro_view"] = "repasse" if "repasse" in _enabled_modules else "frete"

if "faturamento" not in _enabled_modules and st.session_state.get(SESSION_FDL_PRODUCT_AREA_KEY) in (
    FDL_PRODUCT_AREA_FATURAMENTO_DRE,
    FDL_PRODUCT_AREA_COMERCIAL_PEDIDOS,
    FDL_PRODUCT_AREA_APURACAO_FISCAL,
):
    st.session_state[SESSION_FDL_PRODUCT_AREA_KEY] = FDL_PRODUCT_AREA_FINANCEIRO

if _user_perfil_acesso_operacional_only() and st.session_state.get(SESSION_FDL_PRODUCT_AREA_KEY) in (
    FDL_PRODUCT_AREA_FATURAMENTO_DRE,
    FDL_PRODUCT_AREA_COMERCIAL_PEDIDOS,
    FDL_PRODUCT_AREA_APURACAO_FISCAL,
):
    st.session_state[SESSION_FDL_PRODUCT_AREA_KEY] = FDL_PRODUCT_AREA_FINANCEIRO

_fdl_global_trace("03: após definir vista financeiro (session_state)")

# Alinhado a PIPELINE_DATA_REVISION (liberações / Valor pago). Subir junto quando mudar o fluxo.
OPERACIONAL_CACHE_REVISION = PIPELINE_DATA_REVISION
REQUIRED_ONEDRIVE_CSV_COLUMNS = {
    "N° de venda",
    "ID do pedido",
    "Número da nota",
    "Plataforma",
    "Situação",
    "Ação sugerida",
    "Valor pago",
    "Valor da nota",
    "Valor a receber",
    "Diferença",
    "Data de pagamento",
}
REQUIRED_ONEDRIVE_SOURCE_FOLDERS = {
    "Vendas - Mercado Livre",
    "Liberações_ML",
    "notas_saida",
    "contas_receber",
}
RETRYABLE_HTTP_CODES = {429, 500, 502, 503, 504}
MAX_HTTP_RETRIES = 4
ONEDRIVE_SYNC_MIN_INTERVAL_SECONDS = 180
# Pedidos curtos na Cloud: evita ecrã em branco ~vários minutos em sequência (SharePoint pode tardar ou pendurar).
PRECOMPUTED_HTTP_TIMEOUT = 35
# Alguns links SharePoint só redirecionam para o binário quando o pedido parece um browser.
_BROWSER_UA_CHROME = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

_BR_TZ = ZoneInfo("America/Sao_Paulo")


def _safe_streamlit_date(value: object, fallback: date) -> date:
    """Evita TypeError ao comparar None com date (`st.date_input` pode devolver None)."""
    if value is None:
        return fallback
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return fallback


def _pd_to_datetime_pedido_br(series: pd.Series) -> pd.Series:
    """
    Datas do export ML / pedidos em **DD/MM/AAAA**. Sem ``dayfirst=True``, o pandas trata
    datas ambíguas como MM/DD (ex.: 05/02/2026 → 2 mai em vez de 5 fev), quebrando filtro por período e totais.
    """
    return pd.to_datetime(series, errors="coerce", dayfirst=True)


def _faturamento_series_bool_mask(series: pd.Series) -> pd.Series:
    """
    Flags booleanas do materializado: Parquet traz ``bool``; o CSV de app pode trazer ``\"True\"`` / ``\"False\"``.
    ``astype(bool)`` em string ``\"False\"`` em pandas devolve **True** (string não vazia) — incorreto para visão NF.
    """
    s = series
    if isinstance(s.dtype, pd.BooleanDtype):
        return s.fillna(False).astype(bool)
    if s.dtype == bool or pd.api.types.is_bool_dtype(s):
        return s.fillna(False).astype(bool)
    if pd.api.types.is_numeric_dtype(s):
        return pd.to_numeric(s, errors="coerce").fillna(0).ne(0)
    x = s.astype(str).str.strip().str.casefold()
    return x.eq("true") | x.eq("1") | x.eq("yes") | x.eq("sim")


def _faturamento_ts_nf_emissao_para_dia_civil(s: pd.Series) -> pd.Series:
    """``Nota_Data_Emissao`` vem em ISO ``YYYY-MM-DD HH:MM:SS`` do join; **não** usar ``dayfirst=True`` em série (pandas infer quebra parte das linhas)."""
    ts = pd.to_datetime(s, errors="coerce", dayfirst=False)
    if ts.empty:
        return ts
    if getattr(ts.dt, "tz", None) is not None:
        ts = ts.dt.tz_convert(_BR_TZ)
    return ts.dt.normalize()


def _faturamento_mask_nf_emissao_no_periodo(s: pd.Series, d_ini: date, d_fim: date) -> pd.Series:
    """Intervalo inclusive no dia civil BR para data de emissão da NF."""
    ts = _faturamento_ts_nf_emissao_para_dia_civil(s)
    if ts.empty:
        return pd.Series(False, index=ts.index)
    dcal = ts.dt.date
    ok = pd.notna(ts)
    ge = pd.Series(dcal, index=ts.index) >= d_ini
    le = pd.Series(dcal, index=ts.index) <= d_fim
    return ok & ge & le


def _faturamento_ts_pedido_para_dia_civil(s: pd.Series) -> pd.Series:
    """Converte coluna **Data** (texto BR, datetime64 ou tz-aware) para datetime64 naive em dia civil BR."""
    ts = pd.to_datetime(s, errors="coerce", dayfirst=True)
    if ts.empty:
        return ts
    if getattr(ts.dt, "tz", None) is not None:
        ts = ts.dt.tz_convert(_BR_TZ)
    return ts.dt.normalize()


def _faturamento_mask_venda_no_periodo(s: pd.Series, d_ini: date, d_fim: date) -> pd.Series:
    """
    Linhas cuja **Data** (venda) cai entre ``d_ini`` e ``d_fim`` inclusive, em dia civil.
    Evita comparações ``Timestamp`` vs meia-noite que falham com Parquet tz-aware ou tipos mistos.
    """
    ts = _faturamento_ts_pedido_para_dia_civil(s)
    if ts.empty:
        return pd.Series(False, index=ts.index)
    dcal = ts.dt.date
    ok = pd.notna(ts)
    ge = pd.Series(dcal, index=ts.index) >= d_ini
    le = pd.Series(dcal, index=ts.index) <= d_fim
    return ok & ge & le


def _series_datetime_bounds_dates(series: pd.Series, *, dayfirst: bool = True) -> tuple[date, date, bool]:
    """
    Min/max em dia civil a partir de coluna parseável como datetime.
    Não chama .min().date() sobre série só NaT (evita NaT/erros em limites).
    Devolve (d_min, d_max, tem_alguma_data_parseável).
    """
    ts = pd.to_datetime(series, errors="coerce", dayfirst=dayfirst)
    if getattr(ts.dt, "tz", None) is not None:
        ts = ts.dt.tz_convert(_BR_TZ)
    t = ts[ts.notna()]
    if t.empty:
        d = datetime.now(_BR_TZ).date()
        return d, d, False
    return t.min().date(), t.max().date(), True


def _faturamento_period_calendar_limits(d_min: date, d_max: date) -> tuple[date, date]:
    """
    Limites do ``date_input`` mais amplos que o min/max dos dados carregados.

    Sem isto, o utilizador não consegue escolher janeiro→hoje quando a base só tem março
    (o widget ficava preso a [d_min, d_max] e o estado era re-clampado a cada rerun).
    O filtro continua a usar só linhas cuja **Data** cai no intervalo escolhido.
    """
    today = datetime.now(_BR_TZ).date()
    cal_max = max(d_max, today)
    cal_min = min(d_min, today - timedelta(days=3 * 365))
    return cal_min, cal_max


def _fdl_rg_recorte_parcial_um_mes_sem_mes_cheio(data_inicio: date, data_fim: date) -> bool:
    """
    Verdadeiro quando o filtro cobre só parte de um único mês civil (ex.: 01/03–15/03).

    Mantido aqui (e não só importado de ``pace_mensal``) para o app não falhar se um deploy
    trouxer ``app_operacional`` sem o commit que adiciona o helper em ``pace_mensal``.
    Espelha ``recorte_parcial_mes_civil_sem_mes_cheio`` em ``processing/faturamento/pace_mensal.py``.
    """
    from calendar import monthrange

    if data_inicio.year != data_fim.year or data_inicio.month != data_fim.month:
        return False
    if data_inicio.day != 1:
        return True
    ult = date(
        data_inicio.year,
        data_inicio.month,
        monthrange(data_inicio.year, data_inicio.month)[1],
    )
    return data_fim != ult


# Convenção de produto (Faturamento & DRE):
# - Vista mínima: bloco **Conferência venda × NF** usa ``compute_comercial_conferencia_stats`` / ``compute_fiscal_nf_conferencia_stats``.
# - Vista completa / agregados comerciais: «receita líquida» pode usar Σ ``Nota_Valor_Liquido_Rateado`` ou ``Valor total``.
_FATURAMENTO_UI_VALOR_NOTA_FISCAL = "Valor Nota Fiscal"
_FATURAMENTO_HELP_VALOR_NOTA_FISCAL = (
    "Materializado V2 com notas: soma de **Nota_Valor_Liquido_Rateado** (valor líquido da nota de saída, rateado por linha). "
    "Sem join fiscal: fallback à coluna **Valor total** do pedido."
)
_FATURAMENTO_HELP_VL_NF_FISCAL_KPI_MIN = (
    "KPI **fiscal**: soma do **valor líquido total da NF** (``Nota_Valor_Liquido_Total``, uma vez por nota), "
    "com **Nota_Data_Emissao** no período acima. Exclui cancelada / denegada / inutilizada. "
    "Respeita só o filtro **Empresa** (a **Plataforma** não corta este total)."
)
_FATURAMENTO_HELP_VL_NF_COL_MIN_TABLE = (
    "Valor **por linha de venda** (rateio ``Nota_Valor_Liquido_Rateado`` quando existir). "
    "O total fiscal do período de emissão está no KPI **Vl. Nota Fiscal** acima, não na soma desta coluna."
)
_FATURAMENTO_HELP_PERIODO_NF_EMISSAO_MIN = (
    "Eixo **fiscal**: filtra pela **data de emissão** da nota (``Nota_Data_Emissao``). "
    "Independente do **período da venda** e da **Plataforma**. "
    "Neste painel o calendário **não considera emissões anteriores a 01/01/2026**."
)
_FATURAMENTO_HELP_PERIODO_DATA_VENDA_RG_MIN = (
    "Filtra pelo período da **data da venda** (coluna **Data** do pedido). "
    "O **imposto** apresentado na DRE é calculado sobre a **base fiscal** (período das NFs correspondentes, "
    "via ponte fiscal) — pode haver pequena **defasagem temporal** entre KPIs e linhas fiscais da DRE."
)
# Piso de emissão NF no painel mínimo Faturamento & DRE (produto / alinhamento fiscal 2026+).
_FDL_FAT_DRE_MIN_PANEL_NF_EMISSAO_DESDE = date(2026, 1, 1)
_FATURAMENTO_HELP_PERIODO_DATA = (
    "Eixo oficial do período: coluna **Data** (pedido / export ML). "
    "**Data do faturamento** na tabela é informativa e pode estar incompleta — não rege este filtro."
)
_FATURAMENTO_HELP_NUMERO_NF_COL = (
    "Com export só de pedidos ML, costuma vazio — **esperado**, não erro. "
    "NF confiável: outra fonte ou join futuro."
)


def _sb_user_initials(display_name: str) -> str:
    """Iniciais para avatar (máx. 2 caracteres)."""
    parts = [p for p in str(display_name).strip().split() if p]
    if len(parts) >= 2:
        return (parts[0][0] + parts[1][0]).upper()[:2]
    if parts:
        return parts[0][:2].upper()
    return "?"


def _sb_nav_set_repasse() -> None:
    st.session_state["op_financeiro_view"] = "repasse"
    st.session_state[SESSION_FDL_PRODUCT_AREA_KEY] = FDL_PRODUCT_AREA_FINANCEIRO


def _sb_nav_set_frete() -> None:
    st.session_state["op_financeiro_view"] = "frete"
    st.session_state[SESSION_FDL_PRODUCT_AREA_KEY] = FDL_PRODUCT_AREA_FINANCEIRO


def _sb_nav_set_devolucoes() -> None:
    st.session_state["op_financeiro_view"] = "devolucoes"
    st.session_state[SESSION_FDL_PRODUCT_AREA_KEY] = FDL_PRODUCT_AREA_FINANCEIRO


def _sb_nav_set_faturamento_dre() -> None:
    st.session_state[SESSION_FDL_PRODUCT_AREA_KEY] = FDL_PRODUCT_AREA_FATURAMENTO_DRE


def _sb_nav_set_apuracao_fiscal() -> None:
    st.session_state[SESSION_FDL_PRODUCT_AREA_KEY] = FDL_PRODUCT_AREA_APURACAO_FISCAL


def _sb_nav_set_comercial_pedidos() -> None:
    st.session_state[SESSION_FDL_PRODUCT_AREA_KEY] = FDL_PRODUCT_AREA_COMERCIAL_PEDIDOS


def _sb_logout_click() -> None:
    logout_operacional_user()
    st.rerun()


def _fdl_sidebar_inject_layout_css() -> None:
    """Tipografia e navegação da sidebar (gerencial vs operacional, estados ativos)."""
    st.markdown(
        dedent(
            """
            <style>
            [data-testid="stSidebar"] .block-container {
              padding-top: 0.5rem;
              padding-bottom: 3.5rem;
              position: relative;
            }
            [data-testid="stSidebar"] div[data-testid="stButton"] {
              margin-bottom: 6px !important;
            }
            .fdl-sb-brand-shell {
              margin: 0.06rem 0 0.35rem 0;
              padding: 1rem 0.75rem 0.85rem;
              border-radius: 12px;
              background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 55%, #e8eef5 100%);
              border: 1px solid #e2e8f0;
              box-shadow: 0 1px 3px rgba(15, 23, 42, 0.05);
            }
            .fdl-sb-brand {
              margin: 0;
              padding: 0;
            }
            .fdl-sb-wordmark {
              display: block;
              margin: 0 auto 0.35rem auto;
              max-width: 100%;
              padding: 0 0.1rem;
              text-align: center;
              font-family: Inter, ui-sans-serif, system-ui, -apple-system, "Segoe UI", Roboto, sans-serif;
              line-height: 1.2;
            }
            .fdl-sb-wordmark-fdl {
              font-size: 1.28rem;
              font-weight: 800;
              letter-spacing: -0.04em;
              color: #1a56db;
            }
            .fdl-sb-wordmark-analytics {
              font-size: 1.05rem;
              font-weight: 600;
              letter-spacing: -0.02em;
              color: #334155;
            }
            .fdl-sb-tagline {
              font-size: 0.6875rem;
              font-weight: 600;
              letter-spacing: 0.045em;
              color: #5c6675;
              line-height: 1.5;
              margin: 0 0 0.65rem 0;
            }
            .fdl-sb-tagline--after-logo {
              margin-top: 0.1rem;
              margin-bottom: 0.5rem;
              padding: 0 0.45rem;
              text-align: center;
              font-size: 0.75rem;
              font-weight: 500;
              font-style: italic;
              letter-spacing: 0.02em;
              color: #64748b;
              line-height: 1.45;
            }
            .fdl-sb-client-row {
              margin: 0;
              padding-top: 0.65rem;
              border-top: 1px solid rgba(226, 232, 240, 0.85);
            }
            .fdl-sb-client-block {
              display: flex;
              flex-direction: row;
              align-items: center;
              gap: 0.45rem;
              padding: 0.45rem 0.6rem;
              margin-top: 0.12rem;
              margin-bottom: 0.05rem;
              background: #f1f5f9;
              border-radius: 8px;
              border: 1px solid rgba(226, 232, 240, 0.9);
            }
            .fdl-sb-brand-shell .fdl-sb-client-block {
              text-align: left;
            }
            .fdl-sb-client-icon {
              font-size: 1.2rem;
              line-height: 1;
              flex-shrink: 0;
            }
            .fdl-sb-client-text {
              display: flex;
              flex-direction: column;
              gap: 0.12rem;
              min-width: 0;
            }
            .fdl-sb-client-tag {
              font-size: 0.5625rem;
              font-weight: 600;
              text-transform: uppercase;
              letter-spacing: 0.12em;
              color: #94a3b8;
              line-height: 1.2;
            }
            .fdl-sb-client-name {
              font-size: 0.85rem;
              font-weight: 600;
              letter-spacing: -0.015em;
              color: #1e293b;
              line-height: 1.35;
              word-break: break-word;
            }
            .fdl-sb-divider {
              height: 1px;
              background: linear-gradient(90deg, transparent, #e5e7eb 14%, #e5e7eb 86%, transparent);
              margin: 0.62rem 0 0.12rem 0;
              opacity: 0.62;
            }
            .fdl-sb-section-label {
              font-size: 0.65rem;
              font-weight: 700;
              text-transform: uppercase;
              letter-spacing: 0.14em;
              color: #94a3b8;
              margin: 1.05rem 0 0.42rem 0;
              padding: 0 0.12rem 0.38rem 0.12rem;
              border-radius: 0;
              background: transparent;
              border: none;
              border-bottom: 1px solid rgba(226, 232, 240, 0.95);
              box-shadow: none;
              line-height: 1.35;
            }
            .fdl-sb-section-label--first {
              margin-top: 0.28rem;
            }
            .fdl-sb-org-hint {
              font-size: 0.53125rem;
              font-weight: 500;
              letter-spacing: 0.16em;
              text-transform: uppercase;
              color: #c5ced8;
              line-height: 1.35;
              margin: 0 0 0.32rem 0;
              padding: 0 0.2rem;
              text-align: center;
            }
            .fdl-sb-nav-item-hint {
              font-size: 0.72rem;
              font-weight: 450;
              color: #94a3b8;
              line-height: 1.35;
              margin: -0.15rem 0 0.55rem 0;
              padding: 0 0.28rem 0 0.35rem;
              letter-spacing: -0.01em;
            }
            [data-testid="stSidebar"] div[data-testid="stButton"] > button[kind="primary"] {
              font-weight: 600 !important;
              letter-spacing: -0.012em !important;
              color: #ffffff !important;
              background: linear-gradient(180deg, #2563eb 0%, #1a56db 100%) !important;
              border: 1px solid #1d4ed8 !important;
              border-radius: 10px !important;
              box-shadow:
                0 1px 0 rgba(255, 255, 255, 0.2) inset,
                0 2px 6px rgba(26, 86, 219, 0.25) !important;
              padding: 10px 14px !important;
              min-height: 2.5rem !important;
              transition: box-shadow 0.18s ease, border-color 0.18s ease, background 0.18s ease !important;
            }
            [data-testid="stSidebar"] div[data-testid="stButton"] > button[kind="primary"]:hover {
              border-color: #1e40af !important;
              background: linear-gradient(180deg, #3b82f6 0%, #1d4ed8 100%) !important;
              box-shadow:
                0 1px 0 rgba(255, 255, 255, 0.22) inset,
                0 3px 10px rgba(26, 86, 219, 0.32) !important;
            }
            [data-testid="stSidebar"] div[data-testid="stButton"] > button[kind="secondary"] {
              font-weight: 500 !important;
              letter-spacing: -0.008em !important;
              color: #64748b !important;
              border: 1px solid rgba(226, 232, 240, 0.55) !important;
              background: linear-gradient(180deg, rgba(255, 255, 255, 0.72) 0%, rgba(248, 250, 252, 0.55) 100%) !important;
              border-radius: 10px !important;
              padding: 10px 14px !important;
              min-height: 2.4rem !important;
              box-shadow: 0 1px 0 rgba(255, 255, 255, 0.55) inset !important;
              transition: background 0.15s ease, border-color 0.18s ease, color 0.18s ease, box-shadow 0.18s ease !important;
            }
            [data-testid="stSidebar"] div[data-testid="stButton"] > button[kind="secondary"]:hover {
              border-color: rgba(203, 213, 225, 0.85) !important;
              background-color: #f1f5f9 !important;
              background-image: none !important;
              color: #475569 !important;
              box-shadow:
                0 1px 0 rgba(255, 255, 255, 0.85) inset,
                0 2px 6px rgba(15, 23, 42, 0.04) !important;
            }
            /* Sair: único botão tertiary na sidebar (não conflita com nav primary/secondary). */
            [data-testid="stSidebar"] div[data-testid="stButton"] > button[kind="tertiary"] {
              width: 100% !important;
              font-weight: 600 !important;
              letter-spacing: -0.01em !important;
              color: #b91c1c !important;
              border: 1px solid rgba(220, 38, 38, 0.42) !important;
              border-radius: 10px !important;
              background: linear-gradient(180deg, #ffffff 0%, #fff1f2 100%) !important;
              padding-top: 0.52rem !important;
              padding-bottom: 0.52rem !important;
              min-height: 2.52rem !important;
              box-shadow: 0 1px 0 rgba(255, 255, 255, 0.55) inset !important;
              transition: background 0.18s ease, border-color 0.18s ease, color 0.18s ease, box-shadow 0.18s ease !important;
            }
            [data-testid="stSidebar"] div[data-testid="stButton"] > button[kind="tertiary"]:hover {
              border-color: rgba(185, 28, 28, 0.72) !important;
              background: linear-gradient(180deg, #fef2f2 0%, #fee2e2 100%) !important;
              color: #991b1b !important;
              box-shadow:
                0 1px 0 rgba(255, 255, 255, 0.75) inset,
                0 2px 6px rgba(185, 28, 28, 0.12) !important;
            }
            .fdl-sb-footer-rule {
              height: 1px;
              margin: 0.78rem 0 0;
              background: linear-gradient(90deg, transparent, rgba(229, 231, 235, 0.45) 10%, rgba(229, 231, 235, 0.72) 50%, rgba(229, 231, 235, 0.45) 90%, transparent);
              opacity: 0.58;
            }
            .fdl-sb-footer {
              margin: 0.48rem 0 0 0;
              padding: 0 0.2rem 0.12rem;
            }
            .fdl-sb-footer-dados {
              display: flex;
              align-items: flex-start;
              gap: 0.45rem;
              margin: 0.15rem 0 0.35rem 0;
              padding: 0.35rem 0.15rem 0;
            }
            .fdl-sb-footer-dados-icon {
              font-size: 0.95rem;
              line-height: 1.2;
              flex-shrink: 0;
              opacity: 0.85;
            }
            .fdl-sb-footer-dados-text {
              display: flex;
              flex-direction: column;
              gap: 0.08rem;
              min-width: 0;
            }
            .fdl-sb-footer-label {
              font-size: 0.5625rem;
              font-weight: 600;
              text-transform: uppercase;
              letter-spacing: 0.1em;
              color: #94a3b8;
              margin: 0;
              line-height: 1.25;
            }
            .fdl-sb-footer-ts {
              font-size: 0.8125rem;
              font-weight: 600;
              font-variant-numeric: tabular-nums;
              font-family: ui-monospace, "Cascadia Code", "Segoe UI Mono", monospace;
              color: #475569;
              margin: 0;
              line-height: 1.35;
              letter-spacing: 0.02em;
            }
            .fdl-sb-footer-dados-line {
              font-size: 0.8125rem;
              font-weight: 500;
              color: #64748b;
              margin: 0;
              line-height: 1.45;
              letter-spacing: 0.01em;
            }
            .fdl-sb-footer-dados-line .fdl-sb-footer-ts {
              font-size: inherit;
              font-weight: 600;
              color: #475569;
            }
            .fdl-sb-footer-admin {
              font-size: 0.65rem;
              font-weight: 400;
              color: #8b95a3;
              margin: 0.1rem 0 0 0;
              line-height: 1.4;
              word-break: break-word;
            }
            .fdl-sb-footer-spacer {
              height: 0.35rem;
            }
            </style>
            """
        ),
        unsafe_allow_html=True,
    )


def _now_ts_br_str() -> str:
    """Carimbo para UI em Brasília (Streamlit Cloud costuma usar UTC — o dia «salta» para o utilizador BR)."""
    return datetime.now(_BR_TZ).strftime("%Y-%m-%d %H:%M:%S")


def _ts_br_from_mtime_ns(mtime_ns: int) -> str:
    dt = datetime.fromtimestamp(mtime_ns / 1e9, tz=timezone.utc).astimezone(_BR_TZ)
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def _ts_br_from_http_last_modified(header: str | None) -> str | None:
    if not header or not str(header).strip():
        return None
    try:
        dt = parsedate_to_datetime(str(header).strip())
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(_BR_TZ).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return None


def _is_admin_mode() -> bool:
    env_mode = os.environ.get("FDL_APP_MODE", "").strip().lower()
    if env_mode == "admin":
        return True
    try:
        return str(st.secrets.get("FDL_APP_MODE", "")).strip().lower() == "admin"
    except Exception:
        return False


def _fdl_rg_pace_debug_enabled() -> bool:
    """Termômetro de pace: captions de diagnóstico (admin ou ``FDL_RG_PACE_DEBUG``)."""
    return _is_admin_mode() or (
        os.environ.get("FDL_RG_PACE_DEBUG", "").strip().lower() in {"1", "true", "yes", "on"}
    )


def _expose_load_errors() -> bool:
    """
    Quando True, falhas em _load_data() mostram st.exception (mensagem + traceback).
    Para desligar em produção perante cliente final: FDL_SHOW_LOAD_ERRORS=false nos secrets.
    """
    env = os.environ.get("FDL_SHOW_LOAD_ERRORS", "").strip().lower()
    if env in {"0", "false", "no", "off"}:
        return False
    if env in {"1", "true", "yes", "on"}:
        return True
    try:
        raw = st.secrets.get("FDL_SHOW_LOAD_ERRORS", None)
        if isinstance(raw, bool):
            return raw
        v = str(raw or "").strip().lower()
        if v in {"0", "false", "no", "off"}:
            return False
        if v in {"1", "true", "yes", "on"}:
            return True
    except Exception:
        pass
    return True


def _data_source_mode() -> str:
    env_source = os.environ.get("FDL_DATA_SOURCE", "").strip().lower()
    if env_source:
        return env_source
    try:
        return str(st.secrets.get("FDL_DATA_SOURCE", "onedrive")).strip().lower()
    except Exception:
        return "onedrive"


_STRICT_MATERIALIZED_USER_MSG = (
    "Base de dados não disponível. Contacte o administrador ou tente mais tarde."
)


def _strict_materialized() -> bool:
    """Produção: sem fallback live quando repasse/frete em modo materialized."""
    raw = os.environ.get("FDL_STRICT_MATERIALIZED", "").strip().lower()
    if raw in {"1", "true", "yes", "on"}:
        return True
    if raw in {"0", "false", "no", "off"}:
        return False
    try:
        sec = st.secrets.get("FDL_STRICT_MATERIALIZED", False)
        if isinstance(sec, bool):
            return sec
        return str(sec).strip().lower() in {"1", "true", "yes", "on"}
    except Exception:
        return False


def _frete_debug_ui_enabled() -> bool:
    """Diagnóstico na página Frete — opt-in (env ou secrets). Omisso = desligado (não afeta repasse)."""
    raw = os.environ.get("FDL_DEBUG_FRETE_UI", "").strip().lower()
    if raw in {"1", "true", "yes", "on"}:
        return True
    if raw in {"0", "false", "no", "off"}:
        return False
    try:
        sec = st.secrets.get("FDL_DEBUG_FRETE_UI", False)
        if isinstance(sec, bool):
            return sec
        return str(sec).strip().lower() in {"1", "true", "yes", "on"}
    except Exception:
        return False


def _frete_stage_error(etapa: int, titulo: str, exc: BaseException) -> None:
    """Erro visível com número da etapa (1–4) para diagnóstico na Cloud."""
    try:
        st.error(f"**Conciliação de Frete — falha na etapa {etapa}/4: {titulo}**")
        st.exception(exc)
    except Exception:
        pass


def _frete_stage_trace(etapa: int, titulo: str, detalhe: str) -> None:
    try:
        if _frete_debug_ui_enabled():
            st.caption(f"Frete [etapa {etapa}/4 — {titulo}] {detalhe}")
    except Exception:
        pass


def _bootstrap_debug_enabled() -> bool:
    """Rerun global: etapas visíveis (opt-in). Env/secrets: FDL_DEBUG_BOOTSTRAP=1."""
    raw = os.environ.get("FDL_DEBUG_BOOTSTRAP", "").strip().lower()
    if raw in {"1", "true", "yes", "on"}:
        return True
    if raw in {"0", "false", "no", "off"}:
        return False
    try:
        sec = st.secrets.get("FDL_DEBUG_BOOTSTRAP", False)
        if isinstance(sec, bool):
            return sec
        return str(sec).strip().lower() in {"1", "true", "yes", "on"}
    except Exception:
        return False


def _inject_fdl_professional_theme() -> None:
    """CSS global: reduz ruído de plataforma (menu, footer, toolbar/fork) — não é lógica de negócio."""
    if st.session_state.get("_fdl_ui_theme_applied") is True:
        st.session_state.pop("_fdl_ui_theme_applied", None)
    if st.session_state.get("_fdl_ui_theme_applied") == "v3":
        return
    st.markdown(
        """
        <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            [data-testid="stToolbar"] {visibility: hidden !important; height: 0 !important; max-height: 0 !important;}
            [data-testid="stDecoration"] {display: none;}
            header[data-testid="stHeader"] {background: rgba(255,255,255,0);}
            header a[href*="fork"] {display: none !important;}
            div[data-testid="stToolbar"] {visibility: hidden !important;}
            section[data-testid="stMain"] {
                max-width: 100%;
            }
            section[data-testid="stMain"] h2 {
                margin-top: 0.45rem;
                margin-bottom: 0.35rem;
                color: #111827;
                font-weight: 600;
                letter-spacing: -0.015em;
                line-height: 1.25;
            }
            section[data-testid="stMain"] h3 {
                margin-top: 0.5rem;
                margin-bottom: 0.35rem;
                color: #111827;
                font-weight: 600;
            }
            section[data-testid="stMain"] hr {
                margin: 0.5rem 0 !important;
                border: none;
                border-top: 1px solid #d1d9e4;
            }
            section[data-testid="stMain"] [data-testid="stCaption"] {
                color: #64748b !important;
                font-size: 0.8125rem !important;
                font-weight: 500 !important;
                line-height: 1.45 !important;
            }
            section[data-testid="stMain"] label[data-testid="stWidgetLabel"] p,
            section[data-testid="stMain"] [data-testid="stWidgetLabel"] p {
                color: #374151 !important;
                font-weight: 600 !important;
                font-size: 0.84rem !important;
            }
            .fdl-ui-gap-section {
                display: block;
                height: 0.55rem;
                min-height: 0.55rem;
            }
            .fdl-ui-gap-section-lg {
                display: block;
                height: 0.75rem;
                min-height: 0.75rem;
            }
            .fdl-ui-gap-tight {
                display: block;
                height: 0.28rem;
                min-height: 0.28rem;
            }
            .fdl-financeiro-header {
                margin: 0 0 0.5rem 0;
                padding: 0 0 0.15rem 0;
            }
            .fdl-financeiro-header .fdl-header-kicker {
                margin: 0 0 0.22rem 0;
                font-size: 0.75rem;
                font-weight: 700;
                letter-spacing: 0.055em;
                text-transform: uppercase;
                color: #475569;
            }
            .fdl-financeiro-header .fdl-header-title {
                margin: 0 0 0.32rem 0;
                font-size: 1.65rem;
                font-weight: 700;
                letter-spacing: -0.025em;
                line-height: 1.18;
                color: #0f172a;
            }
            .fdl-financeiro-header .fdl-header-sub {
                margin: 0;
                font-size: 0.9rem;
                font-weight: 500;
                color: #64748b;
                line-height: 1.45;
                max-width: 44rem;
            }
            .fdl-financeiro-header--compact .fdl-header-title {
                margin-bottom: 0.18rem;
                font-size: 1.52rem;
                letter-spacing: -0.028em;
            }
            .fdl-financeiro-header--compact .fdl-header-sub {
                font-size: 0.8125rem;
                font-weight: 500;
                color: #6b7785;
                line-height: 1.42;
            }
            /* Cartão de métrica (Streamlit 1.35+: stMetricContainer envolve label + valor) */
            [data-testid="stMetricContainer"] {
                border-radius: 0.5rem !important;
                padding: 0.75rem 0.95rem 0.8rem 0.95rem !important;
                min-height: 4.85rem;
                background: linear-gradient(180deg, #ffffff 0%, #f1f5f9 100%) !important;
                border: 1px solid #cbd5e1 !important;
                box-shadow: 0 1px 3px rgba(15, 23, 42, 0.08) !important;
                box-sizing: border-box !important;
            }
            /* Valor: Streamlit 1.35+ usa Markdown dentro de stMetricValue */
            [data-testid="stMetricValue"] [data-testid="stMarkdownContainer"] p {
                font-weight: 700 !important;
                font-size: 1.62rem !important;
                letter-spacing: -0.04em !important;
                line-height: 1.12 !important;
                color: #0f172a !important;
                margin: 0 !important;
            }
            [data-testid="stMetricValue"] > div {
                font-weight: 700 !important;
                font-size: 1.62rem !important;
                letter-spacing: -0.04em;
                line-height: 1.12 !important;
                color: #0f172a !important;
            }
            [data-testid="stMetricLabel"] {
                opacity: 1 !important;
                margin-bottom: 0.35rem !important;
            }
            [data-testid="stMetricLabel"] label,
            [data-testid="stMetricLabel"] p {
                font-size: 0.78rem !important;
                font-weight: 500 !important;
                color: #64748b !important;
                line-height: 1.3 !important;
                letter-spacing: 0.02em;
            }
            div[data-testid="stVerticalBlockBorderWrapper"] {
                border-radius: 0.55rem;
                border-color: #cbd5e1 !important;
                box-shadow: 0 1px 3px rgba(15, 23, 42, 0.06);
            }
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.session_state["_fdl_ui_theme_applied"] = "v3"


def _fdl_ui_gap_section() -> None:
    """Espaço vertical consistente entre blocos (apenas UI)."""
    st.markdown('<div class="fdl-ui-gap-section" aria-hidden="true"></div>', unsafe_allow_html=True)


def _fdl_ui_gap_section_lg() -> None:
    st.markdown('<div class="fdl-ui-gap-section-lg" aria-hidden="true"></div>', unsafe_allow_html=True)


def _fdl_ui_gap_tight() -> None:
    """Espaço vertical reduzido (MVP Faturamento & DRE — compactação)."""
    st.markdown('<div class="fdl-ui-gap-tight" aria-hidden="true"></div>', unsafe_allow_html=True)


def _render_financeiro_header(
    *,
    segment: str,
    title: str,
    subtitle: str = "",
    kicker_area: str = "Financeiro",
    compact_spacing: bool = False,
) -> None:
    """Topo unificado: evita repetir o nome do cliente (já na barra lateral)."""
    esc_seg = html.escape(segment)
    esc_title = html.escape(title)
    esc_ka = html.escape((kicker_area or "Financeiro").strip() or "Financeiro")
    esc_sub = html.escape((subtitle or "").strip())
    sub_html = ""
    if esc_sub:
        sub_html = f'<p class="fdl-header-sub">{esc_sub}</p>'
    _cls = "fdl-financeiro-header" + (" fdl-financeiro-header--compact" if compact_spacing else "")
    st.markdown(
        f'<div class="{_cls}">'
        f'<p class="fdl-header-kicker">{esc_ka} · {esc_seg}</p>'
        f'<h1 class="fdl-header-title">{esc_title}</h1>'
        f"{sub_html}"
        f"</div>",
        unsafe_allow_html=True,
    )
    if compact_spacing:
        _fdl_ui_gap_tight()
    else:
        _fdl_ui_gap_section()
        st.divider()


def _fdl_safe_mode() -> bool:
    """UI mínima: sem Styler, sem data_editor pesado, menos HTML (FDL_SAFE_MODE=1)."""
    raw = os.environ.get("FDL_SAFE_MODE", "").strip().lower()
    if raw in {"1", "true", "yes", "on"}:
        return True
    if raw in {"0", "false", "no", "off"}:
        return False
    try:
        sec = st.secrets.get("FDL_SAFE_MODE", False)
        if isinstance(sec, bool):
            return sec
        return str(sec).strip().lower() in {"1", "true", "yes", "on"}
    except Exception:
        return False


def _fdl_minimal_layout() -> bool:
    """
    Layout nativo Streamlit: sem CSS global injetado, sem HTML customizado nos painéis.
    Omisso ou FDL_MINIMAL_LAYOUT=1 → ativo. Defina FDL_MINIMAL_LAYOUT=0 para restaurar o design.
    """
    raw = os.environ.get("FDL_MINIMAL_LAYOUT", "").strip().lower()
    if raw in {"0", "false", "no", "off"}:
        return False
    if raw in {"1", "true", "yes", "on"}:
        return True
    try:
        sec = st.secrets.get("FDL_MINIMAL_LAYOUT", True)
        if isinstance(sec, bool):
            return sec
        return str(sec).strip().lower() not in {"0", "false", "no", "off"}
    except Exception:
        return True


def _repasse_consume_mode() -> str:
    """Repasse: live = pipeline; materialized = CSV/XLSX. Com FDL_STRICT_MATERIALIZED, sem fallback para live."""
    raw = os.environ.get("FDL_REPASSE_CONSUME_MODE", "").strip().lower()
    if raw in {"materialized", "live"}:
        return raw
    try:
        s = str(st.secrets.get("FDL_REPASSE_CONSUME_MODE", "")).strip().lower()
        if s in {"materialized", "live"}:
            return s
    except Exception:
        pass
    return "live"


def _repasse_use_parquet() -> bool:
    """PR4: FDL_REPASSE_USE_PARQUET=1 → repasse materializado lê dataset.parquet (path dinâmico ou path fixo .parquet)."""
    try:
        sec = st.secrets.get("FDL_REPASSE_USE_PARQUET", "")
    except Exception:
        sec = ""
    return repasse_use_parquet_flag(os.environ, secret_raw=sec)


def _repasse_materialized_path_str() -> str:
    if _materialized_path_mode() == "dynamic":
        return _dynamic_materialized_repasse_rel_path(_active_org.org_id)
    raw = os.environ.get("FDL_REPASSE_MATERIALIZED_PATH", "").strip()
    if raw:
        return raw
    try:
        return str(st.secrets.get("FDL_REPASSE_MATERIALIZED_PATH", "")).strip()
    except Exception:
        return ""


def _repasse_materialized_url_str() -> str:
    if _materialized_path_mode() == "dynamic":
        return ""
    raw = os.environ.get("FDL_REPASSE_MATERIALIZED_URL", "").strip()
    if raw:
        return raw
    try:
        return str(st.secrets.get("FDL_REPASSE_MATERIALIZED_URL", "")).strip()
    except Exception:
        return ""


def _repasse_resolved_materialized_file() -> Path | None:
    """Path absoluto do CSV/XLSX de repasse materializado, se existir (mesma resolução que ``_load_data``)."""
    path_s = (_repasse_materialized_path_str() or "").strip()
    if not path_s:
        return None
    path = Path(path_s).expanduser()
    if not path.is_absolute():
        path = (_REPO_APP_ROOT / path).resolve()
    else:
        path = path.resolve()
    return path if path.is_file() else None


def _repasse_materialized_source_stat_token() -> str:
    """
    mtime e tamanho do ficheiro consolidado na chave de cache do repasse.
    Após rematerializar com o mesmo path, o Streamlit deixa de servir dados antigos (antes só TTL 900s).
    """
    try:
        p = _repasse_resolved_materialized_file()
        if p is None:
            return "repasse_no_file"
        stt = p.stat()
        return f"repasse_file|mtime_ns={stt.st_mtime_ns}|size={stt.st_size}"
    except OSError:
        return "repasse_stat_err"


def _frete_consume_mode() -> str:
    raw = os.environ.get("FDL_FRETE_CONSUME_MODE", "").strip().lower()
    if raw in {"materialized", "live"}:
        return raw
    try:
        s = str(st.secrets.get("FDL_FRETE_CONSUME_MODE", "")).strip().lower()
        if s in {"materialized", "live"}:
            return s
    except Exception:
        pass
    return "live"


def _frete_materialized_path_str() -> str:
    if _materialized_path_mode() == "dynamic":
        return _dynamic_materialized_frete_rel_path(_active_org.org_id)
    raw = os.environ.get("FDL_FRETE_MATERIALIZED_PATH", "").strip()
    if raw:
        return raw
    try:
        return str(st.secrets.get("FDL_FRETE_MATERIALIZED_PATH", "")).strip()
    except Exception:
        return ""


def _frete_materialized_url_str() -> str:
    if _materialized_path_mode() == "dynamic":
        return ""
    raw = os.environ.get("FDL_FRETE_MATERIALIZED_URL", "").strip()
    if raw:
        return raw
    try:
        return str(st.secrets.get("FDL_FRETE_MATERIALIZED_URL", "")).strip()
    except Exception:
        return ""


def _devolucoes_consume_mode() -> str:
    """Devoluções: nesta fase só há consumo por artefato materializado (sem pipeline live no app)."""
    raw = os.environ.get("FDL_DEVOLUCOES_CONSUME_MODE", "").strip().lower()
    if raw in {"materialized", "live"}:
        return raw
    try:
        s = str(st.secrets.get("FDL_DEVOLUCOES_CONSUME_MODE", "")).strip().lower()
        if s in {"materialized", "live"}:
            return s
    except Exception:
        pass
    return "materialized"


def _devolucoes_materialized_path_str() -> str:
    if _materialized_path_mode() == "dynamic":
        return _dynamic_materialized_devolucoes_rel_path(_active_org.org_id)
    raw = os.environ.get("FDL_DEVOLUCOES_MATERIALIZED_PATH", "").strip()
    if raw:
        return raw
    try:
        return str(st.secrets.get("FDL_DEVOLUCOES_MATERIALIZED_PATH", "")).strip()
    except Exception:
        return ""


def _devolucoes_resolved_load_path() -> Path | None:
    """Ficheiro efetivo: ``dataset.parquet`` na mesma pasta do CSV dinâmico, senão o próprio CSV."""
    path_s = (_devolucoes_materialized_path_str() or "").strip()
    if not path_s:
        return None
    p_csv = Path(path_s).expanduser()
    if not p_csv.is_absolute():
        p_csv = (_REPO_APP_ROOT / p_csv).resolve()
    else:
        p_csv = p_csv.resolve()
    pq = p_csv.parent / "dataset.parquet"
    if pq.is_file():
        return pq
    if p_csv.is_file():
        return p_csv
    return None


def _devolucoes_materialized_session_signature() -> str:
    path_s = (_devolucoes_materialized_path_str() or "").strip()
    if not path_s:
        return "devolucoes|empty_path"
    p_eff = _devolucoes_resolved_load_path()
    if p_eff is None:
        return f"devolucoes|missing|{path_s.strip()[:180]}"
    try:
        stt = p_eff.stat()
        return f"devolucoes|{p_eff.resolve()}|{stt.st_mtime_ns}|{stt.st_size}"
    except OSError:
        return "devolucoes|stat_err"


def _load_devolucoes_materialized_dataframe() -> pd.DataFrame:
    p_eff = _devolucoes_resolved_load_path()
    if p_eff is None:
        raise FileNotFoundError(
            "Artefato de devoluções não encontrado. Materialize com "
            "`python processing/materialize_financeiro.py --modulo devolucoes` ou confira o path."
        )
    if p_eff.suffix.lower() == ".parquet":
        return pd.read_parquet(p_eff)
    return pd.read_csv(p_eff, sep=None, engine="python", encoding="utf-8-sig")


def _load_devolucoes_data(org_id: str) -> tuple[pd.DataFrame, dict[str, object], str]:
    """Lê só o materializado (Parquet preferencial)."""
    if _strict_materialized() and _devolucoes_consume_mode() == "materialized":
        path_hint = _devolucoes_materialized_path_str().strip()
        if not path_hint:
            raise ValueError(
                "Devoluções em modo materialized: defina o path dinâmico (slug + org) ou "
                "FDL_DEVOLUCOES_MATERIALIZED_PATH. " + _STRICT_MATERIALIZED_USER_MSG
            )
        p_eff = _devolucoes_resolved_load_path()
        if p_eff is None:
            raise ValueError(
                "Devoluções: ficheiro consolidado não encontrado no disco. " + _STRICT_MATERIALIZED_USER_MSG
            )

    _ss_key = f"_devolucoes_cache_{org_id}"
    _sig = _devolucoes_materialized_session_signature()
    _cached = st.session_state.get(_ss_key)
    if isinstance(_cached, dict) and _cached.get("source_signature") == _sig:
        df_c = _cached.get("df_devolucoes")
        if isinstance(df_c, pd.DataFrame):
            ts = str(_cached.get("ts_proc", _now_ts_br_str()))
            info = _cached.get("meta_devolucoes")
            if isinstance(info, dict):
                return df_c, info, ts

    df = _load_devolucoes_materialized_dataframe()
    p_eff = _devolucoes_resolved_load_path()
    ts_br = _now_ts_br_str()
    if p_eff is not None and p_eff.is_file():
        ts_br = _frete_ts_for_path(p_eff)
    path_disp = _devolucoes_materialized_path_str().strip()
    info: dict[str, object] = {
        "origem": "devolucoes_materializado",
        "devolucoes_consume": "materialized",
        "devolucoes_materialized_target": path_disp[:500] if path_disp else "",
        "devolucoes_path_resolved": str(p_eff.resolve()) if p_eff else "",
        "linhas": int(len(df)),
    }
    st.session_state[_ss_key] = {
        "df_devolucoes": df,
        "meta_devolucoes": info,
        "source_signature": _sig,
        "ts_proc": ts_br,
    }
    return df, info, ts_br


def _faturamento_consume_mode() -> str:
    """Faturamento: nesta fase só materialized é suportado; default materialized."""
    raw = os.environ.get("FDL_FATURAMENTO_CONSUME_MODE", "").strip().lower()
    if raw in {"materialized", "live"}:
        return raw
    try:
        s = str(st.secrets.get("FDL_FATURAMENTO_CONSUME_MODE", "")).strip().lower()
        if s in {"materialized", "live"}:
            return s
    except Exception:
        pass
    return "materialized"


def _faturamento_data_layout() -> str:
    """
    Como interpretar o ficheiro materializado com path/URL explícitos.

    - v1: layout legado (faturamento por pasta de empresa; sem filtro ``org_id``).
    - v2: dataset multi-empresa; filtrar pela org ativa quando existir coluna ``org_id``.
    - auto: inferir (fallback): ``v2`` se existir coluna ``org_id`` com dados, senão ``v1``.

    Descoberta **sem** path explícito: ``v2_canonical`` fixa layout efetivo em v2;
    ``v1_repasse_sibling`` fixa em v1.
    """
    raw = os.environ.get("FDL_FATURAMENTO_DATA_LAYOUT", "").strip().lower()
    if raw in {"v1", "v2", "auto"}:
        return raw
    try:
        s = str(st.secrets.get("FDL_FATURAMENTO_DATA_LAYOUT", "")).strip().lower()
        if s in {"v1", "v2", "auto"}:
            return s
    except Exception:
        pass
    return "auto"


def _faturamento_resolve_disk_path(path_s: str) -> Path:
    p = Path(path_s.strip()).expanduser()
    if not p.is_absolute():
        p = (_REPO_APP_ROOT / p).resolve()
    else:
        p = p.resolve()
    return p


def _faturamento_materialized_cliente_slug_override() -> str:
    """
    Slug só para o path V2 canónico de faturamento (opcional).

    Útil quando ``FDL_MATERIALIZED_CLIENTE_SLUG`` serve repasse/frete multiempresa (ex.: cliente_2)
    mas o faturamento materializado vive em outro segmento (ex.: ``default`` para Antomóveis).
    """
    raw = os.environ.get("FDL_MATERIALIZED_CLIENTE_SLUG_FATURAMENTO", "").strip()
    if raw:
        return raw
    try:
        return str(st.secrets.get("FDL_MATERIALIZED_CLIENTE_SLUG_FATURAMENTO", "")).strip()
    except Exception:
        return ""


def _faturamento_v2_canonical_dataset_path_str() -> str:
    """
    ``data_products/<cliente_slug>/faturamento/current/``.

    Prioridade: ``dataset.parquet`` (artefato principal da materialização, colunas completas),
    depois ``dataset_faturamento_app.csv`` (espelho para export). Se o CSV ficar velho ao lado
    de um Parquet novo, ler o CSV primeiro duplicava comissão/frete após correções no pipeline.

    Ordem de pastas: override ``FDL_MATERIALIZED_CLIENTE_SLUG_FATURAMENTO`` → slug geral
    ``FDL_MATERIALIZED_CLIENTE_SLUG`` →, se o utilizador tem **só** Antomóveis, tentativa extra
    em ``default/`` (materialização típica desse tenant no repositório).
    """
    root = _materialized_data_products_root().strip().strip("/\\")

    def _path_for_slug(slug: str) -> str:
        if not slug:
            return ""
        base = Path(root) / slug / "faturamento" / "current"
        if not base.is_absolute():
            base = (_REPO_APP_ROOT / base).resolve()
        else:
            base = base.resolve()
        for name in ("dataset.parquet", "dataset_faturamento_app.csv"):
            cand = base / name
            if cand.is_file():
                return str(cand.resolve())
        return ""

    slugs: list[str] = []
    ov = _faturamento_materialized_cliente_slug_override().strip()
    if ov:
        slugs.append(ov)
    main = _materialized_cliente_slug().strip()
    if main and main not in slugs:
        slugs.append(main)
    if (
        len(_app_ctx.organizations) == 1
        and _app_ctx.organizations[0].org_id == "antomoveis"
        and "default" not in slugs
    ):
        slugs.append("default")

    for s in slugs:
        got = _path_for_slug(s)
        if got:
            return got
    return ""


def _faturamento_resolve_materialized_target() -> dict[str, str]:
    """
    Ordem: explícito (path/URL) → **V2 canônico** (se existir em disco) → derivado V1 do repasse.

    Sem path explícito, o ficheiro em ``data_products/<slug>/faturamento/current/`` tem prioridade
    sobre o CSV «irmão do repasse» (V1 por empresa), para carregar join fiscal (``Nota_Data_Emissao``, etc.).

    Devolve chaves: path_s, url_s, resolution_source, path_final_resolved, layout_declared.
    """
    declared = _faturamento_data_layout()
    mp = _faturamento_materialized_path_str()
    mu = _faturamento_materialized_url_str()
    if mp or mu:
        final = ""
        if mp:
            try:
                p = _faturamento_resolve_disk_path(mp)
                final = str(p) if p.is_file() else mp.strip()
            except Exception:
                final = mp.strip()
        if not final and mu:
            final = mu.strip()
        return {
            "path_s": mp.strip(),
            "url_s": mu.strip(),
            "resolution_source": "explicit",
            "path_final_resolved": final or mp.strip() or mu.strip(),
            "layout_declared": declared,
        }
    v2s = _faturamento_v2_canonical_dataset_path_str()
    if v2s:
        return {
            "path_s": v2s,
            "url_s": "",
            "resolution_source": "v2_canonical",
            "path_final_resolved": v2s,
            "layout_declared": declared,
        }
    for anchor in (_repasse_materialized_path_str(), _precomputed_path_str()):
        d = _derive_faturamento_materialized_from_repasse_anchor(anchor)
        if d:
            return {
                "path_s": d,
                "url_s": "",
                "resolution_source": "v1_repasse_sibling",
                "path_final_resolved": d,
                "layout_declared": declared,
            }
    return {
        "path_s": "",
        "url_s": "",
        "resolution_source": "none",
        "path_final_resolved": "",
        "layout_declared": declared,
    }


def _faturamento_materialized_path_str() -> str:
    if _materialized_path_mode() == "dynamic":
        return ""
    raw = os.environ.get("FDL_FATURAMENTO_MATERIALIZED_PATH", "").strip()
    if raw:
        return raw
    try:
        return str(st.secrets.get("FDL_FATURAMENTO_MATERIALIZED_PATH", "")).strip()
    except Exception:
        return ""


def _faturamento_materialized_url_str() -> str:
    if _materialized_path_mode() == "dynamic":
        return ""
    raw = os.environ.get("FDL_FATURAMENTO_MATERIALIZED_URL", "").strip()
    if raw:
        return raw
    try:
        return str(st.secrets.get("FDL_FATURAMENTO_MATERIALIZED_URL", "")).strip()
    except Exception:
        return ""


def _derive_faturamento_materialized_from_repasse_anchor(anchor: str) -> str:
    """
    Se o repasse/precomputed apontar para .../<cliente>/<empresa>/repasse/current/dataset_repasse_app.csv,
    tenta o irmão .../<empresa>/faturamento/current/dataset_faturamento_app.csv e, se faltar, dataset.parquet.

    Isto alinha ao layout **V1** (faturamento por pasta de empresa). Faturamento **V2** grava em
    ``data_products/<cliente_slug>/faturamento/current/``; esse caminho exige ``FDL_FATURAMENTO_MATERIALIZED_PATH``
    ou evolução desta derivação — ver ``docs/faturamento_pipeline.md``.
    """
    if not (anchor or "").strip():
        return ""
    path = Path(anchor.strip()).expanduser()
    if not path.is_absolute():
        path = (_REPO_APP_ROOT / path).resolve()
    if not path.is_file():
        return ""
    if path.parent.name != "current" or path.parent.parent.name != "repasse":
        return ""
    # materialize_financeiro: .../<cliente>/<empresa>/repasse/current/ e .../<empresa>/faturamento/current/ (irmãos)
    empresa_dir = path.parent.parent.parent
    base = empresa_dir / "faturamento" / "current"
    pq = base / "dataset.parquet"
    if pq.is_file():
        return str(pq.resolve())
    csv_c = base / "dataset_faturamento_app.csv"
    if csv_c.is_file():
        return str(csv_c.resolve())
    return ""


def _faturamento_classify_layout_effective(
    *,
    resolution_source: str,
    layout_declared: str,
    df: pd.DataFrame,
) -> tuple[str, str]:
    """
    Devolve (layout_efetivo v1|v2, nota_curta_para_metadata).
    """
    if resolution_source == "v2_canonical":
        return "v2", "v2_canonical"
    if resolution_source == "v1_repasse_sibling":
        return "v1", "v1_repasse_sibling"
    if resolution_source == "explicit":
        if layout_declared == "v1":
            return "v1", "explicit_declared_v1"
        if layout_declared == "v2":
            return "v2", "explicit_declared_v2"
        if "org_id" in df.columns and df["org_id"].notna().any():
            return "v2", "explicit_auto_org_id"
        return "v1", "explicit_auto_no_org_id"
    return "v1", "fallback_v1"


def _faturamento_apply_layout_scope(
    df: pd.DataFrame, *, layout_effective: str, org_id: str
) -> tuple[pd.DataFrame, str | None]:
    """
    Para layout v2, restringe à org ativa. Devolve (df, nota_warning ou None).
    """
    if layout_effective != "v2":
        return df, None
    oid = str(org_id).strip()
    if "org_id" not in df.columns:
        return df, "Layout v2: coluna org_id ausente — filtro por org não aplicado."
    out = df.loc[df["org_id"].astype(str).str.strip() == oid].copy()
    return out, None


def _faturamento_apply_layout_scope_consolidado_v2(
    df: pd.DataFrame, *, allowed_org_ids: frozenset[str]
) -> tuple[pd.DataFrame, str | None]:
    """V2: todas as orgs permitidas ao utilizador (não a org ativa da sidebar)."""
    if "org_id" not in df.columns:
        return df.copy(), "Layout v2: coluna org_id ausente — consolidado sem filtro por org."
    if not allowed_org_ids:
        return df.iloc[0:0].copy(), None
    oid_s = df["org_id"].astype(str).str.strip()
    out = df.loc[oid_s.isin(allowed_org_ids)].copy()
    return out, None


def _load_faturamento_file_from_disk(path: Path) -> pd.DataFrame:
    if not path.is_file():
        raise FileNotFoundError(f"Ficheiro de faturamento não encontrado: {path}")
    suf = path.suffix.lower()
    if suf == ".csv":
        return pd.read_csv(path, sep=None, engine="python", encoding="utf-8-sig")
    if suf == ".parquet":
        return pd.read_parquet(path, engine="pyarrow")
    if suf in {".xlsx", ".xls"}:
        return pd.read_excel(path, engine="openpyxl")
    raise ValueError(f"Formato não suportado para faturamento materializado: {path.name!r}")


def _load_faturamento_materialized_dataframe(path_s: str, url_s: str) -> pd.DataFrame:
    if path_s:
        path = Path(path_s).expanduser()
        if not path.is_absolute():
            path = (_REPO_APP_ROOT / path).resolve()
        return _load_faturamento_file_from_disk(path)
    errs: list[str] = []
    for dl_url, hdr in _precomputed_download_attempts(url_s.strip()):
        try:
            payload, filename, _lm = _download_file_bytes(
                dl_url,
                extra_headers=hdr or None,
                timeout=PRECOMPUTED_HTTP_TIMEOUT,
                http_retries=1,
            )
            return _dataframe_from_frete_materialized_bytes(payload, filename)
        except Exception as exc:  # noqa: BLE001
            hint = dl_url if len(dl_url) < 100 else dl_url[:97] + "..."
            errs.append(f"{hint} → {exc}")
    raise ValueError(
        "Não foi possível ler o faturamento materializado a partir do URL. "
        + " | ".join(errs[:5])
        + (" …" if len(errs) > 5 else "")
    )


def _faturamento_nf_parquet_path_from_materialized_path(path_s: str) -> Path | None:
    """
    ``dataset_faturamento_nf.parquet`` no mesmo diretório que o materializado linha ou dentro da pasta apontada.

    Fallback V2: ``data_products/<slug>/faturamento/current/`` quando o path explícito aponta para uma cópia
    sem os Parquets (alinhado a ``_faturamento_fiscal_parquet_resolve``).
    """
    if not (path_s or "").strip():
        return None
    resolved: Path | None = None
    try:
        p = Path(path_s).expanduser()
        if not p.is_absolute():
            p = (_REPO_APP_ROOT / p).resolve()
        if p.is_dir():
            cand = p / "dataset_faturamento_nf.parquet"
            resolved = cand if cand.is_file() else None
        elif p.is_file():
            cand = p.parent / "dataset_faturamento_nf.parquet"
            resolved = cand if cand.is_file() else None
    except OSError:
        resolved = None
    if resolved is not None:
        return resolved
    v2s = _faturamento_v2_canonical_dataset_path_str()
    if v2s:
        try:
            v2file = _faturamento_resolve_disk_path(v2s)
            if v2file.is_file():
                cand = v2file.parent / "dataset_faturamento_nf.parquet"
                if cand.is_file():
                    return cand.resolve()
        except OSError:
            pass
    return None


def _faturamento_nf_panel_parquet_path_from_materialized_path(path_s: str) -> Path | None:
    """``dataset_faturamento_nf_panel.parquet`` (merge fiscal + frete/resultado pré-calculados) ao lado do NF-first."""
    nf = _faturamento_nf_parquet_path_from_materialized_path(path_s)
    if nf is None:
        return None
    from processing.faturamento.nf_panel_materializado import NF_PANEL_PARQUET_FILENAME

    cand = nf.parent / NF_PANEL_PARQUET_FILENAME
    return cand if cand.is_file() else None


def _faturamento_nf_panel_ads_flag_from_disk(panel_parquet: Path) -> bool:
    """
    Lê ``nf_panel_ads`` de ``metadata.json`` junto ao Parquet do painel NF.

    ``True`` = materialização aplicou custo ADS (3,5% + fixo) e descontou do ``resultado``.
    Omisso ou metadata em falta ⇒ ``True`` (retrocompatível com artefactos antigos).
    """
    meta_p = panel_parquet.parent / "metadata.json"
    if not meta_p.is_file():
        return True
    try:
        raw = json.loads(meta_p.read_text(encoding="utf-8"))
        if isinstance(raw, dict) and "nf_panel_ads" in raw:
            return bool(raw["nf_panel_ads"])
    except Exception:
        return True
    return True


def _faturamento_fiscal_parquet_path_from_materialized_path(path_s: str) -> Path | None:
    """``dataset_faturamento_fiscal.parquet`` no mesmo diretório que o materializado linha ou dentro da pasta apontada."""
    if not (path_s or "").strip():
        return None
    try:
        p = Path(path_s).expanduser()
        if not p.is_absolute():
            p = (_REPO_APP_ROOT / p).resolve()
        if p.is_dir():
            cand = p / "dataset_faturamento_fiscal.parquet"
            return cand if cand.is_file() else None
        if not p.is_file():
            return None
        cand = p.parent / "dataset_faturamento_fiscal.parquet"
        return cand if cand.is_file() else None
    except OSError:
        return None


def _faturamento_fiscal_parquet_resolve(path_s: str) -> tuple[Path | None, str]:
    """
    Resolve ``dataset_faturamento_fiscal.parquet`` para o carregamento do painel.

    Ordem: pasta/ficheiro ao lado do path do materializado linha → pasta do Parquet NF-first
    → ``data_products/<slug>/faturamento/current/`` (V2 canónico em disco), quando existir.

    O fallback V2 cobre casos em que o path explícito aponta para uma cópia sem os Parquets,
    mas o pipeline gravou fiscal na pasta canónica.
    """
    if not (path_s or "").strip():
        return None, ""
    fp = _faturamento_fiscal_parquet_path_from_materialized_path(path_s)
    if fp is not None:
        return fp, "sibling_materializado"
    nf_side = _faturamento_nf_parquet_path_from_materialized_path(path_s)
    if nf_side is not None:
        alt = nf_side.parent / "dataset_faturamento_fiscal.parquet"
        if alt.is_file():
            return alt, "nf_parquet_parent"
    v2s = _faturamento_v2_canonical_dataset_path_str()
    if v2s:
        try:
            v2file = _faturamento_resolve_disk_path(v2s)
            if v2file.is_file():
                cand = v2file.parent / "dataset_faturamento_fiscal.parquet"
                if cand.is_file():
                    return cand.resolve(), "v2_canonical_folder"
        except OSError:
            pass
    return None, ""


def _faturamento_devolucoes_parquet_resolve(path_s: str) -> tuple[Path | None, str]:
    """Resolve ``dataset_faturamento_devolucoes.parquet`` junto ao fiscal materializado."""
    fp_f, how = _faturamento_fiscal_parquet_resolve(path_s)
    if fp_f is None:
        return None, ""
    cand = fp_f.parent / "dataset_faturamento_devolucoes.parquet"
    return (cand if cand.is_file() else None), how


def _faturamento_nf_parquet_stat_token(path_s: str) -> str:
    nf = _faturamento_nf_parquet_path_from_materialized_path(path_s)
    if nf is None:
        return "nf_absent"
    try:
        stt = nf.stat()
        return f"{nf.resolve()}|mtime_ns={stt.st_mtime_ns}|size={stt.st_size}"
    except OSError:
        return "nf_unstat"


def _faturamento_nf_panel_parquet_stat_token(path_s: str) -> str:
    pan = _faturamento_nf_panel_parquet_path_from_materialized_path(path_s)
    if pan is None:
        return "nf_panel_absent"
    try:
        stt = pan.stat()
        return f"{pan.resolve()}|mtime_ns={stt.st_mtime_ns}|size={stt.st_size}"
    except OSError:
        return "nf_panel_unstat"


def _faturamento_ts_for_path(path: Path) -> str:
    try:
        return _ts_br_from_mtime_ns(int(path.stat().st_mtime_ns))
    except OSError:
        return _now_ts_br_str()


def _load_faturamento_data(
    active_org_id: str,
    *,
    scope_consolidado: bool = False,
    allowed_org_ids: frozenset[str] | None = None,
) -> tuple[pd.DataFrame, dict[str, object], str]:
    """
    Carrega dataset materializado (path/URL explícitos → V2 canônico em disco → V1 irmão do repasse),
    classifica layout (v1/v2) e aplica escopo por ``org_id`` (empresa ativa) ou,
    em modo consolidado, todas as orgs permitidas ao utilizador quando o layout é v2.
    """
    if _faturamento_consume_mode() != "materialized":
        return (
            pd.DataFrame(),
            {
                "faturamento_consume": "unsupported",
                "faturamento_note": "Nesta fase só está disponível FDL_FATURAMENTO_CONSUME_MODE=materialized.",
            },
            _now_ts_br_str(),
        )

    resolved = _faturamento_resolve_materialized_target()
    path_s = resolved["path_s"]
    url_s = resolved["url_s"]
    resolution_source = resolved["resolution_source"]
    path_final = resolved["path_final_resolved"]
    layout_declared = resolved["layout_declared"]

    if not path_s and not url_s:
        slug_h = _materialized_cliente_slug().strip() or "(defina FDL_MATERIALIZED_CLIENTE_SLUG)"
        return (
            pd.DataFrame(),
            {
                "faturamento_consume": "missing_config",
                "faturamento_resolution_source": resolution_source,
                "faturamento_layout_declared": layout_declared,
                "faturamento_path_final_resolved": path_final,
                "faturamento_note": (
                    "Faturamento materializado não encontrado. Opções: "
                    "(1) FDL_FATURAMENTO_MATERIALIZED_PATH ou URL; "
                    "(2) V2: ficheiro em data_products/"
                    f"{slug_h}/faturamento/current/ "
                    "(csv ou parquet), com FDL_FATURAMENTO_DATA_LAYOUT=v2 se usar path explícito; "
                    "(3) V1: repasse materializado irmão em .../<org>/faturamento/current/. "
                    "Com path/URL explícitos use FDL_FATURAMENTO_DATA_LAYOUT=v1 ou v2 em produção."
                ),
            },
            _now_ts_br_str(),
        )

    target = (path_s or url_s)[:500]
    try:
        df0 = _load_faturamento_materialized_dataframe(path_s, url_s)
        n_loaded = int(len(df0))
        layout_effective, layout_note = _faturamento_classify_layout_effective(
            resolution_source=resolution_source,
            layout_declared=layout_declared,
            df=df0,
        )
        aids = allowed_org_ids or frozenset()
        if scope_consolidado and layout_effective == "v2":
            df_scoped, scope_warn = _faturamento_apply_layout_scope_consolidado_v2(
                df0, allowed_org_ids=aids
            )
        else:
            df_scoped, scope_warn = _faturamento_apply_layout_scope(
                df0, layout_effective=layout_effective, org_id=active_org_id
            )
        ts = _now_ts_br_str()
        if path_s:
            try:
                p = _faturamento_resolve_disk_path(path_s)
                if p.is_file():
                    ts = _faturamento_ts_for_path(p)
                    path_final = str(p.resolve())
            except Exception:
                pass
        info: dict[str, object] = {
            "faturamento_consume": "materialized",
            "faturamento_materialized_target": target,
            "faturamento_resolution_source": resolution_source,
            "faturamento_layout_declared": layout_declared,
            "faturamento_data_layout": layout_effective,
            "faturamento_layout_classification": layout_note,
            "faturamento_path_final_resolved": path_final or (path_s or url_s).strip(),
            "cliente_slug": str(_materialized_cliente_slug()).strip(),
            "faturamento_row_count_loaded": n_loaded,
            "linhas": int(len(df_scoped)),
            "faturamento_nf_first": False,
            "faturamento_nf_df": None,
            "faturamento_nf_panel_baked": False,
            "faturamento_nf_panel_df": None,
            "faturamento_fiscal_first": False,
            "faturamento_fiscal_df": None,
            "faturamento_devolucoes_df": None,
        }
        if path_s:
            nf_p = _faturamento_nf_parquet_path_from_materialized_path(path_s)
            if nf_p is not None:
                try:
                    df_nf0 = pd.read_parquet(nf_p, engine="pyarrow")
                    if nf_first_contract_dataframe_valid(df_nf0):
                        if scope_consolidado and layout_effective == "v2":
                            df_nf_scoped, _nfw = _faturamento_apply_layout_scope_consolidado_v2(
                                df_nf0, allowed_org_ids=aids
                            )
                        else:
                            df_nf_scoped, _nfw = _faturamento_apply_layout_scope(
                                df_nf0, layout_effective=layout_effective, org_id=active_org_id
                            )
                        info["faturamento_nf_df"] = df_nf_scoped
                        info["faturamento_nf_first"] = True
                        info["faturamento_nf_first_path"] = str(nf_p.resolve())
                        info["faturamento_nf_first_row_count_loaded"] = int(len(df_nf0))
                    else:
                        info["faturamento_nf_first_skip"] = "contract_columns_incompletos"
                except Exception as ex_nf:
                    info["faturamento_nf_first_error"] = str(ex_nf).strip() or ex_nf.__class__.__name__
            else:
                info["faturamento_nf_first_skip"] = "ficheiro_ausente"
            panel_p = _faturamento_nf_panel_parquet_path_from_materialized_path(path_s)
            if panel_p is not None:
                try:
                    from processing.faturamento.nf_panel_materializado import (
                        nf_panel_materializado_dataframe_valid,
                    )

                    df_p0 = pd.read_parquet(panel_p, engine="pyarrow")
                    if nf_panel_materializado_dataframe_valid(df_p0):
                        if scope_consolidado and layout_effective == "v2":
                            df_p_scoped, _pw = _faturamento_apply_layout_scope_consolidado_v2(
                                df_p0, allowed_org_ids=aids
                            )
                        else:
                            df_p_scoped, _pw = _faturamento_apply_layout_scope(
                                df_p0, layout_effective=layout_effective, org_id=active_org_id
                            )
                        info["faturamento_nf_panel_df"] = df_p_scoped
                        info["faturamento_nf_panel_baked"] = True
                        info["faturamento_nf_panel_path"] = str(panel_p.resolve())
                        info["faturamento_nf_panel_row_count_loaded"] = int(len(df_p0))
                        info["faturamento_nf_panel_ads"] = _faturamento_nf_panel_ads_flag_from_disk(
                            panel_p.resolve()
                        )
                    elif df_p0.empty:
                        info["faturamento_nf_panel_error"] = (
                            f"O ficheiro **{panel_p.name}** existe em `{panel_p.parent}` mas está **vazio** (0 linhas). "
                            "O agregado NF-first não produziu notas — confirme pedidos com NF ligada, notas na pasta "
                            "configurada e volte a **materializar faturamento**."
                        )
                    else:
                        info["faturamento_nf_panel_error"] = (
                            "O painel NF em disco tem colunas em falta para o contrato atual — "
                            f"rematerialize o faturamento. Ficheiro: "
                            f"`{html.escape(str(panel_p.name))}`."
                        )
                except Exception as ex_p:  # noqa: BLE001
                    info["faturamento_nf_panel_error"] = str(ex_p).strip() or ex_p.__class__.__name__
        else:
            info["faturamento_nf_first_skip"] = "sem_path_local"
        if path_s:
            fp_p, fp_how = _faturamento_fiscal_parquet_resolve(path_s)
            if fp_p is not None:
                info["faturamento_fiscal_path_resolution"] = fp_how
                try:
                    df_fiscal0 = pd.read_parquet(fp_p, engine="pyarrow")
                    if fiscal_contract_dataframe_valid(df_fiscal0):
                        if scope_consolidado and layout_effective == "v2":
                            df_fiscal_scoped, _fw = _faturamento_apply_layout_scope_consolidado_v2(
                                df_fiscal0, allowed_org_ids=aids
                            )
                        else:
                            df_fiscal_scoped, _fw = _faturamento_apply_layout_scope(
                                df_fiscal0, layout_effective=layout_effective, org_id=active_org_id
                            )
                        info["faturamento_fiscal_df"] = df_fiscal_scoped
                        info["faturamento_fiscal_first"] = True
                        info["faturamento_fiscal_first_path"] = str(fp_p.resolve())
                        info["faturamento_fiscal_first_row_count_loaded"] = int(len(df_fiscal0))
                        if df_fiscal_scoped.empty and len(df_fiscal0) > 0:
                            info["faturamento_fiscal_user_hint"] = (
                                "O ficheiro fiscal tem linhas, mas **nenhuma** ficou no escopo da org ativa "
                                "(confira ``org_id`` no Parquet vs organização da sidebar)."
                            )
                    else:
                        info["faturamento_fiscal_first_skip"] = "contract_columns_incompletos_ou_vazio"
                except Exception as ex_f:  # noqa: BLE001
                    info["faturamento_fiscal_first_error"] = str(ex_f).strip() or ex_f.__class__.__name__
            else:
                info["faturamento_fiscal_first_skip"] = "ficheiro_ausente"
        else:
            info["faturamento_fiscal_first_skip"] = "sem_path_local"
        if path_s:
            dv_p, dv_how = _faturamento_devolucoes_parquet_resolve(path_s)
            if dv_p is not None:
                info["faturamento_devolucoes_path_resolution"] = dv_how
                try:
                    from processing.faturamento.fiscal_devolucoes_materializado import (
                        devolucoes_contract_dataframe_valid,
                    )

                    df_dv0 = pd.read_parquet(dv_p, engine="pyarrow")
                    if devolucoes_contract_dataframe_valid(df_dv0):
                        if scope_consolidado and layout_effective == "v2":
                            df_dv_scoped, _dw = _faturamento_apply_layout_scope_consolidado_v2(
                                df_dv0, allowed_org_ids=aids
                            )
                        else:
                            df_dv_scoped, _dw = _faturamento_apply_layout_scope(
                                df_dv0, layout_effective=layout_effective, org_id=active_org_id
                            )
                        info["faturamento_devolucoes_df"] = df_dv_scoped
                        info["faturamento_devolucoes_first"] = True
                        info["faturamento_devolucoes_first_path"] = str(dv_p.resolve())
                        info["faturamento_devolucoes_first_row_count_loaded"] = int(len(df_dv0))
                    else:
                        info["faturamento_devolucoes_first_skip"] = "contract_columns_incompletos_ou_vazio"
                except Exception as ex_dv:  # noqa: BLE001
                    info["faturamento_devolucoes_first_error"] = (
                        str(ex_dv).strip() or ex_dv.__class__.__name__
                    )
        info.update(_faturamento_materialized_fiscal_audit(df_scoped))
        if resolution_source == "explicit" and not bool(info.get("faturamento_fiscal_join_complete")):
            v2_alt = _faturamento_v2_canonical_dataset_path_str()
            if v2_alt:
                info["faturamento_note_v2_canonical_available"] = (
                    "O path/URL explícito atual não inclui colunas completas do join fiscal "
                    "(ex.: Nota_Data_Emissao, Nota_Valor_Liquido_Rateado). "
                    f"Existe materializado em **{v2_alt}** — remova o path explícito ou aponte para esse ficheiro."
                )
        if scope_warn:
            info["faturamento_scope_note"] = scope_warn
        info["faturamento_escopo"] = (
            FAT_DRE_ESCOPO_CONSOLIDADO
            if scope_consolidado
            else FAT_DRE_ESCOPO_EMPRESA
        )
        if "Status_Custo" in df_scoped.columns:
            _vc = (
                df_scoped["Status_Custo"]
                .astype(str)
                .str.strip()
                .value_counts()
                .to_dict()
            )
            info["faturamento_status_custo_counts"] = _vc
        return (df_scoped, info, ts)
    except Exception as exc:
        return (
            pd.DataFrame(),
            {
                "faturamento_consume": "error",
                "faturamento_materialized_error": str(exc).strip() or exc.__class__.__name__,
                "faturamento_materialized_target": target,
                "faturamento_resolution_source": resolution_source,
                "faturamento_layout_declared": layout_declared,
                "faturamento_path_final_resolved": path_final,
            },
            _now_ts_br_str(),
        )


def _faturamento_materialized_source_stat_token() -> str:
    """
    Inclui mtime/tamanho do ficheiro resolvido para invalidar o cache após rematerialização
    (o path costuma ser o mesmo; sem isto o ``@st.cache_data`` pode servir dados antigos até ao TTL).
    """
    resolved = _faturamento_resolve_materialized_target()
    path_s = (resolved.get("path_s") or "").strip()
    if not path_s:
        return "no_path"
    try:
        p = Path(path_s).expanduser()
        if not p.is_absolute():
            p = (_REPO_APP_ROOT / p).resolve()
        if p.is_file():
            st = p.stat()
            return f"{p.resolve()}|mtime_ns={st.st_mtime_ns}|size={st.st_size}"
    except OSError:
        pass
    return f"unstat:{path_s[:240]}"


def _faturamento_load_cache_signature(
    org_id: str, *, consolidado: bool, allowed_org_ids_key: str
) -> str:
    return "|".join(
        [
            str(org_id),
            "1" if consolidado else "0",
            str(allowed_org_ids_key),
            str(OPERACIONAL_CACHE_REVISION),
            _faturamento_consume_mode(),
            _faturamento_data_layout(),
            str(_materialized_cliente_slug()).strip(),
            str(_faturamento_materialized_path_str()).strip(),
            str(_faturamento_materialized_url_str()).strip(),
            str(_repasse_materialized_path_str()).strip(),
            str(_precomputed_path_str()).strip(),
            str(_strict_materialized()),
            _faturamento_materialized_source_stat_token(),
            _faturamento_nf_parquet_stat_token(
                str(_faturamento_resolve_materialized_target().get("path_s") or "").strip()
            ),
            _faturamento_nf_panel_parquet_stat_token(
                str(_faturamento_resolve_materialized_target().get("path_s") or "").strip()
            ),
        ]
    )


@st.cache_data(show_spinner=False, ttl=900)
def _load_faturamento_dataframe_cached(
    load_signature: str,
    active_org_id: str,
    consolidado: bool,
    allowed_org_ids_key: str,
) -> tuple[pd.DataFrame, dict[str, object], str]:
    _ = load_signature
    aids = frozenset(
        x.strip() for x in str(allowed_org_ids_key).split(",") if x.strip()
    )
    return _load_faturamento_data(
        active_org_id, scope_consolidado=consolidado, allowed_org_ids=aids
    )


def _derive_frete_materialized_path_from_repasse_anchor(anchor: str) -> str:
    """
    Pipeline `materialize_financeiro`: repasse em .../repasse/current/ e frete em .../frete/current/dataset_frete_app.csv.
    `anchor` pode ser FDL_REPASSE_MATERIALIZED_PATH ou FDL_PRECOMPUTED_PATH nesse layout (muitos clientes só definem precomputed).
    """
    if not (anchor or "").strip():
        return ""
    path = Path(anchor.strip()).expanduser()
    if not path.is_absolute():
        path = (_REPO_APP_ROOT / path).resolve()
    if not path.is_file():
        return ""
    if path.parent.name != "current" or path.parent.parent.name != "repasse":
        return ""
    candidate = path.parent.parent / "frete" / "current" / "dataset_frete_app.csv"
    if candidate.is_file():
        return str(candidate.resolve())
    return ""


def _derive_frete_materialized_path_from_repasse() -> str:
    for anchor in (_repasse_materialized_path_str(), _precomputed_path_str()):
        d = _derive_frete_materialized_path_from_repasse_anchor(anchor)
        if d:
            return d
    return ""


def _frete_materialized_targets() -> tuple[str, str]:
    """
    PATH/URL explícitos do frete (com FDL_FRETE_CONSUME_MODE=materialized);
    se vazios, tenta dataset_frete_app.csv ao lado do CSV do repasse em .../repasse/current/
    (via FDL_REPASSE_MATERIALIZED_PATH ou FDL_PRECOMPUTED_PATH nesse layout).

    Sem FDL_FRETE_MATERIALIZED_*: usa o derivado se repasse materializado, ou se FDL_DATA_SOURCE=precomputed|ready|table
    (o repasse costuma ler a tabela final por FDL_PRECOMPUTED_PATH — o frete alinha ao mesmo padrão).
    """
    mp = _frete_materialized_path_str()
    mu = _frete_materialized_url_str()
    derived = _derive_frete_materialized_path_from_repasse()

    if mp or mu:
        if _frete_consume_mode() == "materialized":
            return mp, mu
        return "", ""

    if not derived:
        return "", ""

    if _frete_consume_mode() == "materialized":
        return derived, ""

    if _repasse_consume_mode() == "materialized":
        return derived, ""

    if _data_source_mode() in {"precomputed", "ready", "table"}:
        return derived, ""

    return "", ""


def _validate_frete_materialized_schema(df: pd.DataFrame) -> None:
    validate_frete_operacional_dataframe(df)


def _dataframe_from_frete_materialized_bytes(payload: bytes, filename: str) -> pd.DataFrame:
    head = payload.lstrip()[:800]
    if head.startswith(b"<") or head.upper().startswith(b"<!DOCTYPE"):
        raise ValueError(
            "Resposta HTML em vez do CSV do frete materializado. Confirme URL/partilha do ficheiro."
        )
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        return pd.read_csv(BytesIO(payload), sep=None, engine="python", encoding="utf-8-sig")
    if lower.endswith(".xlsx") or lower.endswith(".xls") or zipfile.is_zipfile(BytesIO(payload)):
        return pd.read_excel(BytesIO(payload), engine="openpyxl")
    try:
        return pd.read_csv(BytesIO(payload), sep=None, engine="python", encoding="utf-8-sig")
    except Exception as exc:
        raise ValueError(f"Formato não suportado para frete materializado: {filename!r}") from exc


def _frete_materialized_session_signature(path_s: str, url_s: str) -> str:
    if path_s:
        p = Path(path_s).expanduser()
        if not p.is_absolute():
            p = (_REPO_APP_ROOT / p).resolve()
        if p.is_file():
            return f"mat|p|{p.resolve()}|{int(p.stat().st_mtime_ns)}"
        return f"mat|p|missing|{path_s.strip()[:180]}"
    if url_s:
        return f"mat|u|{url_s.strip()[:240]}"
    return "mat|empty"


def _load_frete_materialized_dataframe(path_s: str, url_s: str) -> pd.DataFrame:
    if path_s:
        path = Path(path_s).expanduser()
        if not path.is_absolute():
            path = (_REPO_APP_ROOT / path).resolve()
        if not path.is_file():
            raise FileNotFoundError(f"FDL_FRETE_MATERIALIZED_PATH não encontrado: {path}")
        if path.suffix.lower() not in {".csv", ".xlsx", ".xls"}:
            raise ValueError("FDL_FRETE_MATERIALIZED_PATH deve ser .csv, .xlsx ou .xls")
        if path.suffix.lower() == ".csv":
            return pd.read_csv(path, sep=None, engine="python", encoding="utf-8-sig")
        return pd.read_excel(path, engine="openpyxl")
    errs: list[str] = []
    for dl_url, hdr in _precomputed_download_attempts(url_s.strip()):
        try:
            payload, filename, _lm = _download_file_bytes(
                dl_url,
                extra_headers=hdr or None,
                timeout=PRECOMPUTED_HTTP_TIMEOUT,
                http_retries=1,
            )
            return _dataframe_from_frete_materialized_bytes(payload, filename)
        except Exception as exc:  # noqa: BLE001
            hint = dl_url if len(dl_url) < 100 else dl_url[:97] + "..."
            errs.append(f"{hint} → {exc}")
    raise ValueError(
        "Não foi possível ler o frete materializado a partir do URL. "
        + " | ".join(errs[:5])
        + (" …" if len(errs) > 5 else "")
    )


def _frete_session_cache_is_valid(payload: object) -> bool:
    if not isinstance(payload, dict):
        return False
    if "df_frete" not in payload or "meta_frete" not in payload or "source_signature" not in payload:
        return False
    df_obj = payload.get("df_frete")
    if not isinstance(df_obj, pd.DataFrame):
        return False
    meta_obj = payload.get("meta_frete")
    if not isinstance(meta_obj, dict):
        return False
    return True


def _frete_loader_source_signature(org_id: str, fontes: FontesFrete) -> str:
    if (fontes.vendas_url or "").strip():
        u = (fontes.vendas_url or "").strip()
        vendas_origin = u
        vendas_sig = str(stable_mtime_ns_for_frete_url(fontes.vendas_url))
    else:
        paths = [p for p in fontes.vendas_paths if p.is_file()]
        if not paths and fontes.vendas_path and fontes.vendas_path.is_file():
            paths = [fontes.vendas_path]
        if paths:
            vendas_origin = ";".join(str(p.resolve()) for p in paths)
            pieces = "|".join(f"{p.resolve()}:{int(p.stat().st_mtime_ns)}" for p in paths)
            vendas_sig = hashlib.sha256(pieces.encode("utf-8")).hexdigest()[:24]
        else:
            vendas_origin = ""
            vendas_sig = "none"

    frete_origin = (fontes.frete_url or "").strip() or (
        str(fontes.frete_path.resolve()) if fontes.frete_path else ""
    )
    if (fontes.frete_url or "").strip():
        frete_sig = str(stable_mtime_ns_for_frete_url(fontes.frete_url))
    elif fontes.frete_path and fontes.frete_path.is_file():
        frete_sig = str(int(fontes.frete_path.stat().st_mtime_ns))
    else:
        frete_sig = "none"

    return "|".join(
        [
            f"org={org_id}",
            f"v_origin={vendas_origin}",
            f"v_sig={vendas_sig}",
            f"f_origin={frete_origin}",
            f"f_sig={frete_sig}",
        ]
    )


def _frete_merge_info_from_meta(meta: dict[str, object], **extra: object) -> dict[str, object]:
    out: dict[str, object] = {**meta, **extra}
    out.setdefault("origem", "frete_operacional")
    return out


def _frete_ts_for_path(path: Path) -> str:
    try:
        return _ts_br_from_mtime_ns(int(path.stat().st_mtime_ns))
    except OSError:
        return _now_ts_br_str()


def _frete_ts_live_from_fontes(fontes: FontesFrete) -> str:
    paths = [p for p in fontes.vendas_paths if p.is_file()]
    if not paths and fontes.vendas_path and fontes.vendas_path.is_file():
        paths = [fontes.vendas_path]
    if not paths:
        return _now_ts_br_str()
    try:
        m = max(int(p.stat().st_mtime_ns) for p in paths)
        return _ts_br_from_mtime_ns(m)
    except OSError:
        return _now_ts_br_str()


def _frete_local_vendas_caption(fontes: FontesFrete) -> str | None:
    if fontes.vendas_paths and len(fontes.vendas_paths) > 1:
        head = fontes.vendas_paths[:6]
        tail = "; ".join(str(p.resolve()) for p in head)
        return tail + (" …" if len(fontes.vendas_paths) > 6 else "")
    if fontes.vendas_path:
        return str(fontes.vendas_path.resolve())
    return None


def _load_frete_data_strict_materialized_only(
    org_id: str,
) -> tuple[pd.DataFrame, dict[str, object], str]:
    """
    FDL_STRICT_MATERIALIZED + FDL_FRETE_CONSUME_MODE=materialized: só lê artefato (dataset_frete_app.csv
    via path/URL ou derivado); nunca descobrir_fontes_frete nem carregar_tabela_final_frete_operacional.
    """
    _f_mp_exp = _frete_materialized_path_str()
    _f_mu_exp = _frete_materialized_url_str()
    _f_mp, _f_mu = _frete_materialized_targets()
    _mat_try = bool(_f_mp) or bool(_f_mu)
    _frete_mat_from_repasse_sibling = bool(_f_mp) and not _f_mp_exp and not _f_mu_exp
    _frete_ss_key = f"_frete_cache_{org_id}"

    if not _mat_try:
        raise ValueError(
            "Frete em modo materialized: defina FDL_FRETE_MATERIALIZED_PATH ou FDL_FRETE_MATERIALIZED_URL, "
            "ou dataset_frete_app.csv em .../frete/current/ (derivado do repasse em .../repasse/current/). "
            + _STRICT_MATERIALIZED_USER_MSG
        )

    _sig_mat = _frete_materialized_session_signature(_f_mp, _f_mu)
    _cached_top = st.session_state.get(_frete_ss_key)
    if _cached_top is not None and not _frete_session_cache_is_valid(_cached_top):
        st.session_state.pop(_frete_ss_key, None)
        _cached_top = None

    df_mat: pd.DataFrame | None = None
    meta_mat: dict[str, object] = {}

    if _frete_session_cache_is_valid(_cached_top) and str(_cached_top.get("source_signature", "")) == _sig_mat:
        try:
            df_mat = _cached_top.get("df_frete", pd.DataFrame())
            meta_mat = _cached_top.get("meta_frete", {})
            if not isinstance(df_mat, pd.DataFrame):
                raise TypeError("df_frete inválido em cache")
            if not isinstance(meta_mat, dict):
                meta_mat = {}
        except Exception:
            st.session_state.pop(_frete_ss_key, None)
            df_mat = None

    if df_mat is None:
        try:
            df_mat = _load_frete_materialized_dataframe(_f_mp, _f_mu)
            validate_frete_operacional_dataframe(df_mat)
            df_mat = normalize_frete_status_conc_display(df_mat)
            meta_mat = {
                "vendas_arquivo": "dataset_frete_app.csv (materializado)",
                "frete_arquivo": None,
                "frete_tabular": FRETE_UI_STATUS_CONC in df_mat.columns,
                "debug_logs": [],
                "avisos": [],
                "linhas": int(len(df_mat)),
            }
            _loaded_at = _now_ts_br_str()
            st.session_state[_frete_ss_key] = {
                "df_frete": df_mat,
                "meta_frete": meta_mat,
                "debug_logs": [],
                "loaded_at": _loaded_at,
                "source_signature": _sig_mat,
            }
        except Exception as exc:
            raise ValueError(
                f"{_STRICT_MATERIALIZED_USER_MSG} Não foi possível concluir o carregamento dos dados de frete."
            ) from exc

    df_mat = normalize_frete_status_conc_display(df_mat)
    ts_mat = _now_ts_br_str()
    if _f_mp:
        p = Path(_f_mp).expanduser()
        if not p.is_absolute():
            p = (_REPO_APP_ROOT / p).resolve()
        if p.is_file():
            ts_mat = _frete_ts_for_path(p)
    info_ok = _frete_merge_info_from_meta(
        meta_mat,
        frete_consume="materialized",
        frete_materialized_target=(_f_mp or _f_mu)[:500],
        linhas=len(df_mat),
        frete_mat_from_repasse_sibling=_frete_mat_from_repasse_sibling,
    )
    return df_mat, info_ok, ts_mat


def _load_frete_data(org_id: str) -> tuple[pd.DataFrame, dict[str, object], str]:
    """
    Ponto único de carregamento do Frete (materializado primeiro, live com fallback).
    Devolve DataFrame operacional, info (frete_consume, metadados de loader) e timestamp BR.
    """
    if _strict_materialized() and _frete_consume_mode() == "materialized":
        return _load_frete_data_strict_materialized_only(org_id)

    _f_mp_exp = _frete_materialized_path_str()
    _f_mu_exp = _frete_materialized_url_str()
    _f_mp, _f_mu = _frete_materialized_targets()
    _mat_try = bool(_f_mp) or bool(_f_mu)
    _frete_mat_from_repasse_sibling = bool(_f_mp) and not _f_mp_exp and not _f_mu_exp
    _frete_ss_key = f"_frete_cache_{org_id}"
    mat_load_error: Exception | None = None

    def _empty_info(**k: object) -> dict[str, object]:
        base: dict[str, object] = {"origem": "frete_operacional", "linhas": 0}
        base.update(k)
        return base

    if _mat_try:
        _sig_mat = _frete_materialized_session_signature(_f_mp, _f_mu)
        _cached_top = st.session_state.get(_frete_ss_key)
        if _cached_top is not None and not _frete_session_cache_is_valid(_cached_top):
            st.session_state.pop(_frete_ss_key, None)
            _cached_top = None

        df_mat: pd.DataFrame | None = None
        meta_mat: dict[str, object] = {}

        if _frete_session_cache_is_valid(_cached_top) and str(_cached_top.get("source_signature", "")) == _sig_mat:
            try:
                df_mat = _cached_top.get("df_frete", pd.DataFrame())
                meta_mat = _cached_top.get("meta_frete", {})
                if not isinstance(df_mat, pd.DataFrame):
                    raise TypeError("df_frete inválido em cache")
                if not isinstance(meta_mat, dict):
                    meta_mat = {}
            except Exception:
                st.session_state.pop(_frete_ss_key, None)
                df_mat = None

        if df_mat is None:
            try:
                df_mat = _load_frete_materialized_dataframe(_f_mp, _f_mu)
                validate_frete_operacional_dataframe(df_mat)
                df_mat = normalize_frete_status_conc_display(df_mat)
                meta_mat = {
                    "vendas_arquivo": "dataset_frete_app.csv (materializado)",
                    "frete_arquivo": None,
                    "frete_tabular": FRETE_UI_STATUS_CONC in df_mat.columns,
                    "debug_logs": [],
                    "avisos": [],
                    "linhas": int(len(df_mat)),
                }
                _loaded_at = _now_ts_br_str()
                st.session_state[_frete_ss_key] = {
                    "df_frete": df_mat,
                    "meta_frete": meta_mat,
                    "debug_logs": [],
                    "loaded_at": _loaded_at,
                    "source_signature": _sig_mat,
                }
            except Exception as exc:
                mat_load_error = exc
                df_mat = None

        if df_mat is not None:
            df_mat = normalize_frete_status_conc_display(df_mat)
            ts_mat = _now_ts_br_str()
            if _f_mp:
                p = Path(_f_mp).expanduser()
                if not p.is_absolute():
                    p = (_REPO_APP_ROOT / p).resolve()
                if p.is_file():
                    ts_mat = _frete_ts_for_path(p)
            info_ok = _frete_merge_info_from_meta(
                meta_mat,
                frete_consume="materialized",
                frete_materialized_target=(_f_mp or _f_mu)[:500],
                linhas=len(df_mat),
                frete_mat_from_repasse_sibling=_frete_mat_from_repasse_sibling,
            )
            return df_mat, info_ok, ts_mat

    try:
        fontes = descobrir_fontes_frete()
    except Exception as exc:
        return pd.DataFrame(), _empty_info(frete_consume="error", frete_fontes_error=str(exc)), _now_ts_br_str()

    vendas_ref, v_ns = frete_vendas_loader_args(fontes)

    if not _mat_try and not vendas_ref:
        return (
            pd.DataFrame(),
            _empty_info(frete_consume="empty", frete_no_vendas_source=True),
            _now_ts_br_str(),
        )

    _vu = (fontes.vendas_url or "").strip()
    if _vu and "..." in _vu:
        return (
            pd.DataFrame(),
            _empty_info(frete_consume="empty", frete_placeholder_vendas_url=True),
            _now_ts_br_str(),
        )

    # Live local: carregamento automático (sem botão), alinhado ao repasse — vendas ML + frete anúncio
    # descobertos por descobrir_fontes_frete() sob FDL_BASE_DIR.

    _sig_desired = _frete_loader_source_signature(org_id, fontes)
    _cached = st.session_state.get(_frete_ss_key)
    if _cached is not None and not _frete_session_cache_is_valid(_cached):
        st.session_state.pop(_frete_ss_key, None)
        _cached = None

    if _frete_session_cache_is_valid(_cached) and str(_cached.get("source_signature", "")) == _sig_desired:
        try:
            df_frete = _cached.get("df_frete", pd.DataFrame())
            meta_frete = _cached.get("meta_frete", {})
            if not isinstance(meta_frete, dict):
                meta_frete = {}
            ts_live = _frete_ts_live_from_fontes(fontes)
            consume = "live_fallback" if (_mat_try and mat_load_error is not None) else "live"
            fb_err = str(mat_load_error).strip() if mat_load_error else ""
            t_fallback = ((_f_mp or _f_mu)[:500] if _mat_try else "") or ""
            info_fb = _frete_merge_info_from_meta(
                meta_frete,
                frete_consume=consume,
                linhas=len(df_frete),
                frete_materialized_target=t_fallback,
                frete_materialized_error=fb_err,
            )
            if (fontes.vendas_url or "").strip():
                info_fb["frete_vendas_from_url"] = True
            else:
                cap = _frete_local_vendas_caption(fontes)
                if cap:
                    info_fb["frete_fonte_local_path"] = cap
            if (
                _is_admin_mode()
                and _frete_consume_mode() == "materialized"
                and not _f_mp_exp
                and not _f_mu_exp
                and not _derive_frete_materialized_path_from_repasse()
            ):
                info_fb["frete_mat_note"] = (
                    "Frete: modo **ficheiro consolidado** ativo, mas não há caminho dedicado nem ficheiro ao lado do repasse "
                    "(`.../repasse/current/` → `.../frete/current/`) — em uso **fonte em tempo real**."
                )
            return df_frete, info_fb, ts_live
        except Exception:
            st.session_state.pop(_frete_ss_key, None)

    if not vendas_ref:
        return (
            pd.DataFrame(),
            _empty_info(frete_consume="empty", frete_no_vendas_source=True),
            _now_ts_br_str(),
        )

    try:
        frete_ref = (fontes.frete_url or "").strip() or (
            str(fontes.frete_path.resolve())
            if fontes.frete_path and fontes.frete_path.is_file()
            else None
        )
        if (fontes.frete_url or "").strip():
            f_ns = stable_mtime_ns_for_frete_url(fontes.frete_url)
        elif fontes.frete_path and fontes.frete_path.is_file():
            f_ns = int(fontes.frete_path.stat().st_mtime_ns)
        else:
            f_ns = None
        df_frete, meta_frete = carregar_tabela_final_frete_operacional(
            org_id, vendas_ref, v_ns, frete_ref, f_ns
        )
    except ValueError as exc:
        return (
            pd.DataFrame(),
            _empty_info(
                frete_consume="error",
                frete_loader_error=str(exc),
                frete_ml_validation_failed=True,
            ),
            _now_ts_br_str(),
        )

    _sig_store = _frete_loader_source_signature(org_id, fontes)
    _loaded_at = _now_ts_br_str()
    st.session_state[_frete_ss_key] = {
        "df_frete": df_frete,
        "meta_frete": meta_frete,
        "debug_logs": list(meta_frete.get("debug_logs") or []),
        "loaded_at": _loaded_at,
        "source_signature": _sig_store,
    }
    ts_live = _frete_ts_live_from_fontes(fontes)
    consume = "live_fallback" if (_mat_try and mat_load_error is not None) else "live"
    fb_err = ""
    if mat_load_error:
        fb_err = str(mat_load_error).strip() or mat_load_error.__class__.__name__
    info_out = _frete_merge_info_from_meta(
        meta_frete,
        frete_consume=consume,
        linhas=len(df_frete),
        frete_materialized_target=((_f_mp or _f_mu)[:500] if _mat_try else ""),
        frete_materialized_error=fb_err,
    )
    if (fontes.vendas_url or "").strip():
        info_out["frete_vendas_from_url"] = True
    else:
        cap = _frete_local_vendas_caption(fontes)
        if cap:
            info_out["frete_fonte_local_path"] = cap
    if (
        _is_admin_mode()
        and _frete_consume_mode() == "materialized"
        and not _f_mp_exp
        and not _f_mu_exp
        and not _derive_frete_materialized_path_from_repasse()
    ):
        info_out["frete_mat_note"] = (
            "Frete: modo **ficheiro consolidado** ativo, mas não há caminho dedicado nem ficheiro ao lado do repasse "
            "(`.../repasse/current/` → `.../frete/current/`) — em uso **fonte em tempo real**."
        )
    return df_frete, info_out, ts_live


def _prepare_uploaded_base(zip_bytes: bytes) -> Path:
    """Extrai pacote ZIP de dados para o BASE_DIR esperado pelo pipeline."""
    base_dir = Path(BASE_DIR)
    expected_dirs = {
        "Vendas - Mercado Livre",
        "Vendas_ML",
        "Liberações_ML",
        "notas_saida",
        "contas_receber",
    }

    if base_dir.exists():
        shutil.rmtree(base_dir)
    base_dir.mkdir(parents=True, exist_ok=True)

    tmp_zip = base_dir / "_upload.zip"
    tmp_zip.write_bytes(zip_bytes)
    with zipfile.ZipFile(tmp_zip, "r") as zf:
        zf.extractall(base_dir)
    tmp_zip.unlink(missing_ok=True)

    # Aceita ZIP com uma pasta raiz extra (ex.: dataset/data/...).
    children = [p for p in base_dir.iterdir() if p.is_dir()]
    if len(children) == 1 and expected_dirs.intersection({p.name for p in children[0].iterdir() if p.is_dir()}):
        nested_root = children[0]
        for item in nested_root.iterdir():
            target = base_dir / item.name
            if target.exists():
                if target.is_dir():
                    shutil.rmtree(target)
                else:
                    target.unlink()
            shutil.move(str(item), str(target))
        nested_root.rmdir()

    return base_dir


def _onedrive_public_url() -> str:
    env_url = os.environ.get("FDL_ONEDRIVE_URL", "").strip()
    if env_url:
        return env_url
    try:
        return str(st.secrets.get("FDL_ONEDRIVE_URL", "")).strip()
    except Exception:
        return ""


def _onedrive_root_folder_name() -> str:
    env_name = os.environ.get("FDL_ONEDRIVE_ROOT_FOLDER", "").strip()
    if env_name:
        return env_name
    try:
        value = str(st.secrets.get("FDL_ONEDRIVE_ROOT_FOLDER", "")).strip()
        if value:
            return value
    except Exception:
        pass
    return "cursor"


def _onedrive_client_folder_name() -> str:
    env_name = os.environ.get("FDL_ONEDRIVE_CLIENT_FOLDER", "").strip()
    if env_name:
        return env_name
    try:
        value = str(st.secrets.get("FDL_ONEDRIVE_CLIENT_FOLDER", "")).strip()
        if value:
            return value
    except Exception:
        pass
    return "cliente_1"


def _microsoft_share_token_from_url(url: str) -> str:
    raw_b64 = base64.b64encode(url.strip().encode("utf-8")).decode("ascii")
    return raw_b64.rstrip("=").replace("+", "-").replace("/", "_")


def _is_m365_sharing_url(url: str) -> bool:
    p = urlparse(url.strip())
    host = p.netloc.lower()
    path_l = (p.path or "").lower()
    if "1drv.ms" in host:
        return True
    if "sharepoint.com" in host or "onedrive.live.com" in host:
        return any(
            marker in path_l
            for marker in (":x:/", ":f:/", ":w:/", ":b:/", ":u:/", ":v:/", ":g:/")
        )
    return False


def _build_onedrive_download_url(public_url: str) -> str:
    parsed = urlparse(public_url.strip())
    host = parsed.netloc.lower()
    path_l = (parsed.path or "").lower()

    def _use_onedrive_shares_api() -> bool:
        if "1drv.ms" in host:
            return True
        if "sharepoint.com" not in host and "onedrive.live.com" not in host:
            return False
        # Links partilhados (Excel :x:/, pasta :f:/, etc.) — ?download=1 costuma devolver HTML, não o binário.
        return any(
            marker in path_l
            for marker in (":x:/", ":f:/", ":w:/", ":b:/", ":u:/", ":v:/", ":g:/")
        )

    if _use_onedrive_shares_api():
        encoded = _microsoft_share_token_from_url(public_url.strip())
        return f"https://api.onedrive.com/v1.0/shares/u!{encoded}/root/content"

    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query["download"] = "1"
    return urlunparse(
        (
            parsed.scheme,
            parsed.netloc,
            parsed.path,
            parsed.params,
            urlencode(query),
            parsed.fragment,
        )
    )


def _encode_share_token(public_url: str) -> str:
    return base64.urlsafe_b64encode(public_url.encode("utf-8")).decode("utf-8").rstrip("=")


def _graph_bearer_token() -> str:
    env_token = os.environ.get("FDL_MS_GRAPH_TOKEN", "").strip()
    if env_token:
        return env_token
    try:
        return str(st.secrets.get("FDL_MS_GRAPH_TOKEN", "")).strip()
    except Exception:
        return ""


def _download_json(url: str, headers: dict[str, str] | None = None) -> dict[str, object]:
    req_headers = {"User-Agent": "FDL-Streamlit-App/1.0"}
    if headers:
        req_headers.update(headers)
    req = Request(url, headers=req_headers)
    try:
        raw = b""
        for attempt in range(MAX_HTTP_RETRIES + 1):
            try:
                with urlopen(req, timeout=60) as resp:
                    raw = resp.read()
                break
            except HTTPError as exc_retry:
                if exc_retry.code not in RETRYABLE_HTTP_CODES or attempt >= MAX_HTTP_RETRIES:
                    raise
                retry_after = str(exc_retry.headers.get("Retry-After", "")).strip()
                sleep_s = float(retry_after) if retry_after.isdigit() else (1.5**attempt)
                time.sleep(min(max(sleep_s, 0.5), 8.0))
    except HTTPError as exc:
        body = ""
        try:
            body = exc.read().decode("utf-8", errors="replace")
        except Exception:
            body = ""
        body = body.strip().replace("\n", " ")
        body_snippet = body[:320] + ("..." if len(body) > 320 else "")
        raise ValueError(
            f"HTTP {exc.code} ao acessar {url}. Resposta: {body_snippet or '(sem corpo)'}"
        ) from exc
    except URLError as exc:
        raise ValueError(f"Erro de rede ao acessar {url}: {exc.reason}") from exc
    except Exception as exc:
        raise ValueError(f"Falha ao acessar {url}: {exc}") from exc
    data = json.loads(raw.decode("utf-8", errors="replace"))
    if not isinstance(data, dict):
        raise ValueError("Resposta JSON inválida ao listar pasta compartilhada.")
    return data


def _fetch_shared_driveitem_metadata(public_url: str) -> tuple[str, dict[str, object]]:
    token = _encode_share_token(public_url)
    bearer = _graph_bearer_token()
    graph_headers = {"Authorization": f"Bearer {bearer}"} if bearer else None

    candidates = [
        (
            f"https://graph.microsoft.com/v1.0/shares/u!{token}/driveItem",
            graph_headers,
        ),
        (
            f"https://api.onedrive.com/v1.0/shares/u!{token}/root",
            None,
        ),
    ]
    errors: list[str] = []
    for url, headers in candidates:
        try:
            payload = _download_json(url, headers=headers)
            return url, payload
        except Exception as exc:  # pragma: no cover - fallback de conectividade/autorizacao
            errors.append(f"- {url}: {exc}")
    raise ValueError(
        "Não foi possível acessar metadados da pasta compartilhada. "
        "Verifique se o link permite leitura pública ou configure FDL_MS_GRAPH_TOKEN. "
        "Tentativas: "
        + " | ".join(errors)
    )


def _download_shared_folder_dataset(public_url: str) -> None:
    base_dir = Path(BASE_DIR)
    mirror_root = base_dir / "_onedrive_shared"
    target_root_name = _onedrive_root_folder_name()
    target_client_name = _onedrive_client_folder_name()
    mirror_root.mkdir(parents=True, exist_ok=True)

    # Evita sincronização remota completa a cada rerun do Streamlit (ex.: troca de filtro).
    sync_meta_file = base_dir / ".onedrive_sync_meta.json"
    sync_context = {
        "url": public_url,
        "root": target_root_name,
        "client": target_client_name,
    }
    if sync_meta_file.exists():
        try:
            meta = json.loads(sync_meta_file.read_text(encoding="utf-8"))
            last_ts = float(meta.get("synced_at_epoch", 0))
            same_context = (
                str(meta.get("url", "")) == sync_context["url"]
                and str(meta.get("root", "")) == sync_context["root"]
                and str(meta.get("client", "")) == sync_context["client"]
            )
            has_local_data = resolve_pasta_vendas_ml(base_dir).is_dir() and all(
                (base_dir / name).is_dir() for name in REQUIRED_ONEDRIVE_SOURCE_FOLDERS if name != "Vendas - Mercado Livre"
            )
            if same_context and has_local_data and (time.time() - last_ts) < ONEDRIVE_SYNC_MIN_INTERVAL_SECONDS:
                return
        except Exception:
            pass

    share_token = _encode_share_token(public_url)
    root_url, root_meta = _fetch_shared_driveitem_metadata(public_url)
    root_is_graph = "graph.microsoft.com" in root_url.lower()
    root_name = str(root_meta.get("name", "")).strip()
    token = _graph_bearer_token()
    graph_headers = {"Authorization": f"Bearer {token}"} if (token and root_is_graph) else None

    def _children_url(item_url: str) -> str:
        return f"{item_url}/children"

    def _encode_rel_path(*parts: str) -> str:
        clean_parts = [p for p in parts if p]
        return "/".join(quote(p, safe="") for p in clean_parts)

    def _graph_item_url_from_child(child: dict[str, object]) -> str:
        item_id = str(child.get("id", "")).strip()
        parent_ref = child.get("parentReference", {})
        drive_id = ""
        if isinstance(parent_ref, dict):
            drive_id = str(parent_ref.get("driveId", "")).strip()
        if item_id and drive_id:
            return f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}"
        return ""

    def _download_url(item: dict[str, object]) -> str:
        for k in ("@microsoft.graph.downloadUrl", "@content.downloadUrl"):
            v = str(item.get(k, "")).strip()
            if v:
                return v
        return ""

    def _iter_children(item_url: str) -> list[dict[str, object]]:
        entries: list[dict[str, object]] = []
        next_url = _children_url(item_url)
        while next_url:
            payload = _download_json(next_url, headers=graph_headers)
            chunk = payload.get("value", [])
            if isinstance(chunk, list):
                for entry in chunk:
                    if isinstance(entry, dict):
                        entries.append(entry)
            next_url = str(payload.get("@odata.nextLink", "")).strip()
        return entries

    path_prefix = target_root_name

    use_link_root = target_root_name.strip().lower() in {"", ".", "__link_root__", "link_root"}

    def _locate_cliente_1_url() -> str:
        if use_link_root:
            return root_url
        if root_name == target_root_name:
            return root_url
        if root_name == "cliente_1":
            if root_is_graph:
                return f"https://graph.microsoft.com/v1.0/shares/u!{share_token}/driveItem:/cliente_1:"
            return f"https://api.onedrive.com/v1.0/shares/u!{share_token}/root:/cliente_1:"
        for entry in _iter_children(root_url):
            entry_name = str(entry.get("name", "")).strip()
            if entry_name == target_root_name:
                if root_is_graph:
                    return f"https://graph.microsoft.com/v1.0/shares/u!{share_token}/driveItem:/{target_root_name}:"
                return f"https://api.onedrive.com/v1.0/shares/u!{share_token}/root:/{target_root_name}:"
            if entry_name == "cliente_1":
                # Endpoint compatível com Graph e OneDrive API.
                if root_is_graph:
                    return f"https://graph.microsoft.com/v1.0/shares/u!{share_token}/driveItem:/cliente_1:"
                return f"https://api.onedrive.com/v1.0/shares/u!{share_token}/root:/cliente_1:"
        return ""

    cliente_1_url = _locate_cliente_1_url()
    if not cliente_1_url:
        raise ValueError(
            f"Pasta '{target_root_name}' não encontrada no link compartilhado."
        )
    if use_link_root or root_name == target_root_name:
        path_prefix = ""

    def _norm_for_filter(value: str) -> str:
        s = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode().lower().strip()
        s = s.replace("-", " ").replace("_", " ")
        return " ".join(s.split())

    required_norm_keys = {
        "vendas mercado livre",
        "vendas ml",
        "liberacoes ml",
        "notas saida",
        "nota saida",
        "contas receber",
        "contas a receber",
    }
    target_client_norm = _norm_for_filter(target_client_name)

    def _sync_folder(
        item_url: str,
        dest: Path,
        relative_parts: tuple[str, ...] = (),
        *,
        fast_mode: bool = True,
    ) -> None:
        dest.mkdir(parents=True, exist_ok=True)
        next_url = _children_url(item_url)
        while next_url:
            payload = _download_json(next_url, headers=graph_headers)
            children = payload.get("value", [])
            if not isinstance(children, list):
                break
            for child in children:
                if not isinstance(child, dict):
                    continue
                child_name = str(child.get("name", "")).strip()
                if not child_name:
                    continue
                child_rel = relative_parts + (child_name,)
                child_norm = _norm_for_filter(child_name)
                first_norm = _norm_for_filter(child_rel[0]) if child_rel else ""
                child_path = dest / child_name
                if "folder" in child:
                    if fast_mode:
                        if len(child_rel) == 1 and target_client_norm and child_norm not in {target_client_norm, *required_norm_keys}:
                            continue
                        if (
                            len(child_rel) == 2
                            and target_client_norm
                            and first_norm == target_client_norm
                            and child_norm not in required_norm_keys
                        ):
                            continue
                        if len(child_rel) >= 2 and first_norm not in {target_client_norm, *required_norm_keys}:
                            continue
                    if root_is_graph:
                        child_url = _graph_item_url_from_child(child)
                    else:
                        rel_path = _encode_rel_path(*(((path_prefix,) if path_prefix else ()) + child_rel))
                        child_url = f"https://api.onedrive.com/v1.0/shares/u!{share_token}/root:/{rel_path}:"
                    if not child_url:
                        continue
                    _sync_folder(child_url, child_path, child_rel, fast_mode=fast_mode)
                    continue

                if "file" not in child:
                    continue
                dl_url = _download_url(child)
                if not dl_url:
                    continue
                remote_size = child.get("size")
                if child_path.exists() and child_path.is_file() and isinstance(remote_size, int):
                    try:
                        if child_path.stat().st_size == remote_size:
                            continue
                    except Exception:
                        pass
                content, _, _ = _download_file_bytes(dl_url)
                child_path.parent.mkdir(parents=True, exist_ok=True)
                child_path.write_bytes(content)
            next_url = str(payload.get("@odata.nextLink", "")).strip()

    root_label = target_root_name if target_root_name.strip() else "link_root"
    dataset_root = mirror_root / root_label

    required_aliases: dict[str, tuple[str, ...]] = {
        "Vendas - Mercado Livre": ("vendas mercado livre", "vendas ml"),
        "Liberações_ML": ("liberacoes ml", "liberacoes_ml"),
        "notas_saida": ("notas saida", "nota saida", "notas_saida"),
        "contas_receber": ("contas receber", "contas a receber", "contas_receber"),
    }
    norm_to_required: dict[str, str] = {}
    for required_name, aliases in required_aliases.items():
        norm_to_required[_norm_for_filter(required_name)] = required_name
        for alias in aliases:
            norm_to_required[_norm_for_filter(alias)] = required_name

    def _all_candidate_roots(root: Path) -> list[Path]:
        out: list[Path] = []
        if root.exists() and root.is_dir():
            out.append(root)
            for child in root.iterdir():
                if child.is_dir():
                    out.append(child)
                    for grandchild in child.iterdir():
                        if grandchild.is_dir():
                            out.append(grandchild)
        return out

    def _resolve_required_subfolders(base_path: Path) -> dict[str, Path]:
        resolved: dict[str, Path] = {}
        if not base_path.exists() or not base_path.is_dir():
            return resolved
        for child in base_path.iterdir():
            if not child.is_dir():
                continue
            key = norm_to_required.get(_norm_for_filter(child.name))
            if key and key not in resolved:
                resolved[key] = child
        return resolved

    def _resolve_selected_mapping() -> dict[str, Path]:
        candidate_roots = [dataset_root]
        preferred = dataset_root / target_client_name
        if preferred.exists() and preferred.is_dir():
            candidate_roots.insert(0, preferred)
        for candidate in _all_candidate_roots(dataset_root):
            if candidate not in candidate_roots:
                candidate_roots.append(candidate)

        for candidate in candidate_roots:
            mapping = _resolve_required_subfolders(candidate)
            if len(mapping) == len(REQUIRED_ONEDRIVE_SOURCE_FOLDERS):
                return mapping
        return {}

    # 1) Tenta sincronização rápida (menos chamadas e menos arquivos).
    if dataset_root.exists():
        shutil.rmtree(dataset_root)
    _sync_folder(cliente_1_url, dataset_root, fast_mode=True)
    selected_mapping = _resolve_selected_mapping()

    # 2) Fallback automático: sincronização ampla (compatibilidade com estruturas diferentes).
    if not selected_mapping:
        if dataset_root.exists():
            shutil.rmtree(dataset_root)
        _sync_folder(cliente_1_url, dataset_root, fast_mode=False)
        selected_mapping = _resolve_selected_mapping()

    if not selected_mapping:
        raise ValueError(
            "Estrutura de dados não encontrada na pasta compartilhada. "
            "Defina FDL_ONEDRIVE_CLIENT_FOLDER para o cliente correto dentro de "
            f"'{target_root_name}'."
        )

    for required in REQUIRED_ONEDRIVE_SOURCE_FOLDERS:
        dst = base_dir / required
        src = selected_mapping.get(required)
        if src is None:
            continue
        if dst.exists():
            shutil.rmtree(dst)
        shutil.copytree(src, dst)

    missing: list[str] = []
    vdir = resolve_pasta_vendas_ml(base_dir)
    if not vdir.is_dir() or not any(vdir.rglob("*")):
        missing.append("Vendas - Mercado Livre")
    for name in REQUIRED_ONEDRIVE_SOURCE_FOLDERS:
        if name == "Vendas - Mercado Livre":
            continue
        if not any((base_dir / name).rglob("*")):
            missing.append(name)
    if missing:
        raise ValueError(
            "Link de pasta acessado, mas sem arquivos em: " + ", ".join(sorted(missing))
        )
    sync_meta_file.write_text(
        json.dumps(
            {
                **sync_context,
                "synced_at_epoch": time.time(),
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )


@st.cache_data(show_spinner=False, ttl=900)
def _sync_shared_folder_cached(
    public_url: str,
    root_folder: str,
    client_folder: str,
    _revisao: int = OPERACIONAL_CACHE_REVISION,
) -> str:
    del _revisao
    del root_folder
    del client_folder
    _download_shared_folder_dataset(public_url)
    return _now_ts_br_str()


def _http_get_file_follow_redirects(
    url: str, *, timeout: int = 60, extra_headers: dict[str, str] | None = None
) -> tuple[bytes, str, str, str | None]:
    """GET binário seguindo 301–308 (SharePoint / Graph devolve 308 «User migrated»)."""
    headers: dict[str, str] = {"User-Agent": "FDL-Streamlit-App/1.0"}
    if extra_headers:
        headers.update(extra_headers)
    current = url.strip()
    for _ in range(24):
        req = Request(current, headers=headers)
        try:
            with urlopen(req, timeout=timeout) as resp:
                payload = resp.read()
                final_url = resp.geturl()
                filename = ""
                cd = resp.headers.get("Content-Disposition", "")
                if "filename=" in cd:
                    filename = cd.split("filename=", 1)[1].strip().strip("\"'")
                if not filename:
                    filename = Path(urlparse(final_url).path).name or "download.bin"
                last_mod = resp.headers.get("Last-Modified")
                return payload, filename, final_url, last_mod
        except HTTPError as exc:
            if exc.code in (301, 302, 303, 307, 308) and exc.headers.get("Location"):
                current = urljoin(current, exc.headers["Location"])
                continue
            raise
    raise ValueError("Muitos redirecionamentos ao baixar o ficheiro.")


@st.cache_data(show_spinner=False, ttl=900)
def _download_file_bytes(
    url: str,
    _revisao: int = OPERACIONAL_CACHE_REVISION,
    *,
    extra_headers: dict[str, str] | None = None,
    timeout: int = 60,
    http_retries: int | None = None,
) -> tuple[bytes, str, str | None]:
    del _revisao
    url_original = url.strip()
    max_attempts = MAX_HTTP_RETRIES if http_retries is None else http_retries
    for attempt in range(max_attempts + 1):
        try:
            payload, filename, _final_url, last_modified = _http_get_file_follow_redirects(
                url_original, timeout=timeout, extra_headers=extra_headers
            )
            return payload, filename, last_modified
        except HTTPError as exc:
            if exc.code in RETRYABLE_HTTP_CODES and attempt < max_attempts:
                retry_after = str(exc.headers.get("Retry-After", "")).strip()
                sleep_s = float(retry_after) if retry_after.isdigit() else (1.5**attempt)
                time.sleep(min(max(sleep_s, 0.5), 8.0))
                continue
            body = ""
            try:
                body = exc.read().decode("utf-8", errors="replace")
            except Exception:
                body = ""
            body_snippet = body.strip().replace("\n", " ")
            if len(body_snippet) > 400:
                body_snippet = body_snippet[:400] + "..."
            raise ValueError(
                f"HTTP {exc.code} ao baixar ficheiro (URL de destino pode exigir login ou partilha anónima). "
                f"Resposta: {body_snippet or '(sem corpo)'}"
            ) from exc
        except URLError:
            if attempt >= max_attempts:
                raise
            time.sleep(min(1.5**attempt, 8.0))
    raise RuntimeError(f"Falha ao baixar arquivo após tentativas: {url_original}")


def _sync_payload_zip_to_base(payload: bytes) -> None:
    base_dir = Path(BASE_DIR)
    base_dir.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha256(payload).hexdigest()
    digest_file = base_dir / ".onedrive_payload.sha256"
    if digest_file.exists() and digest_file.read_text(encoding="utf-8").strip() == digest:
        return
    _prepare_uploaded_base(payload)
    digest_file.write_text(digest, encoding="utf-8")


def _validate_onedrive_csv_schema(df: pd.DataFrame) -> None:
    missing = sorted(REQUIRED_ONEDRIVE_CSV_COLUMNS - set(df.columns))
    if missing:
        raise ValueError(
            "CSV do OneDrive sem colunas obrigatórias: "
            + ", ".join(missing)
        )


def _precomputed_path_str() -> str:
    raw = os.environ.get("FDL_PRECOMPUTED_PATH", "").strip()
    if raw:
        return raw
    try:
        return str(st.secrets.get("FDL_PRECOMPUTED_PATH", "")).strip()
    except Exception:
        return ""


def _precomputed_url_str() -> str:
    raw = os.environ.get("FDL_PRECOMPUTED_URL", "").strip()
    if raw:
        return raw
    try:
        return str(st.secrets.get("FDL_PRECOMPUTED_URL", "")).strip()
    except Exception:
        return ""


def _precomputed_download_attempts(public_url: str) -> list[tuple[str, dict[str, str]]]:
    """Graph (token) → API shares → URL original (browser) → ?download=1 (muitos :x:/ devolvem 404)."""
    u = public_url.strip()
    attempts: list[tuple[str, dict[str, str]]] = []
    if not _is_m365_sharing_url(u):
        attempts.append((_download_url_for_precomputed_table(u), {}))
        return attempts

    tok = _microsoft_share_token_from_url(u)
    bearer = _graph_bearer_token()
    if bearer:
        attempts.append(
            (
                f"https://graph.microsoft.com/v1.0/shares/u!{tok}/driveItem/content",
                {"Authorization": f"Bearer {bearer}"},
            )
        )
    attempts.append((f"https://api.onedrive.com/v1.0/shares/u!{tok}/root/content", {}))

    # Antes de ?download=1: em muitos links :x:/ o parâmetro download=1 responde 404 «resource cannot be found».
    attempts.append((u, {"User-Agent": _BROWSER_UA_CHROME}))
    attempts.append((u, {}))

    parsed = urlparse(u)
    q = dict(parse_qsl(parsed.query, keep_blank_values=True))
    q["download"] = "1"
    dl_page = urlunparse(
        (parsed.scheme, parsed.netloc, parsed.path, parsed.params, urlencode(q), parsed.fragment)
    )
    attempts.append((dl_page, {}))
    attempts.append((dl_page, {"User-Agent": _BROWSER_UA_CHROME}))
    return attempts


def _download_url_for_precomputed_table(url: str) -> str:
    host = urlparse(url).netloc.lower()
    if "1drv.ms" in host or "sharepoint.com" in host or "onedrive.live.com" in host:
        return _build_onedrive_download_url(url)
    return url


def _dataframe_from_precomputed_bytes(payload: bytes, filename: str) -> pd.DataFrame:
    head = payload.lstrip()[:800]
    if head.startswith(b"<") or head.upper().startswith(b"<!DOCTYPE"):
        raise ValueError(
            "O servidor devolveu HTML (página de login ou erro), não o ficheiro de dados. "
            "No OneDrive, use partilha «Qualquer pessoa com a ligação pode ver» e teste o link em janela anónima."
        )
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        return pd.read_csv(BytesIO(payload), sep=None, engine="python", encoding="utf-8-sig")
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        return pd.read_excel(BytesIO(payload), engine="openpyxl")
    # API shares devolve filename «content»; .xlsx é ZIP (OOXML).
    if zipfile.is_zipfile(BytesIO(payload)):
        return pd.read_excel(BytesIO(payload), engine="openpyxl")
    try:
        return pd.read_csv(BytesIO(payload), sep=None, engine="python", encoding="utf-8-sig")
    except Exception:
        pass
    raise ValueError(f"Formato não suportado: {filename!r}. Use .csv ou .xlsx da conciliação.")


def _finalize_precomputed_df(
    tabela: pd.DataFrame,
    origem_arquivo: str,
    base_label: str,
    *,
    ts: str | None = None,
) -> tuple[pd.DataFrame, dict[str, object], str]:
    _validate_onedrive_csv_schema(tabela)
    if "empresa" not in tabela.columns:
        tabela = tabela.copy()
        tabela["empresa"] = _dataset_empresa_label()
    ts_out = ts if ts is not None else _now_ts_br_str()
    info: dict[str, object] = {
        "base_dir": base_label,
        "linhas": int(len(tabela)),
        "origem": "precomputed",
        "arquivo": origem_arquivo,
    }
    return tabela, info, ts_out


@st.cache_data(show_spinner=False, ttl=180)
def _load_precomputed_from_remote(url: str, _revisao: int = OPERACIONAL_CACHE_REVISION) -> tuple[pd.DataFrame, dict[str, object], str]:
    """Tabela já pronta via URL (ex.: ficheiro no OneDrive). Cache TTL evita download a cada filtro."""
    del _revisao
    errs: list[str] = []
    for dl_url, hdr in _precomputed_download_attempts(url.strip()):
        try:
            payload, filename, last_modified = _download_file_bytes(
                dl_url,
                extra_headers=hdr or None,
                timeout=PRECOMPUTED_HTTP_TIMEOUT,
                http_retries=1,
            )
            tabela = _dataframe_from_precomputed_bytes(payload, filename)
            ts = _ts_br_from_http_last_modified(last_modified) or _now_ts_br_str()
            return _finalize_precomputed_df(tabela, filename, "precomputed_url", ts=ts)
        except Exception as exc:  # noqa: BLE001 — agregamos para mensagem única
            hint = dl_url if len(dl_url) < 120 else dl_url[:117] + "..."
            errs.append(f"{hint} → {exc}")
    raise ValueError(
        "Não foi possível ler a tabela a partir do link. "
        "Em links SharePoint longos (:x:/), o servidor muitas vezes não permite download anónimo "
        "(HTTP 403/404 ou HTML em vez do ficheiro). "
        "Use uma destas opções: (1) `FDL_MS_GRAPH_TOKEN` (Bearer Microsoft Graph com acesso ao ficheiro); "
        "(2) link curto **1drv.ms** para o mesmo ficheiro; (3) alojar o .csv/.xlsx noutro URL público. "
        "Confirme partilha «qualquer pessoa com a ligação pode ver». "
        "Detalhes: " + " | ".join(errs[:5])
        + (" …" if len(errs) > 5 else "")
    )


@st.cache_data(show_spinner=True)
def _load_precomputed_from_disk(
    path_str: str, _mtime_ns: int, _revisao: int = OPERACIONAL_CACHE_REVISION
) -> tuple[pd.DataFrame, dict[str, object], str]:
    del _revisao
    path = Path(path_str)
    tabela = (
        pd.read_csv(path, sep=None, engine="python", encoding="utf-8-sig")
        if path.suffix.lower() == ".csv"
        else pd.read_excel(path, engine="openpyxl")
    )
    return _finalize_precomputed_df(
        tabela, path.name, str(path.parent), ts=_ts_br_from_mtime_ns(_mtime_ns)
    )


@st.cache_data(show_spinner=True)
def _load_repasse_parquet_from_disk(
    path_str: str,
    mtime_ns: int,
    org_id: str,
    _revisao: int = OPERACIONAL_CACHE_REVISION,
) -> tuple[pd.DataFrame, dict[str, object], str]:
    """PR4: repasse materializado em Parquet — contrato mínimo + validação de empresa_id (sem validação CSV OneDrive)."""
    del _revisao
    path = Path(path_str).expanduser().resolve()
    tabela = read_repasse_parquet(path)
    tabela = postprocess_repasse_parquet_dataframe(tabela, org_id)
    if "empresa" not in tabela.columns:
        tabela = tabela.copy()
        tabela["empresa"] = _dataset_empresa_label()
    ts_out = _ts_br_from_mtime_ns(mtime_ns)
    info: dict[str, object] = {
        "base_dir": str(path.parent),
        "linhas": int(len(tabela)),
        "origem": "materialized_parquet",
        "arquivo": path.name,
    }
    return tabela, info, ts_out


def load_precomputed_conciliacao() -> tuple[pd.DataFrame, dict[str, object], str]:
    """
    Lê só o ficheiro já gerado (export Power BI / mirror), sem correr o pipeline.
    Configure FDL_PRECOMPUTED_PATH OU URL (ou link direto para .csv/.xlsx em FDL_ONEDRIVE_URL, sem pasta :f:/).
    """
    path_s = _precomputed_path_str()
    url_s = _precomputed_url_str()
    if not url_s:
        od = _onedrive_public_url()
        if od and ":f:/" not in od.lower():
            url_s = od
    if path_s:
        path = Path(path_s).expanduser()
        if not path.is_absolute():
            path = (_REPO_APP_ROOT / path).resolve()
        if not path.is_file():
            raise FileNotFoundError(f"FDL_PRECOMPUTED_PATH não encontrado: {path}")
        if path.suffix.lower() not in {".csv", ".xlsx", ".xls"}:
            raise ValueError("FDL_PRECOMPUTED_PATH deve ser .csv, .xlsx ou .xls")
        mtime_ns = int(path.stat().st_mtime_ns)
        return _load_precomputed_from_disk(str(path.resolve()), mtime_ns)
    if url_s:
        return _load_precomputed_from_remote(url_s.strip())
    raise ValueError(
        "Modo precomputed: defina FDL_PRECOMPUTED_PATH (ficheiro no servidor) ou "
        "FDL_PRECOMPUTED_URL / link direto para .csv ou .xlsx. "
        "Alternativa: FDL_ONEDRIVE_URL apontando só para o ficheiro (não link de pasta :f:/)."
    )


def load_data_from_onedrive() -> tuple[pd.DataFrame, dict[str, object], str]:
    public_url = _onedrive_public_url()
    if not public_url:
        raise ValueError("FDL_ONEDRIVE_URL não configurada.")

    if ":f:/" in public_url.lower():
        _sync_shared_folder_cached(
            public_url=public_url,
            root_folder=_onedrive_root_folder_name(),
            client_folder=_onedrive_client_folder_name(),
        )
        return carregar_tabela_final_operacional_cache()

    download_url = _build_onedrive_download_url(public_url)
    payload, filename, last_modified = _download_file_bytes(download_url)
    lower_name = filename.lower()

    if lower_name.endswith(".zip") or zipfile.is_zipfile(BytesIO(payload)):
        _sync_payload_zip_to_base(payload)
        return carregar_tabela_final_operacional_cache()

    if lower_name.endswith(".csv"):
        tabela = pd.read_csv(BytesIO(payload), sep=None, engine="python", encoding="utf-8-sig")
        _validate_onedrive_csv_schema(tabela)
        if "empresa" not in tabela.columns:
            tabela = tabela.copy()
            tabela["empresa"] = _dataset_empresa_label()
        ts = _ts_br_from_http_last_modified(last_modified) or _now_ts_br_str()
        info = {
            "base_dir": "onedrive",
            "linhas": int(len(tabela)),
            "origem": "onedrive_csv",
            "arquivo": filename,
        }
        return tabela, info, ts

    raise ValueError(f"Formato não suportado no OneDrive: {filename}. Use ZIP/CSV ou link de pasta compartilhada.")


def _render_cloud_data_loader() -> None:
    with st.sidebar.expander("Admin: atualização de dados", expanded=False):
        st.caption("Uso interno. Não disponibilizar para cliente final.")
        uploaded_zip = st.file_uploader("Base de dados (.zip)", type=["zip"], key="fdl_data_zip")
        if uploaded_zip is not None:
            with st.spinner("Processando base enviada..."):
                _prepare_uploaded_base(uploaded_zip.getvalue())
                st.cache_data.clear()
            st.success("Base atualizada. Recarregando app...")
            st.rerun()


def _load_data_live() -> tuple[pd.DataFrame, dict[str, object], str]:
    source = _data_source_mode()
    if source in {"precomputed", "ready", "table"}:
        return load_precomputed_conciliacao()
    if source in {"filesystem", "onedrive"}:
        return load_data_from_onedrive()
    if source == "api":
        raise NotImplementedError(
            "Origem por API ainda não implementada. "
            "Defina FDL_DATA_SOURCE=onedrive até integrar a API (ex.: Bling)."
        )
    if source == "upload_zip":
        return carregar_tabela_final_operacional_cache()
    raise ValueError(f"FDL_DATA_SOURCE inválido: {source}")


def _load_data() -> tuple[pd.DataFrame, dict[str, object], str]:
    if _repasse_consume_mode() != "materialized":
        return _load_data_live()

    path_s = _repasse_materialized_path_str()
    url_s = _repasse_materialized_url_str()
    if _materialized_path_mode() == "dynamic" and not path_s and not url_s:
        _rep_fname = REPASSE_ARTIFACT_FILENAME if _repasse_use_parquet() else "dataset_repasse_app.csv"
        msg = (
            "Repasse em modo FDL_MATERIALIZED_PATH_MODE=dynamic: defina FDL_MATERIALIZED_CLIENTE_SLUG "
            f"(ex.: cliente_2). Esperado: {_materialized_data_products_root().strip()}/<cliente>/{_active_org.org_id}/"
            f"repasse/current/{_rep_fname}"
        )
        if _strict_materialized():
            raise ValueError(msg + " " + _STRICT_MATERIALIZED_USER_MSG)
        tabela, info, ts = _load_data_live()
        if _is_admin_mode():
            info = {
                **info,
                "repasse_consume": "live",
                "repasse_materialized_note": msg,
            }
        return tabela, info, ts

    if not path_s and not url_s:
        if _strict_materialized():
            raise ValueError(
                "Repasse em modo materialized: defina FDL_REPASSE_MATERIALIZED_PATH ou "
                "FDL_REPASSE_MATERIALIZED_URL. "
                + _STRICT_MATERIALIZED_USER_MSG
            )
        tabela, info, ts = _load_data_live()
        if _is_admin_mode():
            info = {
                **info,
                "repasse_consume": "live",
                "repasse_materialized_note": (
                    "Repasse em modo **ficheiro consolidado**, mas os caminhos dedicados estão vazios — "
                    "em uso **fonte em tempo real** (configuração FDL_DATA_SOURCE)."
                ),
            }
        return tabela, info, ts

    if _repasse_use_parquet() and (url_s or "").strip() and not (path_s or "").strip():
        raise ValueError(
            "FDL_REPASSE_USE_PARQUET=1: leitura só por URL não suportada para Parquet nesta fase. "
            "Use FDL_MATERIALIZED_PATH_MODE=dynamic (path local) ou FDL_REPASSE_MATERIALIZED_PATH apontando a um .parquet."
        )

    target_label = path_s or url_s
    try:
        if path_s:
            path = Path(path_s).expanduser()
            if not path.is_absolute():
                path = (_REPO_APP_ROOT / path).resolve()
            if not path.is_file():
                raise FileNotFoundError(f"FDL_REPASSE_MATERIALIZED_PATH não encontrado: {path}")
            suf = path.suffix.lower()
            if suf == ".parquet":
                if not _repasse_use_parquet():
                    raise ValueError(
                        "Ficheiro repasse .parquet: defina FDL_REPASSE_USE_PARQUET=1 para carregar este formato."
                    )
                mtime_ns = int(path.stat().st_mtime_ns)
                tabela, info, ts = _load_repasse_parquet_from_disk(
                    str(path.resolve()), mtime_ns, _active_org.org_id
                )
            elif suf in {".csv", ".xlsx", ".xls"}:
                mtime_ns = int(path.stat().st_mtime_ns)
                tabela, info, ts = _load_precomputed_from_disk(str(path.resolve()), mtime_ns)
            else:
                raise ValueError(
                    "FDL_REPASSE_MATERIALIZED_PATH deve ser .parquet (com FDL_REPASSE_USE_PARQUET=1), .csv, .xlsx ou .xls"
                )
        else:
            tabela, info, ts = _load_precomputed_from_remote(url_s.strip())
        return (
            tabela,
            {
                **info,
                "repasse_consume": "materialized",
                "repasse_materialized_target": target_label[:500],
            },
            ts,
        )
    except Exception as exc:
        if _strict_materialized():
            raise ValueError(
                f"{_STRICT_MATERIALIZED_USER_MSG} Não foi possível concluir o carregamento do repasse."
            ) from exc
        tabela, info, ts = _load_data_live()
        if _is_admin_mode():
            info = {
                **info,
                "repasse_consume": "live_fallback",
                "repasse_materialized_attempted": True,
                "repasse_materialized_target": target_label[:500],
                "repasse_materialized_error": str(exc).strip() or exc.__class__.__name__,
            }
        return tabela, info, ts


def _repasse_load_cache_signature(org_id: str) -> str:
    """Chave de cache do carregamento repasse: muda com org, revisão, caminhos e mtime do ficheiro materializado."""
    return "|".join(
        [
            str(org_id),
            str(OPERACIONAL_CACHE_REVISION),
            _repasse_consume_mode(),
            str(_repasse_materialized_path_str()).strip(),
            str(_repasse_materialized_url_str()).strip(),
            _data_source_mode(),
            str(_strict_materialized()),
            str(_repasse_use_parquet()),
            _repasse_materialized_source_stat_token(),
        ]
    )


@st.cache_data(show_spinner=False, ttl=900)
def _load_repasse_dataframe_cached(load_signature: str) -> tuple[pd.DataFrame, dict[str, object], str]:
    """Evita reler disco/rede a cada interação com filtros (rerun). `load_signature` isola org/config."""
    _ = load_signature
    return _load_data()



@st.cache_data(show_spinner=True)
def carregar_tabela_final_operacional_cache(
    _revisao: int = OPERACIONAL_CACHE_REVISION,
) -> tuple[pd.DataFrame, dict[str, object], str]:
    del _revisao  # só participa da chave de cache do Streamlit
    ts = _now_ts_br_str()
    tabela, info = carregar_tabela_final_operacional(BASE_DIR)
    return tabela, info, ts


def _serie_numero_nota_valida(s: pd.Series) -> pd.Series:
    """True quando há identificador de NF para conciliar (exclui vazio, None literal, NaN)."""
    x = s.fillna("").astype(str).str.strip()
    lower = x.str.lower()
    return x.ne("") & ~lower.isin({"none", "nan", "nat", "<na>", "null"})


def _col_referencia_como_texto(s: pd.Series) -> pd.Series:
    """Venda / pedido / NF como texto literal (incl. numpy.int64; sem «.0» em IDs float)."""
    def _um(v: object) -> str:
        if v is None:
            return ""
        if isinstance(v, str):
            t = v.strip()
            return "" if t.lower() in {"nan", "none", "nat", "<na>", "null"} else t
        try:
            if pd.isna(v):
                return ""
        except TypeError:
            pass
        if isinstance(v, bool):
            return str(v)
        if isinstance(v, numbers.Integral):
            return str(int(v))
        if isinstance(v, numbers.Real):
            fv = float(v)
            if math.isnan(fv):
                return ""
            iv = int(round(fv))
            if abs(fv - iv) < 1e-9:
                return str(iv)
            t = str(v).strip()
            return "" if t.lower() in {"nan", "none"} else t
        t = str(v).strip()
        if t.lower() in {"nan", "none", "nat", "<na>", "null"}:
            return ""
        if t.endswith(".0") and t.replace(".", "", 1).replace("-", "", 1).isdigit():
            return t[:-2]
        return t

    return s.map(_um).astype("string")


def _excluir_linhas_fora_conciliacao(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove linhas sem nota fiscal e taxas ML residuais (ex.: 3,62 / 3,54) típicas de encargos sem NF.

    A exigência «com nota» aplica-se **só ao Mercado Livre**: Shopee/Amazon costumam ficar «sem nota» na
    fila até haver vínculo com notas de saída; não podem ser apagadas só porque o ML já tem NF.

    Se **nenhuma** linha tiver número de nota preenchido (ex.: materialização sem pasta notas_saida),
    não elimina o conjunto inteiro — o cliente ainda precisa da fila operacional por venda/pedido.
    """
    if df.empty:
        return df
    out = df
    if "Número da nota" in out.columns:
        mask_nf = _serie_numero_nota_valida(out["Número da nota"])
        if "Plataforma" in out.columns:
            plat = out["Plataforma"].fillna("").astype(str).str.strip().str.lower()
            is_ml = plat.eq("mercado livre")
            if is_ml.any():
                ml_has_nf = bool((mask_nf & is_ml).any())
                if ml_has_nf:
                    drop = is_ml & ~mask_nf
                    out = out.loc[~drop].copy()
        elif mask_nf.any():
            out = out.loc[mask_nf].copy()
    if out.empty or "Total BRL" not in out.columns:
        return out
    tb = pd.to_numeric(out["Total BRL"], errors="coerce")
    for fee in (3.62, 3.54):
        out = out.loc[~(tb.sub(fee).abs() < 0.005)].copy()
        tb = pd.to_numeric(out["Total BRL"], errors="coerce")
    return out


def _fmt_brl_ptbr_celula(x: object) -> str:
    """Moeda pt-BR para células de grelha (R$ 1.234,56) — evita estilo US do NumberColumn."""
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except TypeError:
        pass
    try:
        v = float(x)
    except (TypeError, ValueError):
        return str(x).strip()
    if math.isnan(v):
        return ""
    neg = v < 0
    v = abs(v)
    cents = int(round(v * 100 + 1e-9))
    inteiro, cent = divmod(cents, 100)
    int_str = f"{inteiro:,}".replace(",", ".")
    corpo = f"{int_str},{cent:02d}"
    if neg:
        return f"R$ -{corpo}"
    return f"R$ {corpo}"


_FRETE_UI_COL_N_VENDA = "N.º venda"


def _fmt_int_ptbr(n: int) -> str:
    """Quantidade com separador de milhar pt-BR."""
    return f"{int(n):,}".replace(",", ".")


def _fmt_pct_ptbr_1(x: object) -> str:
    """Percentual pt-BR com uma casa decimal (ex.: 12,3%). ``x`` já em escala 0–100."""
    if x is None:
        return "—"
    try:
        if pd.isna(x):
            return "—"
    except TypeError:
        pass
    try:
        v = float(x)
    except (TypeError, ValueError):
        return "—"
    if math.isnan(v):
        return "—"
    return f"{v:.1f}".replace(".", ",") + "%"


def _comercial_fmt_qtd_display(x: object) -> str:
    """Quantidade para tabela comercial: inteiro com milhar pt-BR; senão uma casa decimal."""
    try:
        v = float(x)
    except (TypeError, ValueError):
        return "—"
    try:
        if pd.isna(v):
            return "—"
    except TypeError:
        pass
    if math.isnan(v):
        return "—"
    if abs(v - round(v)) < 1e-6:
        return _fmt_int_ptbr(int(round(v)))
    return f"{v:.1f}".replace(".", ",")


def _format_frete_anuncio_tabela_display(df: pd.DataFrame) -> pd.DataFrame:
    """Formata moeda pt-BR, quantidades com milhar e mantém «Recebido?» como texto."""
    if df.empty:
        return df
    out = df.copy()
    for c in ("Valor total (R$)", "Impacto (R$)"):
        if c in out.columns:
            out[c] = out[c].map(lambda x: _fmt_brl_ptbr_celula(x) if pd.notna(x) and x != "" else "")
    if "Qtde ocorrências" in out.columns:
        out["Qtde ocorrências"] = out["Qtde ocorrências"].map(
            lambda n: _fmt_int_ptbr(int(n)) if pd.notna(n) else ""
        )
    return out


# Texto exibido na coluna «Ação sugerida» da Fila operacional (UI apenas; export mantém valores canónicos).
_REPASSE_ACAO_SUGERIDA_EXIBICAO: dict[str, str] = {
    "Ok": "✅ OK",
    "Baixar no Bling": "⬇️ Baixar no Bling",
    "Baixado": "✅ Baixado",
    "Analisar diferença": "🔍 Analisar diferença",
    "Verificar recebimento": "📥 Verificar recebimento",
    "Verificar faturamento": "📄 Verificar faturamento",
    "Revisar venda zerada": "⚠️ Revisar venda zerada",
}


def _repasse_format_situacao_exibicao(val: object) -> str:
    """Prefixos visuais para «Situação» na grelha (UI; export permanece canónico)."""
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except TypeError:
        pass
    s = str(val).strip()
    if s.lower() in {"nan", "none", "<na>", "nat"}:
        return ""
    low = s.lower()
    if "diverg" in low:
        return f"🔍 {s}"
    if "atrasad" in low:
        return f"🔴 {s}"
    if "vencendo" in low and "hoje" in low:
        return f"🟡 {s}"
    if "vencendo" in low:
        return f"🟡 {s}"
    if "em dia" in low or low == "em dia":
        return f"✅ {s}"
    return s


def _repasse_format_acao_sugerida_exibicao(val: object) -> str:
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except TypeError:
        pass
    s = str(val).strip()
    if s.lower() in {"nan", "none", "<na>", "nat"}:
        return ""
    return _REPASSE_ACAO_SUGERIDA_EXIBICAO.get(s, s)


def _dataframe_conciliacao_somente_grid(df: pd.DataFrame) -> pd.DataFrame:
    """Cópia para st.dataframe: moeda em texto pt-BR e «Ação sugerida» com ícones (export continua canónico)."""
    if df.empty:
        return df
    g = df.copy()
    for c in ("Valor da nota", "Valor a receber", "Valor pago", "Diferença"):
        if c in g.columns:
            g[c] = g[c].map(_fmt_brl_ptbr_celula).astype(object)
    if "Ação sugerida" in g.columns:
        g["Ação sugerida"] = g["Ação sugerida"].map(_repasse_format_acao_sugerida_exibicao).astype(object)
    if "Situação" in g.columns:
        g["Situação"] = g["Situação"].map(_repasse_format_situacao_exibicao).astype(object)
    return g




def _column_config_conciliacao(
    df: pd.DataFrame, *, moeda_como_texto: bool = False
) -> dict[str, NumberColumn | DatetimeColumn | TextColumn]:
    """Cabeçalhos curtos + help; moeda em texto pt-BR na grelha ou número no export."""
    cfg: dict[str, NumberColumn | DatetimeColumn | TextColumn] = {}
    for c in ("Valor da nota", "Valor a receber", "Valor pago", "Diferença"):
        if c in df.columns:
            if moeda_como_texto:
                _disp = {
                    "Valor da nota": ("R$ NF", "Valor da nota fiscal."),
                    "Valor a receber": ("A receber", "Valor esperado de repasse / a receber."),
                    "Valor pago": ("Pago", "Valor efetivamente creditado."),
                    "Diferença": ("Δ R$", "Pago menos a receber (quando aplicável)."),
                }.get(c, (c, ""))
                _hl = _disp[1]
                cfg[c] = (
                    TextColumn(_disp[0], width="small", help=_hl)
                    if _hl
                    else TextColumn(_disp[0], width="small")
                )
            else:
                cfg[c] = NumberColumn(c, format="R$ %,.2f")
    if "Número da venda" in df.columns:
        cfg["Número da venda"] = TextColumn(
            "Venda",
            width="small",
            help="Identificador da venda na plataforma.",
        )
    if "Número do pedido" in df.columns:
        cfg["Número do pedido"] = TextColumn(
            "Pedido",
            width="small",
            help="Pedido / order ID.",
        )
    if "Número da nota" in df.columns:
        cfg["Número da nota"] = TextColumn(
            "NF",
            width="small",
            help="Número da nota fiscal.",
        )
    if "Data de emissão" in df.columns:
        cfg["Data de emissão"] = DatetimeColumn(
            "Emissão",
            format="DD/MM/YYYY",
            width="small",
            help="Data de emissão da NF.",
        )
    if "Período repasse" in df.columns:
        cfg["Período repasse"] = DatetimeColumn(
            "Período",
            format="DD/MM/YYYY",
            width="small",
            help="Data do filtro/ordenação: coluna materializada ou pagamento → emissão se pagamento vazio.",
        )
    if "Data de pagamento" in df.columns:
        cfg["Data de pagamento"] = DatetimeColumn(
            "Pagamento",
            format="DD/MM/YYYY HH:mm",
            width="medium",
            help="Data/hora do pagamento na plataforma.",
        )
    if "Situação" in df.columns:
        cfg["Situação"] = TextColumn(
            "Status",
            width="medium",
            help="Situação do ciclo (prazo, divergência, etc.).",
        )
    if "Ação sugerida" in df.columns:
        cfg["Ação sugerida"] = TextColumn(
            "Ação",
            width="large",
            help="Tratativa sugerida pela conciliação (mesma regra de sempre).",
        )
    if "Plataforma" in df.columns:
        cfg["Plataforma"] = TextColumn("Plat.", width="small", help="Canal de venda.")
    return cfg


def _fdl_repasse_inject_panel_styles() -> None:
    """Refino visual Conciliação de Repasse (filtros compactos vêm de ``_fdl_cp_inject_panel_styles``)."""
    st.markdown(
        dedent(
            """
            <style>
            .fdl-repasse-filtros-h {
              font-size: 1rem;
              font-weight: 700;
              color: #0f172a;
              margin: 0 0 4px 0;
              letter-spacing: -0.02em;
            }
            .fdl-repasse-caption {
              font-size: 0.74rem;
              font-weight: 500;
              color: #64748b;
              line-height: 1.38;
              margin: 0 0 10px 0;
            }
            .fdl-repasse-caption strong { color: #334155; font-weight: 600; }
            .fdl-repasse-section-title {
              font-size: 1.14rem;
              font-weight: 800;
              color: #0f172a;
              margin: 0 0 4px 0;
              letter-spacing: -0.025em;
            }
            .fdl-repasse-section-note {
              font-size: 0.74rem;
              font-weight: 500;
              color: #64748b;
              margin: 0 0 10px 0;
              line-height: 1.42;
            }
            .fdl-repasse-kpi-shell.fdl-fat-kpi-shell .fdl-fat-kpi-card--primary {
              border-left: 3px solid #0f172a;
              padding: 18px 20px 20px 20px;
            }
            .fdl-repasse-kpi-shell.fdl-fat-kpi-shell .fdl-fat-kpi-card--primary .fdl-fat-kpi-value {
              font-size: 1.82rem;
            }
            div[data-testid="stVerticalBlockBorderWrapper"]:has(.fdl-repasse-filtros-h) {
              border-color: #cbd5e1 !important;
              box-shadow: 0 2px 10px rgba(15, 23, 42, 0.05) !important;
            }
            </style>
            """
        ),
        unsafe_allow_html=True,
    )


def _fdl_frete_inject_panel_styles() -> None:
    """Refino visual Conciliação de Frete (filtros compactos: ``_fdl_cp_inject_panel_styles``)."""
    st.markdown(
        dedent(
            """
            <style>
            .fdl-frete-filtros-h {
              font-size: 1rem;
              font-weight: 700;
              color: #0f172a;
              margin: 0 0 4px 0;
              letter-spacing: -0.02em;
            }
            .fdl-frete-caption {
              font-size: 0.74rem;
              font-weight: 500;
              color: #64748b;
              line-height: 1.38;
              margin: 0 0 10px 0;
            }
            .fdl-frete-caption strong { color: #334155; font-weight: 600; }
            .fdl-frete-section-title {
              font-size: 1.14rem;
              font-weight: 800;
              color: #0f172a;
              margin: 0 0 4px 0;
              letter-spacing: -0.025em;
            }
            .fdl-frete-section-note {
              font-size: 0.74rem;
              font-weight: 500;
              color: #64748b;
              margin: 0 0 10px 0;
              line-height: 1.42;
            }
            .fdl-frete-kpi-shell.fdl-fat-kpi-shell .fdl-fat-kpi-card--primary {
              border-left: 3px solid #0f172a;
              padding: 18px 20px 20px 20px;
            }
            .fdl-frete-kpi-shell.fdl-fat-kpi-shell .fdl-fat-kpi-card--primary .fdl-fat-kpi-value {
              font-size: 1.82rem;
            }
            div[data-testid="stVerticalBlockBorderWrapper"]:has(.fdl-frete-filtros-h) {
              border-color: #cbd5e1 !important;
              box-shadow: 0 2px 10px rgba(15, 23, 42, 0.05) !important;
            }
            </style>
            """
        ),
        unsafe_allow_html=True,
    )


def _render_frete_indicadores_kpis(kpi_ex: dict[str, float], *, n_linhas: int) -> None:
    """Cards alinhados ao painel Repasse: volume + montantes |Δ| no recorte."""

    def _card(
        label: str,
        value: str,
        *,
        tier: str,
        accent: bool = False,
        title: str | None = None,
    ) -> str:
        classes = f"fdl-fat-kpi-card fdl-fat-kpi-card--{tier}"
        if accent:
            classes += " fdl-fat-kpi-card--accent"
        tattr = ""
        if title:
            tattr = f' title="{html.escape(title, quote=True)}"'
        return (
            f'<div class="{classes}"{tattr}>'
            f'<div class="fdl-fat-kpi-label">{html.escape(label)}</div>'
            f'<div class="fdl-fat-kpi-value">{html.escape(value)}</div>'
            "</div>"
        )

    primary = _card(
        "Vendas no recorte",
        _fmt_int_ptbr(int(n_linhas)),
        tier="primary",
        accent=True,
        title="Linhas após filtros: período, estado, situação do frete e busca.",
    )
    secondary = "".join(
        [
            _card(
                "A recuperar (cobrado a maior)",
                _fmt_brl_ptbr_celula(kpi_ex.get("cobrado_maior", 0.0)),
                tier="secondary",
                title="Soma |Δ| nas linhas «Cobrado a maior».",
            ),
            _card(
                "Repasse a conferir",
                _fmt_brl_ptbr_celula(kpi_ex.get("repasse", 0.0)),
                tier="secondary",
                title="Soma |Δ| nas linhas com repasse de frete a validar.",
            ),
        ]
    )
    st.markdown(
        f'<div class="fdl-fat-kpi-shell fdl-frete-kpi-shell">'
        f'<div class="fdl-fat-kpi-row fdl-fat-kpi-row--primary">{primary}</div>'
        f'<div class="fdl-fat-kpi-row fdl-fat-kpi-row--secondary">{secondary}</div></div>',
        unsafe_allow_html=True,
    )


def _repasse_fila_operacional_styler(df: pd.DataFrame) -> object:
    """Alinhamento e realce leve em Δ negativo (apenas exibição na grelha)."""
    if df.empty:
        return df
    sty = df.style
    try:
        sty = sty.hide(axis="index")
    except (TypeError, ValueError, AttributeError):
        try:
            sty = sty.hide_index()
        except Exception:
            pass

    def _diff_style(v: object) -> str:
        s = str(v).strip()
        base = "text-align: right; font-variant-numeric: tabular-nums;"
        if s.startswith("R$ -"):
            return base + "background-color: #fef2f2; color: #991b1b; font-weight: 600;"
        return base

    money_no_diff = [
        c for c in ("Valor da nota", "Valor a receber", "Valor pago") if c in df.columns
    ]
    if money_no_diff:
        sty = sty.set_properties(
            subset=money_no_diff,
            **{"text-align": "right", "font-variant-numeric": "tabular-nums"},
        )
    if "Diferença" in df.columns:
        _m = getattr(sty, "map", None) or getattr(sty, "applymap", None)
        if _m is not None:
            sty = _m(_diff_style, subset=["Diferença"])
    for c in ("Situação", "Ação sugerida"):
        if c in df.columns:
            sty = sty.set_properties(
                subset=[c],
                **{"font-weight": "600", "white-space": "nowrap"},
            )
    if "Plataforma" in df.columns:
        sty = sty.set_properties(subset=["Plataforma"], **{"font-weight": "500"})
    return sty


def _render_repasse_resumo_por_acao_kpis(contagens: dict[str, int], *, n_linhas: int) -> None:
    """Cards alinhados ao painel Faturamento: fila + distribuição por ação."""

    def _card(
        label: str,
        value: str,
        *,
        tier: str,
        accent: bool = False,
        title: str | None = None,
    ) -> str:
        classes = f"fdl-fat-kpi-card fdl-fat-kpi-card--{tier}"
        if accent:
            classes += " fdl-fat-kpi-card--accent"
        tattr = ""
        if title:
            tattr = f' title="{html.escape(title, quote=True)}"'
        return (
            f'<div class="{classes}"{tattr}>'
            f'<div class="fdl-fat-kpi-label">{html.escape(label)}</div>'
            f'<div class="fdl-fat-kpi-value">{html.escape(value)}</div>'
            "</div>"
        )

    primary = _card(
        "Registos no recorte",
        _fmt_int_ptbr(int(n_linhas)),
        tier="primary",
        accent=True,
        title="Linhas após filtros: período, plataforma, ação, situação e busca.",
    )
    if _repasse_vendas_liberacoes_only():
        secondary = "".join(
            [
                _card("Baixado", _fmt_int_ptbr(int(contagens.get("Baixado", 0))), tier="secondary"),
                _card(
                    "Divergências",
                    _fmt_int_ptbr(int(contagens.get("Analisar diferença", 0))),
                    tier="secondary",
                    title="Contagem com ação «Analisar diferença».",
                ),
                _card(
                    "Sem pagamento",
                    _fmt_int_ptbr(int(contagens.get("Verificar recebimento", 0))),
                    tier="secondary",
                    title="«Verificar recebimento».",
                ),
                _card(
                    "Zerados",
                    _fmt_int_ptbr(int(contagens.get("Zerado", 0))),
                    tier="secondary",
                    title="«Revisar venda zerada».",
                ),
            ]
        )
    else:
        ok = int(contagens.get("Ok", 0))
        bling = (
            int(contagens.get("Baixado", 0))
            if _repasse_sem_bling()
            else int(contagens.get("Baixar no Bling", 0))
        )
        div = int(contagens.get("Analisar diferença", 0))
        zero = int(contagens.get("Zerado", 0))
        c2_l = "Baixado" if _repasse_sem_bling() else "Baixar no Bling"
        secondary = "".join(
            [
                _card("OK", _fmt_int_ptbr(ok), tier="secondary"),
                _card(c2_l, _fmt_int_ptbr(bling), tier="secondary"),
                _card("Divergências", _fmt_int_ptbr(div), tier="secondary"),
                _card("Zerados", _fmt_int_ptbr(zero), tier="secondary"),
            ]
        )
    st.markdown(
        f'<div class="fdl-fat-kpi-shell fdl-repasse-kpi-shell">'
        f'<div class="fdl-fat-kpi-row fdl-fat-kpi-row--primary">{primary}</div>'
        f'<div class="fdl-fat-kpi-row fdl-fat-kpi-row--secondary">{secondary}</div>'
        f"</div>",
        unsafe_allow_html=True,
    )


def _multiselect_stable(
    key: str,
    label: str,
    options: list[str],
    *,
    compact_label: bool = False,
    help: str | None = None,
    placeholder: str = "Escolher…",
    label_visibility: str | None = None,
) -> list[str]:
    """
    Evita `default=` com listas recém-ordenadas a cada rerun (perdia estado / ecrã em branco).
    Estado inicial vazio: o cliente abre a seta e escolhe; vazio = sem filtro nessa dimensão.
    Se o utilizador limpar o último chip, o valor mantém-se vazio — não repor «todas as opções».
    """
    opts = [x for x in options if str(x).strip()]
    if not opts:
        if key not in st.session_state:
            st.session_state[key] = []
        return []
    if key not in st.session_state:
        st.session_state[key] = []
    else:
        prev = st.session_state[key]
        if not isinstance(prev, list):
            st.session_state[key] = []
        else:
            st.session_state[key] = [x for x in prev if x in opts]
    if compact_label:
        st.caption(label)
        return st.multiselect(
            " ",
            opts,
            key=key,
            placeholder=placeholder,
            label_visibility="collapsed",
            help=help,
        )
    _ms_kw: dict[str, Any] = {"key": key, "placeholder": placeholder, "help": help}
    if label_visibility is not None:
        _ms_kw["label_visibility"] = label_visibility
    return st.multiselect(label, opts, **_ms_kw)


def _faturamento_divergencia_tol() -> float:
    try:
        from processing.faturamento.config import DIVERGENCIA_VALOR_TOL

        return float(DIVERGENCIA_VALOR_TOL)
    except Exception:
        return 0.01


def _faturamento_resolve_produto_column(columns: list[str]) -> str | None:
    for c in ("Descrição", "Produto", "Nome do produto", "Título", "Título do anúncio", "Nome"):
        if c in columns:
            return c
    return None


def _faturamento_painel_custo_produto_col(columns: list[str]) -> str | None:
    """Materializado V2 usa ``Custo_Produto_Total``; legado pode usar ``Custo do Produto``."""
    if "Custo_Produto_Total" in columns:
        return "Custo_Produto_Total"
    if "Custo do Produto" in columns:
        return "Custo do Produto"
    return None


def _faturamento_painel_receita_series(df: pd.DataFrame, pl_col: str) -> pd.Series:
    """Soma de receita para KPIs: ``Vl_Venda`` (Qtd×lista), senão ``Receita_Bruta``, senão preço×qtd."""
    if "Vl_Venda" in df.columns:
        return pd.to_numeric(df["Vl_Venda"], errors="coerce")
    if "Receita_Bruta" in df.columns:
        return pd.to_numeric(df["Receita_Bruta"], errors="coerce")
    if "Quantidade" in df.columns and pl_col in df.columns:
        return pd.to_numeric(df[pl_col], errors="coerce") * pd.to_numeric(df["Quantidade"], errors="coerce")
    if pl_col in df.columns:
        return pd.to_numeric(df[pl_col], errors="coerce")
    return pd.Series(float("nan"), index=df.index, dtype=float)


def _faturamento_painel_missing_schema_columns(df: pd.DataFrame) -> list[str]:
    """Colunas mínimas para o painel; custo aceita alias V2."""
    c = set(df.columns)
    miss: list[str] = []
    for col in (
        "Preço de lista",
        "Valor total",
        "Resultado",
        "Situação",
        "Nome da plataforma",
        "Código",
        "Número do pedido",
        "Número do pedido multiloja",
        "Existe Nota Fiscal gerada",
        "Número da nota",
        "Custo de Frete",
        "Taxa de Comissão",
        "Imposto",
        "Despesas Fixas",
    ):
        if col not in c:
            miss.append(col)
    if _faturamento_painel_custo_produto_col(list(df.columns)) is None:
        miss.append("Custo_Produto_Total ou Custo do Produto")
    return miss


def _faturamento_num_col(df: pd.DataFrame, col: str) -> pd.Series:
    if col not in df.columns:
        return pd.Series(0.0, index=df.index, dtype=float)
    return pd.to_numeric(df[col], errors="coerce").fillna(0.0)


def _faturamento_pedido_display_series(df: pd.DataFrame) -> pd.Series:
    """Texto único para coluna «Pedido»: multiloja se preenchido, senão n.º do pedido."""
    ml = df["Número do pedido multiloja"].fillna("").astype(str).str.strip()
    ped = df["Número do pedido"].fillna("").astype(str).str.strip()
    core = ml.mask(ml.eq(""), ped)
    return _faturamento_disp_texto_sem_none(core.astype(str))


def _faturamento_pedido_id_series(df: pd.DataFrame) -> pd.Series:
    """
    Chave estável por linha para contagem distinta de pedidos.
    Usa multiloja quando preenchido; caso contrário ``Número do pedido``.
    Com ``org_id`` (v2 / consolidado), prefixa a org para evitar colisões entre empresas.
    """
    ml = df["Número do pedido multiloja"].fillna("").astype(str).str.strip()
    ped = df["Número do pedido"].fillna("").astype(str).str.strip()
    core = ml.mask(ml.eq(""), ped)
    if "org_id" in df.columns:
        oid = df["org_id"].fillna("").astype(str).str.strip()
        return oid + "|" + core
    return core


def _faturamento_atendido_mask(df: pd.DataFrame) -> pd.Series:
    """Alinhado a ``apply_faturamento_flags``: situação normalizada == «atendido»."""
    if "Situação" not in df.columns:
        return pd.Series(False, index=df.index, dtype=bool)
    situ = df["Situação"].fillna("").astype(str).str.strip().str.casefold()
    return situ.eq("atendido")


def _faturamento_admin_metadata_rowcount_message(path_final: str, row_count_loaded: int) -> str:
    """
    Compara ``faturamento_row_count_loaded`` com ``metadata.json`` na pasta ``current/`` (materializador).
    Ajuda a detetar cache antigo ou ficheiro errado sem hardcodar contagens por cliente.
    """
    raw = (path_final or "").strip()
    if not raw:
        return "Sem path resolvido — nada a contrastar com metadata."
    try:
        p = Path(raw).expanduser()
        if not p.is_absolute():
            p = (_REPO_APP_ROOT / p).resolve()
        else:
            p = p.resolve()
    except Exception:
        return "Path resolvido inválido para leitura de metadata."
    if not p.is_file():
        return f"Ficheiro não existe no disco: `{p}`"
    meta_path = p.parent / "metadata.json"
    if not meta_path.is_file():
        return f"**metadata.json** ausente em `{p.parent}` — confirme manualmente o artefato."
    try:
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
        rc = meta.get("row_count")
        try:
            rc_i = int(rc) if rc is not None else None
        except (TypeError, ValueError):
            rc_i = None
    except Exception as exc:
        return f"**metadata.json** não legível: `{exc}`"
    if rc_i is None:
        return "**metadata.json** sem `row_count` numérico — confirme manualmente."
    if rc_i == int(row_count_loaded):
        return f"**OK:** `metadata.json` `row_count` = **{rc_i}** coincide com **faturamento_row_count_loaded**."
    return (
        f"**ATENÇÃO:** `metadata.json` indica **row_count={rc_i}** mas o carregamento reportou **{row_count_loaded}** linhas. "
        "Possíveis causas: **cache Streamlit** (até TTL), ficheiro substituído após arranque, ou leitura de outro path."
    )


def _faturamento_materialized_fiscal_audit(df: pd.DataFrame) -> dict[str, object]:
    """Metadados sobre colunas fiscais do join notas (materializado V2)."""
    want = (
        "Nota_Data_Emissao",
        "Nota_Valor_Liquido_Rateado",
        "Nota_Valor_Liquido_Total",
        "Nota_Situacao",
        "Nota_Numero_Normalizado",
        "faturamento_nota_vinculada",
    )
    present = [c for c in want if c in df.columns]
    return {
        "faturamento_fiscal_columns_present": present,
        "faturamento_fiscal_join_complete": bool(
            "Nota_Data_Emissao" in df.columns and "Nota_Valor_Liquido_Rateado" in df.columns
        ),
    }


def _faturamento_agg_recorte(df: pd.DataFrame) -> dict[str, Any]:
    """
    Agregações numéricas do recorte (Visão Geral, KPIs). Independente de Streamlit.

    Definições do módulo (totais no recorte):

    * **Receita bruta (venda comercial)** = Σ ``Vl_Venda`` se existir; senão Σ ``Receita_Bruta`` / preço×qtd.
    * **Valor Nota Fiscal** (KPI) = Σ ``Nota_Valor_Liquido_Rateado`` quando o materializado traz join fiscal; senão
      Σ ``Valor total`` (pedido) ou fallback RB − desconto.
    * **Desconto comercial** = receita bruta − Σ ``Valor total`` quando existe; senão Σ ``Desconto proporcional total``.

    O **resultado** soma só valores numéricos de ``Resultado`` (linhas sem custo OK ficam vazias no materializado).

    ``diag_plug_rb_desc_vt``: Σ RB − Σ Desconto − Σ VT (só diagnóstico; ~0 em dados consistentes).
    """
    out: dict[str, Any] = {
        "n_linhas": int(len(df)),
        "receita_bruta": 0.0,
        "desconto_comercial": 0.0,
        "receita_liquida": 0.0,
        "custo_produto": 0.0,
        "frete": 0.0,
        "frete_me": 0.0,
        "frete_tp": 0.0,
        "comissao_plataforma": 0.0,
        "imposto": 0.0,
        "despesas_fixas": 0.0,
        "outras_despesas": 0.0,
        "resultado": 0.0,
        "margem_principal_pct": float("nan"),
        "pedidos_atendidos_distintos": 0,
        "n_linhas_sem_custo_ok": 0,
        "diag_plug_rb_desc_vt": None,
    }
    if df.empty:
        return out
    pl_col = "Preço de lista"
    rb_s = _faturamento_painel_receita_series(df, pl_col).fillna(0.0)
    out["receita_bruta"] = float(rb_s.sum())
    desc_col = "Desconto proporcional total"
    has_vt = "Valor total" in df.columns
    has_desc = desc_col in df.columns
    has_nvlr = "Nota_Valor_Liquido_Rateado" in df.columns
    vt_sum = float(_faturamento_num_col(df, "Valor total").sum()) if has_vt else None
    desc_sum = float(_faturamento_num_col(df, desc_col).sum()) if has_desc else None

    if vt_sum is not None:
        out["desconto_comercial"] = float(out["receita_bruta"] - vt_sum)
    elif desc_sum is not None:
        out["desconto_comercial"] = desc_sum
    else:
        out["desconto_comercial"] = 0.0

    if has_nvlr:
        out["receita_liquida"] = float(
            pd.to_numeric(df["Nota_Valor_Liquido_Rateado"], errors="coerce").fillna(0.0).sum()
        )
    elif vt_sum is not None:
        out["receita_liquida"] = vt_sum
    elif desc_sum is not None:
        out["receita_liquida"] = float(out["receita_bruta"] - desc_sum)
    else:
        out["receita_liquida"] = float(out["receita_bruta"])

    if has_vt and has_desc and vt_sum is not None and desc_sum is not None:
        out["diag_plug_rb_desc_vt"] = float(out["receita_bruta"] - desc_sum - vt_sum)
    ccol = _faturamento_painel_custo_produto_col(list(df.columns))
    if ccol and ccol in df.columns:
        out["custo_produto"] = float(pd.to_numeric(df[ccol], errors="coerce").fillna(0.0).sum())
    if "Frete_Plataforma" in df.columns:
        out["frete"] = float(_faturamento_num_col(df, "Frete_Plataforma").sum())
    else:
        out["frete"] = float(_faturamento_num_col(df, "Custo de Frete").sum())
    if "Frete Mercado Envios" in df.columns and "Frete transportadora própria" in df.columns:
        out["frete_me"] = float(_faturamento_num_col(df, "Frete Mercado Envios").sum())
        out["frete_tp"] = float(_faturamento_num_col(df, "Frete transportadora própria").sum())
    else:
        out["frete_me"] = float(out["frete"])
        out["frete_tp"] = 0.0
    out["comissao_plataforma"] = float(_faturamento_num_col(df, "Taxa de Comissão").sum())
    out["imposto"] = float(_faturamento_num_col(df, "Imposto").sum())
    out["despesas_fixas"] = float(_faturamento_num_col(df, "Despesas Fixas").sum())
    if "Outras Despesas" in df.columns:
        out["outras_despesas"] = float(_faturamento_num_col(df, "Outras Despesas").sum())
    res_s = pd.to_numeric(df["Resultado"], errors="coerce") if "Resultado" in df.columns else pd.Series(
        dtype=float
    )
    out["resultado"] = float(res_s.sum(skipna=True))
    rb = float(out["receita_bruta"])
    if rb not in (0.0, -0.0) and not math.isnan(rb):
        out["margem_principal_pct"] = float(out["resultado"] / rb)
    m_at = _faturamento_atendido_mask(df)
    if m_at.any():
        pids = _faturamento_pedido_id_series(df.loc[m_at]).astype(str).str.strip()
        pids = pids[pids.ne("")]
        out["pedidos_atendidos_distintos"] = int(pids.nunique()) if len(pids) else 0
    try:
        from processing.faturamento.config import STATUS_CUSTO_OK
    except Exception:
        STATUS_CUSTO_OK = "CUSTO_OK"
    if "Status_Custo" in df.columns:
        sc = df["Status_Custo"].astype(str).str.strip()
        out["n_linhas_sem_custo_ok"] = int((~sc.eq(STATUS_CUSTO_OK)).sum())
    return out


def _faturamento_visao_geral_chart_por_plataforma(df: pd.DataFrame) -> pd.DataFrame:
    """Agregação por plataforma (receita bruta); usada na tabela-resumo da Visão Geral."""
    if df.empty or "Nome da plataforma" not in df.columns:
        return pd.DataFrame()
    pl_col = "Preço de lista"
    rb = _faturamento_painel_receita_series(df, pl_col).fillna(0.0)
    plat = (
        df["Nome da plataforma"]
        .fillna("")
        .astype(str)
        .str.strip()
        .replace("", "(sem plataforma)")
    )
    g = (
        pd.DataFrame({"_rb": rb, "_plat": plat})
        .groupby("_plat", sort=False)["_rb"]
        .sum()
        .sort_values(ascending=False)
        .head(12)
    )
    return pd.DataFrame({"Plataforma": g.index.astype(str), "Receita bruta": g.values})


def _faturamento_visao_geral_tabela_plataforma(df: pd.DataFrame) -> pd.DataFrame:
    """Tabela-resumo por plataforma: receita bruta e participação % (mais legível que barras no MVP)."""
    base = _faturamento_visao_geral_chart_por_plataforma(df)
    if base.empty:
        return base
    tot = float(base["Receita bruta"].sum())
    base = base.copy()
    if tot not in (0.0, -0.0):
        prc = (base["Receita bruta"] / tot * 100.0).round(1)
        base["Part. do recorte"] = prc.map(
            lambda x: f"{x:.1f} %".replace(".", ",") if pd.notna(x) else "—"
        )
    else:
        base["Part. do recorte"] = "—"
    return base


def _render_faturamento_dre_visao_geral(df: pd.DataFrame, load_info: dict[str, object]) -> None:
    """Bloco MVP Visão Geral — mesmo recorte global que o painel detalhado."""
    st.subheader("Visão geral")
    escopo = str(load_info.get("faturamento_escopo", "") or "").strip()
    escopo_lbl = (
        "Consolidado (orgs permitidas)"
        if escopo == FAT_DRE_ESCOPO_CONSOLIDADO
        else "Empresa ativa"
        if escopo == FAT_DRE_ESCOPO_EMPRESA
        else escopo or "—"
    )
    st.caption(
        f"Mesmo **recorte global** que o detalhamento (antes de **Visão** / alertas / busca no bloco abaixo) · "
        f"escopo de carga: **{escopo_lbl}**."
    )
    _fdl_ui_gap_tight()
    if df.empty:
        st.info(
            "Sem dados para o resumo executivo. Confirme a **carga** de faturamento e o **recorte** acima quando existir base."
        )
        return
    if _faturamento_painel_missing_schema_columns(df):
        if _is_admin_mode():
            st.warning(
                "Visão geral indisponível: faltam colunas esperadas no ficheiro de faturamento. "
                "Verifique o materializado (layout V2)."
            )
        else:
            st.warning("Dados insuficientes para a Visão geral. Contacte o suporte.")
        return
    agg = _faturamento_agg_recorte(df)
    n_lin = int(agg["n_linhas"])
    n_sem = int(agg["n_linhas_sem_custo_ok"])
    if n_lin == 0:
        st.info("Sem linhas neste recorte — alargue período, situação ou plataforma.")
        return

    with st.container(border=True):
        st.caption("**Indicadores principais**")
        if "Nota_Valor_Liquido_Rateado" in df.columns:
            st.caption(
                f"**{_FATURAMENTO_UI_VALOR_NOTA_FISCAL}** = soma de **Nota_Valor_Liquido_Rateado** "
                "(valor líquido da nota de saída, rateado por linha; join pedidos↔notas)."
            )
        else:
            st.caption(
                f"**{_FATURAMENTO_UI_VALOR_NOTA_FISCAL}** = soma de **Valor total** do pedido (materializado sem coluna fiscal de nota; reprocesse V2 com notas)."
            )
        rp = st.columns(5)
        with rp[0]:
            st.metric("Receita bruta", _fmt_brl_ptbr_celula(agg["receita_bruta"]))
        with rp[1]:
            st.metric(_FATURAMENTO_UI_VALOR_NOTA_FISCAL, _fmt_brl_ptbr_celula(agg["receita_liquida"]))
        with rp[2]:
            st.metric("Resultado", _fmt_brl_ptbr_celula(agg["resultado"]))
        with rp[3]:
            mp = agg["margem_principal_pct"]
            if isinstance(mp, float) and not math.isnan(mp):
                st.metric("Margem principal", f"{mp * 100:.2f}%".replace(".", ","))
            else:
                st.metric("Margem principal", "—")
        with rp[4]:
            st.metric("Pedidos atendidos", _fmt_int_ptbr(agg["pedidos_atendidos_distintos"]))

    _n_ok_custo_vg = max(0, n_lin - n_sem)
    if n_sem > 0:
        ratio = (n_sem / n_lin) if n_lin else 0.0
        _msg_custo = (
            f"**{_n_ok_custo_vg}** linhas com custo alocado; **{n_sem}** exceções sem custo.\n\n"
            "O resultado total considera apenas linhas com custo."
        )
        if ratio > 0.10:
            st.info(_msg_custo)
        else:
            st.caption(_msg_custo)
    elif n_lin > 0:
        st.caption(
            f"**{_n_ok_custo_vg}** linhas com custo alocado. O resultado total considera apenas linhas com custo."
        )

    with st.expander("Outras métricas e definições", expanded=False):
        _n_ic = (
            int(_faturamento_series_bool_mask(df["faturamento_consolidado"]).sum())
            if "faturamento_consolidado" in df.columns
            else 0
        )
        st.metric("Itens consolidados (visão NF)", _fmt_int_ptbr(_n_ic))
        st.caption("Contagem de linhas em visão consolidada de NF (filtros do detalhamento).")
        o1 = st.columns(4)
        with o1[0]:
            st.metric("Desconto comercial", _fmt_brl_ptbr_celula(agg["desconto_comercial"]))
        with o1[1]:
            st.metric("Linhas no recorte", _fmt_int_ptbr(n_lin))
        with o1[2]:
            st.metric("Linhas sem custo OK", _fmt_int_ptbr(n_sem))
        with o1[3]:
            st.metric("Custo do produto", _fmt_brl_ptbr_celula(agg["custo_produto"]))
        o2 = st.columns(4)
        with o2[0]:
            if "Frete Mercado Envios" in df.columns and "Frete transportadora própria" in df.columns:
                st.metric("Frete (Mercado Envios)", _fmt_brl_ptbr_celula(agg["frete_me"]))
                st.metric("Frete (transp. própria)", _fmt_brl_ptbr_celula(agg["frete_tp"]))
            else:
                st.metric("Frete", _fmt_brl_ptbr_celula(agg["frete"]))
        with o2[1]:
            st.metric("Comissão plataforma", _fmt_brl_ptbr_celula(agg["comissao_plataforma"]))
        with o2[2]:
            st.metric("Imposto", _fmt_brl_ptbr_celula(agg["imposto"]))
        with o2[3]:
            st.metric("Despesas fixas", _fmt_brl_ptbr_celula(agg["despesas_fixas"]))
        o3 = st.columns(2)
        with o3[0]:
            if "Outras Despesas" in df.columns:
                st.metric("Outras despesas", _fmt_brl_ptbr_celula(agg["outras_despesas"]))
            else:
                st.metric("Outras despesas", "—")
        st.caption(
            "**Definições:** receita bruta = Σ **Vl_Venda** / **Receita_Bruta** conforme materializado. "
            "**Desconto comercial** = receita bruta − Σ **Valor total** (quando existir). "
            "**Valor Nota Fiscal** (este bloco) = Σ **Nota_Valor_Liquido_Rateado** com join fiscal; sem essa coluna, o KPI usa Σ **Valor total**."
            " **NF vazia** no fluxo ML: esperado sem vínculo."
        )
        _plug = agg.get("diag_plug_rb_desc_vt")
        if (
            _plug is not None
            and _is_admin_mode()
            and abs(float(_plug)) > max(1.0, 0.001 * max(n_lin, 1))
        ):
            st.caption(
                f"**Admin:** Σ RB − Σ Desconto prop. − Σ Valor total = **{_plug:,.2f}** (esperado ~0)."
            )

    tpm = _faturamento_visao_geral_tabela_plataforma(df)
    if not tpm.empty:
        st.caption("**Por plataforma** (top 12 · receita bruta e peso no recorte).")
        _tc_pl: dict[str, NumberColumn | TextColumn] = {
            "Plataforma": TextColumn("Plataforma", width="medium"),
            "Receita bruta": NumberColumn("Receita bruta", format="R$ %,.2f"),
            "Part. do recorte": TextColumn("Part. do recorte", width="small"),
        }
        _h_tbl = min(300, 52 + 34 * len(tpm))
        st.dataframe(
            tpm,
            use_container_width=True,
            hide_index=True,
            height=_h_tbl,
            column_config=_tc_pl,
        )

    _fdl_ui_gap_tight()


def _faturamento_dre_etiquetas_empresa_recorte(df: pd.DataFrame) -> list[str]:
    """Rótulos únicos de empresa no recorte (coluna ``empresa``; senão ``org_id``)."""
    if df.empty:
        return []
    if "empresa" in df.columns:
        u = sorted({str(x).strip() for x in df["empresa"].dropna().unique() if str(x).strip()})
        if u:
            return u
    if "org_id" in df.columns:
        return sorted({str(x).strip() for x in df["org_id"].dropna().unique() if str(x).strip()})
    return []


def _faturamento_dre_filtrar_por_etiquetas_empresa(df: pd.DataFrame, labels: list[str]) -> pd.DataFrame:
    if df.empty or not labels:
        return df
    if "empresa" in df.columns:
        em = df["empresa"].fillna("").astype(str).str.strip()
        return df.loc[em.isin(labels)].copy()
    if "org_id" in df.columns:
        oid = df["org_id"].astype(str).str.strip()
        return df.loc[oid.isin(labels)].copy()
    return df


def _faturamento_dre_default_empresa_labels(
    df: pd.DataFrame, org_id: str, org_display_name: str
) -> list[str]:
    """Rótulos iniciais do multiselect **Empresa**: vazio = todas quando há várias marcas na base."""
    _ = org_id, org_display_name
    opts = _faturamento_dre_etiquetas_empresa_recorte(df)
    if not opts:
        return []
    if len(opts) == 1:
        return list(opts)
    return []


def _render_faturamento_dre_bloco_por_empresa(
    df_recorte: pd.DataFrame,
    load_info: dict[str, object],
    ts_proc: str,
    org_id: str,
) -> None:
    """Bloco operacional: grelha, export e filtros sobre o mesmo ``df_recorte`` do recorte global."""
    st.subheader("Detalhamento operacional")
    st.caption(
        "Tabela, indicadores deste bloco e **CSV** partilham o **recorte do módulo** (acima). "
        "Aqui só pode refinar por **Visão** (NF comercial do pedido), **alertas** e **busca** — "
        "empresa, datas e fiscal ficam só no **Recorte do módulo**."
    )
    _fdl_ui_gap_tight()

    _painel_faturamento(
        df_recorte,
        load_info,
        ts_proc,
        org_id,
        use_modulo_recorte=True,
        mvp_rotulos_bloco_dre=True,
    )


def _df_get_series_column(df: pd.DataFrame, col: str) -> pd.Series:
    """
    Garante uma única ``pd.Series`` por nome de coluna.
    Se existirem colunas duplicadas, ``df[col]`` pode ser ``DataFrame`` — aqui usa-se a primeira série.
    """
    obj = df[col]
    if isinstance(obj, pd.DataFrame):
        return obj.iloc[:, 0].copy()
    return obj.copy()


def _series_empty_str_to_dash(s: pd.Series, *, dash: str = "—") -> pd.Series:
    """Normaliza texto: vazio / placeholders → traço (células de tabela NF-first)."""
    t = s.fillna("").map(lambda v: str(v).strip())
    return t.mask(t.eq("") | t.str.lower().isin({"nan", "none", "nat", "<na>", "null"}), dash)


def _series_nf_emissao_pt_br(s: pd.Series) -> pd.Series:
    """Datas de emissão como dd/mm/aaaa (aceita string, Timestamp, date vindo do Parquet)."""
    ts = pd.to_datetime(s, errors="coerce", dayfirst=False)
    fmt = ts.dt.strftime("%d/%m/%Y")
    return fmt.where(ts.notna(), "—")


def _faturamento_nf_platform_display_series(df_nf: pd.DataFrame) -> pd.Series:
    if "plataforma" in df_nf.columns:
        s = _df_get_series_column(df_nf, "plataforma").fillna("").astype(str)
    elif "plataforma_resumo" in df_nf.columns:
        s = _df_get_series_column(df_nf, "plataforma_resumo").fillna("").astype(str)
    else:
        return pd.Series("", index=df_nf.index, dtype=str)
    return s.map(nf_grain_plataforma_label_for_ui)


def _faturamento_nf_apply_minimal_recorte(
    df_nf: pd.DataFrame,
    *,
    empresas_sel: tuple[str, ...],
    plataformas_sel: tuple[str, ...],
    nf_d_ini: date,
    nf_d_fim: date,
    ok_nf_dates: bool,
) -> pd.DataFrame:
    if df_nf.empty:
        return df_nf
    out = df_nf.copy()
    emp_opts = _faturamento_dre_etiquetas_empresa_recorte(out)
    sel_emp = [str(x).strip() for x in empresas_sel if str(x).strip()]
    if emp_opts and sel_emp and "empresa" in out.columns:
        sel_cf = {x.casefold() for x in sel_emp}
        em_cf = out["empresa"].fillna("").astype(str).str.strip().str.casefold()
        out = out.loc[em_cf.isin(sel_cf)].copy()
    sel_plat = [str(x).strip() for x in plataformas_sel if str(x).strip()]
    _plat_col = (
        "plataforma"
        if "plataforma" in out.columns
        else (
            "plataforma_resumo"
            if "plataforma_resumo" in out.columns
            else ("Nome da plataforma" if "Nome da plataforma" in out.columns else "")
        )
    )
    if sel_plat and _plat_col:
        want = {nf_grain_plataforma_match_key(x) for x in sel_plat}
        want.discard("")
        if want:
            got = out[_plat_col].map(nf_grain_plataforma_match_key)
            out = out.loc[got.isin(want)].copy()
    if ok_nf_dates and nf_d_fim >= nf_d_ini and "Nota_Data_Emissao" in out.columns:
        m = _fdl_fr_mask_nf_emissao_no_periodo(out["Nota_Data_Emissao"], nf_d_ini, nf_d_fim)
        out = out.loc[m].copy()
    return out


def _faturamento_nf_filter_by_situacao(
    df_nf: pd.DataFrame,
    situacoes_sel: tuple[str, ...],
) -> pd.DataFrame:
    """Restringe por ``Nota_Situacao`` (case-insensitive), alinhado a ``build_faturamento_fiscal_base_slice``."""
    if df_nf.empty or not situacoes_sel:
        return df_nf
    want = {str(x).strip().casefold() for x in situacoes_sel if str(x).strip()}
    if not want or "Nota_Situacao" not in df_nf.columns:
        return df_nf
    s = df_nf["Nota_Situacao"].fillna("").astype(str).str.strip()
    return df_nf.loc[s.str.casefold().isin(want)].copy()


def _faturamento_fiscal_apply_minimal_recorte(
    df_fiscal: pd.DataFrame,
    *,
    empresas_sel: tuple[str, ...],
    nf_d_ini: date,
    nf_d_fim: date,
    ok_nf_dates: bool,
) -> pd.DataFrame:
    """Recorte fiscal: empresa + emissão. Sem filtro de plataforma (artefato fiscal não tem canal)."""
    if df_fiscal.empty:
        return df_fiscal
    out = df_fiscal.copy()
    emp_opts = _faturamento_dre_etiquetas_empresa_recorte(out)
    sel_emp = [str(x).strip() for x in empresas_sel if str(x).strip()]
    if emp_opts and sel_emp and "empresa" in out.columns:
        sel_cf = {x.casefold() for x in sel_emp}
        em_cf = out["empresa"].fillna("").astype(str).str.strip().str.casefold()
        out = out.loc[em_cf.isin(sel_cf)].copy()
    if ok_nf_dates and nf_d_fim >= nf_d_ini and "Nota_Data_Emissao" in out.columns:
        m = _fdl_fr_mask_nf_emissao_no_periodo(out["Nota_Data_Emissao"], nf_d_ini, nf_d_fim)
        out = out.loc[m].copy()
    return out


def _nf_panel_filter_merged_fiscal_by_plataforma_resumo(
    df_nf: pd.DataFrame,
    plataformas_sel: tuple[str, ...],
) -> pd.DataFrame:
    """
    Com Parquet fiscal + merge left join, o universo da tabela era **todas** as NFs fiscais; o filtro de
    plataforma só cortava o comercial antes do join, deixando linhas «só fiscais» com «—» e venda 0.

    Quando o utilizador escolhe plataforma(s), restringe o quadro **após** o merge às linhas cujo
    ``plataforma_resumo`` (texto antes de «(+N)») coincide com a seleção.
    """
    if df_nf.empty:
        return df_nf
    sel_plat = [str(x).strip() for x in plataformas_sel if str(x).strip()]
    if not sel_plat or "plataforma_resumo" not in df_nf.columns:
        return df_nf
    want = {nf_grain_plataforma_match_key(x) for x in sel_plat}
    want.discard("")
    if not want:
        return df_nf

    def _key_from_resumo(raw: object) -> str:
        s = str(raw).strip() if raw is not None else ""
        if not s or s == "—":
            return ""
        base = s.split("(", 1)[0].strip()
        return nf_grain_plataforma_match_key(base)

    got = df_nf["plataforma_resumo"].map(_key_from_resumo)
    return df_nf.loc[got.isin(want)].copy()


def _merge_fiscal_base_with_commercial_nf(
    df_fiscal: pd.DataFrame,
    df_commercial: pd.DataFrame,
) -> pd.DataFrame:
    """
    Uma linha por NF do **recorte fiscal**; ``valor_faturado_nf`` = ``Valor_Liquido_NF`` (Bling/export notas).
    Colunas comerciais por ``(org_id, empresa, NF normalizada)``; ver ``fiscal_commercial_nf_merge`` para fallback
    quando o comercial tem ``org_id`` vazio e o fiscal tem org preenchida.
    """
    out = merge_fiscal_base_with_commercial_nf_dataframe(df_fiscal, df_commercial)
    if not out.empty and "Nota_Situacao" in out.columns:
        out = out.copy()
        out["Nota_Situacao"] = _series_empty_str_to_dash(out["Nota_Situacao"])
    return out


def _fmt_pct_ptbr_ratio(ratio: float, *, decimals: int = 1) -> str:
    """Ex.: 0,123 → «12,3%» (apenas apresentação)."""
    if math.isnan(ratio) or math.isinf(ratio):
        return "—"
    p = ratio * 100.0
    body = f"{p:.{decimals}f}".replace(".", ",")
    return f"{body}%"


def _margem_sobre_venda_str(resultado: float, valor_venda: float) -> str:
    """Margem % = Σ Resultado ÷ Σ Valor da venda (mesmos totais do painel; só exibição)."""
    if (
        valor_venda == 0
        or math.isnan(valor_venda)
        or math.isnan(resultado)
        or math.isinf(resultado)
    ):
        return "—"
    return _fmt_pct_ptbr_ratio(resultado / valor_venda, decimals=1)


def _nf_row_margem_resultado_venda_ratio(valor_venda: pd.Series, resultado: pd.Series) -> pd.Series:
    """Por NF: resultado ÷ valor da venda (lista); NaN quando indefinido — mesma regra do painel."""
    vv = pd.to_numeric(valor_venda, errors="coerce")
    rr = pd.to_numeric(resultado, errors="coerce")
    denom = vv.mask(vv == 0, other=pd.NA)
    return (rr / denom).where(vv.notna() & rr.notna())



def _render_fdl_fat_dre_nf_kpi_cards(
    *,
    kp: dict[str, float | int],
    ok_nf_dates: bool,
    use_nf_materializado: bool,
    valor_faturado_from_fiscal_parquet: bool = False,
    fat_dre_faturado_mode: str = "nf_first",
    nf_panel_ads: bool = True,
) -> None:
    """
    Cards executivos NF-first (legado NF / fallback): hierarquia em 3 níveis.

    Preferir ``_render_resultado_gerencial_kpi_cards`` no Resultado Gerencial (âncora **Data** venda).
    """
    _ = valor_faturado_from_fiscal_parquet, fat_dre_faturado_mode, use_nf_materializado, nf_panel_ads
    if not _FAT_DRE_UI_V2 or build_kpi_nf_premium_shell_html is None:
        n_ped = int(kp.get("n_nf", 0))
        vv = float(kp["valor_venda"])
        tm_s = _fmt_brl_ptbr_celula(vv / float(n_ped)) if ok_nf_dates and n_ped > 0 else "—"
        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            st.metric("Valor da venda", _fmt_brl_ptbr_celula(kp["valor_venda"]) or "R$ 0,00")
        with c2:
            st.metric("Pedidos", _fmt_int_ptbr(n_ped) if ok_nf_dates else "—")
        with c3:
            st.metric("Ticket médio", tm_s or "—")
        with c4:
            st.metric(
                "Resultado",
                _fmt_brl_ptbr_celula(kp["resultado"]) or "—",
                help=(
                    "Resultado consolidado por NF no recorte selecionado (empresa, período, situação, plataforma). "
                    "Pode diferir do Painel de Saúde, que usa grão de linha de pedido e não aplica filtro de plataforma."
                ),
            )
        with c5:
            st.metric("Margem %", _margem_sobre_venda_str(float(kp["resultado"]), float(kp["valor_venda"])))
        return

    vv = float(kp["valor_venda"])
    res = float(kp["resultado"])
    n_ped_i = int(kp.get("n_nf", 0))
    pedidos_fmt = _fmt_int_ptbr(n_ped_i) if ok_nf_dates else "—"
    ticket_fmt = (
        (_fmt_brl_ptbr_celula(vv / float(n_ped_i)) or "—") if ok_nf_dates and n_ped_i > 0 else "—"
    )
    margem_str = _margem_sobre_venda_str(res, vv)
    venda_fmt = _fmt_brl_ptbr_celula(kp["valor_venda"]) or "R$ 0,00"
    res_fmt = _fmt_brl_ptbr_celula(kp["resultado"]) or "—"

    st.markdown(
        build_kpi_nf_premium_shell_html(
            valor_venda_fmt=venda_fmt,
            pedidos_fmt=pedidos_fmt,
            ticket_medio_fmt=ticket_fmt,
            resultado_fmt=res_fmt,
            margem_str=margem_str,
            valor_venda=vv,
            resultado=res,
            chips=[],
            mode_pill_html="",
        ),
        unsafe_allow_html=True,
    )


def _render_resultado_gerencial_kpi_cards(
    *,
    kp_rg: dict[str, float | int],
    ok_dates: bool,
    use_nf_materializado: bool,
    valor_faturado_from_fiscal_parquet: bool = False,
    fat_dre_faturado_mode: str = "nf_first",
    nf_panel_ads: bool = True,
    comparacao_temporal: object | None = None,
) -> None:
    """
    Cards superiores do Resultado Gerencial — totais por **data da venda** (grão linha) + imposto fiscal injetado.
    """
    _ = valor_faturado_from_fiscal_parquet, fat_dre_faturado_mode, use_nf_materializado, nf_panel_ads
    vv = float(kp_rg["valor_venda_lista"])
    res = float(kp_rg["resultado"])
    n_ped = int(kp_rg["pedidos"])
    _RG_KPI_RESULTADO_TITLE = (
        "Resultado gerencial das vendas no período filtrado (por data da venda). Considera receita, comissão, custo, "
        "frete, imposto (via Apuração Fiscal / ponte fiscal), despesas fixas e ADS quando presentes nas linhas."
    )
    if not _FAT_DRE_UI_V2 or build_kpi_nf_premium_shell_html is None:
        tm_s = "—"
        if ok_dates and n_ped > 0:
            tr = kp_rg["ticket_medio"]
            if isinstance(tr, float) and math.isnan(tr):
                tm_s = _fmt_brl_ptbr_celula(vv / float(n_ped)) or "—"
            else:
                tm_s = _fmt_brl_ptbr_celula(float(tr)) or "—"
        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            st.metric("Valor da venda", _fmt_brl_ptbr_celula(vv) or "R$ 0,00")
        with c2:
            st.metric("Pedidos", _fmt_int_ptbr(n_ped) if ok_dates else "—")
        with c3:
            st.metric("Ticket médio", tm_s or "—")
        with c4:
            st.metric(
                "Resultado",
                _fmt_brl_ptbr_celula(res) or "—",
                help=(
                    "Resultado no recorte por **data da venda**, com imposto da ponte fiscal. "
                    "Pode diferir dos blocos que ainda usam grão NF ou linhas sem este recorte."
                ),
            )
        with c5:
            st.metric("Margem %", _margem_sobre_venda_str(res, vv))
        return

    pedidos_fmt = _fmt_int_ptbr(n_ped) if ok_dates else "—"
    tm_raw = kp_rg["ticket_medio"]
    if ok_dates and n_ped > 0:
        tm_ok = float(tm_raw) if not (isinstance(tm_raw, float) and math.isnan(tm_raw)) else vv / float(n_ped)
        ticket_fmt = _fmt_brl_ptbr_celula(tm_ok) or "—"
    else:
        ticket_fmt = "—"
    margem_str = _margem_sobre_venda_str(res, vv)
    venda_fmt = _fmt_brl_ptbr_celula(vv) or "R$ 0,00"
    res_fmt = _fmt_brl_ptbr_celula(res) or "—"

    _cap_res = ""
    _cap_mg = ""
    if comparacao_temporal is not None:
        from processing.faturamento.comparacao_temporal_kpis import build_temporal_kpi_captions_html

        _cap_res, _cap_mg = build_temporal_kpi_captions_html(comparacao_temporal)

    st.markdown(
        build_kpi_nf_premium_shell_html(
            valor_venda_fmt=venda_fmt,
            pedidos_fmt=pedidos_fmt,
            ticket_medio_fmt=ticket_fmt,
            resultado_fmt=res_fmt,
            margem_str=margem_str,
            valor_venda=vv,
            resultado=res,
            chips=[],
            mode_pill_html="",
            resultado_title=_RG_KPI_RESULTADO_TITLE,
            omit_hero_meta=True,
            tier_b_layout=True,
            hero_caption_resultado_html=_cap_res,
            hero_caption_margem_html=_cap_mg,
        ),
        unsafe_allow_html=True,
    )


def _fmt_brl_ptbr_encargo_dre(v: object) -> str:
    """Total já somado no painel; só exibição como saída (−R$ …) na DRE gerencial."""
    try:
        x = float(v)
    except (TypeError, ValueError):
        return "—"
    if math.isnan(x):
        return "—"
    s = _fmt_brl_ptbr_celula(abs(x))
    if not s:
        return "—"
    if s.startswith("R$ "):
        return "−" + s
    return "−" + s



def _render_fdl_fat_dre_nf_gerencial(
    *,
    kp: dict[str, float | int],
    ok_nf_dates: bool,
    valor_faturado_from_fiscal_parquet: bool = False,
    periodo_label: str = "",
    nf_panel_ads: bool = True,
    fiscal_base_stats: FaturamentoFiscalBaseStats | None = None,
) -> None:
    """DRE gerencial (totais de ``compute_nf_panel_kpis``), layout demonstração financeira.

    Linha Imposto: ver ``dre_imposto_para_linha_dre_gerencial`` (ponte fiscal única).
    """
    if not _FAT_DRE_UI_V2 or build_dre_gerencial_premium_html is None:
        st.subheader("DRE gerencial")
        st.write(
            {
                "Venda": float(kp["valor_venda"]),
                "Resultado": float(kp["resultado"]),
                "Margem %": _margem_sobre_venda_str(float(kp["resultado"]), float(kp["valor_venda"])),
            }
        )
        return

    vv = float(kp["valor_venda"])
    res = float(kp["resultado"])
    imp_raw = float(kp["imposto"])
    # Ponte entre módulos: ver ``dre_imposto_para_linha_dre_gerencial`` em ``faturamento_dre_recorte_minimo``.
    imp_nf = dre_imposto_para_linha_dre_gerencial(
        kp,
        fiscal_base_stats=fiscal_base_stats,
        aplicar_ponte_base_liquida=(
            fiscal_base_stats is not None and valor_faturado_from_fiscal_parquet
        ),
    )
    res_nf = res + (imp_raw - imp_nf)
    rec_frete_num = float(kp.get("receita_frete_tp", 0.0))
    rec_venda = _fmt_brl_ptbr_celula(kp["valor_venda"]) or "R$ 0,00"
    margem_s = _margem_sobre_venda_str(res_nf, vv)
    res_disp = _fmt_brl_ptbr_celula(res_nf) or "—"
    enc_com = _fmt_brl_ptbr_encargo_dre(kp["comissao"])
    enc_custo = _fmt_brl_ptbr_encargo_dre(kp.get("custo_produto", 0.0))
    rec_frete_disp = _fmt_brl_ptbr_celula(float(kp.get("receita_frete_tp", 0.0))) or "R$ 0,00"
    enc_frete_plat = _fmt_brl_ptbr_encargo_dre(float(kp.get("custo_frete_plataforma", 0.0)))
    enc_repasse_tp = _fmt_brl_ptbr_encargo_dre(float(kp.get("repasse_frete_transportadora_propria", 0.0)))
    enc_imp = _fmt_brl_ptbr_encargo_dre(imp_nf)
    enc_df = _fmt_brl_ptbr_encargo_dre(kp["despesa_fixa"])
    ads_sum = float(kp.get("custo_ads_variavel", 0.0)) + float(kp.get("custo_ads_fixo", 0.0))
    enc_ads = _fmt_brl_ptbr_encargo_dre(ads_sum)

    enc_rows = [
        ("Comissão", enc_com),
        ("Custo produto", enc_custo),
        ("Frete plataforma", enc_frete_plat),
        ("Frete transp. própria", enc_repasse_tp),
        ("Imposto", enc_imp),
        ("Despesa fixa", enc_df),
    ]
    if nf_panel_ads:
        enc_rows.append(("ADS (3,5% + fixo)", enc_ads))

    total_rec_num = vv + rec_frete_num
    total_rec_fmt = _fmt_brl_ptbr_celula(total_rec_num) or "R$ 0,00"
    _ded_sum_num = (
        float(kp["comissao"])
        + float(kp.get("custo_produto", 0.0))
        + float(kp.get("custo_frete_plataforma", 0.0))
        + float(kp.get("repasse_frete_transportadora_propria", 0.0))
        + float(imp_nf)
        + float(kp["despesa_fixa"])
    )
    if nf_panel_ads:
        _ded_sum_num += ads_sum
    total_ded_fmt = _fmt_brl_ptbr_encargo_dre(_ded_sum_num)

    _marg_base = (
        "Margem = soma do resultado ÷ soma da receita de venda (lista). "
        "O valor faturado fiscal não entra neste cálculo."
    )
    _marg_simple = "Margem = resultado ÷ receita de venda (lista)."
    if valor_faturado_from_fiscal_parquet:
        if nf_panel_ads:
            tt_res = (
                "Total consolidado por NF no mesmo recorte dos cards. "
                "Inclui ADS (3,5% sobre receita lista + valor fixo por NF)."
            )
            tt_marg = _marg_base
        else:
            tt_res = (
                "Total consolidado por NF no mesmo recorte dos cards. "
                "Neste cliente não há linha de ADS neste quadro."
            )
            tt_marg = _marg_base
    elif nf_panel_ads:
        tt_res = (
            "Total consolidado por NF no mesmo recorte dos cards. "
            "Inclui ADS (3,5% + fixo por NF)."
        )
        tt_marg = _marg_simple
    else:
        tt_res = (
            "Total consolidado por NF no mesmo recorte dos cards. "
            "Sem custo de ADS neste cliente."
        )
        tt_marg = _marg_simple
    if valor_faturado_from_fiscal_parquet:
        if nf_panel_ads:
            foot = (
                "Resultado e margem consideram ADS (3,5% sobre receita lista e valor fixo por NF) quando aplicável. "
                "Margem = resultado ÷ receita de venda (lista); o valor faturado fiscal não entra no denominador."
            )
        else:
            foot = (
                "Margem = resultado ÷ receita de venda (lista); o valor faturado fiscal não entra no denominador. "
                "Sem linha de ADS neste quadro."
            )
    elif nf_panel_ads:
        foot = (
            "Resultado e margem consideram ADS quando aplicável. "
            "Margem = resultado ÷ receita de venda (lista); totais consolidados por nota fiscal."
        )
    else:
        foot = "Margem = resultado ÷ receita de venda (lista). Sem custo de ADS neste cliente."

    per = (periodo_label or "").strip() or ("Emissão NF no filtro" if ok_nf_dates else "Período indisponível")

    st.markdown(
        build_dre_gerencial_premium_html(
            period_caption=per,
            valor_venda_fmt=rec_venda,
            rec_frete_fmt=rec_frete_disp,
            total_receita_fmt=total_rec_fmt,
            enc_rows=enc_rows,
            total_deducoes_fmt=total_ded_fmt,
            resultado_fmt=res_disp,
            resultado_value=res_nf,
            margem_str=margem_s,
            resultado_tooltip=tt_res,
            margem_tooltip=tt_marg,
            footnote_plain=foot,
        ),
        unsafe_allow_html=True,
    )


def _render_fdl_fat_dre_gerencial_linha(
    *,
    stats: object,
    kp_rg: dict[str, float | int],
    imp_nf: float,
    ok_nf_dates: bool,
    valor_faturado_from_fiscal_parquet: bool,
    periodo_label: str,
    nf_panel_ads: bool,
    rg_header_subtitle: str = "",
    show_resultado_discreto: bool = False,
) -> None:
    """DRE gerencial no **grão linha**, mesmo recorte dos KPIs (**Data** venda) + imposto da ponte fiscal."""
    vv = float(kp_rg["valor_venda_lista"])
    res_nf = float(kp_rg["resultado"])
    rec_frete_num = float(stats.frete_transportadora_propria_total)
    rec_venda = _fmt_brl_ptbr_celula(vv) or "R$ 0,00"
    margem_s = _margem_sobre_venda_str(res_nf, vv)
    res_disp = _fmt_brl_ptbr_celula(res_nf) or "—"
    enc_com = _fmt_brl_ptbr_encargo_dre(stats.comissao_total)
    enc_custo = _fmt_brl_ptbr_encargo_dre(stats.cmv_total)
    rec_frete_disp = _fmt_brl_ptbr_celula(rec_frete_num) or "R$ 0,00"
    enc_frete_plat = _fmt_brl_ptbr_encargo_dre(stats.frete_plataforma_total)
    enc_repasse_tp = _fmt_brl_ptbr_encargo_dre(stats.frete_transportadora_propria_total)
    enc_imp = _fmt_brl_ptbr_encargo_dre(imp_nf)
    enc_df = _fmt_brl_ptbr_encargo_dre(stats.despesa_fixa_total)
    ads_sum = float(stats.ads_total)
    enc_ads = _fmt_brl_ptbr_encargo_dre(ads_sum)

    enc_rows = [
        ("Comissão", enc_com),
        ("Custo produto", enc_custo),
        ("Frete plataforma", enc_frete_plat),
        ("Frete transp. própria", enc_repasse_tp),
        ("Imposto", enc_imp),
        ("Despesa fixa", enc_df),
    ]
    if nf_panel_ads:
        enc_rows.append(("ADS (3,5% + fixo)", enc_ads))

    total_rec_num = vv + rec_frete_num
    total_rec_fmt = _fmt_brl_ptbr_celula(total_rec_num) or "R$ 0,00"
    _ded_sum_num = float(kp_rg["total_deducoes"])
    total_ded_fmt = _fmt_brl_ptbr_encargo_dre(_ded_sum_num)

    _marg_base = (
        "Margem = soma do resultado ÷ soma da receita de venda (lista). "
        "O valor faturado fiscal não entra neste cálculo."
    )
    _marg_simple = "Margem = resultado ÷ receita de venda (lista)."
    if valor_faturado_from_fiscal_parquet:
        if nf_panel_ads:
            tt_res = (
                "Totais consolidados no **grão linha** por **data da venda** (mesmo filtro dos KPIs). "
                "Imposto na linha «Imposto» vem da **ponte fiscal** (base de emissão). "
                "Inclui ADS quando aplicável ao materializado."
            )
            tt_marg = _marg_base
        else:
            tt_res = (
                "Totais consolidados no **grão linha** por **data da venda**. "
                "Imposto via ponte fiscal. Sem linha de ADS neste quadro."
            )
            tt_marg = _marg_base
    elif nf_panel_ads:
        tt_res = (
            "Totais consolidados no **grão linha** por **data da venda** (coluna Data). "
            "Inclui ADS quando presente nas linhas."
        )
        tt_marg = _marg_simple
    else:
        tt_res = (
            "Totais consolidados no **grão linha** por **data da venda**. Sem custo de ADS neste quadro."
        )
        tt_marg = _marg_simple

    if valor_faturado_from_fiscal_parquet:
        if nf_panel_ads:
            foot = (
                "Receita e deduções no recorte por **data da venda**; imposto exibido segue a **ponte fiscal**. "
                "Margem = resultado ÷ receita de venda (lista)."
            )
        else:
            foot = (
                "Recorte por **data da venda**; imposto via ponte fiscal. Sem linha de ADS neste quadro. "
                "Margem = resultado ÷ receita de venda (lista)."
            )
    elif nf_panel_ads:
        foot = (
            "Totais no grão linha por **data da venda**, alinhados aos KPIs superiores. "
            "Margem = resultado ÷ receita de venda (lista)."
        )
    else:
        foot = "Recorte por **data da venda**. Margem = resultado ÷ receita de venda (lista). Sem ADS neste quadro."

    per = (periodo_label or "").strip() or ("Data da venda no filtro" if ok_nf_dates else "Período indisponível")

    if not _FAT_DRE_UI_V2 or build_dre_gerencial_premium_html is None:
        st.subheader("DRE gerencial")
        st.write(
            {
                "Venda (lista)": vv,
                "Resultado": res_nf,
                "Margem %": margem_s,
                "Total deduções": _ded_sum_num,
            }
        )
        return

    st.markdown(
        build_dre_gerencial_premium_html(
            period_caption=per,
            valor_venda_fmt=rec_venda,
            rec_frete_fmt=rec_frete_disp,
            total_receita_fmt=total_rec_fmt,
            enc_rows=enc_rows,
            total_deducoes_fmt=total_ded_fmt,
            resultado_fmt=res_disp,
            resultado_value=res_nf,
            margem_str=margem_s,
            resultado_tooltip=tt_res,
            margem_tooltip=tt_marg,
            footnote_plain=foot,
            hide_period_in_header=True,
            hide_footnote=True,
            hide_resultado_margem_block=True,
            show_resultado_discreto=show_resultado_discreto,
            rg_header_subtitle=rg_header_subtitle,
        ),
        unsafe_allow_html=True,
    )


def _fdl_fat_min_inject_ui_styles() -> None:
    """Textos auxiliares NF-first mais discretos (apenas UI)."""
    st.markdown(
        dedent(
            """
            <style>
            .fdl-fat-min-aside {
              color: #4b5563;
              font-size: 0.76rem;
              line-height: 1.48;
              margin: 0 0 9px 0;
              max-width: 52rem;
            }
            .fdl-fat-min-aside strong { color: #1e293b; font-weight: 600; }
            .fdl-fat-min-aside--tight { margin-bottom: 4px; font-size: 0.72rem; color: #64748b; }
            .fdl-fat-min-aside--recorte {
              font-size: 0.72rem;
              color: #64748b;
              line-height: 1.42;
              margin: 0 0 6px 0;
            }
            .fdl-fat-min-aside--recorte strong { color: #334155; font-weight: 600; }
            .fdl-fat-min-table-cap {
              color: #64748b;
              font-size: 0.72rem;
              line-height: 1.4;
              margin: 0 0 10px 0;
              max-width: 52rem;
            }
            .fdl-fat-min-nf-h {
              margin: 0 0 0.15rem 0;
              font-size: 1.05rem;
              font-weight: 700;
              letter-spacing: -0.02em;
              color: #0f172a;
              line-height: 1.25;
            }
            .fdl-fat-min-nf-sub {
              margin: 0 0 0.5rem 0;
              font-size: 0.72rem;
              font-weight: 500;
              color: #64748b;
              letter-spacing: 0.02em;
              text-transform: uppercase;
            }
            .fdl-fat-min-vsp-sm {
              display: block;
              height: 0.5rem;
              min-height: 0.5rem;
            }
            .fdl-fat-min-vsp-md {
              display: block;
              height: 1rem;
              min-height: 1rem;
            }
            .fdl-fat-min-vsp-lg {
              display: block;
              height: 1.35rem;
              min-height: 1.35rem;
            }
            hr.fdl-divider-simple {
              border: none;
              border-top: 1px solid #e2e8f0;
              margin: 1.65rem 0 1.15rem 0;
              opacity: 0.82;
            }
            .fdl-fat-dre-wrap {
              max-width: min(48rem, 100%);
              width: 100%;
              margin: 0;
              padding: 16px 18px 18px 18px;
              border: 1px solid #b8c4d0;
              border-radius: 14px;
              background: linear-gradient(165deg, #f8fafc 0%, #f1f5f9 48%, #eef2f6 100%);
              box-shadow: 0 2px 10px rgba(15, 23, 42, 0.055);
              font-family: var(--font, "Source Sans Pro", sans-serif);
              box-sizing: border-box;
            }
            .fdl-fat-dre-title {
              font-size: 1.02rem;
              font-weight: 800;
              color: #0f172a;
              margin: 0 0 4px 0;
              letter-spacing: -0.025em;
            }
            .fdl-fat-dre-sub {
              font-size: 0.7rem;
              font-weight: 500;
              color: #64748b;
              margin: 0 0 12px 0;
              line-height: 1.4;
            }
            .fdl-fat-dre-block-h {
              font-size: 0.66rem;
              font-weight: 700;
              text-transform: uppercase;
              letter-spacing: 0.09em;
              color: #475569;
              margin: 16px 0 6px 0;
            }
            .fdl-fat-dre-block-h:first-of-type { margin-top: 0; }
            .fdl-fat-dre-block-h--a {
              margin-bottom: 8px;
            }
            .fdl-fat-dre-a-shell {
              border: 1px solid #c5d0da;
              border-radius: 10px;
              padding: 8px 14px 10px 14px;
              margin: 0 0 6px 0;
              background: rgba(255, 255, 255, 0.88);
            }
            .fdl-fat-dre-block-h--enc {
              margin-top: 14px;
              margin-bottom: 8px;
            }
            .fdl-fat-dre-row {
              display: grid;
              grid-template-columns: minmax(0, 1fr) minmax(8rem, max-content);
              column-gap: 1.4rem;
              align-items: baseline;
              padding: 9px 0;
              border-bottom: 1px solid #e2e8f0;
              font-size: 0.875rem;
              color: #1e293b;
            }
            .fdl-fat-dre-a-shell .fdl-fat-dre-row--lead {
              border-radius: 6px;
            }
            .fdl-fat-dre-row--lead {
              padding: 12px 2px 13px 2px;
              margin-bottom: 0;
              border-bottom: 2px solid #d8dde4;
              background: linear-gradient(180deg, #ffffff 0%, #f7f8fa 100%);
            }
            .fdl-fat-dre-row--lead .fdl-fat-dre-lab {
              font-weight: 600;
              font-size: 0.96rem;
              color: #0f172a;
              letter-spacing: -0.018em;
            }
            .fdl-fat-dre-row--lead .fdl-fat-dre-val {
              font-weight: 700;
              font-size: 1.08rem;
              color: #0f172a;
            }
            .fdl-fat-dre-row--ref {
              margin-top: 0;
              padding-top: 5px;
              padding-bottom: 6px;
              border-top: none;
              border-bottom: 1px dotted #e2e6ec;
              background: transparent;
            }
            .fdl-fat-dre-row--ref .fdl-fat-dre-lab {
              color: #c5ccd6;
              font-weight: 400;
              font-size: 0.625rem;
              font-style: italic;
              letter-spacing: 0.03em;
            }
            .fdl-fat-dre-row--ref .fdl-fat-dre-val {
              color: #c5ccd6;
              font-weight: 400;
              font-size: 0.625rem;
            }
            .fdl-fat-dre-row--bridge {
              margin-top: 5px;
              margin-bottom: 2px;
              padding: 11px 12px 12px 12px;
              border-bottom: none;
              background: #e2e8f0;
              border-radius: 8px;
              border: 1px solid #cbd5e1;
            }
            .fdl-fat-dre-row--bridge .fdl-fat-dre-lab {
              font-weight: 600;
              font-size: 0.9rem;
              color: #111827;
            }
            .fdl-fat-dre-row--bridge .fdl-fat-dre-val {
              font-weight: 700;
              font-size: 0.95rem;
              color: #0f172a;
            }
            .fdl-fat-dre-foot-a-note {
              font-size: 0.55rem;
              color: #d8dee6;
              line-height: 1.32;
              margin: 6px 0 0 0;
              padding: 5px 0 2px 10px;
              border-left: 2px solid #eef2f6;
              max-width: none;
            }
            .fdl-fat-dre-row--enc {
              padding: 11px 4px;
              border-bottom: 1px solid #f1f2f4;
            }
            .fdl-fat-dre-row--enc-last {
              border-bottom: 1px solid #e2e6ed;
              padding-bottom: 12px;
            }
            .fdl-fat-dre-lab {
              min-width: 0;
              line-height: 1.38;
            }
            .fdl-fat-dre-val {
              font-variant-numeric: tabular-nums;
              text-align: right;
              white-space: nowrap;
              font-weight: 500;
              color: #111827;
              justify-self: end;
            }
            .fdl-fat-dre-val--out {
              color: #525a63;
              font-weight: 600;
              letter-spacing: 0.02em;
              font-variant-numeric: tabular-nums;
            }
            .fdl-fat-dre-foot {
              font-size: 0.63rem;
              color: #9ca3af;
              line-height: 1.4;
              margin: 3px 0 0 0;
              max-width: 42rem;
            }
            .fdl-fat-dre-foot--inline {
              margin-top: 2px;
              margin-bottom: 2px;
            }
            .fdl-fat-dre-foot--final {
              margin-top: 30px;
              margin-bottom: 0;
              padding-top: 4px;
              font-size: 0.47rem;
              color: #dde2e9;
              line-height: 1.34;
            }
            .fdl-fat-dre-close {
              margin-top: 14px;
              border-radius: 12px;
              border: 1px solid #94a3b8;
              background: linear-gradient(165deg, #ffffff 0%, #f1f5f9 55%, #e8edf3 100%);
              overflow: hidden;
              box-shadow:
                0 2px 4px rgba(15, 23, 42, 0.05),
                0 8px 22px rgba(15, 23, 42, 0.07);
            }
            .fdl-fat-dre-row--result {
              display: grid;
              grid-template-columns: minmax(0, 1fr) minmax(8rem, max-content);
              column-gap: 1.4rem;
              align-items: baseline;
              padding: 17px 20px;
              border-bottom: 1px solid #cdd2d9;
              background: transparent;
            }
            .fdl-fat-dre-row--result .fdl-fat-dre-lab {
              font-weight: 600;
              font-size: 0.95rem;
              color: #111827;
              letter-spacing: -0.012em;
            }
            .fdl-fat-dre-row--result .fdl-fat-dre-val {
              font-weight: 700;
              font-size: 1.2rem;
              color: #0f172a;
              font-variant-numeric: tabular-nums;
              text-align: right;
              justify-self: end;
            }
            .fdl-fat-dre-row--margem {
              display: grid;
              grid-template-columns: minmax(0, 1fr) minmax(8rem, max-content);
              column-gap: 1.4rem;
              align-items: baseline;
              padding: 13px 20px 15px 20px;
              background: rgba(255, 255, 255, 0.62);
            }
            .fdl-fat-dre-row--margem .fdl-fat-dre-lab {
              font-weight: 500;
              font-size: 0.83rem;
              color: #64748b;
            }
            .fdl-fat-dre-row--margem .fdl-fat-dre-val {
              font-weight: 600;
              font-size: 1rem;
              color: #334155;
              font-variant-numeric: tabular-nums;
              text-align: right;
              justify-self: end;
            }
            /* Faturamento & DRE: filtros mais compactos (só quando esta vista injeta o bloco) */
            section[data-testid="stMain"] [data-baseweb="select"] > div {
              min-height: 38px !important;
            }
            section[data-testid="stMain"] [data-baseweb="input"] > div {
              min-height: 38px !important;
            }
            section[data-testid="stMain"] [data-baseweb="input"] input {
              padding-top: 0.42rem !important;
              padding-bottom: 0.42rem !important;
              font-size: 0.875rem !important;
            }
            section[data-testid="stMain"] [data-testid="stMultiSelect"] [data-baseweb="select"] > div {
              min-height: 38px !important;
            }
            /* Cobertura comercial (painel NF-first): colapsável + badge de alerta */
            .fdl-fat-cobertura {
              margin: 1rem 0;
              border: 1px solid var(--fdl-neutral-200, #e2e8f0);
              border-radius: 8px;
              background: var(--fdl-neutral-50, #f8fafc);
            }
            .fdl-fat-cobertura summary {
              padding: 0.75rem 1rem;
              cursor: pointer;
              display: flex;
              align-items: center;
              gap: 0.75rem;
              flex-wrap: wrap;
              font-weight: 500;
              color: var(--fdl-neutral-700, #374151);
              list-style: none;
            }
            .fdl-fat-cobertura summary::-webkit-details-marker {
              display: none;
            }
            .fdl-fat-cobertura summary::marker {
              content: "";
            }
            .fdl-fat-cobertura summary::before {
              content: "▶";
              font-size: 0.7rem;
              transition: transform 0.2s ease;
              flex-shrink: 0;
            }
            .fdl-fat-cobertura[open] summary::before {
              transform: rotate(90deg);
            }
            .fdl-fat-cobertura > div {
              padding: 1rem;
              border-top: 1px solid var(--fdl-neutral-200, #e2e8f0);
            }
            .fdl-fat-cobertura-caption {
              font-size: 0.75rem;
              color: #64748b;
              line-height: 1.45;
              margin: 0 0 0.85rem 0;
              max-width: 52rem;
            }
            .fdl-fat-cobertura-caption strong {
              color: #334155;
              font-weight: 600;
            }
            .fdl-fat-cobertura-grid {
              display: grid;
              grid-template-columns: repeat(3, minmax(0, 1fr));
              gap: 0.75rem 1rem;
            }
            .fdl-fat-cobertura-cell {
              min-width: 0;
            }
            .fdl-fat-cobertura-lab {
              font-size: 0.68rem;
              font-weight: 600;
              color: #64748b;
              text-transform: uppercase;
              letter-spacing: 0.04em;
              margin-bottom: 0.25rem;
              line-height: 1.25;
            }
            .fdl-fat-cobertura-val {
              font-size: 1.05rem;
              font-weight: 700;
              font-variant-numeric: tabular-nums;
              color: #0f172a;
            }
            .fdl-fat-cobertura--rg-premium {
              margin: 0;
              padding: 14px 16px;
              border: 1px solid var(--fdl-neutral-200, #e2e8f0);
              border-radius: 12px;
              background: #ffffff;
              box-sizing: border-box;
            }
            .fdl-fat-cob-rg-head {
              display: flex;
              justify-content: space-between;
              align-items: center;
              gap: 10px;
              flex-wrap: wrap;
              margin-bottom: 12px;
            }
            .fdl-fat-cob-rg-title {
              font-weight: 600;
              font-size: 0.95rem;
              color: var(--fdl-neutral-800, #0f172a);
            }
            .fdl-fat-cob-rg-meta {
              display: flex;
              align-items: center;
              gap: 8px;
            }
            .fdl-fat-cob-rg-info {
              cursor: help;
              font-size: 1rem;
              line-height: 1;
              opacity: 0.6;
            }
            .fdl-fat-cob-v2-grid {
              display: grid;
              grid-template-columns: 1fr 1fr;
              gap: 10px 14px;
              margin-top: 4px;
            }
            .fdl-fat-cob-v2-cell {
              min-width: 0;
              border: 1px solid var(--fdl-neutral-200, #e5e7eb);
              border-radius: 10px;
              padding: 10px 12px;
              box-sizing: border-box;
              background: #fafafa;
            }
            .fdl-fat-cob-v2-lab {
              font-size: 0.68rem;
              font-weight: 600;
              color: #64748b;
              line-height: 1.35;
              margin-bottom: 6px;
            }
            .fdl-fat-cob-v2-lab2 {
              font-size: 0.62rem;
              font-weight: 500;
              color: #94a3b8;
            }
            .fdl-fat-cob-v2-val {
              font-size: 1.05rem;
              font-weight: 700;
              font-variant-numeric: tabular-nums;
              color: #0f172a;
            }
            @media (max-width: 520px) {
              .fdl-fat-cob-v2-grid { grid-template-columns: 1fr; }
            }
            .fdl-fat-cobertura-admin {
              margin-top: 0.75rem;
              font-size: 0.72rem;
              color: #64748b;
            }
            .fdl-cobertura-legenda {
              font-size: 0.8rem;
              color: var(--fdl-neutral-600, #4b5563);
              margin-bottom: 1rem;
              line-height: 1.4;
              max-width: 52rem;
            }
            .fdl-cobertura-cta {
              font-size: 0.8rem;
              color: var(--fdl-warning-700, #b45309);
              background: var(--fdl-warning-50, #fffbeb);
              padding: 0.5rem 0.75rem;
              border-radius: 4px;
              margin-top: 1rem;
              max-width: 52rem;
            }
            .fdl-fat-cobertura--embedded {
              margin-top: 0.5rem;
              padding-top: 0.85rem;
              border-top: 1px solid var(--fdl-neutral-200, #e2e8f0);
            }
            .fdl-fat-cobertura-embed-title {
              font-size: 0.78rem;
              font-weight: 700;
              text-transform: uppercase;
              letter-spacing: 0.06em;
              color: #64748b;
              margin: 0 0 0.65rem 0;
              display: flex;
              align-items: center;
              gap: 0.5rem;
              flex-wrap: wrap;
            }
            /* Filtros + Base fiscal — vista mínima Faturamento & DRE */
            .fdl-fat-filtros-periodo-tit {
              font-size: 0.72rem;
              font-weight: 700;
              text-transform: uppercase;
              letter-spacing: 0.06em;
              color: #64748b;
              margin: 0.5rem 0 0.35rem 0;
            }
            .fdl-base-fiscal-card {
              border: 1px solid var(--fdl-neutral-200, #e2e8f0);
              border-radius: 12px;
              background: var(--fdl-neutral-50, #f8fafc);
              overflow: hidden;
              margin: 0.75rem 0 0.35rem 0;
            }
            .fdl-base-fiscal-header {
              display: flex;
              justify-content: space-between;
              align-items: center;
              padding: 0.75rem 1.25rem;
              border-bottom: 1px solid var(--fdl-neutral-200, #e2e8f0);
              background: #ffffff;
              flex-wrap: wrap;
              gap: 0.35rem 1rem;
            }
            .fdl-base-fiscal-title {
              font-weight: 600;
              font-size: 1rem;
              color: var(--fdl-neutral-800, #1e293b);
            }
            .fdl-base-fiscal-periodo {
              font-size: 0.875rem;
              color: var(--fdl-neutral-500, #64748b);
              font-family: ui-monospace, "Cascadia Code", "Segoe UI Mono", monospace;
            }
            .fdl-base-fiscal-body {
              display: flex;
              justify-content: space-around;
              flex-wrap: wrap;
              gap: 1rem 2rem;
              padding: 1.35rem 1.25rem;
              background: #ffffff;
            }
            .fdl-base-fiscal-kpi {
              text-align: center;
              min-width: 10rem;
            }
            .fdl-base-fiscal-valor {
              display: block;
              font-size: 1.5rem;
              font-weight: 700;
              color: #1e293b;
              font-family: ui-monospace, "Cascadia Code", "Segoe UI Mono", monospace;
              font-variant-numeric: tabular-nums;
              line-height: 1.15;
            }
            .fdl-base-fiscal-count {
              display: block;
              font-size: 1.25rem;
              font-weight: 600;
              color: #1e293b;
              font-family: ui-monospace, "Cascadia Code", "Segoe UI Mono", monospace;
              font-variant-numeric: tabular-nums;
              line-height: 1.15;
            }
            .fdl-base-fiscal-label {
              display: block;
              font-size: 0.75rem;
              text-transform: uppercase;
              letter-spacing: 0.05em;
              color: #64748b;
              margin-top: 0.35rem;
              font-weight: 600;
            }
            .fdl-base-fiscal-footer {
              padding: 0.55rem 1.25rem;
              border-top: 1px solid var(--fdl-neutral-200, #e2e8f0);
              background: var(--fdl-neutral-50, #f8fafc);
            }
            .fdl-base-fiscal-contexto {
              font-size: 0.75rem;
              color: var(--fdl-neutral-500, #64748b);
              line-height: 1.4;
            }
            /* Cabeçalho da página — Resultado Gerencial (vista mínima) */
            .fdl-page-breadcrumb {
              font-size: 0.75rem;
              font-weight: 600;
              letter-spacing: 0.06em;
              text-transform: uppercase;
              color: var(--fdl-neutral-400, #94a3b8);
              margin: 0 0 0.35rem 0;
            }
            .fdl-page-header {
              padding: 1.25rem 0 0.85rem 0;
              margin-bottom: 0.35rem;
              border-bottom: 1px solid var(--fdl-neutral-200, #e2e8f0);
            }
            .fdl-page-header-main {
              margin-bottom: 0.65rem;
            }
            .fdl-page-title {
              font-size: 1.85rem;
              font-weight: 700;
              color: var(--fdl-neutral-800, #1e293b);
              margin: 0 0 0.35rem 0;
              font-family: Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
              letter-spacing: -0.02em;
              line-height: 1.15;
            }
            .fdl-page-title-hint {
              font-size: 1rem;
              font-weight: 500;
              opacity: 0.55;
              cursor: help;
              margin-left: 0.35rem;
              vertical-align: middle;
            }
            .fdl-page-subtitle {
              font-size: 0.95rem;
              color: var(--fdl-neutral-500, #64748b);
              margin: 0;
              font-weight: 400;
              line-height: 1.45;
              max-width: 48rem;
            }
            .fdl-page-header-meta {
              display: flex;
              justify-content: flex-start;
              align-items: center;
              flex-wrap: wrap;
              gap: 0.35rem 1rem;
            }
            .fdl-page-updated {
              font-size: 0.8rem;
              color: var(--fdl-neutral-400, #94a3b8);
              font-family: ui-monospace, "Cascadia Code", "Segoe UI Mono", monospace;
              font-variant-numeric: tabular-nums;
            }
            .fdl-badge {
              font-size: 0.75rem;
              padding: 2px 8px;
              border-radius: 4px;
              font-weight: 500;
            }
            .fdl-badge-warning {
              background: #fef3c7;
              color: #92400e;
            }
            .fdl-badge-danger {
              background: #fee2e2;
              color: #991b1b;
            }
            </style>
            """
        )
        + (fat_dre_premium_css() if _FAT_DRE_UI_V2 else ""),
        unsafe_allow_html=True,
    )


def _fdl_fat_section_rule(label: str) -> None:
    """Divisor com rótulo (Faturamento & DRE premium)."""
    if _FAT_DRE_UI_V2:
        st.markdown(faturamento_section_rule_html(label), unsafe_allow_html=True)
    else:
        st.divider()
        st.caption(label)


def _fdl_fat_divider_simple() -> None:
    """Linha divisória discreta (sem rótulo) entre blocos do painel NF-first."""
    st.markdown(
        '<hr class="fdl-divider-simple" aria-hidden="true">',
        unsafe_allow_html=True,
    )


def _fdl_fat_min_vsp(*, size: str = "md") -> None:
    st.markdown(
        f'<div class="fdl-fat-min-vsp-{size}" aria-hidden="true"></div>',
        unsafe_allow_html=True,
    )


def _fdl_fat_min_aside(
    html_body: str, *, tight: bool = False, recorte: bool = False
) -> None:
    cls = "fdl-fat-min-aside"
    if tight:
        cls += " fdl-fat-min-aside--tight"
    if recorte:
        cls += " fdl-fat-min-aside--recorte"
    st.markdown(f'<div class="{cls}">{html_body}</div>', unsafe_allow_html=True)


def _fdl_cp_inject_panel_styles() -> None:
    """CSS compartilhado com KPIs Faturamento (fdl-fat-kpi-*) + refinamento Comercial & pedidos."""
    st.markdown(
        dedent(
            """
            <style>
            .fdl-fat-kpi-shell {
              font-family: var(--font, "Source Sans Pro", sans-serif);
              margin: 0 0 20px 0;
            }
            .fdl-fat-kpi-row {
              display: flex;
              flex-wrap: wrap;
              gap: 12px;
              margin-bottom: 14px;
            }
            .fdl-fat-kpi-row--secondary {
              gap: 10px;
              margin-bottom: 0;
            }
            .fdl-fat-kpi-card {
              flex: 1 1 0;
              min-width: 148px;
              background: #ffffff;
              border: 1px solid #cbd5e1;
              border-radius: 12px;
              padding: 14px 16px;
              box-sizing: border-box;
              box-shadow: 0 2px 6px rgba(15, 23, 42, 0.055);
            }
            .fdl-fat-kpi-card--primary {
              padding: 16px 18px;
              min-width: 158px;
            }
            .fdl-fat-kpi-card--primary.fdl-fat-kpi-card--accent {
              background: linear-gradient(180deg, #f8fafc 0%, #f1f5f9 100%);
              border-color: #94a3b8;
              box-shadow: 0 2px 8px rgba(15, 23, 42, 0.07);
            }
            .fdl-fat-kpi-card--secondary {
              padding: 11px 13px;
              min-width: 118px;
            }
            .fdl-fat-kpi-label {
              font-size: 0.68rem;
              font-weight: 600;
              color: #475569;
              line-height: 1.3;
              margin: 0 0 8px 0;
              letter-spacing: 0.04em;
              text-transform: uppercase;
            }
            .fdl-fat-kpi-card--secondary .fdl-fat-kpi-label {
              font-size: 0.625rem;
              margin-bottom: 7px;
              letter-spacing: 0.035em;
            }
            .fdl-fat-kpi-value {
              font-size: 1.48rem;
              font-weight: 700;
              color: #0f172a;
              line-height: 1.12;
              font-variant-numeric: tabular-nums;
              letter-spacing: -0.03em;
            }
            .fdl-fat-kpi-card--primary .fdl-fat-kpi-value {
              font-size: 1.62rem;
              font-weight: 800;
            }
            .fdl-fat-kpi-card--primary.fdl-fat-kpi-card--accent .fdl-fat-kpi-value {
              font-size: 1.75rem;
            }
            .fdl-fat-kpi-card--secondary .fdl-fat-kpi-value {
              font-size: 1.12rem;
              font-weight: 700;
              color: #1e293b;
            }
            .fdl-cp-kpi-shell.fdl-fat-kpi-shell .fdl-fat-kpi-card {
              display: flex;
              flex-direction: column;
              align-items: stretch;
            }
            .fdl-cp-kpi-shell.fdl-fat-kpi-shell .fdl-fat-kpi-label {
              font-size: 0.66rem;
              font-weight: 600;
              color: #475569;
              margin: 0 0 10px 0;
              letter-spacing: 0.04em;
              text-transform: uppercase;
            }
            .fdl-cp-kpi-shell.fdl-fat-kpi-shell .fdl-fat-kpi-card--secondary .fdl-fat-kpi-label {
              font-size: 0.625rem;
              margin-bottom: 8px;
              letter-spacing: 0.035em;
            }
            .fdl-cp-kpi-shell.fdl-fat-kpi-shell .fdl-fat-kpi-value {
              text-align: right;
              align-self: stretch;
            }
            .fdl-cp-kpi-shell.fdl-fat-kpi-shell .fdl-fat-kpi-card--primary .fdl-fat-kpi-value {
              font-size: 1.82rem;
              font-weight: 800;
              letter-spacing: -0.035em;
            }
            .fdl-cp-kpi-shell.fdl-fat-kpi-shell .fdl-fat-kpi-card--primary.fdl-fat-kpi-card--accent .fdl-fat-kpi-value {
              font-size: 1.95rem;
              color: #0f172a;
            }
            .fdl-cp-kpi-shell.fdl-fat-kpi-shell .fdl-fat-kpi-card--secondary .fdl-fat-kpi-value {
              font-size: 1.14rem;
              font-weight: 700;
              color: #1e293b;
            }
            .fdl-cp-kpi-shell.fdl-fat-kpi-shell .fdl-fat-kpi-card--primary {
              padding: 20px 22px 22px 22px;
              border-left: 3px solid #0f172a;
            }
            .fdl-cp-kpi-shell.fdl-fat-kpi-shell .fdl-fat-kpi-card--secondary {
              padding: 12px 14px;
            }
            div[data-testid="stVerticalBlockBorderWrapper"]:has(.fdl-cp-title-main) {
              border-color: #94a3b8 !important;
              box-shadow: 0 3px 16px rgba(15, 23, 42, 0.075) !important;
            }
            div[data-testid="stVerticalBlockBorderWrapper"]:has(.fdl-cp-title-decision) {
              border-color: #cbd5e1 !important;
              box-shadow: 0 2px 12px rgba(15, 23, 42, 0.06) !important;
            }
            section[data-testid="stMain"] [data-baseweb="select"] > div {
              min-height: 38px !important;
            }
            section[data-testid="stMain"] [data-baseweb="input"] > div {
              min-height: 38px !important;
            }
            section[data-testid="stMain"] [data-baseweb="input"] input {
              padding-top: 0.42rem !important;
              padding-bottom: 0.42rem !important;
              font-size: 0.875rem !important;
            }
            section[data-testid="stMain"] [data-testid="stMultiSelect"] [data-baseweb="select"] > div {
              min-height: 38px !important;
            }
            .fdl-cp-filtros-h {
              font-size: 1rem;
              font-weight: 700;
              color: #0f172a;
              margin: 0 0 5px 0;
              letter-spacing: -0.02em;
            }
            .fdl-cp-caption {
              font-size: 0.74rem;
              font-weight: 500;
              color: #64748b;
              line-height: 1.42;
              margin: 0 0 8px 0;
            }
            .fdl-cp-caption strong { color: #334155; font-weight: 600; }
            .fdl-cp-title-main {
              font-size: 1.38rem;
              font-weight: 800;
              color: #0f172a;
              margin: 0 0 5px 0;
              letter-spacing: -0.032em;
              line-height: 1.2;
            }
            .fdl-cp-title-sub {
              font-size: 0.74rem;
              font-weight: 500;
              color: #6b7785;
              margin: 0 0 14px 0;
              line-height: 1.45;
              max-width: 46rem;
            }
            .fdl-cp-title-sec {
              font-size: 0.875rem;
              font-weight: 600;
              color: #94a3b8;
              margin: 0 0 3px 0;
              letter-spacing: 0.02em;
              text-transform: uppercase;
            }
            .fdl-cp-title-sec-note {
              font-size: 0.7rem;
              font-weight: 400;
              color: #b8c2cc;
              margin: 0 0 10px 0;
              line-height: 1.4;
            }
            .fdl-cp-title-decision {
              font-size: 1.2rem;
              font-weight: 800;
              color: #0f172a;
              margin: 0 0 4px 0;
              letter-spacing: -0.028em;
              line-height: 1.22;
            }
            .fdl-cp-exec {
              background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%);
              border: 1px solid #cbd5e1;
              border-radius: 12px;
              padding: 14px 16px 14px 16px;
              margin: 0 0 14px 0;
              box-shadow: 0 1px 4px rgba(15, 23, 42, 0.04);
            }
            .fdl-cp-exec-h {
              font-size: 0.625rem;
              font-weight: 700;
              text-transform: uppercase;
              letter-spacing: 0.1em;
              color: #64748b;
              margin: 0 0 9px 0;
            }
            .fdl-cp-exec-row {
              display: flex;
              flex-wrap: wrap;
              gap: 10px 22px;
              margin-bottom: 9px;
              font-size: 0.875rem;
              font-weight: 500;
              color: #475569;
              line-height: 1.5;
            }
            .fdl-cp-exec-row strong { color: #0f172a; font-weight: 700; font-variant-numeric: tabular-nums; }
            .fdl-cp-exec-top {
              font-size: 0.8rem;
              font-weight: 500;
              color: #475569;
              margin: 6px 0 0 0;
              padding-top: 10px;
              border-top: 1px solid #e2e8f0;
              line-height: 1.5;
            }
            .fdl-cp-exec-top strong { color: #0f172a; font-weight: 700; }
            .fdl-cp-trend-summary {
              display: flex;
              flex-wrap: wrap;
              gap: 10px 22px;
              align-items: baseline;
              font-size: 0.8rem;
              font-weight: 500;
              color: #475569;
              margin: 0 0 12px 0;
              padding: 12px 15px;
              background: linear-gradient(180deg, #f8fafc 0%, #f1f5f9 100%);
              border-radius: 10px;
              border: 1px solid #cbd5e1;
            }
            .fdl-cp-trend-summary span strong {
              color: #0f172a;
              font-weight: 700;
              font-variant-numeric: tabular-nums;
            }
            </style>
            """
        ),
        unsafe_allow_html=True,
    )


def _render_comercial_pedidos_kpi_cards(kpis: dict[str, float | int]) -> None:
    """Cards de KPI no mesmo espírito visual do painel Faturamento & DRE."""

    def _card(
        label: str,
        value: str,
        *,
        tier: str,
        accent: bool = False,
        title: str | None = None,
    ) -> str:
        classes = f"fdl-fat-kpi-card fdl-fat-kpi-card--{tier}"
        if accent:
            classes += " fdl-fat-kpi-card--accent"
        tattr = ""
        if title:
            tattr = f' title="{html.escape(title, quote=True)}"'
        return (
            f'<div class="{classes}"{tattr}>'
            f'<div class="fdl-fat-kpi-label">{html.escape(label)}</div>'
            f'<div class="fdl-fat-kpi-value">{html.escape(value)}</div>'
            "</div>"
        )

    vcom = _fmt_brl_ptbr_celula(float(kpis["valor_comercial_lista"])) or "R$ 0,00"
    qtd = _fmt_int_ptbr(int(round(float(kpis["quantidade_total"]))))
    ped = _fmt_int_ptbr(int(kpis["pedidos_atendidos_distintos"]))
    skus = _fmt_int_ptbr(int(kpis["skus_distintos"]))
    primary = "".join(
        [
            _card(
                "Receita (lista)",
                vcom,
                tier="primary",
                accent=True,
                title="Valor comercial: Σ Vl_Venda (materializado) ou lista×qtd · só pedidos atendidos · sem NF.",
            ),
        ]
    )
    secondary = "".join(
        [
            _card("Quantidade (unidades)", qtd, tier="secondary"),
            _card(
                "Pedidos atendidos (distintos)",
                ped,
                tier="secondary",
                title="Pedidos únicos (multiloja/org), não linhas.",
            ),
            _card("SKUs distintos", skus, tier="secondary"),
        ]
    )
    st.markdown(
        f'<div class="fdl-fat-kpi-shell fdl-cp-kpi-shell">'
        f'<div class="fdl-fat-kpi-row fdl-fat-kpi-row--primary">{primary}</div>'
        f'<div class="fdl-fat-kpi-row fdl-fat-kpi-row--secondary">{secondary}</div>'
        f"</div>",
        unsafe_allow_html=True,
    )


def _comercial_abc_valor_table_styler(abc_v: pd.DataFrame):
    """Tabela ABC valor só para UI: moeda e % pt-BR, alinhamento, sem alterar ``abc_v`` de origem."""
    if abc_v.empty:
        try:
            return abc_v.style.hide(axis="index")
        except Exception:
            return abc_v.style
    d = abc_v.copy()
    if "Valor comercial (lista)" in d.columns:
        d["Valor comercial (lista)"] = d["Valor comercial (lista)"].map(
            lambda x: _fmt_brl_ptbr_celula(x) or "—"
        )
    if "Part %" in d.columns:
        part_raw = pd.to_numeric(d["Part %"], errors="coerce").fillna(0.0) * 100.0
        d["Part %"] = part_raw.map(_fmt_pct_ptbr_1)
    if "Acum %" in d.columns:
        acum_raw = pd.to_numeric(d["Acum %"], errors="coerce").fillna(0.0) * 100.0
        d["Acum %"] = acum_raw.map(_fmt_pct_ptbr_1)
    sty = d.style.hide(axis="index")
    _right = [c for c in ("Valor comercial (lista)", "Part %", "Acum %") if c in d.columns]
    if _right:
        sty = sty.set_properties(
            subset=_right,
            **{"text-align": "right", "font-variant-numeric": "tabular-nums"},
        )
    if "Produto" in d.columns:
        sty = sty.set_properties(
            subset=["Produto"],
            **{
                "text-align": "left",
                "max-width": "24rem",
                "white-space": "nowrap",
                "overflow": "hidden",
                "text-overflow": "ellipsis",
            },
        )
    if "SKU" in d.columns:
        sty = sty.set_properties(
            subset=["SKU"],
            **{"text-align": "left", "font-variant-numeric": "tabular-nums"},
        )
    if "Classe" in d.columns:
        sty = sty.set_properties(subset=["Classe"], **{"text-align": "center", "font-weight": "600"})
    return sty


def _comercial_abc_quantidade_table_styler(abc_q: pd.DataFrame):
    """Tabela ABC quantidade só para UI: qtd e % pt-BR."""
    if abc_q.empty:
        try:
            return abc_q.style.hide(axis="index")
        except Exception:
            return abc_q.style
    d = abc_q.copy()
    if "Quantidade" in d.columns:
        d["Quantidade"] = d["Quantidade"].map(_comercial_fmt_qtd_display)
    if "Part %" in d.columns:
        pr = pd.to_numeric(d["Part %"], errors="coerce").fillna(0.0) * 100.0
        d["Part %"] = pr.map(_fmt_pct_ptbr_1)
    if "Acum %" in d.columns:
        ar = pd.to_numeric(d["Acum %"], errors="coerce").fillna(0.0) * 100.0
        d["Acum %"] = ar.map(_fmt_pct_ptbr_1)
    sty = d.style.hide(axis="index")
    _right = [c for c in ("Quantidade", "Part %", "Acum %") if c in d.columns]
    if _right:
        sty = sty.set_properties(
            subset=_right,
            **{"text-align": "right", "font-variant-numeric": "tabular-nums"},
        )
    if "Produto" in d.columns:
        sty = sty.set_properties(
            subset=["Produto"],
            **{
                "text-align": "left",
                "max-width": "20rem",
                "white-space": "nowrap",
                "overflow": "hidden",
                "text-overflow": "ellipsis",
            },
        )
    if "SKU" in d.columns:
        sty = sty.set_properties(
            subset=["SKU"],
            **{"text-align": "left", "font-variant-numeric": "tabular-nums"},
        )
    if "Classe" in d.columns:
        sty = sty.set_properties(subset=["Classe"], **{"text-align": "center", "font-weight": "600"})
    return sty


def _comercial_abc_valor_exec_html(abc_v: pd.DataFrame) -> str:
    """Leitura executiva derivada apenas do DataFrame ABC valor (sem alterar lógica)."""
    if abc_v.empty or "Classe" not in abc_v.columns:
        return ""
    cls = abc_v["Classe"].astype(str).str.strip().str.upper()
    part = pd.to_numeric(abc_v.get("Part %"), errors="coerce").fillna(0.0)
    share_a = float(part.loc[cls.eq("A")].sum())
    share_a_pct = share_a * 100.0
    n_a = int(cls.eq("A").sum())
    n_b = int(cls.eq("B").sum())
    n_c = int(cls.eq("C").sum())
    top_lines: list[str] = []
    if "Produto" in abc_v.columns and "Valor comercial (lista)" in abc_v.columns:
        top = abc_v.head(3)
        for _, row in top.iterrows():
            lab = str(row.get("Produto", "—")).strip() or "—"
            if len(lab) > 56:
                lab = lab[:53] + "…"
            vl = row.get("Valor comercial (lista)")
            try:
                vs = _fmt_brl_ptbr_celula(float(vl)) or "—"
            except (TypeError, ValueError):
                vs = "—"
            top_lines.append(f"<strong>{html.escape(lab)}</strong> · {html.escape(vs)}")
    top_html = "<br/>".join(top_lines) if top_lines else "—"
    share_disp = _fmt_pct_ptbr_1(share_a_pct)
    return (
        '<div class="fdl-cp-exec">'
        '<div class="fdl-cp-exec-h">Leitura executiva</div>'
        '<div class="fdl-cp-exec-row">'
        f"<span>Classe <strong>A</strong> concentra <strong>{html.escape(share_disp)}</strong> "
        "da receita comercial (lista) no recorte.</span>"
        "</div>"
        '<div class="fdl-cp-exec-row">'
        f"<span>SKUs <strong>A</strong>: <strong>{html.escape(_fmt_int_ptbr(n_a))}</strong></span>"
        f"<span>· <strong>B</strong>: <strong>{html.escape(_fmt_int_ptbr(n_b))}</strong></span>"
        f"<span>· <strong>C</strong>: <strong>{html.escape(_fmt_int_ptbr(n_c))}</strong></span>"
        "</div>"
        '<div class="fdl-cp-exec-top">'
        "<strong>Top 3 por valor comercial</strong><br/>"
        f"{top_html}"
        "</div></div>"
    )


def _comercial_trend_summary_html(trend_tbl: pd.DataFrame) -> str:
    if trend_tbl.empty or "Tendência" not in trend_tbl.columns:
        return ""
    tnorm = trend_tbl["Tendência"].fillna("").astype(str).str.strip().str.casefold()
    vc = tnorm.value_counts()
    cresc = int(vc.get("crescente", 0))
    decr = int(vc.get("decrescente", 0))
    est = int(vc.get("estável", 0))
    ins = int(vc.get("insuficiente para tendência", 0))
    _ni = _fmt_int_ptbr
    parts = [
        f"<span><strong>{html.escape(_ni(cresc))}</strong> crescente</span>",
        f"<span><strong>{html.escape(_ni(decr))}</strong> decrescente</span>",
        f"<span><strong>{html.escape(_ni(est))}</strong> estável</span>",
        f"<span><strong>{html.escape(_ni(ins))}</strong> volume insuficiente</span>",
    ]
    if "Sugestão de compra" in trend_tbl.columns:
        s = trend_tbl["Sugestão de compra"].fillna("").astype(str).str.strip().str.casefold()
        n_prio = int(s.str.contains("priorizar", na=False).sum())
        n_red = int(s.str.contains("reduzir", na=False).sum())
        n_caut = int(s.str.contains("evitar", na=False).sum())
        parts.append(
            f"<span><strong>{html.escape(_ni(n_prio))}</strong> priorizar · "
            f"<strong>{html.escape(_ni(n_red))}</strong> reduzir · "
            f"<strong>{html.escape(_ni(n_caut))}</strong> cautela</span>"
        )
    return '<div class="fdl-cp-trend-summary">' + "".join(parts) + "</div>"


def _cp_trend_cell_style(v: object) -> str:
    _base = (
        "text-align: left; font-variant-numeric: tabular-nums; padding: 6px 10px; "
        "border-radius: 6px; border-left: 3px solid transparent; "
    )
    t = str(v).strip().casefold()
    if t == "crescente":
        return _base + "background-color: #ecfdf5; color: #14532d; font-weight: 700; border-left-color: #16a34a"
    if t == "decrescente":
        return _base + "background-color: #fef2f2; color: #7f1d1d; font-weight: 700; border-left-color: #dc2626"
    if t == "estável":
        return _base + "background-color: #f1f5f9; color: #1e293b; font-weight: 600; border-left-color: #64748b"
    if t == "insuficiente para tendência":
        return _base + "background-color: #f8fafc; color: #64748b; font-weight: 500; border-left-color: #cbd5e1"
    return _base + "font-weight: 500; color: #334155"


def _cp_sug_cell_style(v: object) -> str:
    _base = "text-align: left; padding: 6px 10px; border-radius: 6px; border-left: 3px solid transparent; "
    s = str(v).strip().casefold()
    if "priorizar" in s:
        return _base + "background-color: #ecfdf5; color: #14532d; font-weight: 700; border-left-color: #16a34a"
    if "reduzir" in s:
        return _base + "background-color: #fef2f2; color: #7f1d1d; font-weight: 700; border-left-color: #dc2626"
    if "evitar" in s:
        return _base + "background-color: #fffbeb; color: #92400e; font-weight: 700; border-left-color: #d97706"
    if "manter" in s:
        return _base + "background-color: #f1f5f9; color: #334155; font-weight: 600; border-left-color: #64748b"
    if "testar" in s:
        return _base + "background-color: #f8fafc; color: #334155; font-weight: 600; border-left-color: #94a3b8"
    return _base + "font-weight: 500; color: #334155"


def _comercial_trend_styler(trend_tbl: pd.DataFrame):
    """Formatação pt-BR nas colunas numéricas + realce existente em Tendência / Sugestão."""
    if trend_tbl.empty:
        try:
            return trend_tbl.style.hide(axis="index")
        except Exception:
            return trend_tbl.style

    df = trend_tbl.copy()
    qty_cols = [c for c in df.columns if str(c).startswith("Qtd mês")]
    val_cols = [c for c in df.columns if str(c).startswith("Valor lista")]
    for c in qty_cols:
        df[c] = df[c].map(_comercial_fmt_qtd_display)
    for c in val_cols:
        df[c] = df[c].map(lambda x: _fmt_brl_ptbr_celula(x) or "—")

    sty = df.style
    try:
        sty = sty.hide(axis="index")
    except (TypeError, ValueError, AttributeError):
        try:
            sty = sty.hide_index()
        except Exception:
            pass

    def _elem_map(sty_obj: object, fn: object, subset: list[str]) -> object:
        m = getattr(sty_obj, "map", None) or getattr(sty_obj, "applymap", None)
        if m is None:
            return sty_obj
        return m(fn, subset=subset)

    _right = [c for c in qty_cols + val_cols if c in df.columns]
    if _right:
        sty = sty.set_properties(
            subset=_right,
            **{"text-align": "right", "font-variant-numeric": "tabular-nums"},
        )
    if "Produto" in df.columns:
        sty = sty.set_properties(
            subset=["Produto"],
            **{
                "text-align": "left",
                "max-width": "17rem",
                "white-space": "nowrap",
                "overflow": "hidden",
                "text-overflow": "ellipsis",
            },
        )
    if "SKU" in df.columns:
        sty = sty.set_properties(
            subset=["SKU"],
            **{"text-align": "left", "font-variant-numeric": "tabular-nums"},
        )
    if "Tendência" in df.columns:
        sty = _elem_map(sty, _cp_trend_cell_style, ["Tendência"])
    if "Sugestão de compra" in df.columns:
        sty = _elem_map(sty, _cp_sug_cell_style, ["Sugestão de compra"])
    return sty


def _comercial_trend_tbl_apply_ui_filters(
    trend_tbl: pd.DataFrame,
    *,
    produtos_sel: tuple[str, ...],
    tendencias_sel: tuple[str, ...],
    sugestoes_sel: tuple[str, ...],
) -> pd.DataFrame:
    """Recorta a tabela de tendência pelos multiselects (vazio = sem filtro nessa dimensão)."""
    if trend_tbl.empty:
        return trend_tbl
    out = trend_tbl
    sel_p = tuple(str(x).strip() for x in produtos_sel if str(x).strip())
    if sel_p and "Produto" in out.columns:
        pr = out["Produto"].fillna("").astype(str).str.strip()
        out = out.loc[pr.isin(sel_p)].copy()
    sel_t = tuple(str(x).strip() for x in tendencias_sel if str(x).strip())
    if sel_t and "Tendência" in out.columns:
        tt = out["Tendência"].fillna("").astype(str).str.strip()
        out = out.loc[tt.isin(sel_t)].copy()
    sel_s = tuple(str(x).strip() for x in sugestoes_sel if str(x).strip())
    if sel_s and "Sugestão de compra" in out.columns:
        ss = out["Sugestão de compra"].fillna("").astype(str).str.strip()
        out = out.loc[ss.isin(sel_s)].copy()
    return out


def _render_comercial_pedidos_analise(
    df: pd.DataFrame,
    load_info: dict[str, object],
    ts_proc: str,
) -> None:
    """Vista comercial: atendidos, coluna Data, sem NF — lógica em ``comercial_pedidos_analise``."""
    _ = load_info, ts_proc
    if df.empty:
        st.info(
            "Sem dados de pedidos para este escopo. Confirme **materialização** e o módulo **faturamento**."
        )
        return
    if not cpa.data_column(df):
        if _is_admin_mode():
            st.warning(
                "Falta a coluna **Data** na base de faturamento — a vista Comercial & pedidos precisa dela para filtros e tendência."
            )
        else:
            st.warning("Dados incompletos para a vista comercial. Contacte o suporte.")
        return

    df_atend = cpa.filter_atendidos(df)
    if df_atend.empty:
        st.info("Não há linhas com **Situação = atendido** neste carregamento.")
        return

    _fdl_cp_inject_panel_styles()
    _fdl_ui_gap_tight()

    d_min, d_max = cpa.bounds_dates_atendidos(df)
    _today = datetime.now(_BR_TZ).date()
    _sig_k = "fdl_cp_bounds_sig"
    if d_min is not None and d_max is not None:
        _bs = (d_min.isoformat(), d_max.isoformat())
        if st.session_state.get(_sig_k) != _bs:
            st.session_state[_sig_k] = _bs
            _di = d_min
            _df = min(d_max, _today)
            if _df < _di:
                _df = _di
            st.session_state["fdl_cp_d_ini"] = _di
            st.session_state["fdl_cp_d_fim"] = _df
        if "fdl_cp_d_ini" not in st.session_state:
            _di = d_min
            _df = min(d_max, _today)
            if _df < _di:
                _df = _di
            st.session_state["fdl_cp_d_ini"] = _di
            st.session_state["fdl_cp_d_fim"] = _df
        st.session_state["fdl_cp_d_ini"] = _safe_streamlit_date(st.session_state["fdl_cp_d_ini"], d_min)
        st.session_state["fdl_cp_d_fim"] = _safe_streamlit_date(st.session_state["fdl_cp_d_fim"], d_max)
        _d_ini_ui = st.session_state["fdl_cp_d_ini"]
        _d_fim_ui = st.session_state["fdl_cp_d_fim"]
        if _d_fim_ui < _d_ini_ui:
            st.session_state["fdl_cp_d_fim"] = _d_ini_ui
            _d_fim_ui = _d_ini_ui
    else:
        _d_ini_ui = None
        _d_fim_ui = None

    emp_opts = _faturamento_dre_etiquetas_empresa_recorte(df_atend)
    plats = (
        nf_grain_plataforma_ui_options(df_atend["Nome da plataforma"])
        if "Nome da plataforma" in df_atend.columns
        else []
    )

    with st.container(border=True):
        st.markdown('<p class="fdl-cp-filtros-h">Filtros</p>', unsafe_allow_html=True)
        st.markdown(
            '<p class="fdl-cp-caption">Só <strong>pedidos atendidos</strong> · valores = coluna <strong>Vl_Venda</strong> '
            "da tabela materializada (se existir); senão <strong>preço lista × qtd</strong> · sem NF.</p>",
            unsafe_allow_html=True,
        )
        if emp_opts:
            if "fdl_cp_emp" not in st.session_state:
                st.session_state["fdl_cp_emp"] = []
            else:
                prev_e = st.session_state["fdl_cp_emp"]
                if isinstance(prev_e, list):
                    st.session_state["fdl_cp_emp"] = [x for x in prev_e if x in emp_opts]
                else:
                    st.session_state["fdl_cp_emp"] = []
            st.multiselect(
                "Empresa",
                emp_opts,
                key="fdl_cp_emp",
                help="**Vazio** = todas. Recorte por marca (mesma coluna que Resultado Gerencial).",
                placeholder="Todas",
            )
        _multiselect_stable("fdl_cp_plat", "Plataforma", plats)
        if d_min is not None and d_max is not None:
            r_d = st.columns((1, 1))
            with r_d[0]:
                st.date_input(
                    "Data pedido — início",
                    min_value=d_min,
                    max_value=d_max,
                    format="DD/MM/YYYY",
                    key="fdl_cp_d_ini",
                )
            with r_d[1]:
                st.date_input(
                    "Data pedido — fim",
                    min_value=d_min,
                    max_value=d_max,
                    format="DD/MM/YYYY",
                    key="fdl_cp_d_fim",
                )
        else:
            st.markdown(
                '<p class="fdl-cp-caption">Datas de <strong>Data</strong> indisponíveis no recorte atendido.</p>',
                unsafe_allow_html=True,
            )
        _cp_clr_l, _cp_clr_r = st.columns((1.55, 1))
        with _cp_clr_r:
            if st.button(
                "Limpar filtros desta vista",
                key="fdl_cp_reset",
                use_container_width=True,
                help="Repor marca, plataforma, período e filtros da tabela Tendência e compra.",
            ):
                for _k in (
                    "fdl_cp_emp",
                    "fdl_cp_plat",
                    "fdl_cp_d_ini",
                    "fdl_cp_d_fim",
                    "fdl_cp_bounds_sig",
                    "fdl_cp_tr_prod",
                    "fdl_cp_tr_tend",
                    "fdl_cp_tr_sug",
                ):
                    st.session_state.pop(_k, None)
                st.rerun()

    emp_sel = tuple(str(x).strip() for x in (st.session_state.get("fdl_cp_emp") or []) if str(x).strip())
    plat_sel = tuple(str(x).strip() for x in (st.session_state.get("fdl_cp_plat") or []) if str(x).strip())

    if d_min is not None and d_max is not None:
        d_ini_f = _safe_streamlit_date(st.session_state.get("fdl_cp_d_ini"), d_min)
        d_fim_f = _safe_streamlit_date(st.session_state.get("fdl_cp_d_fim"), d_max)
        if d_fim_f < d_ini_f:
            d_fim_f = d_ini_f
    else:
        d_ini_f, d_fim_f = None, None

    filtrado = cpa.filter_ui(
        df_atend,
        empresas_sel=emp_sel,
        plataformas_sel=plat_sel,
        d_ini=d_ini_f,
        d_fim=d_fim_f,
    )
    period_end_trend = d_fim_f if d_fim_f is not None else (_today if d_max is None else min(d_max, _today))

    _fdl_ui_gap_section()
    kpis = cpa.compute_kpis(filtrado)
    _render_comercial_pedidos_kpi_cards(kpis)

    abc_v = cpa.compute_abc_valor(filtrado)
    abc_q = cpa.compute_abc_quantidade(filtrado)

    _fdl_ui_gap_section()
    with st.container(border=True):
        st.markdown(
            '<p class="fdl-cp-title-main">ABC por receita</p>'
            '<p class="fdl-cp-title-sub">Onde está a receita comercial · Pareto 80% / 95% · base = Vl_Venda materializado (ou lista×qtd)</p>',
            unsafe_allow_html=True,
        )
        if abc_v.empty:
            st.markdown(
                '<p class="fdl-cp-caption">Sem SKU com código no recorte.</p>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(_comercial_abc_valor_exec_html(abc_v), unsafe_allow_html=True)
            _abc_cfg: dict[str, object] = {
                "SKU": TextColumn("SKU", width="small", help="Código do SKU."),
                "Produto": TextColumn("Produto", width="large", help="Descrição resumida."),
                "Valor comercial (lista)": TextColumn(
                    "Receita (lista)",
                    width="medium",
                    help="Σ Vl_Venda no recorte (pedidos atendidos); senão lista×qtd.",
                ),
                "Part %": TextColumn(
                    "Part. %",
                    width="small",
                    help="Participação do SKU na receita comercial do recorte.",
                ),
                "Acum %": TextColumn(
                    "Acum. %",
                    width="small",
                    help="Participação acumulada ordenada por receita (Pareto).",
                ),
                "Classe": TextColumn("ABC", width="small", help="Classe A/B/C (80% / 95% acumulado)."),
            }
            _abc_cfg = {k: v for k, v in _abc_cfg.items() if k in abc_v.columns}
            st.dataframe(
                _comercial_abc_valor_table_styler(abc_v),
                use_container_width=True,
                height=min(400, 48 + min(len(abc_v), 14) * 34),
                column_config=_abc_cfg or None,
            )

    _fdl_ui_gap_section()
    st.markdown(
        '<p class="fdl-cp-title-sec">ABC por unidades</p>'
        '<p class="fdl-cp-title-sec-note">Complementar ao bloco acima · classes por quantidade vendida, não por receita</p>',
        unsafe_allow_html=True,
    )
    if abc_q.empty:
        st.markdown(
            '<p class="fdl-cp-caption">Sem quantidade por SKU no recorte.</p>',
            unsafe_allow_html=True,
        )
    else:
        _q_cfg: dict[str, object] = {
            "SKU": TextColumn("SKU", width="small"),
            "Produto": TextColumn("Produto", width="medium"),
            "Quantidade": TextColumn("Unidades", width="small", help="Quantidade vendida no período."),
            "Part %": TextColumn("Part. %", width="small", help="Participação em unidades."),
            "Acum %": TextColumn("Acum. %", width="small", help="Acumulado por unidades (Pareto)."),
            "Classe": TextColumn("ABC", width="small", help="Classe por giro em unidades."),
        }
        _q_cfg = {k: v for k, v in _q_cfg.items() if k in abc_q.columns}
        st.dataframe(
            _comercial_abc_quantidade_table_styler(abc_q),
            use_container_width=True,
            height=min(280, 36 + min(len(abc_q), 10) * 32),
            column_config=_q_cfg or None,
        )

    _fdl_ui_gap_section()
    with st.container(border=True):
        st.markdown(
            '<p class="fdl-cp-title-decision">Tendência e compra</p>'
            '<p class="fdl-cp-title-sub">Três meses civis fechados · mesmos filtros · tendência de qtd e sugestão já calculadas</p>',
            unsafe_allow_html=True,
        )
        _triple = cpa.three_closed_months_trend_bounds(period_end_trend, as_of=_today)[2]
        _tw0 = f"{_triple[0][1]:02d}/{_triple[0][0]}"
        _tw1 = f"{_triple[2][1]:02d}/{_triple[2][0]}"
        _tfim = f"{period_end_trend.day:02d}/{period_end_trend.month:02d}/{period_end_trend.year}"
        st.markdown(
            f'<p class="fdl-cp-caption">Janela <strong>{html.escape(_tw0)}</strong>–<strong>{html.escape(_tw1)}</strong> · '
            f'corte do filtro <strong>{html.escape(_tfim)}</strong> · sem mês em aberto.</p>',
            unsafe_allow_html=True,
        )
        df_trend = cpa.filter_trend_window(
            df_atend,
            empresas_sel=emp_sel,
            plataformas_sel=plat_sel,
            period_end=period_end_trend,
            as_of=_today,
        )
        trend_tbl = cpa.compute_trend_and_suggestion(
            df_trend, abc_v, period_end=period_end_trend, as_of=_today
        )
        if trend_tbl.empty:
            st.markdown(
                '<p class="fdl-cp-caption">Sem linhas com SKU e quantidade na janela de tendência.</p>',
                unsafe_allow_html=True,
            )
        else:
            _tr_prod_opts = sorted(
                {
                    str(x).strip()
                    for x in trend_tbl["Produto"].dropna().unique()
                    if str(x).strip()
                }
            )
            _tr_tend_opts = sorted(
                {
                    str(x).strip()
                    for x in trend_tbl["Tendência"].dropna().unique()
                    if str(x).strip()
                }
            ) if "Tendência" in trend_tbl.columns else []
            _tr_sug_opts = sorted(
                {
                    str(x).strip()
                    for x in trend_tbl["Sugestão de compra"].dropna().unique()
                    if str(x).strip()
                }
            ) if "Sugestão de compra" in trend_tbl.columns else []

            st.markdown(
                '<p class="fdl-cp-caption">Filtros abaixo aplicam-se <strong>só</strong> a esta tabela e à barra de resumo.</p>',
                unsafe_allow_html=True,
            )
            _f_tr = st.columns((1, 1, 1))
            with _f_tr[0]:
                _multiselect_stable(
                    "fdl_cp_tr_prod",
                    "Produto (tabela tendência)",
                    _tr_prod_opts,
                    help="**Vazio** = todos os produtos. Correspondência exata ao texto da coluna Produto.",
                )
            with _f_tr[1]:
                _multiselect_stable(
                    "fdl_cp_tr_tend",
                    "Tendência",
                    _tr_tend_opts,
                    help="**Vazio** = todas. Filtra pela classificação de tendência (ex.: crescente, decrescente).",
                )
            with _f_tr[2]:
                _multiselect_stable(
                    "fdl_cp_tr_sug",
                    "Sugestão de compra",
                    _tr_sug_opts,
                    help="**Vazio** = todas. Filtra pelo texto completo da sugestão.",
                )

            _sel_tr_p = tuple(
                str(x).strip()
                for x in (st.session_state.get("fdl_cp_tr_prod") or [])
                if str(x).strip()
            )
            _sel_tr_t = tuple(
                str(x).strip()
                for x in (st.session_state.get("fdl_cp_tr_tend") or [])
                if str(x).strip()
            )
            _sel_tr_s = tuple(
                str(x).strip()
                for x in (st.session_state.get("fdl_cp_tr_sug") or [])
                if str(x).strip()
            )
            trend_disp = _comercial_trend_tbl_apply_ui_filters(
                trend_tbl,
                produtos_sel=_sel_tr_p,
                tendencias_sel=_sel_tr_t,
                sugestoes_sel=_sel_tr_s,
            )

            if trend_disp.empty:
                st.warning(
                    "Nenhuma linha com a combinação de filtros de **Produto**, **Tendência** e **Sugestão de compra**. "
                    "Limpe um dos filtros ou repor com **Limpar filtros desta vista**."
                )
            else:
                st.caption(
                    f"**{len(trend_disp)}** linha(s) após filtros da tabela (total sem filtro de tabela: **{len(trend_tbl)}**)."
                )
                st.markdown(_comercial_trend_summary_html(trend_disp), unsafe_allow_html=True)
                _tr_cfg: dict[str, object] = {
                    "SKU": TextColumn("SKU", width="small", help="Código do SKU."),
                    "Produto": TextColumn("Produto", width="medium", help="Descrição resumida."),
                    "Qtd mês -2": TextColumn(
                        "Unid. M-2",
                        width="small",
                        help="Unidades vendidas no 1º mês civil fechado da janela (mais antigo).",
                    ),
                    "Qtd mês -1": TextColumn(
                        "Unid. M-1",
                        width="small",
                        help="Unidades vendidas no mês intermediário.",
                    ),
                    "Qtd mês atual": TextColumn(
                        "Unid. último",
                        width="small",
                        help="Unidades vendidas no último mês civil fechado da janela.",
                    ),
                    "Valor lista mês -2": TextColumn(
                        "R$ lista M-2",
                        width="small",
                        help="Σ Vl_Venda (ou lista×qtd) no 1º mês fechado.",
                    ),
                    "Valor lista mês -1": TextColumn(
                        "R$ lista M-1",
                        width="small",
                        help="Σ Vl_Venda (ou lista×qtd) no mês intermediário.",
                    ),
                    "Valor lista mês atual": TextColumn(
                        "R$ lista últ.",
                        width="small",
                        help="Σ Vl_Venda (ou lista×qtd) no último mês fechado.",
                    ),
                    "Tendência": TextColumn(
                        "Tendência",
                        width="medium",
                        help="Combina evolução de unidades e de receita (lista) nos 3 meses; «insuficiente» = pouco volume para ler tendência.",
                    ),
                    "Sugestão de compra": TextColumn(
                        "Sugestão compra",
                        width="large",
                        help="Orientação a partir da classe ABC (receita) e da tendência de giro — mesma regra de cálculo.",
                    ),
                }
                _tr_cfg = {k: v for k, v in _tr_cfg.items() if k in trend_disp.columns}
                st.dataframe(
                    _comercial_trend_styler(trend_disp),
                    use_container_width=True,
                    height=min(480, 44 + min(len(trend_disp), 18) * 34),
                    column_config=_tr_cfg or None,
                )


def _faturamento_nf_quantidade_itens_por_nf(df_line: pd.DataFrame, df_nf: pd.DataFrame) -> pd.Series:
    """Soma ``Quantidade`` (unidades) no grão linha por (org_id, NF normalizada), alinhada a ``df_nf``."""
    if df_nf.empty:
        return pd.Series(dtype=float)
    if df_line is None or df_line.empty or "Quantidade" not in df_line.columns:
        return pd.Series(0.0, index=df_nf.index, dtype=float)
    from processing.faturamento.normalize import normalize_nf_fiscal_commercial_join_key_scalar as _nk_nf

    d = df_line.copy()
    d["_q"] = pd.to_numeric(d["Quantidade"], errors="coerce").fillna(0.0)
    oid = (
        d["org_id"].fillna("").astype(str).str.strip()
        if "org_id" in d.columns
        else pd.Series("", index=d.index, dtype=str)
    )
    nn = d["Nota_Numero_Normalizado"] if "Nota_Numero_Normalizado" in d.columns else pd.Series("", index=d.index)
    d["_kk"] = oid + "|" + nn.astype(str).map(_nk_nf)
    sums = d.groupby("_kk", sort=False)["_q"].sum()

    if "org_id" in df_nf.columns:
        o2 = df_nf["org_id"].fillna("").astype(str).str.strip()
    else:
        o2 = pd.Series("", index=df_nf.index, dtype=str)
    n2 = (
        df_nf["Nota_Numero_Normalizado"]
        if "Nota_Numero_Normalizado" in df_nf.columns
        else pd.Series("", index=df_nf.index, dtype=str)
    )
    kk = o2 + "|" + n2.astype(str).map(_nk_nf)
    return kk.map(lambda k: float(sums.get(k, 0.0))).astype(float)


def _faturamento_dre_apply_produto_e_sinal_venda(
    df_nf: pd.DataFrame,
    *,
    produtos_sel: tuple[str, ...],
    sinais_resultado: tuple[str, ...],
) -> pd.DataFrame:
    """
    Refina o quadro NF por ``produto_resumo`` e pelo **sinal do resultado comercial** (lucro / prejuízo / empate).

    ``sinais_resultado`` pode incluir ``lucro``, ``prejuizo``, ``empate``. **Vazio** = não filtra por sinal (mostra todas as NFs).
    Com **lucro** e **prejuízo** ambos na seleção (com ou sem empate), **não** se filtra por sinal — mesmo efeito que vazio na UI legada.
    Caso contrário, faz a união das faixas escolhidas.
    Valores vêm do materializado — sem recalcular DRE.
    """
    if df_nf.empty:
        return df_nf
    out = df_nf
    sel_p = tuple(str(x).strip() for x in produtos_sel if str(x).strip())
    if sel_p and "produto_resumo" in out.columns:
        pr = out["produto_resumo"].fillna("").astype(str).str.strip()
        out = out.loc[pr.isin(sel_p)].copy()
    raw = [str(x).strip().lower() for x in sinais_resultado if str(x).strip()]
    sel = {x for x in raw if x in {"lucro", "prejuizo", "empate"}}
    if not sel:
        return out
    if "resultado" not in out.columns:
        return out
    # Lucro **e** prejuízo na mesma seleção ⇒ não filtra por sinal (equivale ao multiselect “cheio” legado).
    if "lucro" in sel and "prejuizo" in sel:
        return out
    res = pd.to_numeric(out["resultado"], errors="coerce")
    _eps = 1e-9
    mask = pd.Series(False, index=out.index)
    if "lucro" in sel:
        mask |= res.notna() & (res > _eps)
    if "prejuizo" in sel:
        mask |= res.notna() & (res < -_eps)
    if "empate" in sel:
        mask |= res.notna() & (res >= -_eps) & (res <= _eps)
    return out.loc[mask].copy()


def _fdl_fat_min_base_fiscal_card_html(
    *,
    total_faturado: str,
    n_nfs: str,
    periodo: str,
    contexto: str,
    primary_label: str = "Total faturado (fiscal)",
    primary_valor_title: str | None = None,
) -> str:
    """Card HTML do topo fiscal (valores já formatados em pt-BR)."""
    _tt_fat = html.escape(
        primary_valor_title
        or "Soma de Valor_Liquido_NF (1× por NF) no Parquet fiscal, após empresa + emissão + situação válida."
    )
    _tt_nf = html.escape("Contagem de notas distintas no mesmo conjunto base fiscal.")
    _lbl = html.escape(primary_label.strip() or "Total faturado (fiscal)")
    return (
        '<div class="fdl-base-fiscal-card">'
        '<div class="fdl-base-fiscal-header">'
        '<span class="fdl-base-fiscal-title">📊 Base fiscal</span>'
        f'<span class="fdl-base-fiscal-periodo">{html.escape(periodo)}</span>'
        "</div>"
        '<div class="fdl-base-fiscal-body">'
        '<div class="fdl-base-fiscal-kpi">'
        f'<span class="fdl-base-fiscal-valor" title="{_tt_fat}">'
        f"{html.escape(total_faturado)}</span>"
        f'<span class="fdl-base-fiscal-label">{_lbl}</span>'
        "</div>"
        '<div class="fdl-base-fiscal-kpi">'
        f'<span class="fdl-base-fiscal-count" title="{_tt_nf}">'
        f"{html.escape(n_nfs)}</span>"
        '<span class="fdl-base-fiscal-label">NFs emitidas</span>'
        "</div>"
        "</div>"
        '<div class="fdl-base-fiscal-footer">'
        f'<span class="fdl-base-fiscal-contexto">ℹ️ {html.escape(contexto)}</span>'
        "</div>"
        "</div>"
    )


def _fdl_fat_base_fiscal_composition_block_html(*, stats: FaturamentoFiscalBaseStats) -> str:
    """Bloco detalhado emitidas − devoluções = base líquida (auditoria)."""
    emitido = float(stats.valor_liquido_fiscal_sum)
    devolvido = float(stats.total_devolvido)
    liquida = float(stats.base_fiscal_liquida)
    nfs_emit = int(stats.n_nf)
    nfs_dev = int(stats.nfs_devolucao)
    s_emit = html.escape(_fmt_brl_ptbr_celula(emitido) or "R$ 0,00")
    s_dev = html.escape(_fmt_brl_ptbr_celula(devolvido) or "R$ 0,00")
    s_liq = html.escape(_fmt_brl_ptbr_celula(liquida) or "R$ 0,00")
    s_ne = html.escape(_fmt_int_ptbr(nfs_emit))
    s_nd = html.escape(_fmt_int_ptbr(nfs_dev))
    return (
        '<div style="'
        "background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:12px 16px;"
        "margin-top:12px;font-size:0.85rem;font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace;"
        "color:#374151;line-height:1.8;"
        '">'
        '<span style="color:#16a34a">(+)</span> '
        f"{s_emit} &nbsp;&nbsp;"
        f'<span style="color:#6b7280">{s_ne} emitidas</span><br>'
        '<span style="color:#dc2626">(−)</span> '
        f"{s_dev} &nbsp;&nbsp;"
        f'<span style="color:#6b7280">{s_nd} devoluções</span><br>'
        '<hr style="border:none;border-top:1px solid #e2e8f0;margin:6px 0">'
        '<span style="color:#1e293b;font-weight:600">(=)</span> '
        f"<strong>{s_liq}</strong> &nbsp;&nbsp;"
        '<span style="color:#6b7280">Base Fiscal Líquida</span>'
        "</div>"
    )


def _render_faturamento_dre_fiscal_base_top(
    *,
    stats: FaturamentoFiscalBaseStats,
    ok_nf_dates: bool,
    empresas_sel: tuple[str, ...],
    emp_opts: list[str],
    nf_d_ini: date,
    nf_d_fim: date,
    fiscal_parquet_ok: bool,
    situacoes_nf_sel: tuple[str, ...] = (),
) -> None:
    """Topo do painel: conjunto base fiscal (empresa + emissão), comparável ao Bling — sem plataforma/produto comercial."""
    if ok_nf_dates:
        _per = f"{nf_d_ini.strftime('%d/%m/%Y')} — {nf_d_fim.strftime('%d/%m/%Y')}"
    else:
        _per = "Período indisponível"
    _emp_ctx = (
        " · ".join(str(x).strip() for x in empresas_sel if str(x).strip())
        if empresas_sel
        else ("Todas as empresas" if emp_opts else "—")
    )
    _sit_ctx = (
        ", ".join(str(x).strip() for x in situacoes_nf_sel if str(x).strip())
        if situacoes_nf_sel
        else "Situações válidas (exc. cancelada/denegada/inutilizada)"
    )
    _contexto_compacto = f"{_emp_ctx} · {_sit_ctx}"
    if not fiscal_parquet_ok:
        st.info(
            "Conjunto base fiscal indisponível: publique **`dataset_faturamento_fiscal.parquet`** junto do materializado "
            "para alinhar totais ao Bling."
        )
        return
    if not ok_nf_dates:
        st.warning("Datas de emissão da NF não utilizáveis — recorte fiscal vazio.")
        return
    _has_dev_abatimento = float(stats.total_devolvido) > 0.0
    _tot_val = float(stats.base_fiscal_liquida) if _has_dev_abatimento else float(stats.valor_liquido_fiscal_sum)
    _tot = _fmt_brl_ptbr_celula(_tot_val) or "R$ 0,00"
    _nn = _fmt_int_ptbr(stats.n_nf)
    _lbl_princ = "BASE FISCAL LÍQUIDA" if _has_dev_abatimento else "Total faturado (fiscal)"
    _tt_princ = (
        "Valor principal = base fiscal líquida (emitidas − devoluções no recorte); detalhe no quadro abaixo."
        if _has_dev_abatimento
        else None
    )
    st.html(
        _fdl_fat_min_base_fiscal_card_html(
            total_faturado=_tot,
            n_nfs=_nn,
            periodo=_per,
            contexto=_contexto_compacto,
            primary_label=_lbl_princ,
            primary_valor_title=_tt_princ,
        )
    )
    if _has_dev_abatimento:
        st.markdown(_fdl_fat_base_fiscal_composition_block_html(stats=stats), unsafe_allow_html=True)
    _fdl_fat_min_vsp(size="sm")


def _render_faturamento_dre_commercial_complement_banner(
    *,
    coverage: CommercialCoverageStats,
    n_fiscal_base: int,
    aligned_to_fiscal_base: bool,
    ok_nf_dates: bool,
    fiscal_parquet_ok: bool,
    kpi_subset_by_platform: bool = False,
    embedded_in_sobre_expander: bool = False,
    rg_premium_single_expander: bool = False,
) -> None:
    """
    Bloco imediatamente abaixo do topo fiscal: deixa explícito que os KPIs/DRE comerciais são complemento
    sobre o mesmo período/empresa, com cobertura parcial possível. Colapsável (fechado por defeito) + badge de alerta.

    Quando ``embedded_in_sobre_expander=True``, o bloco é só o conteúdo técnico (sem ``<details>`` nem divisor final),
    para uso dentro do expander «Sobre este módulo» no Resultado Gerencial.
    """

    def _md_bold_to_html(s: str) -> str:
        parts = s.split("**")
        if len(parts) == 1:
            return html.escape(s)
        chunks: list[str] = []
        for i, part in enumerate(parts):
            if i % 2 == 0:
                chunks.append(html.escape(part))
            else:
                chunks.append(f"<strong>{html.escape(part)}</strong>")
        return "".join(chunks)

    _nt = max(int(coverage.n_total), 0)
    _pct_sem = (100.0 * float(coverage.n_sem_resultado) / float(_nt)) if _nt else 0.0
    if _pct_sem <= 5.0:
        _badge_html = ""
    elif _pct_sem <= 10.0:
        _badge_html = (
            f'<span class="fdl-badge fdl-badge-warning">{_pct_sem:.0f}% sem custo</span>'
        )
    else:
        _badge_html = (
            f'<span class="fdl-badge fdl-badge-danger">{_pct_sem:.0f}% sem custo</span>'
        )

    if aligned_to_fiscal_base and fiscal_parquet_ok and ok_nf_dates:
        if kpi_subset_by_platform:
            _cap_txt = (
                "Usa o **mesmo período**, **empresa(s)** e **situação NF** (se filtrada) do topo **Base fiscal**. "
                "Com **Plataforma** selecionada, os KPIs somam **só** NFs desse canal (o **topo fiscal** permanece **sem** plataforma). "
                "Dados de venda, custos e resultado vêm do **enriquecimento comercial** e **podem faltar** em parte das notas."
            )
        else:
            _cap_txt = (
                "Usa o **mesmo período de emissão**, **empresa(s)** e **situação NF** (se filtrada) do topo **Base fiscal**. "
                "Cada linha = uma NF do conjunto **N_base**; dados de venda, custos e resultado vêm do **enriquecimento comercial** "
                "(painel materializado) e **podem faltar** em parte das notas — por isso os totais comerciais são interpretados "
                "com cobertura parcial, sem retirar a NF do universo."
            )
    elif fiscal_parquet_ok and not ok_nf_dates:
        _cap_txt = (
            "Período de emissão indisponível — indicadores comerciais seguem o recorte possível no painel NF "
            "(empresa), sem alinhamento ao topo fiscal."
        )
    else:
        _cap_txt = (
            "Sem **Parquet fiscal** publicado: os indicadores comerciais abaixo referem-se às NFs do **painel** "
            "no recorte **empresa + emissão** (mesmas datas selecionadas quando aplicável), **não** ao conjunto "
            "**N_base** do Bling. Publique `dataset_faturamento_fiscal.parquet` para alinhar o universo ao topo."
        )

    _h_univ = (
        "Número de linhas usadas em cards e DRE comerciais: com fiscal ativo, coincide com N_base do topo quando "
        "Plataforma está vazia; com Plataforma selecionada, é o subconjunto comercial nesse canal (ainda alinhado ao fiscal por NF). "
        "Sem Parquet fiscal: NFs do painel após empresa + emissão."
    )
    _h_vinc = "Notas com faturamento_nota_vinculada no materializado (há pedido comercial ligado)."
    _h_so_fiscal = "Sem vínculo comercial ou com venda (lista) ~0 — totais comerciais podem ser neutros nessas linhas."
    _h_venda = "Notas com valor de venda em lista estritamente positivo no materializado."
    _h_sem_res = (
        "NF comercial incompleta: tipicamente SKUs sem custo mapeado (Status_Custo ≠ CUSTO_OK). "
        "O resultado consolidado não é calculado e não entra na soma da DRE até o cadastro de custos ser corrigido."
    )
    _h_com_res = (
        "NFs com custo e encargos comerciais suficientes para calcular o resultado consolidado (pode ser 0)."
    )

    def _cell(lab: str, val: int, title: str) -> str:
        return (
            f'<div class="fdl-fat-cobertura-cell" title="{html.escape(title, quote=True)}">'
            f'<div class="fdl-fat-cobertura-lab">{html.escape(lab)}</div>'
            f'<div class="fdl-fat-cobertura-val">{html.escape(_fmt_int_ptbr(val))}</div></div>'
        )

    _grid_html = ""
    if ok_nf_dates and (aligned_to_fiscal_base or not fiscal_parquet_ok):
        _grid_html = (
            '<div class="fdl-fat-cobertura-grid">'
            + _cell("Notas no universo dos KPIs", coverage.n_total, _h_univ)
            + _cell("Com vínculo pedido–NF", coverage.n_com_vinculo_pedido_nf, _h_vinc)
            + _cell("Só fiscal / sem vínculo útil", coverage.n_sem_vinculo_ou_so_fiscal, _h_so_fiscal)
            + _cell("Com venda (lista) > 0", coverage.n_com_venda_lista, _h_venda)
            + _cell("Sem custo (SKU não mapeado)", coverage.n_sem_resultado, _h_sem_res)
            + _cell("Com custo calculado", coverage.n_com_resultado_numerico, _h_com_res)
            + "</div>"
        )

    _legenda_html = (
        '<p class="fdl-cobertura-legenda">'
        "NFs marcadas como &quot;sem custo&quot; têm SKUs não encontrados na base de custo. "
        "O resultado dessas notas não entra na DRE até que o cadastro seja corrigido."
        "</p>"
    )
    _cta_html = ""
    if _pct_sem > 10.0:
        _cta_html = (
            '<p class="fdl-cobertura-cta">'
            "💡 Para reduzir esse percentual, revise o cadastro de custos dos SKUs pendentes."
            "</p>"
        )

    _admin_html = ""
    if (
        aligned_to_fiscal_base
        and ok_nf_dates
        and n_fiscal_base != coverage.n_total
        and _is_admin_mode()
        and not kpi_subset_by_platform
    ):
        _admin_html = (
            '<p class="fdl-fat-cobertura-admin">'
            "<strong>Admin:</strong> "
            f"N_base no topo = <strong>{html.escape(_fmt_int_ptbr(n_fiscal_base))}</strong> · "
            f"linhas no frame alinhado = <strong>{html.escape(_fmt_int_ptbr(coverage.n_total))}</strong> — "
            "esperado igual; investigar chaves de merge."
            "</p>"
        )

    if rg_premium_single_expander:

        def _cell_rg(top: str, bot: str, val: int, tt: str) -> str:
            return (
                f'<div class="fdl-fat-cob-v2-cell" title="{html.escape(tt, quote=True)}">'
                f'<div class="fdl-fat-cob-v2-lab">{html.escape(top)}<br/>'
                f'<span class="fdl-fat-cob-v2-lab2">{html.escape(bot)}</span></div>'
                f'<div class="fdl-fat-cob-v2-val">{html.escape(_fmt_int_ptbr(val))}</div></div>'
            )

        _tip_chunks = [
            _cap_txt.replace("**", "").strip(),
            "NFs «sem custo»: SKUs não encontrados na base de custo; resultado não entra na DRE até correção.",
        ]
        if _pct_sem > 10.0:
            _tip_chunks.append("Revise cadastro de custos dos SKUs pendentes.")
        _tip_esc = html.escape(" ".join(_tip_chunks).strip(), quote=True)
        _grid_rg = ""
        if ok_nf_dates and (aligned_to_fiscal_base or not fiscal_parquet_ok):
            _grid_rg = (
                '<div class="fdl-fat-cob-v2-grid">'
                + _cell_rg("Notas no universo", "dos KPIs", coverage.n_total, _h_univ)
                + _cell_rg("Com vínculo", "pedido–NF", coverage.n_com_vinculo_pedido_nf, _h_vinc)
                + _cell_rg("Só fiscal /", "sem vínculo útil", coverage.n_sem_vinculo_ou_so_fiscal, _h_so_fiscal)
                + _cell_rg("Com venda (lista)", "> 0", coverage.n_com_venda_lista, _h_venda)
                + _cell_rg("Sem custo", "(SKU não mapeado)", coverage.n_sem_resultado, _h_sem_res)
                + _cell_rg("Com custo", "calculado", coverage.n_com_resultado_numerico, _h_com_res)
                + "</div>"
            )
        _badge_tail_rg = f" {_badge_html}" if _badge_html else ""
        st.html(
            '<div class="fdl-fat-cobertura fdl-fat-cobertura--rg-premium">'
            '<div class="fdl-fat-cob-rg-head">'
            '<span class="fdl-fat-cob-rg-title">Cobertura comercial</span>'
            '<span class="fdl-fat-cob-rg-meta">'
            f"{_badge_tail_rg}"
            f'<span class="fdl-fat-cob-rg-info" title="{_tip_esc}" aria-label="Detalhes técnicos">ℹ️</span>'
            "</span></div>"
            f"{_grid_rg}{_admin_html}"
            "</div>"
        )
        return

    _inner = (
        f'<p class="fdl-fat-cobertura-caption">{_md_bold_to_html(_cap_txt)}</p>'
        f"{_legenda_html}{_grid_html}{_cta_html}{_admin_html}"
    )
    # ``st.markdown(..., unsafe_allow_html=True)`` passa pelo parser GFM e pode degradar/sanitizar
    # ``<details>``/aninhamento; ``st.html`` envia o fragmento como HTML nativo (CSS já injetado em
    # ``_fdl_fat_min_inject_ui_styles`` no início desta vista).
    if embedded_in_sobre_expander:
        _badge_tail = f" {_badge_html}" if _badge_html else ""
        st.html(
            f'<div class="fdl-fat-cobertura fdl-fat-cobertura--embedded"><p class="fdl-fat-cobertura-embed-title">'
            f"Cobertura comercial{_badge_tail}</p>{_inner}</div>"
        )
        return
    st.html(
        '<details class="fdl-fat-cobertura">'
        f'<summary>Cobertura comercial{" " if _badge_html else ""}{_badge_html}</summary>'
        f"<div>{_inner}</div>"
        "</details>"
    )
    _fdl_fat_min_vsp(size="sm")
    st.divider()


def _fdl_parse_ts_proc(ts_raw: str) -> datetime | None:
    """Interpreta carimbos típicos de ``ts_proc`` (ISO, ``YYYY-mm-dd HH:MM[:SS][.µs]``, ``T``)."""
    s = (ts_raw or "").strip()
    if not s:
        return None
    candidates = [s]
    if " " in s and "T" not in s[:12]:
        candidates.append(s.replace(" ", "T", 1))
    for c in candidates:
        try:
            d = datetime.fromisoformat(c.replace("Z", "+00:00"))
            if d.tzinfo is not None:
                d = d.astimezone(_BR_TZ).replace(tzinfo=None)
            return d
        except ValueError:
            continue
    ts = pd.to_datetime(s, errors="coerce")
    if isinstance(ts, pd.Timestamp) and not pd.isna(ts):
        if ts.tz is not None:
            ts = ts.tz_convert(_BR_TZ)
        return ts.to_pydatetime().replace(tzinfo=None)
    return None


def _fdl_fat_min_format_updated_at(ts_raw: str) -> str:
    """Formata ``ts_proc`` para «Última atualização: DD/mmm às HH:MM»."""
    s = (ts_raw or "").strip()
    dt = _fdl_parse_ts_proc(s)
    if dt is None:
        return s if s else "—"
    _meses = ("jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez")
    return f"Última atualização: {dt.day}/{_meses[dt.month - 1]} às {dt.strftime('%H:%M')}"


def _build_faturamento_dre_page_header_html(
    *,
    updated_at: str | None = None,
    sobre_tooltip: str = "",
) -> str:
    """Cabeçalho premium do módulo Resultado Gerencial (HTML; ``updated_at`` já seguro para inserção)."""
    meta = ""
    if (updated_at or "").strip():
        meta = (
            '<div class="fdl-page-header-meta">'
            f'<span class="fdl-page-updated">{html.escape(str(updated_at).strip())}</span>'
            "</div>"
        )
    tip = html.escape(sobre_tooltip.strip(), quote=True) if sobre_tooltip.strip() else ""
    ico = (
        f'<span class="fdl-page-title-hint" title="{tip}" aria-label="Sobre este módulo">ℹ️</span>'
        if tip
        else ""
    )
    return (
        '<div class="fdl-page-header">'
        '<div class="fdl-page-header-main">'
        '<p class="fdl-page-breadcrumb">Gerencial &gt; Resultado Gerencial</p>'
        '<h1 class="fdl-page-title">📊 Resultado Gerencial '
        f"{ico}</h1>"
        "<p class=\"fdl-page-subtitle\">"
        "DRE, margem e desempenho da operação"
        "</p>"
        "</div>"
        f"{meta}"
        "</div>"
    )


def _render_faturamento_dre_nf_table_section(
    *,
    df_nf_pre: pd.DataFrame,
    df: pd.DataFrame,
    df_fiscal_pre: pd.DataFrame,
    load_info: dict[str, object],
    _min_state: FaturamentoRecorteMinState,
    _nf_kpi_ini: object,
    _nf_kpi_fim: object,
    ok_nf_dates: bool,
    use_fiscal_kpi: bool,
    use_nf_materializado: bool,
    use_fiscal_parquet: bool,
    _nf_panel_ads_ui: bool,
    _df_fiscal_base: pd.DataFrame,
    _fiscal_base_stats: FaturamentoFiscalBaseStats,
    _kp_cards: dict[str, float | int],
    org_id: str,
    prefix_main: str,
    prefix_nf: str,
    csv_file_name: str = "faturamento_recorte_minimo_nf.csv",
    table_heading: str = "### Tabela por NF",
    nf_table_ui_mode: Literal["gerencial", "fiscal"] = "gerencial",
) -> None:
    """Tabela por NF (filtros inline, CSV, paginação).

    ``nf_table_ui_mode='fiscal'`` — colunas e filtros orientados à Apuração Fiscal (prefixos ``prefix_main`` / ``prefix_nf``).
    ``'gerencial'`` — vista completa para o Resultado Gerencial (quando essa função voltar a ser chamada de lá).
    """
    _oid = str(org_id)
    _ui_fiscal = nf_table_ui_mode == "fiscal"
    # Só chegamos aqui com painel materializado válido: recorte = filtrar linhas já agregadas (sem recomputar DRE).
    df_nf_lines = _faturamento_nf_apply_minimal_recorte(
        df_nf_pre,
        empresas_sel=_min_state.empresas,
        plataformas_sel=_min_state.plataformas,
        nf_d_ini=_nf_kpi_ini,
        nf_d_fim=_nf_kpi_fim,
        ok_nf_dates=ok_nf_dates,
    )
    df_nf_lines = _faturamento_nf_filter_by_situacao(df_nf_lines, _min_state.situacoes_nf)
    df_nf_commercial = df_nf_lines.copy()
    if "plataforma_resumo" not in df_nf_commercial.columns:
        if "plataforma" in df_nf_commercial.columns:
            df_nf_commercial["plataforma_resumo"] = (
                df_nf_commercial["plataforma"].fillna("").astype(str)
            )
        else:
            df_nf_commercial["plataforma_resumo"] = "—"

    df_nf = df_nf_commercial.copy()
    if use_fiscal_kpi and _min_state.plataformas:
        df_nf = _nf_panel_filter_merged_fiscal_by_plataforma_resumo(
            df_nf, _min_state.plataformas
        )

    # df_nf = merge fiscal + comercial após empresa / emissão / plataforma; df_nf_panel = idem + produto / sinal.
    _prod_opts: list[str] = []
    if not df_nf.empty and "produto_resumo" in df_nf.columns:
        _prod_opts = sorted(
            {
                str(x).strip()
                for x in df_nf["produto_resumo"].dropna().unique()
                if str(x).strip() and str(x).strip() != "—"
            }
        )

    _k_sinais = f"{prefix_main}_sinais_resultado"
    if _k_sinais not in st.session_state:
        _leg = st.session_state.get(f"{prefix_main}_sinal_resultado")
        _leg_vs = st.session_state.get(f"{prefix_main}_venda_sinal")
        if isinstance(_leg, str):
            _s = _leg.strip().lower()
            if _s == "lucro":
                st.session_state[_k_sinais] = ["lucro"]
            elif _s == "prejuizo":
                st.session_state[_k_sinais] = ["prejuizo"]
            else:
                st.session_state[_k_sinais] = []
            st.session_state.pop(f"{prefix_main}_sinal_resultado", None)
        elif isinstance(_leg_vs, str):
            _m = {"positiva": "lucro", "negativa": "prejuizo"}
            _one = _m.get(_leg_vs.strip().lower())
            st.session_state[_k_sinais] = (
                [_one] if _one else []
            )
        else:
            st.session_state[_k_sinais] = []
    _prev_s = st.session_state.get(_k_sinais)
    if not isinstance(_prev_s, list):
        st.session_state[_k_sinais] = []
    else:
        _filt = [x for x in _prev_s if x in ("lucro", "prejuizo", "empate")]
        st.session_state[_k_sinais] = _filt

    st.markdown(
        """
    <style>
    .tabela-nf-contador {
    color: #64748b;
    font-size: 0.85rem;
    margin: 8px 0 16px 0;
    }
    </style>
    """,
        unsafe_allow_html=True,
    )

    _col_tit, _col_acoes = st.columns([3, 1])
    with _col_tit:
        st.markdown(table_heading)
    with _col_acoes:
        _col_cfg, _col_csv = st.columns(2)
        with _col_cfg:
            with st.popover("⚙️"):
                st.caption(
                    "Colunas opcionais e «Diferença» pelas caixas abaixo; o CSV exportado segue as colunas visíveis."
                    if _ui_fiscal
                    else "Por defeito mostram-se receita, deduções principais e resultado. Marque abaixo para acrescentar ao quadro e ao CSV."
                )
                st.checkbox(
                    "Diferença (lista − fiscal)",
                    value=False,
                    key=f"{prefix_nf}_show_diferenca",
                    help="Receita de venda (lista) menos valor faturado na NF.",
                )
                st.markdown("**Mais colunas**")
                st.checkbox("Plataforma", key=f"{prefix_nf}_opt_plat", value=False)
                if not _ui_fiscal:
                    st.checkbox("Situação da NF", key=f"{prefix_nf}_opt_sit", value=False)
                st.checkbox("Pedido", key=f"{prefix_nf}_opt_ped", value=False)
                st.checkbox("Linhas", key=f"{prefix_nf}_opt_linhas", value=False)
                st.checkbox("Quantidade", key=f"{prefix_nf}_opt_qtd", value=False)
                if not _ui_fiscal:
                    st.checkbox("Faturado (NF)", key=f"{prefix_nf}_opt_vf", value=False)
                st.checkbox("Receita de Frete", key=f"{prefix_nf}_opt_rf", value=False)
                st.checkbox("Repasse transp.", key=f"{prefix_nf}_opt_rp", value=False)
                st.checkbox("Frete pedido (Σ)", key=f"{prefix_nf}_opt_tar", value=False)
                st.checkbox("Despesa fixa", key=f"{prefix_nf}_opt_df", value=False)
                if _nf_panel_ads_ui:
                    st.checkbox("ADS (3,5% + fixo)", key=f"{prefix_nf}_opt_ads", value=False)
                st.checkbox("Alertas", key=f"{prefix_nf}_opt_alert", value=False)
        with _col_csv:
            _nf_dl_hdr_slot = st.empty()

    if _ui_fiscal:
        _fz1, _fz2 = st.columns([2, 2])
        with _fz1:
            _ps_plat = (
                _faturamento_nf_platform_display_series(df_nf).fillna("").astype(str).str.strip()
                if not df_nf.empty
                else pd.Series(dtype=str)
            )
            _plat_opts = sorted({x for x in _ps_plat if x and x != "—"})
            if _plat_opts:
                if f"{prefix_nf}_tbl_plataforma" not in st.session_state:
                    st.session_state[f"{prefix_nf}_tbl_plataforma"] = []
                else:
                    _po_prev = st.session_state[f"{prefix_nf}_tbl_plataforma"]
                    if isinstance(_po_prev, list):
                        st.session_state[f"{prefix_nf}_tbl_plataforma"] = [
                            x for x in _po_prev if x in _plat_opts
                        ]
                st.multiselect(
                    "Plataforma",
                    options=_plat_opts,
                    key=f"{prefix_nf}_tbl_plataforma",
                    placeholder="Plataforma…",
                    label_visibility="collapsed",
                    help="Vazio = todas as plataformas do recorte. Refina só a tabela.",
                )
            else:
                st.caption("Sem plataforma no recorte.")
        with _fz2:
            st.text_input(
                "Buscar",
                key=f"{prefix_nf}_tbl_busca",
                placeholder="🔍 Buscar NF ou Pedido…",
                label_visibility="collapsed",
            )
        df_nf_panel = df_nf.copy()
    else:
        _f1, _f2, _f3, _f4 = st.columns([1.5, 2, 1.5, 2])
        with _f1:
            st.multiselect(
                "Status",
                options=("lucro", "prejuizo", "empate"),
                format_func=lambda x: {
                    "lucro": "Lucro",
                    "prejuizo": "Prejuízo",
                    "empate": "Neutro",
                }[x],
                key=_k_sinais,
                placeholder="Status…",
                label_visibility="collapsed",
                help=(
                    "**Vazio** = todas as NFs. «Neutro» = resultado ~0. "
                    "**Lucro** e **Prejuízo** juntos ⇒ sem filtro por sinal; caso contrário união das faixas escolhidas."
                ),
            )
        with _f2:
            if _prod_opts:
                _multiselect_stable(
                    f"{prefix_main}_prod",
                    "Produto",
                    _prod_opts,
                    compact_label=False,
                    placeholder="Filtrar por produto…",
                    label_visibility="collapsed",
                    help=(
                        "Vazio = todos. Filtra pela coluna «Produtos» (resumo na NF). "
                        "Não altera o topo fiscal nem os cards/DRE."
                    ),
                )
            else:
                st.caption("Sem produto no recorte — filtro indisponível.")
        with _f3:
            _ps_plat = (
                _faturamento_nf_platform_display_series(df_nf).fillna("").astype(str).str.strip()
                if not df_nf.empty
                else pd.Series(dtype=str)
            )
            _plat_opts = sorted({x for x in _ps_plat if x and x != "—"})
            if _plat_opts:
                if f"{prefix_nf}_tbl_plataforma" not in st.session_state:
                    st.session_state[f"{prefix_nf}_tbl_plataforma"] = []
                else:
                    _po_prev = st.session_state[f"{prefix_nf}_tbl_plataforma"]
                    if isinstance(_po_prev, list):
                        st.session_state[f"{prefix_nf}_tbl_plataforma"] = [
                            x for x in _po_prev if x in _plat_opts
                        ]
                st.multiselect(
                    "Plataforma",
                    options=_plat_opts,
                    key=f"{prefix_nf}_tbl_plataforma",
                    placeholder="Plataforma…",
                    label_visibility="collapsed",
                    help="Vazio = todas as plataformas do recorte. Refina só a tabela.",
                )
            else:
                st.caption("Sem plataforma no recorte.")
        with _f4:
            st.text_input(
                "Buscar",
                key=f"{prefix_nf}_tbl_busca",
                placeholder="🔍 Buscar NF ou Pedido…",
                label_visibility="collapsed",
            )

        _prod_sel = tuple(
            str(x).strip()
            for x in (st.session_state.get(f"{prefix_main}_prod") or [])
            if str(x).strip()
        )
        _sinais_ui = st.session_state.get(f"{prefix_main}_sinais_resultado")
        _sinais_tuple = (
            tuple(str(x).strip().lower() for x in _sinais_ui if str(x).strip())
            if isinstance(_sinais_ui, list)
            else ()
        )
        df_nf_panel = _faturamento_dre_apply_produto_e_sinal_venda(
            df_nf,
            produtos_sel=_prod_sel,
            sinais_resultado=_sinais_tuple,
        )
    # Cards e DRE: N_base fiscal + situação + enriquecimento; plataforma opcional; tabela = + produto/sinal.
    _kp_table = compute_nf_panel_kpis(df_nf_panel)
    _df_fiscal_kpi_anchor: pd.DataFrame | None = (
        _df_fiscal_base.copy() if use_fiscal_kpi and not _df_fiscal_base.empty else None
    )

    if _is_admin_mode():
        with st.expander("Diagnóstico materializado (admin)", expanded=False):
            _fdl_fat_min_aside(
                "<strong>Base fiscal</strong>: <code>dataset_faturamento_fiscal.parquet</code> — empresa + emissão + "
                "situação válida; filtro UI <strong>Situação da NF</strong> opcional. <strong>Painel NF</strong>: "
                "<code>dataset_faturamento_nf_panel.parquet</code> — <strong>cards/DRE</strong> = <strong>N_base</strong> "
                "+ situação + <strong>plataforma</strong> (se filtrada); <strong>tabela</strong> = + produto e sinal."
            )
            if len(df_nf_panel) != len(df_nf):
                _kp_pre_produto = compute_nf_panel_kpis(df_nf)
                _fdl_fat_min_aside(
                    "<strong>Conferência (tabela)</strong> — após plataforma vs após produto / resultado: "
                    f"venda lista <strong>{float(_kp_table['valor_venda']):.2f}</strong> vs "
                    f"<strong>{float(_kp_pre_produto['valor_venda']):.2f}</strong>; "
                    f"faturado NF <strong>{float(_kp_table['valor_faturado_nf']):.2f}</strong> vs "
                    f"<strong>{float(_kp_pre_produto['valor_faturado_nf']):.2f}</strong>; "
                    f"Σ resultado <strong>{float(_kp_table['resultado']):.2f}</strong> vs "
                    f"<strong>{float(_kp_pre_produto['resultado']):.2f}</strong>.",
                    tight=True,
                )
            if use_fiscal_parquet:
                _fdl_fat_min_aside(
                    "<strong>Plataforma</strong> **não** altera o **topo fiscal**; altera <strong>cards/DRE</strong> e "
                    "<strong>tabela</strong> (antes de produto/sinal).",
                    tight=True,
                )
            if load_info.get("faturamento_nf_panel_path"):
                _pp = html.escape(str(load_info.get("faturamento_nf_panel_path")))
                _fdl_fat_min_aside(f"Path painel NF: <code>{_pp}</code>", tight=True)
            if use_nf_materializado and load_info.get("faturamento_nf_first_path"):
                _p = html.escape(str(load_info.get("faturamento_nf_first_path")))
                _fdl_fat_min_aside(f"Path Parquet NF-first: <code>{_p}</code>", tight=True)
            elif load_info.get("faturamento_nf_first_skip") or load_info.get("faturamento_nf_first_error"):
                _sk = load_info.get("faturamento_nf_first_skip")
                _e = load_info.get("faturamento_nf_first_error")
                _parts = ["NF-first não ativo."]
                if _sk:
                    _parts.append(f"Motivo: <code>{html.escape(str(_sk))}</code>.")
                if _e:
                    _parts.append(f"Erro: <code>{html.escape(str(_e))}</code>.")
                _fdl_fat_min_aside(" ".join(_parts), tight=True)
            if load_info.get("faturamento_fiscal_path_resolution"):
                _pr = html.escape(str(load_info.get("faturamento_fiscal_path_resolution")))
                _fdl_fat_min_aside(f"Parquet fiscal resolvido via: <code>{_pr}</code>", tight=True)
            if use_fiscal_parquet and load_info.get("faturamento_fiscal_first_path"):
                _pf = html.escape(str(load_info.get("faturamento_fiscal_first_path")))
                _fdl_fat_min_aside(f"Path Parquet fiscal: <code>{_pf}</code>", tight=True)
            elif load_info.get("faturamento_fiscal_first_skip") or load_info.get("faturamento_fiscal_first_error"):
                _skf = load_info.get("faturamento_fiscal_first_skip")
                _ef = load_info.get("faturamento_fiscal_first_error")
                _parts_f = ["Parquet fiscal não ativo no carregamento (fallback ao faturado NF comercial)."]
                if _skf:
                    _parts_f.append(f"Motivo: <code>{html.escape(str(_skf))}</code>.")
                if _ef:
                    _parts_f.append(f"Erro: <code>{html.escape(str(_ef))}</code>.")
                _fdl_fat_min_aside(" ".join(_parts_f), tight=True)
            if use_nf_materializado and not use_fiscal_parquet:
                _fiscal_why: list[str] = []
                if load_info.get("faturamento_fiscal_user_hint"):
                    _fiscal_why.append(str(load_info["faturamento_fiscal_user_hint"]))
                elif load_info.get("faturamento_fiscal_first_error"):
                    _fiscal_why.append(f"Erro ao ler: {load_info['faturamento_fiscal_first_error']}")
                elif load_info.get("faturamento_fiscal_first_skip"):
                    _sk = str(load_info["faturamento_fiscal_first_skip"])
                    _fiscal_why.append(
                        "ficheiro ausente na pasta do materializado"
                        if _sk == "ficheiro_ausente"
                        else (
                            "materializado só por URL sem pasta local — não dá para ler o Parquet fiscal"
                            if _sk == "sem_path_local"
                            else _sk
                        )
                    )
                else:
                    _fiscal_why.append("Parquet fiscal não validado ou vazio após escopo")
                _fdl_fat_min_aside(
                    "<strong>Valor faturado (NF)</strong> neste ecrã está em <strong>NF-first (pedidos ligados)</strong>. "
                    "Para alinhar ao Bling, publique <code>dataset_faturamento_fiscal.parquet</code> junto do materializado. "
                    f"<strong>Estado fiscal:</strong> {' · '.join(html.escape(str(x)) for x in _fiscal_why)}"
                )
            if use_fiscal_kpi and _df_fiscal_kpi_anchor is not None:
                _aud_sum = float(
                    pd.to_numeric(_df_fiscal_kpi_anchor["Valor_Liquido_NF"], errors="coerce")
                    .fillna(0.0)
                    .sum()
                )
                _top_fiscal = float(_fiscal_base_stats.valor_liquido_fiscal_sum)
                _match = abs(_aud_sum - _top_fiscal) < 0.02
                _kp_vf_cards = float(_kp_cards["valor_faturado_nf"])
                _plat_empty = not _min_state.plataformas
                _vf_match_cards = abs(_kp_vf_cards - _top_fiscal) < 0.02 if _plat_empty else None
                _cards_vf_line = (
                    f"Σ <code>valor_faturado_nf</code> nos **cards/DRE** = <strong>{_kp_vf_cards:.2f}</strong> — "
                    f"com **Plataforma** vazia, deve coincidir com o topo: **{'sim' if _vf_match_cards else 'NÃO'}**."
                    if _plat_empty
                    else (
                        f"Σ <code>valor_faturado_nf</code> nos **cards/DRE** = <strong>{_kp_vf_cards:.2f}</strong> — com "
                        "**Plataforma** filtrada, **não** deve igualar o topo fiscal (subconjunto comercial por canal)."
                    )
                )
                _fdl_fat_min_aside(
                    "Auditoria fiscal (admin): "
                    f"<code>faturamento_fiscal_first</code>={load_info.get('faturamento_fiscal_first')!s}; "
                    f"Σ <code>Valor_Liquido_NF</code> no <strong>slice base</strong> (topo) = <strong>{_aud_sum:.2f}</strong> "
                    f"(deve coincidir com o valor principal **base fiscal líquida** no cartão): **{'sim' if _match else 'NÃO'}**. "
                    + _cards_vf_line,
                    tight=True,
                )
            if use_fiscal_parquet and isinstance(df_fiscal_pre, pd.DataFrame):
                _fdl_fat_min_aside(
                    f"Parquet fiscal (escopo org no carregamento): <strong>{len(df_fiscal_pre)}</strong> NF(s).",
                    tight=True,
                )

    _FAT_NF_TABLE_STYLER_MAX_ROWS = 500

    _show_col_diferenca = bool(st.session_state.get(f"{prefix_nf}_show_diferenca", False))
    _vf_opt_label = "Valor Fiscal" if _ui_fiscal else "Faturado (NF)"
    if _ui_fiscal:
        _nat_ok = bool(
            not df_nf_panel.empty
            and any(c in df_nf_panel.columns for c in ("Natureza_operacao", "Natureza"))
        )
        _nf_vis = [
            "Emissão",
            "Situação",
            "Empresa",
            "NF",
        ]
        if _nat_ok:
            _nf_vis.insert(_nf_vis.index("NF") + 1, "Natureza")
        _nf_vis.extend(["Valor Fiscal", "Base tributável", "Imposto"])
        if _show_col_diferenca:
            _nf_vis.append("Diferença")
    else:
        _nf_vis = [
            "Emissão",
            "Status",
            "Empresa",
            "NF",
            "Produtos",
            "Receita de Venda",
            "Comissão",
            "Custo produto",
            "Frete plataforma",
            "Imposto",
            "Resultado",
            "Margem %",
        ]
        if _show_col_diferenca:
            _nf_vis.append("Diferença")
    _nf_opt_cols: list[tuple[str, str]] = [
        (f"{prefix_nf}_opt_plat", "Plataforma"),
        (f"{prefix_nf}_opt_sit", "Situação"),
        (f"{prefix_nf}_opt_ped", "Pedido"),
        (f"{prefix_nf}_opt_linhas", "Linhas"),
        (f"{prefix_nf}_opt_qtd", "Quantidade"),
        (f"{prefix_nf}_opt_vf", _vf_opt_label),
        (f"{prefix_nf}_opt_rf", "Receita de Frete"),
        (f"{prefix_nf}_opt_rp", "Repasse transp."),
        (f"{prefix_nf}_opt_tar", "Frete pedido (Σ)"),
        (f"{prefix_nf}_opt_df", "Despesa fixa"),
    ]
    for _ok, _colname in _nf_opt_cols:
        if _ui_fiscal and _ok == f"{prefix_nf}_opt_sit":
            continue
        if _ui_fiscal and _ok == f"{prefix_nf}_opt_vf":
            continue
        if bool(st.session_state.get(_ok, False)):
            _nf_vis.append(_colname)
    if _nf_panel_ads_ui and bool(st.session_state.get(f"{prefix_nf}_opt_ads", False)):
        _nf_vis.extend(["ADS 3,5%", "ADS fixo"])
    if bool(st.session_state.get(f"{prefix_nf}_opt_alert", False)):
        _nf_vis.append("Alertas")

    _nf_table_cols_order_ui: list[str] = []
    for _c in _nf_vis:
        if _c in ("ADS 3,5%", "ADS fixo") and not _nf_panel_ads_ui:
            continue
        _nf_table_cols_order_ui.append(_c)

    _df_nf_table = df_nf_panel
    if not df_nf_panel.empty and "Nota_Data_Emissao" in df_nf_panel.columns:
        _tmp_sort = df_nf_panel.copy()
        _tmp_sort["_fdl_nf_emi_ord"] = pd.to_datetime(
            _df_get_series_column(_tmp_sort, "Nota_Data_Emissao"),
            errors="coerce",
            dayfirst=False,
        )
        _df_nf_table = (
            _tmp_sort.sort_values("_fdl_nf_emi_ord", ascending=False, na_position="last")
            .drop(columns=["_fdl_nf_emi_ord"], errors="ignore")
            .reset_index(drop=True)
        )

    _disp_nf_full = pd.DataFrame()
    _disp_nf_ui = pd.DataFrame()
    if not _df_nf_table.empty:
        _plat_s = _faturamento_nf_platform_display_series(_df_nf_table).astype(str)
        _marg_ratio = _nf_row_margem_resultado_venda_ratio(
            _df_nf_table["valor_venda"],
            _df_nf_table["resultado"],
        )
        _custo_s = (
            _df_nf_table["custo_produto"]
            if "custo_produto" in _df_nf_table.columns
            else pd.Series(0.0, index=_df_nf_table.index, dtype=float)
        )
        _inc_flag = (
            _df_nf_table["comercial_incompleto"].fillna(False).astype(bool)
            if "comercial_incompleto" in _df_nf_table.columns
            else pd.Series(False, index=_df_nf_table.index, dtype=bool)
        )
        _ads_v_s = (
            pd.to_numeric(_df_nf_table["custo_ads_variavel"], errors="coerce").fillna(0.0)
            if "custo_ads_variavel" in _df_nf_table.columns
            else pd.Series(0.0, index=_df_nf_table.index, dtype=float)
        )
        _ads_f_s = (
            pd.to_numeric(_df_nf_table["custo_ads_fixo"], errors="coerce").fillna(0.0)
            if "custo_ads_fixo" in _df_nf_table.columns
            else pd.Series(0.0, index=_df_nf_table.index, dtype=float)
        )
        _res_line_nf = pd.to_numeric(_df_nf_table["resultado"], errors="coerce")
        _vv_num = pd.to_numeric(_df_nf_table["valor_venda"], errors="coerce").fillna(0.0)
        _cm_num = pd.to_numeric(_df_nf_table["comissao"], errors="coerce").fillna(0.0)
        _imp_num = pd.to_numeric(_df_nf_table["imposto"], errors="coerce").fillna(0.0)
        _cp_num = pd.to_numeric(_custo_s, errors="coerce").fillna(0.0)
        _eps_z = 1e-6
        _sem_mov = (
            (_vv_num.abs() <= _eps_z)
            & (_cp_num.abs() <= _eps_z)
            & (_res_line_nf.fillna(0.0).abs() <= _eps_z)
            & (_cm_num.abs() <= _eps_z)
            & (_imp_num.abs() <= _eps_z)
        )
        if _is_admin_mode() and bool(_sem_mov.any()):
            st.caption(f"Admin: {_fmt_int_ptbr(int(_sem_mov.sum()))} NF(s) só com zeros nos principais valores comerciais.")

        def _nf_alert_txt(i: int) -> str:
            _parts: list[str] = []
            if bool(_sem_mov.iloc[i]):
                _parts.append("NF sem movimento comercial")
            if bool(_inc_flag.iloc[i]):
                _parts.append("Falta custo / dados")
            return " · ".join(_parts) if _parts else "—"

        _alertas_col = pd.Series(
            [_nf_alert_txt(i) for i in range(len(_df_nf_table))],
            index=_df_nf_table.index,
            dtype=object,
        )
        _qtd_itens = _faturamento_nf_quantidade_itens_por_nf(df, _df_nf_table)

        def _nf_status_label_nf(r: object) -> str:
            try:
                x = float(r)
            except (TypeError, ValueError):
                return "—"
            if pd.isna(x):
                return "—"
            if x > 0:
                return "Lucro"
            if x < 0:
                return "Prejuízo"
            return "Neutro"

        _disp_nf_full = pd.DataFrame(
            {
                "Emissão": _series_nf_emissao_pt_br(
                    _df_get_series_column(_df_nf_table, "Nota_Data_Emissao")
                ),
                "Status": _res_line_nf.map(_nf_status_label_nf),
                "Empresa": _series_empty_str_to_dash(_df_get_series_column(_df_nf_table, "empresa")),
                "Plataforma": _plat_s,
                "NF": _df_get_series_column(_df_nf_table, "Nota_Numero_Normalizado")
                .fillna("")
                .map(lambda v: str(v).strip()),
                "Situação": _series_empty_str_to_dash(
                    _df_get_series_column(_df_nf_table, "Nota_Situacao")
                ),
                "Pedido": _faturamento_disp_texto_sem_none(_df_nf_table["pedido_resumo"]),
                "Produtos": _faturamento_disp_texto_sem_none(_df_nf_table["produto_resumo"]),
                "Linhas": _df_nf_table["n_linhas_pedido"].astype(int),
                "Quantidade": _qtd_itens,
                "Receita de Venda": pd.to_numeric(_df_nf_table["valor_venda"], errors="coerce"),
                "Faturado (NF)": pd.to_numeric(_df_nf_table["valor_faturado_nf"], errors="coerce"),
                **(
                    {"Diferença": pd.to_numeric(_df_nf_table["diferenca"], errors="coerce")}
                    if _show_col_diferenca
                    else {}
                ),
                "Comissão": pd.to_numeric(_df_nf_table["comissao"], errors="coerce"),
                "Custo produto": pd.to_numeric(_custo_s, errors="coerce").fillna(0.0),
                "Receita de Frete": pd.to_numeric(
                    _df_nf_table["receita_frete_tp"], errors="coerce"
                )
                if "receita_frete_tp" in _df_nf_table.columns
                else pd.Series(0.0, index=_df_nf_table.index),
                "Frete plataforma": pd.to_numeric(
                    _df_nf_table["custo_frete_plataforma"], errors="coerce"
                )
                if "custo_frete_plataforma" in _df_nf_table.columns
                else pd.Series(0.0, index=_df_nf_table.index),
                "Repasse transp.": pd.to_numeric(
                    _df_nf_table["repasse_frete_transportadora_propria"], errors="coerce"
                )
                if "repasse_frete_transportadora_propria" in _df_nf_table.columns
                else pd.Series(0.0, index=_df_nf_table.index),
                "Frete pedido (Σ)": pd.to_numeric(
                    _df_nf_table["tarifa_custo_envio"], errors="coerce"
                )
                if "tarifa_custo_envio" in _df_nf_table.columns
                else pd.Series(0.0, index=_df_nf_table.index),
                "Imposto": pd.to_numeric(_df_nf_table["imposto"], errors="coerce"),
                "Despesa fixa": pd.to_numeric(_df_nf_table["despesa_fixa"], errors="coerce"),
                "ADS 3,5%": _ads_v_s,
                "ADS fixo": _ads_f_s,
                "Resultado": _res_line_nf,
                "Alertas": _alertas_col.astype(str),
                "Margem %": (_marg_ratio * 100.0),
            }
        )
        _nat_col_tb = next(
            (c for c in ("Natureza_operacao", "Natureza") if c in _df_nf_table.columns),
            None,
        )
        if _nat_col_tb:
            _disp_nf_full["Natureza"] = _series_empty_str_to_dash(_df_nf_table[_nat_col_tb])
        if _ui_fiscal and "Faturado (NF)" in _disp_nf_full.columns:
            _disp_nf_full = _disp_nf_full.rename(columns={"Faturado (NF)": "Valor Fiscal"})
            _disp_nf_full["Base tributável"] = _disp_nf_full["Valor Fiscal"]
        _disp_nf_full = _disp_nf_full[_nf_table_cols_order_ui]
        _disp_nf_ui = _disp_nf_full.copy()

        def _fat_min_trunc_text_cell(v: object, max_len: int = 72) -> str:
            t = str(v).strip()
            if t in ("", "—", "nan") or len(t) <= max_len:
                return t if t else "—"
            return t[: max_len - 1] + "…"

        def _nf_tbl_money_str(x: object) -> str:
            if x is None:
                return "—"
            try:
                if pd.isna(x):
                    return "—"
            except TypeError:
                pass
            s = _fmt_brl_ptbr_celula(x)
            return s if s else "—"

        def _nf_tbl_linhas_str(x: object) -> str:
            try:
                if pd.isna(x):
                    return "—"
            except TypeError:
                pass
            try:
                n = int(round(float(x)))
            except (TypeError, ValueError):
                return "—"
            return _fmt_int_ptbr(n)

        def _nf_tbl_margem_str(ratio_times_100: object) -> str:
            """``Margem %`` no export numérico = ratio×100; reconstrói ratio para o mesmo formato do painel."""
            try:
                if pd.isna(ratio_times_100):
                    return "—"
            except TypeError:
                return "—"
            try:
                pct = float(ratio_times_100)
            except (TypeError, ValueError):
                return "—"
            if math.isnan(pct) or math.isinf(pct):
                return "—"
            return _fmt_pct_ptbr_ratio(pct / 100.0, decimals=1)

        for _money_col in (
            "Receita de Venda",
            "Faturado (NF)",
            "Valor Fiscal",
            "Base tributável",
            "Diferença",
            "Comissão",
            "Custo produto",
            "Receita de Frete",
            "Frete plataforma",
            "Repasse transp.",
            "Frete pedido (Σ)",
            "Imposto",
            "Despesa fixa",
            "ADS 3,5%",
            "ADS fixo",
            "Resultado",
        ):
            if _money_col in _disp_nf_ui.columns:
                _disp_nf_ui[_money_col] = _disp_nf_ui[_money_col].map(_nf_tbl_money_str)
        if "Linhas" in _disp_nf_ui.columns:
            _disp_nf_ui["Linhas"] = _disp_nf_ui["Linhas"].map(_nf_tbl_linhas_str)
        if "Quantidade" in _disp_nf_ui.columns:
            _disp_nf_ui["Quantidade"] = _disp_nf_ui["Quantidade"].map(_nf_tbl_linhas_str)
        if "Margem %" in _disp_nf_ui.columns:
            _disp_nf_ui["Margem %"] = _disp_nf_full["Margem %"].map(_nf_tbl_margem_str)
        if "Pedido" in _disp_nf_ui.columns:
            _disp_nf_ui["Pedido"] = _disp_nf_ui["Pedido"].map(
                lambda x: _fat_min_trunc_text_cell(x, 72)
            )
        if "Produtos" in _disp_nf_ui.columns:
            _disp_nf_ui["Produtos"] = _disp_nf_ui["Produtos"].map(
                lambda x: _fat_min_trunc_text_cell(x, 72)
            )

        _plat_filt = st.session_state.get(f"{prefix_nf}_tbl_plataforma") or []
        if not isinstance(_plat_filt, list):
            _plat_filt = []
        _ps_pf = (
            _faturamento_nf_platform_display_series(df_nf).fillna("").astype(str).str.strip()
            if not df_nf.empty
            else pd.Series(dtype=str)
        )
        _plat_avail_nf = {x for x in _ps_pf if x and x != "—"}
        _plat_filt = [x for x in _plat_filt if x in _plat_avail_nf]
        _busca_filt = str(st.session_state.get(f"{prefix_nf}_tbl_busca") or "")
        _nf_tbl_n_antes_extra = len(_disp_nf_full)
        _nf_tbl_mask = _fdl_nf_table_filter_mask(
            _disp_nf_full,
            plataformas_sel=_plat_filt,
            busca=_busca_filt,
        )
        _disp_nf_full = _disp_nf_full.loc[_nf_tbl_mask].reset_index(drop=True)
        _disp_nf_ui = _disp_nf_ui.loc[_nf_tbl_mask].reset_index(drop=True)
    else:
        _nf_tbl_n_antes_extra = 0

    _cfg_nf: dict[str, NumberColumn | TextColumn] = {}
    _nf_col_help: dict[str, str | None] = {
        "Emissão": None,
        "Status": "Lucro, prejuízo ou neutro (resultado ~0) conforme o resultado consolidado da NF.",
        "Empresa": None,
        "Plataforma": "«—» = NF sem canal comercial associado neste recorte.",
        "NF": None,
        "Situação": (
            "Situação da NF no export fiscal (emitida, cancelada, autorizada pendente…)."
            if _ui_fiscal
            else None
        ),
        "Natureza": (
            "Natureza da operação quando existir no export ou materializado."
            if _ui_fiscal
            else None
        ),
        "Pedido": "«—» = sem pedido comercial resolvido para esta NF. Texto completo no CSV.",
        "Produtos": "«—» = sem produto agregado na NF. Texto completo no CSV.",
        "Linhas": "Quantidade de linhas de pedido agregadas nesta NF (comercial).",
        "Quantidade": "Soma de **Quantidade** (unidades) nas linhas de pedido ligadas a esta NF (materializado linha).",
        "Receita de Venda": (
            "Comercial: Σ Quantidade × Preço de lista dos pedidos ligados a esta NF (0 se não houver vínculo)."
            if use_fiscal_kpi
            else "Σ Quantidade × Preço de lista (pedidos ligados à NF)."
        ),
        "Faturado (NF)": (
            "Valor líquido da NF na referência fiscal (Bling / export), 1× por nota no período."
            if use_fiscal_kpi
            else "Valor líquido da NF (uma vez por nota) na base materializada."
        ),
        "Valor Fiscal": (
            "Valor líquido da NF (referência fiscal), uma vez por nota no período."
            if _ui_fiscal
            else None
        ),
        "Base tributável": (
            "Base de referência nesta linha (no painel atual coincide com **Valor Fiscal** quando não há coluna de base granular)."
            if _ui_fiscal
            else None
        ),
        "Diferença": (
            "Comercial − fiscal nesta linha: Receita de Venda − Faturado (NF). "
            "Interpretar como ponte lista↔nota, não como erro automático."
            if use_fiscal_kpi
            else "Receita de Venda − Faturado (NF)."
        ),
        "Comissão": "Comercial: soma das comissões das linhas de pedido ligadas à NF." if use_fiscal_kpi else None,
        "Custo produto": (
            "Comercial: Σ **Custo_Produto_Total** (ou «Custo do Produto») das linhas de pedido ligadas à NF."
            if use_fiscal_kpi
            else "Σ custo do produto nas linhas de pedido desta NF."
        ),
        "Receita de Frete": (
            "Frete destacado na **nota fiscal** (``Frete_Nota_Export`` no merge fiscal), por NF."
            if use_fiscal_kpi
            else "Receita de frete na NF / gap comercial quando sem fiscal."
        ),
        "Frete plataforma": (
            "Custo de logística da **plataforma** (ME / «Frete_Plataforma»), após separar do repasse TP quando aplicável."
            if use_fiscal_kpi
            else "Parcela plataforma do frete no pedido."
        ),
        "Repasse transp.": (
            "Repasse à **transportadora própria** (parcela TP do «Custo de Frete»); se o pedido não separa modalidade "
            "mas a NF cobra frete, imputa-se pass-through até o teto da tarifa da NF."
            if use_fiscal_kpi
            else "Repasse TP / imputação alinhada à receita NF."
        ),
        "Frete pedido (Σ)": (
            "Σ **Custo de Frete** (ou «Frete_Plataforma») no pedido — conferência (plataforma + repasse após coerência)."
            if use_fiscal_kpi
            else "Σ frete no pedido."
        ),
        "Imposto": (
            "Imposto alocado a esta NF no materializado; total do período alinha-se à Apuração Fiscal via **base fiscal líquida** × taxa efetiva."
            if _ui_fiscal
            else ("Comercial: soma do imposto das linhas de pedido ligadas à NF." if use_fiscal_kpi else None)
        ),
        "Despesa fixa": (
            "Comercial: 5% sobre valor da venda (lista) agregado à NF."
            if use_fiscal_kpi
            else "5% sobre valor da venda agregado à NF."
        ),
        "ADS 3,5%": "3,5% × receita de venda (lista) nesta NF — custo de mídia (materializado).",
        "ADS fixo": "R$ 2,00 quando a receita de venda (lista) > 0 nesta NF (materializado).",
        "Resultado": (
            "Comercial: resultado consolidado por NF **já líquido de ADS** (materializado)."
            if use_fiscal_kpi
            else "Resultado consolidado por NF **já líquido de ADS** (materializado)."
        ),
        "Margem %": (
            None
            if _ui_fiscal
            else (
                "Comercial: Resultado ÷ Receita de Venda nesta NF; não usa valor faturado fiscal."
                if use_fiscal_kpi
                else "Resultado ÷ Receita de Venda; alinhado ao KPI «Margem %» do painel."
            )
        ),
        "Alertas": (
            "«NF sem movimento comercial» quando receita, custo, comissão, imposto e resultado são ~0. "
            "«Falta custo / dados» quando o materializado sinaliza dados incompletos."
        ),
    }
    _nf_col_width: dict[str, str] = {
        "Emissão": "small",
        "Status": "small",
        "Empresa": "medium",
        "Plataforma": "small",
        "NF": "small",
        "Situação": "small",
        "Pedido": "large",
        "Produtos": "large",
        "Linhas": "small",
        "Quantidade": "small",
        "Receita de Venda": "medium",
        "Faturado (NF)": "medium",
        "Valor Fiscal": "medium",
        "Base tributável": "medium",
        "Natureza": "small",
        "Diferença": "small",
        "Comissão": "small",
        "Custo produto": "medium",
        "Receita de Frete": "small",
        "Frete plataforma": "small",
        "Repasse transp.": "small",
        "Frete pedido (Σ)": "small",
        "Imposto": "small",
        "Despesa fixa": "small",
        "ADS 3,5%": "small",
        "ADS fixo": "small",
        "Resultado": "medium",
        "Alertas": "medium",
        "Margem %": "small",
    }
    for _cn in _nf_table_cols_order_ui:
        if _cn not in _disp_nf_ui.columns:
            continue
        _w = _nf_col_width.get(_cn, "medium")
        _h = _nf_col_help.get(_cn)
        _cfg_nf[_cn] = (
            TextColumn(_cn, width=_w, help=_h) if _h else TextColumn(_cn, width=_w)
        )

    _nf_dl_n = len(_disp_nf_ui)
    _nf_dl_scope = _nf_tbl_n_antes_extra if _nf_tbl_n_antes_extra else _nf_dl_n
    _nf_dl_hdr_slot.download_button(
        "📥 CSV",
        _disp_nf_full.to_csv(index=False).encode("utf-8-sig") if not _disp_nf_full.empty else b"",
        file_name=csv_file_name,
        mime="text/csv",
        key=f"{prefix_main}_dl_hdr_{_oid}",
        disabled=_disp_nf_full.empty,
    )

    with st.container(border=True):
        _nf_cap_txt = ""
        if _nf_dl_scope and _nf_dl_n == _nf_dl_scope:
            _nf_cap_txt = (
                f"{_nf_dl_scope:,} notas · emissão em ordem decrescente · CSV alinhado às colunas visíveis."
                if use_fiscal_kpi
                else f"{_nf_dl_scope:,} notas · emissão decrescente · CSV alinhado às colunas visíveis."
            )
        elif _nf_dl_scope:
            _nf_cap_txt = (
                f"Mostrando {_nf_dl_n:,} de {_nf_dl_scope:,} notas · emissão decrescente."
            )
        else:
            _nf_cap_txt = "Sem linhas para exibir com os filtros atuais."
        st.markdown(f"<p class='tabela-nf-contador'>{html.escape(_nf_cap_txt)}</p>", unsafe_allow_html=True)
        if _disp_nf_ui.empty:
            if _ui_fiscal:
                st.info(
                    "Sem linhas na **tabela** com os filtros atuais (plataforma ou busca). "
                    "Confira **período de emissão** e **empresa** nos filtros acima."
                )
            else:
                st.info(
                    (
                        "Sem linhas na **tabela** com os filtros atuais (status, produto, plataforma, busca). "
                        "O **topo fiscal** e os **cards/DRE** podem ainda ter **N_base** > 0 — confira período e empresa."
                    )
                    if use_fiscal_kpi
                    else (
                        "Sem linhas no recorte (confirme **período de emissão**, **empresa** e **plataforma**)."
                    )
                )
        else:
            _nf_page_sz = 25
            _nf_total_rows = len(_disp_nf_ui)
            _nf_pages = max(1, (_nf_total_rows + _nf_page_sz - 1) // _nf_page_sz)
            _pg_sel = 1
            if _nf_pages > 1:
                _pg_a, _pg_b = st.columns((1, 2))
                with _pg_a:
                    _pg_sel = int(
                        st.number_input(
                            "Página",
                            min_value=1,
                            max_value=int(_nf_pages),
                            value=1,
                            step=1,
                            key=f"{prefix_main}_nf_pg",
                        )
                    )
                with _pg_b:
                    _i0p = (int(_pg_sel) - 1) * _nf_page_sz
                    _i1p = min(_i0p + _nf_page_sz, _nf_total_rows)
                    st.caption(
                        f"Mostrando **{_i0p + 1}**–**{_i1p}** de **{_nf_total_rows}** notas. "
                        "Use o cabeçalho da tabela para ordenar a página atual."
                    )
            else:
                st.caption(f"**{_nf_total_rows}** nota(s) no recorte (ordenar pelo cabeçalho da coluna).")
            _i0 = (int(_pg_sel) - 1) * _nf_page_sz
            _slice_ui = _disp_nf_ui.iloc[_i0 : _i0 + _nf_page_sz].copy()
            _slice_num = _disp_nf_full.iloc[_i0 : _i0 + _nf_page_sz].reset_index(drop=True)
            _slice_ui_r = _slice_ui.reset_index(drop=True)

            def _nf_style_status_col(s: pd.Series) -> list[str]:
                out: list[str] = []
                for v in s.astype(object):
                    vs = str(v).strip()
                    if vs == "Lucro":
                        c = CORES_STATUS["Lucro"]
                        out.append(
                            f"background-color: #dcfce7; color: {c}; font-weight: 500; font-size: 0.75rem"
                        )
                    elif vs == "Prejuízo":
                        c = CORES_STATUS["Prejuízo"]
                        out.append(
                            f"background-color: #fee2e2; color: {c}; font-weight: 500; font-size: 0.75rem"
                        )
                    elif vs == "Neutro":
                        c = CORES_STATUS["Neutro"]
                        out.append(
                            f"background-color: #f3f4f6; color: {c}; font-weight: 500; font-size: 0.75rem"
                        )
                    else:
                        out.append("")
                return out

            def _nf_row_highlight_fat(r: pd.Series) -> list[str]:
                ri = r.name
                try:
                    res = float(pd.to_numeric(_slice_num.loc[ri, "Resultado"], errors="coerce"))
                except Exception:
                    res = 0.0
                c = "background-color: #fef2f2" if res < 0 else ""
                return [c] * len(r)

            _h_tbl = min(440, 132 + 34 * min(len(_slice_ui_r), 14))
            _df_arg: object = _slice_ui_r
            _nf_styler_ok = (not _fdl_safe_mode()) and int(_slice_ui_r.size) < 200_000
            if _nf_styler_ok:
                try:
                    if "Status" in _slice_ui_r.columns:
                        _st_obj = _slice_ui_r.style.apply(_nf_style_status_col, subset=["Status"], axis=0)
                    else:
                        _st_obj = _slice_ui_r.style
                    if _nf_total_rows <= _FAT_NF_TABLE_STYLER_MAX_ROWS and "Resultado" in _slice_num.columns:
                        _st_obj = _st_obj.apply(_nf_row_highlight_fat, axis=1)
                    _df_arg = _st_obj
                except Exception:
                    _df_arg = _slice_ui_r
            if _nf_total_rows > _FAT_NF_TABLE_STYLER_MAX_ROWS and _nf_styler_ok:
                st.caption("Destaque de linha para prejuízo ativo até **500** notas.")
            st.dataframe(
                _df_arg,
                use_container_width=True,
                hide_index=True,
                height=_h_tbl,
                column_config=_cfg_nf,
            )

    _fdl_fat_min_vsp(size="md")

FDL_RG_MSG_SEM_EMPRESA = "Selecione pelo menos uma empresa para visualizar o resultado."
_FDL_RG_TOOLTIP_MODULO = (
    "DRE e KPIs filtram pela data da venda; o imposto na DRE segue a ponte fiscal (apuração por NF). "
    "Canceladas/denegadas/inutilizadas já estão excluídas da base materializada."
)


def _fdl_rg_resumo_filtros_linha() -> str:
    """Linha única para rótulo do expander de filtros (estado atual da sessão)."""
    parts: list[str] = []
    em = st.session_state.get("fdl_fat_min_emp")
    if isinstance(em, list) and em:
        parts.append(" · ".join(str(x) for x in em))
    else:
        parts.append("Todas as empresas")
    pl = st.session_state.get("fdl_fat_min_plat")
    if isinstance(pl, list) and pl:
        parts.append(" · ".join(str(x) for x in pl))
    else:
        parts.append("Todas as plataformas")
    d0 = st.session_state.get("fdl_fat_min_nf_d_ini")
    d1 = st.session_state.get("fdl_fat_min_nf_d_fim")
    if isinstance(d0, date) and isinstance(d1, date):
        parts.append(f"{d0.strftime('%d/%m/%Y')} — {d1.strftime('%d/%m/%Y')}")
    else:
        parts.append("Período")
    return " · ".join(parts)


def _fdl_rg_header_context(min_state: FaturamentoRecorteMinState) -> str:
    """Subtítulo unificado (DRE + Painel): empresa selecionada, consolidado ou recorte aberto."""
    ne = len(min_state.empresas)
    if ne >= 2:
        return f"Consolidado · {ne} empresas"
    if ne == 1:
        return str(min_state.empresas[0])
    return "Todas as empresas no recorte"


def _fdl_health_panel_rg_benchmark_margins(
    *,
    df_linha: pd.DataFrame,
    df_nf_pre: pd.DataFrame,
    df_commercial_nf_kpi: pd.DataFrame,
    df_fiscal_pre: pd.DataFrame,
    df_devolucoes_pre: pd.DataFrame | None,
    devolucoes_ok: bool,
    use_fiscal_parquet: bool,
    use_fiscal_kpi: bool,
    ok_nf_dates: bool,
    fiscal_base_stats: FaturamentoFiscalBaseStats | None,
    nf_d_ini: date,
    nf_d_fim: date,
    ano_civil: int,
    mes_civil: int,
    empresas_sel: tuple[str, ...],
    plataformas_sel: tuple[str, ...],
    situacoes_nf: tuple[str, ...],
    org_sidebar: str,
) -> float | None:
    """Margem % do período anterior — mesma fonte dos KPIs/DRE (``compute_resultado_gerencial_kpis``)."""
    from calendar import monthrange

    from processing.faturamento.resultado_gerencial_slice import (
        build_resultado_gerencial_slice,
        compute_resultado_gerencial_kpis,
    )

    margem_ant: float | None = None
    _ = df_commercial_nf_kpi, org_sidebar

    try:
        if mes_civil == 1:
            py, pm = ano_civil - 1, 12
        else:
            py, pm = ano_civil, mes_civil - 1
        p_ini = date(py, pm, 1)
        p_fim = date(py, pm, monthrange(py, pm)[1])

        _df_fb_p, fst_p = build_faturamento_fiscal_base_slice(
            df_fiscal_pre if use_fiscal_parquet else pd.DataFrame(),
            empresas_sel=empresas_sel,
            nf_d_ini=p_ini,
            nf_d_fim=p_fim,
            ok_nf_dates=ok_nf_dates,
            situacoes_sel=situacoes_nf,
            df_devolucoes=df_devolucoes_pre if devolucoes_ok else None,
        )
        df_nf_p = _faturamento_nf_apply_minimal_recorte(
            df_nf_pre,
            empresas_sel=empresas_sel,
            plataformas_sel=(),
            nf_d_ini=p_ini,
            nf_d_fim=p_fim,
            ok_nf_dates=ok_nf_dates,
        )
        df_nf_p = _faturamento_nf_filter_by_situacao(df_nf_p, situacoes_nf)
        if _df_fb_p.empty:
            df_nf_p_kpi = df_nf_p.copy()
        else:
            df_nf_p_kpi = build_nf_panel_aligned_to_fiscal_base(_df_fb_p, df_nf_p)
        kp_p = compute_nf_panel_kpis(df_nf_p_kpi)
        imp_p = dre_imposto_para_linha_dre_gerencial(
            kp_p,
            fiscal_base_stats=fst_p if use_fiscal_parquet else None,
            aplicar_ponte_base_liquida=(fst_p is not None and use_fiscal_kpi),
        )
        slice_prev = build_resultado_gerencial_slice(
            df_linha,
            empresas_sel=empresas_sel,
            plataformas_sel=plataformas_sel,
            data_venda_ini=p_ini,
            data_venda_fim=p_fim,
        )
        kp_rg_p = compute_resultado_gerencial_kpis(slice_prev, fiscal_imposto_valor=imp_p)
        margem_ant = float(kp_rg_p["margem"]) * 100.0
    except Exception:
        margem_ant = None

    return margem_ant


def _render_faturamento_dre_minimal(
    df: pd.DataFrame,
    load_info: dict[str, object],
    ts_proc: str,
    *,
    org_id: str,
    org_display_name: str,
) -> None:
    """
    Painel **NF-first** só a partir de ``dataset_faturamento_nf_panel.parquet``: merge fiscal↔comercial, frete e
    **resultado** já calculados na materialização — **sem** grão, merge ou ajustes comerciais no Streamlit.
    """
    _fdl_fat_min_inject_ui_styles()
    _upd_disp = _fdl_fat_min_format_updated_at(ts_proc)
    _h_periodo = (
        f"{_FATURAMENTO_HELP_PERIODO_DATA_VENDA_RG_MIN} "
        f"Última atualização dos dados carregados: {_upd_disp}."
    )
    st.html(
        _build_faturamento_dre_page_header_html(
            updated_at=None,
            sobre_tooltip=_FDL_RG_TOOLTIP_MODULO,
        )
    )
    _fdl_fat_min_vsp(size="sm")
    _oid = str(org_id)
    _ = org_display_name, load_info
    use_nf_panel_baked = bool(load_info.get("faturamento_nf_panel_baked"))
    _df_nf_panel = load_info.get("faturamento_nf_panel_df")
    _df_nf_contract = load_info.get("faturamento_nf_df")
    use_nf_panel_baked_effective = (
        use_nf_panel_baked
        and isinstance(_df_nf_panel, pd.DataFrame)
        and nf_panel_materializado_dataframe_valid(_df_nf_panel)
    )
    df_nf_pre = _df_nf_panel if use_nf_panel_baked_effective else _df_nf_contract
    use_nf_materializado = False
    if use_nf_panel_baked_effective:
        use_nf_materializado = isinstance(df_nf_pre, pd.DataFrame) and not df_nf_pre.empty
    elif (
        bool(load_info.get("faturamento_nf_first"))
        and isinstance(df_nf_pre, pd.DataFrame)
        and nf_first_contract_dataframe_valid(df_nf_pre)
    ):
        use_nf_materializado = True
    if use_nf_materializado and isinstance(df_nf_pre, pd.DataFrame) and df_nf_pre.empty and not df.empty:
        use_nf_materializado = False

    _nf_panel_ads_ui = bool(load_info.get("faturamento_nf_panel_ads", True))

    df_fiscal_pre = load_info.get("faturamento_fiscal_df")
    df_devolucoes_pre = load_info.get("faturamento_devolucoes_df")
    _df_dev_ok = isinstance(df_devolucoes_pre, pd.DataFrame)
    use_fiscal_parquet = (
        bool(load_info.get("faturamento_fiscal_first"))
        and isinstance(df_fiscal_pre, pd.DataFrame)
        and fiscal_contract_dataframe_valid(df_fiscal_pre)
    )
    use_fiscal_kpi = bool(
        use_fiscal_parquet and isinstance(df_fiscal_pre, pd.DataFrame) and fiscal_contract_dataframe_valid(df_fiscal_pre)
    )

    if not use_nf_panel_baked_effective:
        st.error(
            "**Dados por nota fiscal indisponíveis.** "
            "Esta área usa a base consolidada **já publicada** para a sua organização."
        )
        st.caption(
            "Peça a atualização pelo processo habitual de fecho ou aguarde a próxima publicação de dados."
        )
        _pe = load_info.get("faturamento_nf_panel_error")
        if _pe:
            _pe_line = (
                f"Erro técnico: `{html.escape(str(_pe))}`"
                if _is_admin_mode()
                else "Não foi possível carregar a base consolidada. Tente novamente mais tarde ou contacte o suporte."
            )
            st.caption(_pe_line)
        elif not use_nf_panel_baked:
            st.caption(
                "A base consolidada ainda não está disponível neste ambiente ou não foi encontrada. "
                "Volte mais tarde ou contacte o suporte."
            )
        elif not isinstance(_df_nf_panel, pd.DataFrame):
            st.caption("Não foi possível preparar a tabela neste momento. Recarregue a página ou tente mais tarde.")
        elif _df_nf_panel.empty:
            st.info(
                "Não há linhas para o período e empresa selecionados. "
                "Verifique filtros ou o escopo na barra lateral."
            )
        elif _is_admin_mode():
            st.warning(
                "Contrato do painel incompleto (faltam colunas obrigatórias). "
                "Rematerialize o faturamento com a versão atual do pipeline."
            )
        return

    if df.empty and not use_nf_materializado:
        st.info(
            "Sem dados de faturamento para este escopo. Confirme **materialização**, **slug** do cliente "
            "e o **escopo** (empresa ativa / consolidado) na barra lateral."
        )
        return

    _bounds_parts: list[pd.DataFrame] = []
    _base_bounds = df_nf_pre if use_nf_materializado else df
    if isinstance(_base_bounds, pd.DataFrame) and not _base_bounds.empty:
        _bounds_parts.append(_base_bounds)
    if use_fiscal_parquet and isinstance(df_fiscal_pre, pd.DataFrame) and not df_fiscal_pre.empty:
        _bounds_parts.append(df_fiscal_pre)
    _df_bounds = (
        pd.concat(_bounds_parts, ignore_index=True)
        if len(_bounds_parts) > 1
        else (_bounds_parts[0] if _bounds_parts else pd.DataFrame())
    )
    if use_nf_materializado and isinstance(df_nf_pre, pd.DataFrame) and df_nf_pre.empty:
        st.info(
            "Sem notas fiscais neste recorte. Confirme filtros de data, empresa e consolidado na barra lateral "
            "e que a base consolidada está atualizada."
        )
        return
    nf_min, nf_max, ok_nf_dates = faturamento_min_series_nf_emissao_bounds_dates(_df_bounds)
    _emit_floor = _FDL_FAT_DRE_MIN_PANEL_NF_EMISSAO_DESDE
    if ok_nf_dates:
        nf_cal_min, nf_cal_max = _min_cal_limits(nf_min, nf_max)
        nf_cal_min = max(nf_cal_min, _emit_floor)
        if nf_max >= _emit_floor:
            nf_min = max(nf_min, _emit_floor)
        else:
            nf_min, nf_max = _emit_floor, _emit_floor
        nf_cal_max = max(nf_cal_max, nf_max, nf_min, nf_cal_min)
    else:
        nf_cal_min, nf_cal_max = (nf_min, nf_max)
    _nf_sig_k = "fdl_fat_min_nf_bounds_sig"
    _today = datetime.now(_BR_TZ).date()
    if ok_nf_dates:
        _nf_bs = (nf_min.isoformat(), nf_max.isoformat())
        if st.session_state.get(_nf_sig_k) != _nf_bs:
            st.session_state[_nf_sig_k] = _nf_bs
            _nfi = nf_min
            _nff = min(nf_max, _today)
            _nfi = min(max(_nfi, nf_cal_min), nf_cal_max)
            _nff = min(max(_nff, nf_cal_min), nf_cal_max)
            if _nff < _nfi:
                _nff = _nfi
            st.session_state["fdl_fat_min_nf_d_ini"] = _nfi
            st.session_state["fdl_fat_min_nf_d_fim"] = _nff
        if "fdl_fat_min_nf_d_ini" not in st.session_state:
            _nfi = nf_min
            _nff = min(nf_max, _today)
            _nfi = min(max(_nfi, nf_cal_min), nf_cal_max)
            _nff = min(max(_nff, nf_cal_min), nf_cal_max)
            if _nff < _nfi:
                _nff = _nfi
            st.session_state["fdl_fat_min_nf_d_ini"] = _nfi
            st.session_state["fdl_fat_min_nf_d_fim"] = _nff
        st.session_state["fdl_fat_min_nf_d_ini"] = min(
            max(_safe_streamlit_date(st.session_state["fdl_fat_min_nf_d_ini"], nf_min), nf_cal_min),
            nf_cal_max,
        )
        st.session_state["fdl_fat_min_nf_d_fim"] = min(
            max(_safe_streamlit_date(st.session_state["fdl_fat_min_nf_d_fim"], nf_max), nf_cal_min),
            nf_cal_max,
        )

    emp_opts = _faturamento_dre_etiquetas_empresa_recorte(_df_bounds)
    if use_nf_materializado and isinstance(df_nf_pre, pd.DataFrame) and "plataforma" in df_nf_pre.columns:
        # Opções alinhadas ao grão NF (Parquet); evita misturar com linhas fiscais sem ``plataforma``
        # e garante o mesmo rótulo que o filtro ``plataforma`` / ``nf_grain_plataforma_match_key``.
        plats = nf_grain_plataforma_ui_options(df_nf_pre["plataforma"])
    elif use_nf_materializado and isinstance(df_nf_pre, pd.DataFrame) and "plataforma_resumo" in df_nf_pre.columns:
        plats = nf_grain_plataforma_ui_options(df_nf_pre["plataforma_resumo"])
    elif use_nf_materializado and "plataforma" in _df_bounds.columns:
        plats = nf_grain_plataforma_ui_options(_df_bounds["plataforma"])
    elif "Nome da plataforma" in df.columns:
        plats = nf_grain_plataforma_ui_options(df["Nome da plataforma"])
    else:
        plats = []

    _plat_expl = (
        "Filtra notas pela plataforma consolidada no grão NF (materializado)."
        if use_nf_materializado
        else "Restringe linhas de pedido no enriquecimento (venda, comissão, frete, etc.) nas NFs já filtradas por emissão."
    )
    if use_fiscal_parquet:
        _plat_expl += (
            " Com Parquet fiscal ativo, **Plataforma** restringe **cards** e **DRE** ao subconjunto comercial nesse canal "
            "(o conjunto fiscal de referência para imposto permanece na **Apuração Fiscal**)."
        )
    _plat_help = "Plataforma: " + _plat_expl

    def _fdl_rg_cb_emp_changed() -> None:
        st.session_state["fdl_rg_emp_multiselect_dirty"] = True

    _exp_filtros = not st.session_state.get("fdl_rg_filtros_ever_shown", False)
    with st.expander(
        f"▸ Filtros · {_fdl_rg_resumo_filtros_linha()}",
        expanded=_exp_filtros,
    ):
        with st.container(border=True):
            _fh_t, _fh_b = st.columns((4, 1))
            with _fh_t:
                st.markdown("**Filtros**")
            with _fh_b:
                if st.button(
                    "Limpar filtros",
                    key="fdl_fat_min_reset",
                    type="secondary",
                    use_container_width=True,
                    help="Repor empresa, plataforma, situação NF e período (data da venda) ao padrão (inclui preferências antigas da tabela NF no estado).",
                ):
                    for _k in (
                        "fdl_fat_min_emp",
                        "fdl_fat_min_plat",
                        "fdl_fat_min_nf_sit",
                        "fdl_fat_min_nf_d_ini",
                        "fdl_fat_min_nf_d_fim",
                        "fdl_fat_min_nf_bounds_sig",
                        "fdl_fat_min_prod",
                        "fdl_fat_min_venda_sinal",
                        "fdl_fat_min_sinal_resultado",
                        "fdl_fat_min_sinais_resultado",
                        "fdl_fat_nf_show_diferenca",
                        "fdl_fat_nf_opt_plat",
                        "fdl_fat_nf_opt_sit",
                        "fdl_fat_nf_opt_ped",
                        "fdl_fat_nf_opt_linhas",
                        "fdl_fat_nf_opt_qtd",
                        "fdl_fat_nf_opt_vf",
                        "fdl_fat_nf_opt_rf",
                        "fdl_fat_nf_opt_rp",
                        "fdl_fat_nf_opt_tar",
                        "fdl_fat_nf_opt_df",
                        "fdl_fat_nf_opt_ads",
                        "fdl_fat_nf_opt_alert",
                        "fdl_fat_nf_opt_com",
                        "fdl_fat_nf_opt_fp",
                        "fdl_fat_nf_opt_imp",
                        "fdl_rg_sess_emp_default_feito",
                        "fdl_rg_emp_multiselect_dirty",
                    ):
                        st.session_state.pop(_k, None)
                    st.rerun()
            _fc1, _fc2 = st.columns(2)
            with _fc1:
                if emp_opts:
                    if not st.session_state.get("fdl_rg_sess_emp_default_feito"):
                        cur_e = st.session_state.get("fdl_fat_min_emp")
                        if not isinstance(cur_e, list) or len(cur_e) == 0:
                            _pick_e = _rg_pick_empresa_maior_receita_mes_fechado(df, emp_opts)
                            st.session_state["fdl_fat_min_emp"] = (
                                [_pick_e] if _pick_e else [sorted(emp_opts)[0]]
                            )
                        st.session_state["fdl_rg_sess_emp_default_feito"] = True
                    if "fdl_fat_min_emp" not in st.session_state:
                        st.session_state["fdl_fat_min_emp"] = []
                    else:
                        prev_e = st.session_state["fdl_fat_min_emp"]
                        if isinstance(prev_e, list):
                            st.session_state["fdl_fat_min_emp"] = [x for x in prev_e if x in emp_opts]
                        else:
                            st.session_state["fdl_fat_min_emp"] = []
                    st.multiselect(
                        "Empresa",
                        emp_opts,
                        key="fdl_fat_min_emp",
                        help=(
                            "Selecione uma ou mais marcas. Para consolidar várias empresas, selecione-as ao mesmo tempo."
                        ),
                        placeholder="Selecione",
                        on_change=_fdl_rg_cb_emp_changed,
                    )
                else:
                    st.caption("Empresa: sem opções distintas no recorte atual.")
            with _fc2:
                if plats:
                    _multiselect_stable(
                        "fdl_fat_min_plat", "Plataforma", plats, help=_plat_help, placeholder="Todas"
                    )
                else:
                    st.caption("Plataforma: sem opções no recorte.")
            if ok_nf_dates:
                st.markdown(
                    '<p class="fdl-fat-filtros-periodo-tit">Período da venda</p>',
                    unsafe_allow_html=True,
                )
                r_nf = st.columns((1, 1))
                with r_nf[0]:
                    st.date_input(
                        "Início",
                        min_value=nf_cal_min,
                        max_value=nf_cal_max,
                        format="DD/MM/YYYY",
                        key="fdl_fat_min_nf_d_ini",
                        help=_h_periodo,
                    )
                with r_nf[1]:
                    st.date_input(
                        "Fim",
                        min_value=nf_cal_min,
                        max_value=nf_cal_max,
                        format="DD/MM/YYYY",
                        key="fdl_fat_min_nf_d_fim",
                        help=_h_periodo,
                    )
            elif "Nota_Data_Emissao" in _df_bounds.columns:
                st.caption(
                    "Período da venda indisponível (limites de calendário não derivados corretamente das datas do carregamento)."
                )
            else:
                st.caption("Período da venda indisponível (referência temporal ausente neste carregamento).")

    st.session_state["fdl_rg_filtros_ever_shown"] = True

    if (
        emp_opts
        and isinstance(st.session_state.get("fdl_fat_min_emp"), list)
        and len(st.session_state["fdl_fat_min_emp"]) == 0
        and st.session_state.get("fdl_rg_emp_multiselect_dirty")
    ):
        st.info(FDL_RG_MSG_SEM_EMPRESA)
        return

    _fdl_ui_gap_section()
    _fdl_fat_min_vsp(size="md")

    _min_state = faturamento_recorte_min_state_from_session(st.session_state)
    if _min_state.empresas:
        st.session_state["fdl_rg_ultima_empresa_selecionada"] = _min_state.empresas
    if len(_min_state.empresas) >= 2:
        st.caption("📊 Visão consolidada · " + " + ".join(_min_state.empresas))
    _nf_kpi_ini = _safe_streamlit_date(st.session_state.get("fdl_fat_min_nf_d_ini"), nf_min)
    _nf_kpi_fim = _safe_streamlit_date(st.session_state.get("fdl_fat_min_nf_d_fim"), nf_max)
    if ok_nf_dates:
        _nf_kpi_ini = min(max(_nf_kpi_ini, nf_cal_min), nf_cal_max)
        _nf_kpi_fim = min(max(_nf_kpi_fim, nf_cal_min), nf_cal_max)
        if _nf_kpi_fim < _nf_kpi_ini:
            _nf_kpi_fim = _nf_kpi_ini

    _df_fiscal_base, _fiscal_base_stats = build_faturamento_fiscal_base_slice(
        df_fiscal_pre
        if use_fiscal_parquet and isinstance(df_fiscal_pre, pd.DataFrame)
        else pd.DataFrame(),
        empresas_sel=_min_state.empresas,
        nf_d_ini=_nf_kpi_ini,
        nf_d_fim=_nf_kpi_fim,
        ok_nf_dates=ok_nf_dates,
        situacoes_sel=_min_state.situacoes_nf,
        df_devolucoes=df_devolucoes_pre if _df_dev_ok else None,
    )

    df_nf_scope_emissao = _faturamento_nf_apply_minimal_recorte(
        df_nf_pre,
        empresas_sel=_min_state.empresas,
        plataformas_sel=(),
        nf_d_ini=_nf_kpi_ini,
        nf_d_fim=_nf_kpi_fim,
        ok_nf_dates=ok_nf_dates,
    )
    df_nf_scope_emissao = _faturamento_nf_filter_by_situacao(df_nf_scope_emissao, _min_state.situacoes_nf)
    _commercial_kpi_aligned_fiscal = bool(not _df_fiscal_base.empty)
    if _commercial_kpi_aligned_fiscal:
        df_nf_commercial_kpi = build_nf_panel_aligned_to_fiscal_base(_df_fiscal_base, df_nf_scope_emissao)
    else:
        df_nf_commercial_kpi = df_nf_scope_emissao.copy()
    if _min_state.plataformas:
        df_nf_commercial_kpi = _nf_panel_filter_merged_fiscal_by_plataforma_resumo(
            df_nf_commercial_kpi, _min_state.plataformas
        )
    _commercial_coverage = compute_commercial_coverage_stats(df_nf_commercial_kpi)
    _kp_cards = compute_nf_panel_kpis(df_nf_commercial_kpi)
    _imp_rg_kpis = dre_imposto_para_linha_dre_gerencial(
        _kp_cards,
        fiscal_base_stats=_fiscal_base_stats if use_fiscal_parquet else None,
        aplicar_ponte_base_liquida=(
            (_fiscal_base_stats if use_fiscal_parquet else None) is not None and use_fiscal_kpi
        ),
    )

    with st.expander("Cobertura comercial", expanded=False):
        _render_faturamento_dre_commercial_complement_banner(
            coverage=_commercial_coverage,
            n_fiscal_base=int(_fiscal_base_stats.n_nf),
            aligned_to_fiscal_base=_commercial_kpi_aligned_fiscal,
            ok_nf_dates=ok_nf_dates,
            fiscal_parquet_ok=use_fiscal_parquet,
            kpi_subset_by_platform=bool(_min_state.plataformas),
            embedded_in_sobre_expander=False,
            rg_premium_single_expander=True,
        )

    _fdl_fat_min_vsp(size="md")
    try:
        from processing.faturamento.resultado_gerencial_slice import REQUIRED_LINE_COLUMNS

        _rg_kpi_cols_ok = REQUIRED_LINE_COLUMNS.issubset(df.columns)
    except ImportError:
        _rg_kpi_cols_ok = False

    _slice_rg = None
    _kp_rg = None
    _rg_kpis_rendered = False
    if _rg_kpi_cols_ok:
        try:
            from datetime import date as _date_pace

            from app.components.rg_cached_compute import (
                cached_comparacao_kpis_temporal,
                cached_rg_slice_kpis_tabela,
                pipeline_version as _rg_pipeline_version,
            )
            from app.components.termometro_pace import render_termometro_pace
            from processing.faturamento.ficha_pedido_rg import load_resultado_gerencial_config
            from processing.faturamento.pace_mensal import (
                compute_pace_mensal,
                compute_trailing_monthly_revenues,
                explicar_motivo_pace_none,
            )
            from processing.faturamento.rg_cache_keys import normalize_sorted_str_tuple

            _emp_norm = normalize_sorted_str_tuple(_min_state.empresas)
            _plat_norm = normalize_sorted_str_tuple(_min_state.plataformas)
            _slice_rg, _kp_rg, _rg_linhas_ped_tab = cached_rg_slice_kpis_tabela(
                df,
                _emp_norm,
                _plat_norm,
                _nf_kpi_ini,
                _nf_kpi_fim,
                float(_imp_rg_kpis),
                _rg_pipeline_version(),
                str(_oid).strip() if _oid else "",
            )
            _pace_dbg_show = _fdl_rg_pace_debug_enabled()
            _pace_log_motivo: str | None = None
            if ok_nf_dates and _slice_rg is not None:
                try:
                    _rg_conf = load_resultado_gerencial_config(str(_oid).strip() if _oid else None)
                    _hist_cons = compute_trailing_monthly_revenues(
                        df,
                        empresas_sel=_emp_norm,
                        plataformas_sel=_plat_norm,
                        mes_referencia=(_nf_kpi_fim.year, _nf_kpi_fim.month),
                    )
                    _hist_pe: dict[str, list[float]] = {}
                    for _emp in _min_state.empresas:
                        _hist_pe[str(_emp)] = compute_trailing_monthly_revenues(
                            df,
                            empresas_sel=(str(_emp),),
                            plataformas_sel=_plat_norm,
                            mes_referencia=(_nf_kpi_fim.year, _nf_kpi_fim.month),
                        )
                    _pace_today = _date_pace.today()
                    _pace = compute_pace_mensal(
                        _slice_rg,
                        _hist_cons,
                        _rg_conf,
                        list(_min_state.empresas),
                        _nf_kpi_ini,
                        _nf_kpi_fim,
                        _pace_today,
                        historico_por_empresa=_hist_pe,
                    )
                    if _pace is None:
                        _pace_log_motivo = explicar_motivo_pace_none(
                            n_linhas=int(_slice_rg.stats.n_linhas),
                            data_inicio=_nf_kpi_ini,
                            data_fim=_nf_kpi_fim,
                            hoje=_pace_today,
                        )
                    elif _pace.modo == "recorte_parcial":
                        _pace_log_motivo = (
                            "render omitido · modo=recorte_parcial · "
                            f"ini={_nf_kpi_ini.isoformat()} · fim={_nf_kpi_fim.isoformat()}"
                        )
                        if _fdl_rg_recorte_parcial_um_mes_sem_mes_cheio(_nf_kpi_ini, _nf_kpi_fim):
                            st.caption(
                                "📊 Termômetro de pace disponível apenas para mês civil completo. "
                                "Filtro atual abrange período parcial — use os KPIs acima para leitura."
                            )
                    else:
                        if _pace_dbg_show:
                            st.caption(
                                f"🔍 pace pré-render: modo={_pace.modo} · "
                                f"receita={float(_pace.receita_realizada):.2f}"
                            )
                        render_termometro_pace(_pace)
                        if _pace_dbg_show:
                            st.caption("🔍 pace pós-render: chamada concluída")
                        _pace_log_motivo = f"renderizado · modo={_pace.modo}"
                except Exception as exc:
                    _pace_log_motivo = f"exceção: {type(exc).__name__}: {exc}"
                if _pace_dbg_show and _pace_log_motivo:
                    st.caption(f"🔍 pace debug: {_pace_log_motivo}")
            elif _pace_dbg_show:
                if not ok_nf_dates:
                    st.caption("🔍 pace debug: bloco não executado · ok_nf_dates=False")
                elif _slice_rg is None:
                    st.caption("🔍 pace debug: bloco não executado · slice_rg=None")
            _comp_temporal = None
            try:
                _comp_temporal = cached_comparacao_kpis_temporal(
                    df,
                    _emp_norm,
                    _plat_norm,
                    _nf_kpi_ini,
                    _nf_kpi_fim,
                    float(_kp_rg["valor_venda_lista"]),
                    float(_kp_rg["resultado"]),
                    _rg_pipeline_version(),
                    str(_oid).strip() if _oid else "",
                )
            except Exception as exc:
                if _fdl_rg_pace_debug_enabled():
                    st.caption(f"🔍 comparacao temporal debug: exceção {type(exc).__name__}: {exc}")
            _fdl_fat_divider_simple()
            _render_resultado_gerencial_kpi_cards(
                kp_rg=_kp_rg,
                ok_dates=ok_nf_dates,
                use_nf_materializado=use_nf_materializado,
                valor_faturado_from_fiscal_parquet=use_fiscal_kpi,
                fat_dre_faturado_mode=("fiscal" if use_fiscal_kpi else "nf_first"),
                nf_panel_ads=_nf_panel_ads_ui,
                comparacao_temporal=_comp_temporal,
            )
            _rg_kpis_rendered = True
        except (ValueError, KeyError, TypeError) as _exc_rg_kpi:
            if _is_admin_mode():
                st.warning(f"KPIs por data da venda indisponíveis ({_exc_rg_kpi!s}); usando visão NF.")
            _fdl_fat_divider_simple()
            _render_fdl_fat_dre_nf_kpi_cards(
                kp=_kp_cards,
                ok_nf_dates=ok_nf_dates,
                use_nf_materializado=use_nf_materializado,
                valor_faturado_from_fiscal_parquet=use_fiscal_kpi,
                fat_dre_faturado_mode=("fiscal" if use_fiscal_kpi else "nf_first"),
                nf_panel_ads=_nf_panel_ads_ui,
            )
    else:
        if _is_admin_mode():
            st.caption(
                "Colunas insuficientes no grão linha para KPIs por **data da venda** — exibindo KPIs na visão NF."
            )
        _fdl_fat_divider_simple()
        _render_fdl_fat_dre_nf_kpi_cards(
            kp=_kp_cards,
            ok_nf_dates=ok_nf_dates,
            use_nf_materializado=use_nf_materializado,
            valor_faturado_from_fiscal_parquet=use_fiscal_kpi,
            fat_dre_faturado_mode=("fiscal" if use_fiscal_kpi else "nf_first"),
            nf_panel_ads=_nf_panel_ads_ui,
        )

    _fdl_fat_min_vsp(size="md")
    _periodo_dre_lbl = ""
    if ok_nf_dates:
        _periodo_dre_lbl = f"{_nf_kpi_ini.strftime('%d/%m/%Y')} — {_nf_kpi_fim.strftime('%d/%m/%Y')}"

    _hp_mrg_ant: float | None = None
    if _rg_kpis_rendered and _slice_rg is not None and _kp_rg is not None:
        try:
            from app.components.health_score import periodo_mes_de_datas

            _ano_h, _mes_h, _ = periodo_mes_de_datas(_nf_kpi_ini, _nf_kpi_fim)
            _hp_mrg_ant = _fdl_health_panel_rg_benchmark_margins(
                df_linha=df,
                df_nf_pre=df_nf_pre,
                df_commercial_nf_kpi=df_nf_commercial_kpi,
                df_fiscal_pre=df_fiscal_pre if isinstance(df_fiscal_pre, pd.DataFrame) else pd.DataFrame(),
                df_devolucoes_pre=df_devolucoes_pre if _df_dev_ok else None,
                devolucoes_ok=_df_dev_ok,
                use_fiscal_parquet=use_fiscal_parquet,
                use_fiscal_kpi=use_fiscal_kpi,
                ok_nf_dates=ok_nf_dates,
                fiscal_base_stats=_fiscal_base_stats if use_fiscal_parquet else None,
                nf_d_ini=_nf_kpi_ini,
                nf_d_fim=_nf_kpi_fim,
                ano_civil=_ano_h,
                mes_civil=_mes_h,
                empresas_sel=tuple(_min_state.empresas),
                plataformas_sel=tuple(_min_state.plataformas),
                situacoes_nf=_min_state.situacoes_nf,
                org_sidebar=_oid,
            )
        except Exception:
            pass

    _col_dre, _col_hp = st.columns([5, 7])
    with _col_dre:
        st.markdown(
            '<span class="fdl-rg-col-mark-dre" aria-hidden="true"></span>',
            unsafe_allow_html=True,
        )
        if _rg_kpis_rendered and _slice_rg is not None and _kp_rg is not None:
            _render_fdl_fat_dre_gerencial_linha(
                stats=_slice_rg.stats,
                kp_rg=_kp_rg,
                imp_nf=float(_imp_rg_kpis),
                ok_nf_dates=ok_nf_dates,
                valor_faturado_from_fiscal_parquet=use_fiscal_kpi,
                periodo_label=_periodo_dre_lbl,
                nf_panel_ads=_nf_panel_ads_ui,
                rg_header_subtitle=_fdl_rg_header_context(_min_state),
                show_resultado_discreto=True,
            )
        else:
            _render_fdl_fat_dre_nf_gerencial(
                kp=_kp_cards,
                ok_nf_dates=ok_nf_dates,
                valor_faturado_from_fiscal_parquet=use_fiscal_kpi,
                periodo_label=_periodo_dre_lbl,
                nf_panel_ads=_nf_panel_ads_ui,
                fiscal_base_stats=_fiscal_base_stats if use_fiscal_parquet else None,
            )
    with _col_hp:
        st.markdown(
            '<span class="fdl-rg-col-mark-hp" aria-hidden="true"></span>',
            unsafe_allow_html=True,
        )
        try:
            from app.components.health_panel_ui import render_faturamento_health_panel_if_enabled

            render_faturamento_health_panel_if_enabled(
                df,
                nf_d_ini=_nf_kpi_ini,
                nf_d_fim=_nf_kpi_fim,
                empresas_sel=tuple(_min_state.empresas),
                org_sidebar=_oid,
                plataformas_sel=tuple(_min_state.plataformas),
                coluna_temporal="Data",
                kpis_rg=_kp_rg if _rg_kpis_rendered else None,
                cmv_total_rg=float(_slice_rg.stats.cmv_total) if _slice_rg is not None else None,
                margem_anterior_pct=_hp_mrg_ant,
                margem_grupo_pct=None,
                rg_streamlined=True,
                rg_header_context=_fdl_rg_header_context(_min_state),
            )
        except Exception as _exc_hp:
            if _is_admin_mode():
                st.caption(f"Painel de saúde financeira indisponível: `{_exc_hp}`")

    _fdl_fat_min_vsp(size="md")

    if _rg_kpis_rendered and _slice_rg is not None and _kp_rg is not None:
        try:
            from app.components.analise_plataforma_ui import render_analise_plataforma
            from app.components.rg_cached_compute import cached_analise_plataforma, pipeline_version as _rg_pv_plat
            from processing.faturamento.rg_cache_keys import normalize_sorted_str_tuple

            _ap_emp = normalize_sorted_str_tuple(_min_state.empresas)
            _ap_plat = normalize_sorted_str_tuple(_min_state.plataformas)
            _analise_plat = cached_analise_plataforma(
                df,
                _ap_emp,
                _ap_plat,
                _nf_kpi_ini,
                _nf_kpi_fim,
                float(_imp_rg_kpis),
                _rg_pv_plat(),
                str(_oid).strip() if _oid else "",
            )
            render_analise_plataforma(_analise_plat, debug_enabled=_fdl_rg_pace_debug_enabled())
        except Exception as exc:
            if _fdl_rg_pace_debug_enabled():
                st.caption(f"🔍 analise_plat debug: exceção {type(exc).__name__}: {exc}")

    if _rg_kpis_rendered and _slice_rg is not None and _kp_rg is not None:
        try:
            from app.components.curva_abc_ui import render_curva_abc
            from app.components.rg_cached_compute import cached_curva_abc, pipeline_version as _rg_pv_abc
            from processing.faturamento.rg_cache_keys import normalize_sorted_str_tuple

            _abc_emp = normalize_sorted_str_tuple(_min_state.empresas)
            _abc_plat = normalize_sorted_str_tuple(_min_state.plataformas)
            curva_abc_obj = cached_curva_abc(
                df,
                _abc_emp,
                _abc_plat,
                _nf_kpi_ini,
                _nf_kpi_fim,
                float(_imp_rg_kpis),
                _rg_pv_abc(),
                str(_oid).strip() if _oid else "",
            )
            _rg_tbl_label = f"{str(_oid).strip()}_{_nf_kpi_ini.year:04d}-{_nf_kpi_ini.month:02d}"
            if curva_abc_obj is not None and len(curva_abc_obj.linhas) >= 2:
                render_curva_abc(
                    curva_abc_obj,
                    debug_enabled=_fdl_rg_pace_debug_enabled(),
                    checkbox_key=f"fdl_curva_abc_ver_{_rg_tbl_label}",
                )
        except Exception as exc:
            if _fdl_rg_pace_debug_enabled():
                st.caption(f"🔍 curva_abc debug: exceção {type(exc).__name__}: {exc}")

    if _rg_kpis_rendered and _slice_rg is not None and _kp_rg is not None:
        try:
            from app.components.tabela_pedidos_gerencial import render_tabela_pedidos_rg

            _rg_tbl_label = f"{str(_oid).strip()}_{_nf_kpi_ini.year:04d}-{_nf_kpi_ini.month:02d}"
            render_tabela_pedidos_rg(
                _slice_rg,
                _kp_rg,
                fiscal_imposto_valor=float(_imp_rg_kpis),
                export_label=_rg_tbl_label,
                debug_coerencia=_is_admin_mode(),
                cliente_slug=str(_oid).strip() if _oid else None,
                linhas_full=_rg_linhas_ped_tab,
            )
        except Exception as _exc_tp:
            st.warning(
                "Não foi possível carregar a **tabela por pedido** / ficha neste recorte. "
                "Confirme que está no **Faturamento (DRE)** com KPIs por **data da venda** ativos e reinicie o app após atualizar o código."
            )
            if _is_admin_mode():
                st.exception(_exc_tp)

    _fdl_fat_min_vsp(size="md")

    _fdl_ui_gap_section()
    _fdl_fat_min_vsp(size="sm")


def _faturamento_disp_texto_sem_none(s: pd.Series, *, placeholder: str = "—") -> pd.Series:
    """Evita que None/NaN apareçam como o texto «None» no ``st.dataframe`` (colunas opcionais do CSV de pedidos)."""

    def _cell(v: object) -> str:
        if v is None:
            return placeholder
        if isinstance(v, float) and math.isnan(v):
            return placeholder
        try:
            if pd.isna(v):
                return placeholder
        except (ValueError, TypeError):
            pass
        xs = str(v).strip()
        if not xs or xs.casefold() in {"nan", "none", "nat", "<na>"}:
            return placeholder
        return xs

    return s.map(_cell)


def _faturamento_disp_data_pedidos(s: pd.Series) -> pd.Series:
    """Datas do export ML (dia/mês/ano); placeholders tipo 00/00/0000 → «—»."""

    def _one(v: object) -> str:
        if v is None:
            return "—"
        if isinstance(v, float) and math.isnan(v):
            return "—"
        if isinstance(v, (pd.Timestamp, datetime)):
            if pd.isna(v):
                return "—"
            ts = pd.Timestamp(v)
            if ts.year < 1900:
                return "—"
            return ts.strftime("%d/%m/%Y")
        try:
            if pd.isna(v):
                return "—"
        except (ValueError, TypeError):
            pass
        xs = str(v).strip()
        if not xs or xs.casefold() in {"nan", "none", "nat"}:
            return "—"
        if xs in {"00/00/0000", "0/0/0000"} or xs.replace("0", "").replace("/", "").strip() == "":
            return "—"
        t = pd.to_datetime(xs, errors="coerce", dayfirst=True)
        if pd.isna(t) or t.year < 1900:
            return xs
        return t.strftime("%d/%m/%Y")

    return s.map(_one)


def _faturamento_compute_alert_bools(df: pd.DataFrame) -> pd.DataFrame:
    """Colunas auxiliares _ab_* para KPIs, filtros e texto de alertas."""
    try:
        from processing.faturamento.normalize import to_numeric_br as _fat_to_num
    except Exception:  # noqa: BLE001
        _fat_to_num = None

    def _num(s: pd.Series) -> pd.Series:
        if _fat_to_num is not None:
            return _fat_to_num(s)
        return pd.to_numeric(s, errors="coerce")

    out = df.copy()
    pl, vt = "Preço de lista", "Valor total"
    pln = _num(out[pl]) if pl in out.columns else pd.Series(float("nan"), index=out.index)
    vtn = _num(out[vt]) if vt in out.columns else pd.Series(float("nan"), index=out.index)
    tol = _faturamento_divergencia_tol()
    out["_ab_pl_zero"] = pln.notna() & (pln == 0)
    desc_col = "Desconto proporcional total"
    if "Receita_Bruta" in out.columns:
        rbn = _num(out["Receita_Bruta"])
        base_ok = rbn.notna() & vtn.notna()
        if desc_col in out.columns:
            dcn = _num(out[desc_col])
            residual = (rbn - dcn - vtn).abs()
            out["_ab_div"] = base_ok & (
                dcn.notna() & (residual > tol)
                | dcn.isna() & ((rbn - vtn).abs() > tol)
            )
        else:
            out["_ab_div"] = base_ok & ((rbn - vtn).abs() > tol)
    else:
        out["_ab_div"] = pln.notna() & vtn.notna() & ((pln - vtn).abs() > tol)
    situ = (
        out["Situação"].fillna("").astype(str).str.strip().str.casefold()
        if "Situação" in out.columns
        else pd.Series("", index=out.index, dtype=object)
    )
    nf = (
        out["Existe Nota Fiscal gerada"].fillna("").astype(str).str.strip().str.casefold()
        if "Existe Nota Fiscal gerada" in out.columns
        else pd.Series("", index=out.index, dtype=object)
    )
    atendido = situ == "atendido"
    sem_nf = nf.eq("não") | nf.eq("nao")
    if "faturamento_consolidado" in out.columns:
        fc = _faturamento_series_bool_mask(out["faturamento_consolidado"])
    else:
        fc = pd.Series(False, index=out.index)
    out["_ab_sem_nf_np"] = atendido & sem_nf & ~fc
    return out


def _faturamento_alertas_text(s: pd.Series) -> str:
    parts: list[str] = []
    if bool(s.get("_ab_pl_zero")):
        parts.append("Preço lista zero")
    if bool(s.get("_ab_div")):
        parts.append("Divergência receita × valor (não explicada por desconto)")
    if bool(s.get("_ab_sem_nf_np")):
        parts.append("Sem NF não permitido")
    return " · ".join(parts)


def _faturamento_dre_global_filter_keys() -> list[str]:
    return [
        "fdl_fat_dre_emp",
        "fdl_fat_dre_sit",
        "fdl_fat_dre_d_ini",
        "fdl_fat_dre_d_fim",
        "fdl_fat_dre_plat",
        "fdl_fat_dre_presenca_nf",
        "fdl_fat_dre_nf_emi_use",
        "fdl_fat_dre_nf_emi_ini",
        "fdl_fat_dre_nf_emi_fim",
        "fdl_fat_dre_nf_sit",
        "fdl_fat_dre_data_bounds_sig",
        "fdl_fat_dre_nf_bounds_sig",
    ]


def _render_faturamento_dre_linha_confianca(
    *,
    n_base: int,
    df_recorte: pd.DataFrame,
    ok_dates: bool,
    d_min: date,
    d_max: date,
    has_nf_emi: bool,
    nf_ok_dates: bool,
    nf_d_min: date,
    nf_d_max: date,
) -> None:
    """Resumo curto: linhas no recorte vs. base, eixos comercial/fiscal ativos (sem novo filtro)."""
    n_rec = int(len(df_recorte))
    _pres = str(st.session_state.get("fdl_fat_dre_presenca_nf") or "Todos").strip()
    _emi = bool(st.session_state.get("fdl_fat_dre_nf_emi_use"))
    _nf_sit = st.session_state.get("fdl_fat_dre_nf_sit") or []
    _nf_sit_lbl = ", ".join(str(x) for x in _nf_sit) if _nf_sit else "todas"

    _head = f"**{n_rec}** linha(s) no recorte"
    if n_base != n_rec:
        _head += f" · base neste escopo: **{n_base}** linha(s)"

    if ok_dates:
        _di = _safe_streamlit_date(st.session_state.get("fdl_fat_dre_d_ini"), d_min)
        _df = _safe_streamlit_date(st.session_state.get("fdl_fat_dre_d_fim"), d_max)
        _venda = f"{_di.strftime('%d/%m/%Y')} → {_df.strftime('%d/%m/%Y')}"
    else:
        _venda = "período por **Data** indisponível nesta base"

    if not _emi:
        _emi_txt = "desligado"
    elif has_nf_emi and nf_ok_dates:
        _ni = _safe_streamlit_date(st.session_state.get("fdl_fat_dre_nf_emi_ini"), nf_d_min)
        _nf = _safe_streamlit_date(st.session_state.get("fdl_fat_dre_nf_emi_fim"), nf_d_max)
        _emi_txt = f"ligado · {_ni.strftime('%d/%m/%Y')} → {_nf.strftime('%d/%m/%Y')}"
    else:
        _emi_txt = "ligado (sem datas utilizáveis — recorte pode ficar vazio)"

    st.caption(
        f"**Confiança do recorte:** {_head}. "
        f"Venda: **{_venda}**. Vínculo fiscal (join): **{_pres}**. "
        f"Emissão NF: **{_emi_txt}**. Situação NF: **{_nf_sit_lbl}**."
    )


def _render_faturamento_dre_recorte_global(
    df: pd.DataFrame,
    *,
    org_id: str,
    org_display_name: str,
) -> pd.DataFrame:
    """
    Recorte global (comercial + fiscal) do módulo Faturamento & DRE, alinhado a
    ``docs/faturamento_recorte_modulo_aprovado.md``. Devolve o DataFrame já recortado.
    """
    if df.empty:
        with st.container(border=True):
            st.subheader("Recorte do módulo")
            st.info(
                "Sem linhas na base de faturamento. Confirme **materialização**, **escopo** na barra lateral e permissões. "
                "Com dados carregados, os blocos de recorte comercial e fiscal aparecem aqui."
            )
        return df
    out = df
    has_data = "Data" in out.columns
    if has_data:
        d_min, d_max, ok_dates = _series_datetime_bounds_dates(out["Data"])
    else:
        d_min = d_max = datetime.now(_BR_TZ).date()
        ok_dates = False
    cal_min, cal_max = (
        _faturamento_period_calendar_limits(d_min, d_max) if ok_dates else (d_min, d_max)
    )
    _fat_dre_bounds_sig_key = "fdl_fat_dre_data_bounds_sig"
    if ok_dates:
        _bs = (d_min.isoformat(), d_max.isoformat())
        if st.session_state.get(_fat_dre_bounds_sig_key) != _bs:
            st.session_state[_fat_dre_bounds_sig_key] = _bs
            st.session_state["fdl_fat_dre_d_ini"] = d_min
            st.session_state["fdl_fat_dre_d_fim"] = min(d_max, datetime.now(_BR_TZ).date())
    has_nf_emi = "Nota_Data_Emissao" in out.columns
    if has_nf_emi:
        nf_d_min, nf_d_max, nf_ok_dates = _series_datetime_bounds_dates(
            out["Nota_Data_Emissao"], dayfirst=False
        )
    else:
        nf_d_min = nf_d_max = datetime.now(_BR_TZ).date()
        nf_ok_dates = False
    nf_cal_min, nf_cal_max = (
        _faturamento_period_calendar_limits(nf_d_min, nf_d_max) if nf_ok_dates else (nf_d_min, nf_d_max)
    )
    _fat_dre_nf_bounds_sig_key = "fdl_fat_dre_nf_bounds_sig"
    if nf_ok_dates:
        _nf_bs = (nf_d_min.isoformat(), nf_d_max.isoformat())
        if st.session_state.get(_fat_dre_nf_bounds_sig_key) != _nf_bs:
            st.session_state[_fat_dre_nf_bounds_sig_key] = _nf_bs
            st.session_state["fdl_fat_dre_nf_emi_ini"] = nf_d_min
            st.session_state["fdl_fat_dre_nf_emi_fim"] = min(nf_d_max, datetime.now(_BR_TZ).date())
    sits = sorted({str(x).strip() for x in out["Situação"].dropna().unique() if str(x).strip()})
    plats = (
        nf_grain_plataforma_ui_options(out["Nome da plataforma"])
        if "Nome da plataforma" in out.columns
        else []
    )
    emp_opts = _faturamento_dre_etiquetas_empresa_recorte(out)
    nf_sit_vals = (
        sorted({str(x).strip() for x in out["Nota_Situacao"].dropna().unique() if str(x).strip()})
        if "Nota_Situacao" in out.columns
        else []
    )
    nf_sit_opts = sorted(set(nf_sit_vals) | {"Cancelada", "Denegada", "Inutilizada"})

    with st.container(border=True):
        st.subheader("Recorte do módulo")
        st.caption(
            "Dois eixos independentes: **comercial** (venda / empresa / plataforma) e **fiscal** (emissão NF / situação da NF). "
            "Com ambos os intervalos de datas preenchidos, o resultado é a **interseção**. "
            "A **Visão geral**, o **detalhamento** e o **CSV exportado** usam o mesmo recorte."
        )
        with st.container(border=True):
            st.caption("**Recorte comercial**")
            if emp_opts:
                st.caption(
                    "A carga do módulo é **consolidada** (orgs permitidas a si); a **Empresa** na barra lateral **não** corta esse universo. "
                    "Use o multiselect **Empresa** abaixo para Esquilo, Wood, ambas ou **vazio** = todas as marcas na base."
                )
            if emp_opts:
                if "fdl_fat_dre_emp" not in st.session_state:
                    st.session_state["fdl_fat_dre_emp"] = _faturamento_dre_default_empresa_labels(
                        out, org_id, org_display_name
                    )
                else:
                    prev_e = st.session_state["fdl_fat_dre_emp"]
                    if isinstance(prev_e, list):
                        st.session_state["fdl_fat_dre_emp"] = [x for x in prev_e if x in emp_opts]
                    else:
                        st.session_state["fdl_fat_dre_emp"] = _faturamento_dre_default_empresa_labels(
                            out, org_id, org_display_name
                        )
                    if not st.session_state["fdl_fat_dre_emp"]:
                        st.session_state["fdl_fat_dre_emp"] = _faturamento_dre_default_empresa_labels(
                            out, org_id, org_display_name
                        )
                st.multiselect(
                    "Empresa",
                    emp_opts,
                    key="fdl_fat_dre_emp",
                    help=(
                        "**Recorte principal** deste módulo. **Vazio** = todas as marcas na base carregada. "
                        "Uma ou mais opções isolam/comparam dentro do consolidado (independente da empresa selecionada na sidebar)."
                    ),
                    placeholder="Todas",
                )
            if "fdl_fat_dre_sit" not in st.session_state:
                st.session_state["fdl_fat_dre_sit"] = (
                    ["Atendido"] if "Atendido" in sits else ([] if not sits else [sits[0]])
                )
            else:
                prev = st.session_state["fdl_fat_dre_sit"]
                if isinstance(prev, list):
                    st.session_state["fdl_fat_dre_sit"] = [x for x in prev if x in sits]
                else:
                    st.session_state["fdl_fat_dre_sit"] = ["Atendido"] if "Atendido" in sits else []
            r_sp = st.columns((1, 1))
            with r_sp[0]:
                st.multiselect(
                    "Situação do pedido",
                    sits,
                    key="fdl_fat_dre_sit",
                    help="Padrão **Atendido**. Vazio = todas as situações.",
                )
            with r_sp[1]:
                _multiselect_stable("fdl_fat_dre_plat", "Plataforma", plats)
            if ok_dates:
                if "fdl_fat_dre_d_ini" not in st.session_state:
                    st.session_state["fdl_fat_dre_d_ini"] = d_min
                if "fdl_fat_dre_d_fim" not in st.session_state:
                    st.session_state["fdl_fat_dre_d_fim"] = min(d_max, datetime.now(_BR_TZ).date())
                st.session_state["fdl_fat_dre_d_ini"] = min(
                    max(_safe_streamlit_date(st.session_state["fdl_fat_dre_d_ini"], d_min), cal_min),
                    cal_max,
                )
                st.session_state["fdl_fat_dre_d_fim"] = min(
                    max(_safe_streamlit_date(st.session_state["fdl_fat_dre_d_fim"], d_max), cal_min),
                    cal_max,
                )
                r0 = st.columns((1, 1))
                with r0[0]:
                    st.date_input(
                        "Data da venda — início",
                        min_value=cal_min,
                        max_value=cal_max,
                        format="DD/MM/YYYY",
                        key="fdl_fat_dre_d_ini",
                        help=_FATURAMENTO_HELP_PERIODO_DATA,
                    )
                with r0[1]:
                    st.date_input(
                        "Data da venda — fim",
                        min_value=cal_min,
                        max_value=cal_max,
                        format="DD/MM/YYYY",
                        key="fdl_fat_dre_d_fim",
                        help=_FATURAMENTO_HELP_PERIODO_DATA,
                    )
                st.caption(
                    "Eixo **Data** do pedido (venda). Pode escolher datas fora do intervalo dos dados; só entram linhas cuja **Data** "
                    "cai entre início (inclusive) e fim (inclusive)."
                )
            elif has_data:
                st.caption("**Data** sem valores utilizáveis — filtro por data da venda desativado.")
            _pres_opts = ("Todos", "Com NF vinculada", "Sem NF vinculada")
            st.selectbox(
                "Com / sem nota (vínculo fiscal na base)",
                options=list(_pres_opts),
                key="fdl_fat_dre_presenca_nf",
                help="Refinamento comercial alinhado a **faturamento_nota_vinculada** (join pedidos↔notas). "
                "A **Visão** da tabela (Consolidado / Com NF (pedido) / …) continua separada, abaixo.",
            )

        with st.container(border=True):
            st.caption("**Recorte fiscal**")
            if "fdl_fat_dre_nf_emi_use" not in st.session_state:
                st.session_state["fdl_fat_dre_nf_emi_use"] = False
            st.checkbox(
                "Filtrar por data de emissão da NF",
                key="fdl_fat_dre_nf_emi_use",
                help="Desligado por omissão (**recorte fiscal inativo**). Ligado: aplica intervalo sobre **Nota_Data_Emissao** e "
                "**exclui linhas sem nota vinculada**.",
            )
            _emi_on = bool(st.session_state.get("fdl_fat_dre_nf_emi_use"))
            if _emi_on:
                if nf_ok_dates:
                    if "fdl_fat_dre_nf_emi_ini" not in st.session_state:
                        st.session_state["fdl_fat_dre_nf_emi_ini"] = nf_d_min
                    if "fdl_fat_dre_nf_emi_fim" not in st.session_state:
                        st.session_state["fdl_fat_dre_nf_emi_fim"] = min(
                            nf_d_max, datetime.now(_BR_TZ).date()
                        )
                    st.session_state["fdl_fat_dre_nf_emi_ini"] = min(
                        max(
                            _safe_streamlit_date(st.session_state["fdl_fat_dre_nf_emi_ini"], nf_d_min),
                            nf_cal_min,
                        ),
                        nf_cal_max,
                    )
                    st.session_state["fdl_fat_dre_nf_emi_fim"] = min(
                        max(
                            _safe_streamlit_date(st.session_state["fdl_fat_dre_nf_emi_fim"], nf_d_max),
                            nf_cal_min,
                        ),
                        nf_cal_max,
                    )
                    r_nf = st.columns((1, 1))
                    with r_nf[0]:
                        st.date_input(
                            "Data de emissão da NF — início",
                            min_value=nf_cal_min,
                            max_value=nf_cal_max,
                            format="DD/MM/YYYY",
                            key="fdl_fat_dre_nf_emi_ini",
                        )
                    with r_nf[1]:
                        st.date_input(
                            "Data de emissão da NF — fim",
                            min_value=nf_cal_min,
                            max_value=nf_cal_max,
                            format="DD/MM/YYYY",
                            key="fdl_fat_dre_nf_emi_fim",
                        )
                elif has_nf_emi:
                    st.caption("**Nota_Data_Emissao** sem datas utilizáveis — não é possível aplicar o intervalo.")
                else:
                    st.caption("Coluna **Nota_Data_Emissao** ausente — filtro fiscal por emissão indisponível.")
            if "fdl_fat_dre_nf_sit" not in st.session_state:
                st.session_state["fdl_fat_dre_nf_sit"] = []
            else:
                prev_nf = st.session_state["fdl_fat_dre_nf_sit"]
                if isinstance(prev_nf, list):
                    st.session_state["fdl_fat_dre_nf_sit"] = [x for x in prev_nf if x in nf_sit_opts]
                else:
                    st.session_state["fdl_fat_dre_nf_sit"] = []
            st.multiselect(
                "Situação da NF",
                nf_sit_opts,
                key="fdl_fat_dre_nf_sit",
                help="Vazio = todas. Valores vêm do export de notas (materialização); **Cancelada** mantém-se como opção mesmo sem linhas.",
                placeholder="Todas",
            )

        if st.button(
            "Redefinir recorte",
            key="fdl_fat_dre_reset_recorte",
            help="Repõe empresa, comercial, fiscal e refinamentos deste bloco",
        ):
            for _k in _faturamento_dre_global_filter_keys():
                st.session_state.pop(_k, None)
            for _rk in list(st.session_state.keys()):
                if isinstance(_rk, str) and _rk.startswith("fdl_fat_dre_painel_emp_"):
                    st.session_state.pop(_rk, None)
            st.rerun()

    _rec_state = faturamento_recorte_state_from_session(st.session_state)
    _rec_res = apply_recorte_modulo(out, _rec_state)
    for _w in _rec_res.warnings:
        st.warning(_w)
    for _a in _rec_res.admin_messages:
        if _is_admin_mode():
            st.warning(_a)
    sliced = _rec_res.df

    _render_faturamento_dre_linha_confianca(
        n_base=len(out),
        df_recorte=sliced,
        ok_dates=ok_dates,
        d_min=d_min,
        d_max=d_max,
        has_nf_emi=has_nf_emi,
        nf_ok_dates=nf_ok_dates,
        nf_d_min=nf_d_min,
        nf_d_max=nf_d_max,
    )

    _emi_active = bool(st.session_state.get("fdl_fat_dre_nf_emi_use"))
    if _emi_active and (not has_nf_emi or not nf_ok_dates):
        _msg_nf = (
            "esta base não inclui a coluna **Nota_Data_Emissao**"
            if not has_nf_emi
            else "a coluna **Nota_Data_Emissao** não tem datas de emissão utilizáveis (apenas vazias ou inválidas)"
        )
        st.info(
            f"O filtro **por data de emissão da NF** está ligado, mas {_msg_nf}. "
            "Por isso o **recorte fica vazio** — a Visão geral e a tabela não mostram linhas até desligar esse filtro "
            "ou corrigir o materializado / export de notas de saída."
        )

    return sliced


def _faturamento_filter_keys(org_id: str) -> list[str]:
    oid = str(org_id)
    return [
        f"fat_visao_{oid}",
        f"fat_d_ini_{oid}",
        f"fat_d_fim_{oid}",
        f"fat_d_bounds_sig_{oid}",
        f"fat_ms_plat_{oid}",
        f"fat_ms_sit_{oid}",
        f"fat_busca_{oid}",
        f"fat_ms_alert_{oid}",
    ]


def _faturamento_dre_vl_venda_series(df: pd.DataFrame, pl_col: str) -> pd.Series:
    """Vl. Venda comercial: coluna ``Vl_Venda`` ou mesma lógica da receita bruta do painel."""
    if "Vl_Venda" in df.columns:
        return pd.to_numeric(df["Vl_Venda"], errors="coerce")
    return _faturamento_painel_receita_series(df, pl_col)


def _faturamento_dre_vl_nota_fiscal_series(df: pd.DataFrame) -> pd.Series:
    """Valor líquido fiscal por linha: nota rateada; senão ``Valor total`` do pedido (legado)."""
    if "Nota_Valor_Liquido_Rateado" in df.columns:
        return pd.to_numeric(df["Nota_Valor_Liquido_Rateado"], errors="coerce")
    if "Valor total" in df.columns:
        return pd.to_numeric(df["Valor total"], errors="coerce")
    return pd.Series(float("nan"), index=df.index)


def _faturamento_dre_nf_coluna_display(df: pd.DataFrame) -> pd.Series:
    """Uma coluna NF: join ``Nota_Numero_Normalizado``, fallback «Número da nota» do pedido."""
    best = pd.Series("", index=df.index, dtype=object)
    if "Nota_Numero_Normalizado" in df.columns:
        best = df["Nota_Numero_Normalizado"].fillna("").astype(str).str.strip()
    if "Número da nota" in df.columns:
        ml = df["Número da nota"].fillna("").astype(str).str.strip()
        best = best.mask(best.eq(""), ml)
    return _faturamento_disp_texto_sem_none(best.astype(str))


def _faturamento_dre_frete_display_series(df: pd.DataFrame) -> pd.Series:
    """Um frete para leitura principal: plataforma; senão custo total de frete."""
    if "Frete_Plataforma" in df.columns:
        return pd.to_numeric(df["Frete_Plataforma"], errors="coerce")
    if "Custo de Frete" in df.columns:
        return pd.to_numeric(df["Custo de Frete"], errors="coerce")
    return pd.Series(float("nan"), index=df.index)


_FATURAMENTO_PAINEL_EM_CONSTRUCAO = False


def _painel_faturamento(
    df: pd.DataFrame,
    _load_info: dict[str, object],
    ts_proc: str,
    org_id: str,
    *,
    use_modulo_recorte: bool = False,
    mvp_rotulos_bloco_dre: bool = False,
) -> None:
    """
    Faturamento: filtros, indicadores, tabela e export CSV.

    Com ``use_modulo_recorte=True``, empresa / período de venda / plataforma / situação / fiscal vêm **só**
    do **Recorte do módulo** (``apply_recorte_modulo``); não se reaplica filtro de **Data** aqui.
    Com ``mvp_rotulos_bloco_dre=True``, rótulos alinham ao **Detalhamento operacional** (MVP DRE).
    """
    if _FATURAMENTO_PAINEL_EM_CONSTRUCAO:
        with st.container(border=True):
            st.caption("Módulo")
            st.info(
                "Em preparação. Em breve, as mesmas funções de análise e exportação da conciliação."
            )
        return
    _oid = str(org_id)
    if df.empty:
        st.info("Sem dados para este recorte. Verifique a empresa ou contacte o suporte.")
        return

    missing = _faturamento_painel_missing_schema_columns(df)
    if missing:
        if _is_admin_mode():
            st.warning(
                "Dados de faturamento sem colunas esperadas pelo painel. "
                f"Faltam: {', '.join(missing[:12])}{'…' if len(missing) > 12 else ''}."
            )
        else:
            st.warning(
                "Não foi possível apresentar o faturamento com a estrutura esperada. "
                "Contacte o administrador."
            )
        return

    work = _faturamento_compute_alert_bools(df)
    for c in (
        "faturamento_consolidado",
        "faturamento_com_nf",
        "faturamento_sem_nf",
    ):
        if c not in work.columns:
            work[c] = False

    if "Status_Custo" in work.columns:
        _vc = work["Status_Custo"].astype(str).str.strip()
        _n_sem = int(_vc.eq("SKU_SEM_CORRESPONDENCIA").sum())
        _n_ok_all = int(_vc.eq("CUSTO_OK").sum())
        _n_tot = len(work)
        _ratio_sem = (_n_sem / _n_tot) if _n_tot else 0.0
        if _n_sem > 0:
            if _ratio_sem <= 0.10:
                st.caption(
                    f"**Custo:** a maioria das linhas já tem custo alocado (**{_n_ok_all}** CUSTO_OK). "
                    f"Restam **{_n_sem}** exceções (SKU ausente na planilha, código inválido ou cadastro divergente). "
                    "Nessas linhas, **Resultado** pode ficar vazio."
                )
            else:
                st.info(
                    f"**Custo:** **{_n_sem}** de **{_n_tot}** linhas ainda sem custo na referência "
                    "(SKU sem correspondência ou cadastro). Onde não há custo, **Resultado** fica vazio."
                )

    pl_col, res_col = "Preço de lista", "Resultado"
    custo_prod_col = _faturamento_painel_custo_produto_col(list(work.columns))
    assert custo_prod_col is not None  # garantido por _faturamento_painel_missing_schema_columns

    has_data_col = "Data" in work.columns
    if use_modulo_recorte:
        has_usable_dates = False
        d_min = d_max = datetime.now(_BR_TZ).date()
        plats: list[str] = []
        sits: list[str] = []
    else:
        if has_data_col:
            d_min, d_max, has_usable_dates = _series_datetime_bounds_dates(work["Data"])
        else:
            d_min = d_max = datetime.now(_BR_TZ).date()
            has_usable_dates = False
        plats = (
            nf_grain_plataforma_ui_options(work["Nome da plataforma"])
            if "Nome da plataforma" in work.columns
            else []
        )
        sits = sorted({str(x).strip() for x in work["Situação"].dropna().unique() if str(x).strip()})

    _opt_alertas = (
        "Preço lista zero",
        "Divergência receita × valor (não explicada por desconto)",
        "Sem NF não permitido",
    )

    with st.container(border=True):
        st.subheader("Filtros")
        if use_modulo_recorte:
            st.caption(
                "Neste bloco: visão NF, alertas e busca. **Recorte comercial** e **recorte fiscal** (empresa, datas venda/emissão, plataforma, situações) vêm do **Recorte do módulo** acima."
                if not mvp_rotulos_bloco_dre
                else "**Detalhamento:** **Visão** (NF do pedido no materializado), **alertas** e **busca**. "
                "O **recorte** (empresa, datas, fiscal) está **só** no bloco acima — **sem segundo filtro de data** aqui."
            )
        else:
            st.caption("Período, visão, critérios e busca para refinar o recorte.")
        _fdl_ui_gap_section()
        _k_visao = f"fat_visao_{_oid}"
        if st.session_state.get(_k_visao) == _FAT_PAINEL_VISAO_COM_NF_LEGACY:
            st.session_state[_k_visao] = _FAT_PAINEL_VISAO_COM_NF
        _fat_visao_opts = (
            "Todos",
            "Consolidado",
            _FAT_PAINEL_VISAO_COM_NF,
            "Sem NF permitido",
        )
        visao = st.selectbox(
            "Visão",
            _fat_visao_opts,
            key=_k_visao,
            help=(
                "**Consolidado**: linhas com `faturamento_consolidado` (com NF permitido **ou** sem NF permitido). "
                "Não é “soma contábil” entre empresas. "
                f"**{_FAT_PAINEL_VISAO_COM_NF}**: coluna `faturamento_com_nf` — critério do **export de pedidos** "
                "(ex.: “Existe Nota Fiscal gerada”); **não** é o mesmo que **nota vinculada** na base fiscal "
                "(`faturamento_nota_vinculada`), filtrada em “Com / sem nota” acima. "
                "**Sem NF permitido**: `faturamento_sem_nf`."
            ),
        )
        if not use_modulo_recorte:
            r0 = st.columns((1.15, 1.15))
            if has_usable_dates:
                cal_min, cal_max = _faturamento_period_calendar_limits(d_min, d_max)
                k_ini = f"fat_d_ini_{_oid}"
                k_fim = f"fat_d_fim_{_oid}"
                _sig_k = f"fat_d_bounds_sig_{_oid}"
                _sig = (d_min.isoformat(), d_max.isoformat())
                if st.session_state.get(_sig_k) != _sig:
                    st.session_state[_sig_k] = _sig
                    st.session_state[k_ini] = d_min
                    st.session_state[k_fim] = min(d_max, datetime.now(_BR_TZ).date())
                if k_ini not in st.session_state:
                    st.session_state[k_ini] = d_min
                if k_fim not in st.session_state:
                    st.session_state[k_fim] = min(d_max, datetime.now(_BR_TZ).date())
                st.session_state[k_ini] = min(
                    max(_safe_streamlit_date(st.session_state[k_ini], d_min), cal_min),
                    cal_max,
                )
                st.session_state[k_fim] = min(
                    max(_safe_streamlit_date(st.session_state[k_fim], d_max), cal_min),
                    cal_max,
                )
                with r0[0]:
                    d_ini = st.date_input(
                        "Período — início (Data)",
                        min_value=cal_min,
                        max_value=cal_max,
                        format="DD/MM/YYYY",
                        key=k_ini,
                        help=_FATURAMENTO_HELP_PERIODO_DATA,
                    )
                with r0[1]:
                    d_fim = st.date_input(
                        "Período — fim (Data)",
                        min_value=cal_min,
                        max_value=cal_max,
                        format="DD/MM/YYYY",
                        key=k_fim,
                        help=_FATURAMENTO_HELP_PERIODO_DATA,
                    )
                st.caption(
                    "Datas fora do intervalo dos dados são permitidas; só entram linhas com **Data** no período escolhido."
                )
            elif has_data_col:
                st.caption(
                    "A coluna **Data** existe mas não tem valores parseáveis — o filtro por período está desativado."
                )
            r1 = st.columns((1.15, 1.15))
            with r1[0]:
                sel_plat = _multiselect_stable(f"fat_ms_plat_{_oid}", "Plataforma", plats)
            with r1[1]:
                sel_sit = _multiselect_stable(f"fat_ms_sit_{_oid}", "Situação do pedido", sits)
        else:
            sel_plat = []
            sel_sit = []
        busca = st.text_input(
            "Busca (pedido, multiloja, SKU, n.º da nota)",
            key=f"fat_busca_{_oid}",
            placeholder="Texto livre…",
        ).strip().lower()
        sel_alerts = st.multiselect(
            "Alertas",
            list(_opt_alertas),
            key=f"fat_ms_alert_{_oid}",
            placeholder="Nenhum filtro por alerta",
        )
        st.caption(
            "Divergência **receita × valor pago**: só quando a diferença não fecha com "
            "**Receita_Bruta − Desconto proporcional total − Valor total** (desconto comercial da fonte). "
            "Sem coluna de desconto, mantém-se a comparação direta receita de lista vs. valor pago."
        )
        if st.button("Limpar filtros", key=f"fat_clear_{_oid}"):
            for _k in _faturamento_filter_keys(_oid):
                st.session_state.pop(_k, None)
            st.rerun()

    if not use_modulo_recorte and has_usable_dates:
        d_ini = _safe_streamlit_date(d_ini, d_min)
        d_fim = _safe_streamlit_date(d_fim, d_max)

    _fdl_ui_gap_tight() if use_modulo_recorte and mvp_rotulos_bloco_dre else _fdl_ui_gap_section()

    filt = work.copy()
    if visao == "Consolidado":
        filt = filt.loc[_faturamento_series_bool_mask(filt["faturamento_consolidado"])].copy()
    elif visao == _FAT_PAINEL_VISAO_COM_NF:
        filt = filt.loc[_faturamento_series_bool_mask(filt["faturamento_com_nf"])].copy()
    elif visao == "Sem NF permitido":
        filt = filt.loc[_faturamento_series_bool_mask(filt["faturamento_sem_nf"])].copy()

    if not use_modulo_recorte and has_usable_dates:
        if d_fim < d_ini:
            st.warning("A data final não pode ser anterior à inicial.")
            d_fim = d_ini
        filt = filt.loc[_faturamento_mask_venda_no_periodo(filt["Data"], d_ini, d_fim)].copy()

    if not use_modulo_recorte and sel_plat:
        filt = filt[filt["Nome da plataforma"].isin(sel_plat)]
    if not use_modulo_recorte and sel_sit:
        filt = filt[filt["Situação"].isin(sel_sit)]
    if busca:
        m_bus = pd.Series(False, index=filt.index)
        for col in (
            "Número do pedido",
            "Número do pedido multiloja",
            "Código",
            "Número da nota",
            "Número",
            "Data",
            "Data do faturamento",
        ):
            if col in filt.columns:
                m_bus = m_bus | filt[col].fillna("").astype(str).str.lower().str.contains(busca, regex=False)
        filt = filt.loc[m_bus].copy()

    if sel_alerts:
        m_a = pd.Series(False, index=filt.index)
        if "Preço lista zero" in sel_alerts:
            m_a = m_a | filt["_ab_pl_zero"]
        if "Divergência receita × valor (não explicada por desconto)" in sel_alerts:
            m_a = m_a | filt["_ab_div"]
        if "Sem NF não permitido" in sel_alerts:
            m_a = m_a | filt["_ab_sem_nf_np"]
        filt = filt.loc[m_a].copy()

    if "Data" in filt.columns:
        _sort_dt = _faturamento_ts_pedido_para_dia_civil(filt["Data"])
        filt = (
            filt.assign(_fdl_sort_dt=_sort_dt)
            .sort_values("_fdl_sort_dt", ascending=False, na_position="last")
            .drop(columns=["_fdl_sort_dt"])
        )
    else:
        filt = filt.sort_values(res_col, ascending=True, na_position="last")

    receita_s = _faturamento_painel_receita_series(filt, pl_col)
    receita_sum = float(receita_s.fillna(0).sum())
    res_sum = float(pd.to_numeric(filt[res_col], errors="coerce").fillna(0).sum())
    margem_total = (res_sum / receita_sum) if receita_sum not in (0.0, -0.0) else float("nan")
    n_cons = (
        int(_faturamento_series_bool_mask(filt["faturamento_consolidado"]).sum())
        if "faturamento_consolidado" in filt.columns
        else 0
    )
    any_alert = filt["_ab_pl_zero"] | filt["_ab_div"] | filt["_ab_sem_nf_np"]
    n_alert = int(any_alert.sum())

    st.subheader("📊 Indicadores")
    if mvp_rotulos_bloco_dre:
        st.caption(
            "Sobre o **recorte do módulo**, depois de **Visão** / **alertas** / **busca** (se usar). "
            "**Visão geral** (topo) resume o recorte **antes** desses três. "
            "**Receita bruta** na grelha = **Receita_Bruta** por linha. **Resultado** e **Margem %** respeitam linhas sem custo."
        )
    else:
        st.caption(
            "Valores sobre o recorte filtrado. **Receita por Produtos** = soma da receita por linha "
            "(Receita_Bruta do materializado, ou preço × quantidade); alinhado à coluna **Receita (produto)** na tabela. "
            "Com custo alocado na maior parte das linhas, **Resultado total** e **Margem %** passam a refletir margem real."
        )
    _fdl_ui_gap_tight() if mvp_rotulos_bloco_dre else _fdl_ui_gap_section()
    if mvp_rotulos_bloco_dre:
        fk1, fk2, fk3, fk4 = st.columns(4)
        _fk = [fk1, fk2, fk3, fk4]
    else:
        fk1, fk2, fk3, fk4, fk5 = st.columns(5)
        _fk = [fk1, fk2, fk3, fk4, fk5]
    with _fk[0]:
        st.metric(
            "Receita bruta" if mvp_rotulos_bloco_dre else "Receita por Produtos",
            _fmt_brl_ptbr_celula(receita_sum),
        )
    with _fk[1]:
        st.metric("Resultado Total", _fmt_brl_ptbr_celula(res_sum))
    with _fk[2]:
        if receita_sum == 0 or (isinstance(margem_total, float) and math.isnan(margem_total)):
            st.metric("Margem Total %", "—")
        else:
            st.metric("Margem Total %", f"{margem_total * 100:.2f}%".replace(".", ","))
    if mvp_rotulos_bloco_dre:
        with _fk[3]:
            st.metric("Alertas Ativos", _fmt_int_ptbr(n_alert))
    else:
        with _fk[3]:
            st.metric("Itens Consolidados", _fmt_int_ptbr(n_cons))
        with _fk[4]:
            st.metric("Alertas Ativos", _fmt_int_ptbr(n_alert))

    if "Status_Custo" in filt.columns:
        _stc = filt["Status_Custo"].astype(str).str.strip()
        _n_ok_f = int(_stc.eq("CUSTO_OK").sum())
        _n_exc_f = int(len(filt) - _n_ok_f)
        if _n_exc_f == 0:
            st.caption(
                f"**{_n_ok_f}** linhas com custo alocado. O resultado total considera apenas linhas com custo."
            )
        else:
            st.caption(
                f"**{_n_ok_f}** linhas com custo alocado; **{_n_exc_f}** exceções sem custo.\n\n"
                "O resultado total considera apenas linhas com custo."
            )
    else:
        st.caption("O mesmo recorte aplica-se à tabela seguinte.")

    prod_col = _faturamento_resolve_produto_column(list(filt.columns))
    rpct = pd.to_numeric(filt["Resultado_Pct"], errors="coerce") if "Resultado_Pct" in filt.columns else pd.Series(float("nan"), index=filt.index)
    receita_linha = _faturamento_painel_receita_series(filt, pl_col)
    _ix = filt.index
    if "empresa" in filt.columns:
        _em = filt["empresa"].fillna("").astype(str).str.strip()
        empresa_s = _em.mask(_em.eq(""), "—")
    elif "org_id" in filt.columns:
        empresa_s = filt["org_id"].astype(str).str.strip()
    else:
        empresa_s = pd.Series("—", index=_ix)
    pedido_s = _faturamento_pedido_display_series(filt)

    if mvp_rotulos_bloco_dre:
        _data_s = (
            _faturamento_disp_data_pedidos(filt["Data"])
            if "Data" in filt.columns
            else pd.Series("—", index=_ix)
        )
        _prod_s = filt[prod_col].astype(str) if prod_col else pd.Series("", index=_ix)
        _qtd_s = (
            pd.to_numeric(filt["Quantidade"], errors="coerce")
            if "Quantidade" in filt.columns
            else pd.Series(float("nan"), index=_ix)
        )
        _vl_v = _faturamento_dre_vl_venda_series(filt, pl_col)
        _vl_nf = _faturamento_dre_vl_nota_fiscal_series(filt)
        _nf_s = _faturamento_dre_nf_coluna_display(filt)
        _fr_s = _faturamento_dre_frete_display_series(filt)
        _main_cols: list[tuple[str, pd.Series]] = [
            ("Data", _data_s),
            ("Empresa", empresa_s),
            ("Plataforma", filt["Nome da plataforma"]),
            ("Pedido", pedido_s),
            ("NF", _nf_s),
            ("Produto", _prod_s),
            ("SKU", filt["Código"]),
            ("Qtd", _qtd_s),
            ("Vl. Venda", _vl_v),
            ("Vl. Nota Fiscal", _vl_nf),
            ("Frete", _fr_s),
            ("Comissão", pd.to_numeric(filt["Taxa de Comissão"], errors="coerce")),
            ("Custo Total", pd.to_numeric(filt[custo_prod_col], errors="coerce")),
            ("Despesa fixa", pd.to_numeric(filt["Despesas Fixas"], errors="coerce")),
            ("Imposto", pd.to_numeric(filt["Imposto"], errors="coerce")),
            ("Resultado", pd.to_numeric(filt[res_col], errors="coerce")),
        ]
        disp = pd.DataFrame(dict(_main_cols))
        disp["Alertas"] = filt.apply(_faturamento_alertas_text, axis=1)

        _extra_export: dict[str, pd.Series] = {}
        if "Resultado_Pct" in filt.columns:
            _extra_export["Resultado %"] = rpct * 100.0
        if "Status_Custo" in filt.columns:
            _extra_export["Status custo"] = filt["Status_Custo"].astype(str)
        _extra_export["Receita bruta (auditoria)"] = receita_linha
        if "Valor total" in filt.columns:
            _extra_export["Valor total (pedido)"] = pd.to_numeric(filt["Valor total"], errors="coerce")
        _extra_export["Situação do pedido"] = filt["Situação"].astype(str)
        _extra_export["N.º do pedido"] = _faturamento_disp_texto_sem_none(filt["Número do pedido"])
        _extra_export["N.º pedido multiloja"] = _faturamento_disp_texto_sem_none(filt["Número do pedido multiloja"])
        if "Data do faturamento" in filt.columns:
            _extra_export["Data do faturamento"] = _faturamento_disp_data_pedidos(filt["Data do faturamento"])
        _extra_export["NF emitida? (pedido)"] = _faturamento_disp_texto_sem_none(filt["Existe Nota Fiscal gerada"])
        if "Frete Mercado Envios" in filt.columns and "Frete transportadora própria" in filt.columns:
            _extra_export["Frete Mercado Envios"] = pd.to_numeric(filt["Frete Mercado Envios"], errors="coerce")
            _extra_export["Frete transp. própria"] = pd.to_numeric(filt["Frete transportadora própria"], errors="coerce")
            _extra_export["Custo frete (total)"] = pd.to_numeric(filt["Custo de Frete"], errors="coerce")
        if "Número" in filt.columns:
            _extra_export["Ref. ML (col. Número)"] = _faturamento_disp_texto_sem_none(filt["Número"])
        if "Base_Imposto" in filt.columns:
            _extra_export["Base imposto"] = pd.to_numeric(filt["Base_Imposto"], errors="coerce")
        if "Nota_Situacao" in filt.columns:
            _extra_export["Situação da NF"] = filt["Nota_Situacao"].fillna("").astype(str)
        if "Nota_Data_Emissao" in filt.columns:
            _extra_export["Data emissão NF"] = _faturamento_disp_data_pedidos(filt["Nota_Data_Emissao"])
        _export = pd.concat([disp, pd.DataFrame(_extra_export)], axis=1)

        _cfg = {}
        for c in (
            "Vl. Venda",
            "Vl. Nota Fiscal",
            "Frete",
            "Comissão",
            "Custo Total",
            "Despesa fixa",
            "Imposto",
            "Resultado",
        ):
            if c in disp.columns:
                _hc = (
                    _FATURAMENTO_HELP_VALOR_NOTA_FISCAL
                    if c == "Vl. Nota Fiscal"
                    else None
                )
                _cfg[c] = (
                    NumberColumn(c, format="R$ %,.2f", help=_hc)
                    if _hc
                    else NumberColumn(c, format="R$ %,.2f")
                )
        if "Qtd" in disp.columns:
            _cfg["Qtd"] = NumberColumn("Qtd", format="%.2f")
        for c in ("Data", "Empresa", "Plataforma", "Pedido", "NF", "Produto", "SKU", "Alertas"):
            if c in disp.columns:
                _tc_kw: dict[str, str | bool] = {"width": "large" if c == "Alertas" else "medium"}
                if c == "Data":
                    _tc_kw["help"] = _FATURAMENTO_HELP_PERIODO_DATA
                elif c == "NF":
                    _tc_kw["help"] = "N.º a partir do join com notas de saída; se vazio, texto do pedido quando existir."
                _cfg[c] = TextColumn(c, **_tc_kw)

        st.subheader("Tabela principal")
        st.caption(
            f"{len(disp)} linhas · ordenação por **Data da venda** (mais recente primeiro). "
            "**CSV** exporta **as mesmas linhas** que a grelha (após **Visão** / alertas / busca). "
            "**Vl. Venda** = comercial (**Vl_Venda** ou equivalente). **Vl. Nota Fiscal** = "
            "**Nota_Valor_Liquido_Rateado** quando há join fiscal; senão **Valor total** do pedido. "
            "**Frete** = **Frete_Plataforma** ou **Custo de Frete**. "
            "O CSV inclui ainda colunas extra (pedido, frete detalhado, base fiscal, etc.)."
        )
    else:
        _core: list[tuple[str, pd.Series]] = [
            (
                "Data",
                _faturamento_disp_data_pedidos(filt["Data"])
                if "Data" in filt.columns
                else pd.Series("—", index=_ix),
            ),
            ("Empresa", empresa_s),
            ("Plataforma", filt["Nome da plataforma"]),
            ("Pedido", pedido_s),
            ("SKU", filt["Código"]),
            ("Produto", filt[prod_col].astype(str) if prod_col else pd.Series("", index=_ix)),
            ("Receita bruta", receita_linha),
            (_FATURAMENTO_UI_VALOR_NOTA_FISCAL, pd.to_numeric(filt["Valor total"], errors="coerce")),
            ("Resultado", pd.to_numeric(filt[res_col], errors="coerce")),
            ("Resultado %", rpct * 100.0),
        ]
        if "Status_Custo" in filt.columns:
            _core.append(("Status custo", filt["Status_Custo"].astype(str)))
        disp = pd.DataFrame(dict(_core))
        disp["Alertas"] = filt.apply(_faturamento_alertas_text, axis=1)

        _tail: list[tuple[str, pd.Series]] = [
            ("Situação do pedido", filt["Situação"].astype(str)),
            ("N.º do pedido", _faturamento_disp_texto_sem_none(filt["Número do pedido"])),
            ("N.º pedido multiloja", _faturamento_disp_texto_sem_none(filt["Número do pedido multiloja"])),
            (
                "Data do faturamento",
                _faturamento_disp_data_pedidos(filt["Data do faturamento"])
                if "Data do faturamento" in filt.columns
                else pd.Series("—", index=_ix),
            ),
            ("NF emitida?", _faturamento_disp_texto_sem_none(filt["Existe Nota Fiscal gerada"])),
            ("N.º da nota", _faturamento_disp_texto_sem_none(filt["Número da nota"])),
        ]
        if "Quantidade" in filt.columns:
            _tail.append(("Quantidade", pd.to_numeric(filt["Quantidade"], errors="coerce")))
        _tail.append(("Custo do produto", pd.to_numeric(filt[custo_prod_col], errors="coerce")))
        if "Frete Mercado Envios" in filt.columns and "Frete transportadora própria" in filt.columns:
            _tail.append(("Frete Mercado Envios", pd.to_numeric(filt["Frete Mercado Envios"], errors="coerce")))
            _tail.append(
                ("Frete transp. própria", pd.to_numeric(filt["Frete transportadora própria"], errors="coerce"))
            )
            _tail.append(("Custo frete (total)", pd.to_numeric(filt["Custo de Frete"], errors="coerce")))
        else:
            _tail.append(("Frete", pd.to_numeric(filt["Custo de Frete"], errors="coerce")))
        _tail.extend(
            [
                ("Comissão Plataforma", pd.to_numeric(filt["Taxa de Comissão"], errors="coerce")),
                ("Imposto", pd.to_numeric(filt["Imposto"], errors="coerce")),
                ("Despesas fixas", pd.to_numeric(filt["Despesas Fixas"], errors="coerce")),
            ]
        )
        _tail.append(
            (
                "Ref. ML (col. Número)",
                _faturamento_disp_texto_sem_none(filt["Número"])
                if "Número" in filt.columns
                else pd.Series("—", index=_ix),
            )
        )
        disp = pd.concat([disp, pd.DataFrame(dict(_tail))], axis=1)
        _export = disp.copy()

        _cfg = {}
        money_cols = (
            "Receita bruta",
            _FATURAMENTO_UI_VALOR_NOTA_FISCAL,
            "Custo do produto",
            "Frete Mercado Envios",
            "Frete transp. própria",
            "Custo frete (total)",
            "Frete",
            "Comissão Plataforma",
            "Imposto",
            "Despesas fixas",
            "Resultado",
        )
        for c in money_cols:
            if c in disp.columns:
                if c == _FATURAMENTO_UI_VALOR_NOTA_FISCAL:
                    _cfg[c] = NumberColumn(
                        c,
                        format="R$ %,.2f",
                        help=_FATURAMENTO_HELP_VALOR_NOTA_FISCAL,
                    )
                elif c == "Custo frete (total)":
                    _cfg[c] = NumberColumn(
                        c,
                        format="R$ %,.2f",
                        help="Soma ME + transportadora própria; continua a entrar no **Resultado**.",
                    )
                else:
                    _cfg[c] = NumberColumn(c, format="R$ %,.2f")
        if "Resultado %" in disp.columns:
            _cfg["Resultado %"] = NumberColumn("Resultado %", format="%.2f%%")
        if "Quantidade" in disp.columns:
            _cfg["Quantidade"] = NumberColumn("Quantidade", format="%.2f")
        if "Status custo" in disp.columns:
            _cfg["Status custo"] = TextColumn("Status custo", width="small")
        for c in (
            "Data",
            "Empresa",
            "Plataforma",
            "Pedido",
            "SKU",
            "Produto",
            "Alertas",
            "Situação do pedido",
            "N.º do pedido",
            "N.º pedido multiloja",
            "Data do faturamento",
            "Ref. ML (col. Número)",
            "NF emitida?",
            "N.º da nota",
        ):
            if c in disp.columns:
                _tc_kw = {"width": "large" if c == "Alertas" else "medium"}
                if c == "Data":
                    _tc_kw["help"] = _FATURAMENTO_HELP_PERIODO_DATA
                elif c == "Data do faturamento":
                    _tc_kw["help"] = (
                        "Campo secundário / futuro (competência fiscal). Pode estar vazio ou inconsistente; "
                        "o período do módulo usa a coluna **Data**."
                    )
                elif c == "N.º da nota":
                    _tc_kw["help"] = _FATURAMENTO_HELP_NUMERO_NF_COL
                _cfg[c] = TextColumn(c, **_tc_kw)

        st.subheader("Tabela principal")
        st.caption(
            f"{len(disp)} linhas · ordenação por **Data da venda** (mais recente primeiro). "
            "À esquerda: operação do dia; à direita: NF, quantidade e composição de custos/taxas (**Ref. ML** no fim). "
            f"**{_FATURAMENTO_UI_VALOR_NOTA_FISCAL}** = **Valor total** (convenção; ver ajuda da coluna — não é NF-e legal). "
            "**N.º da nota** vazio: esperado sem integração fiscal. "
            "Frete repartido (ME / transp. própria) quando o CSV tiver modalidade; total em **Custo frete (total)**. "
            "**Pedido** = multiloja se existir, senão n.º do pedido."
        )
        _export = disp.copy()

    st.dataframe(
        disp,
        use_container_width=True,
        hide_index=True,
        height=440,
        column_config=_cfg,
    )
    if "SKU_Normalizado" in filt.columns:
        _export["SKU_Normalizado"] = filt["SKU_Normalizado"].astype(str)
    if mvp_rotulos_bloco_dre:
        st.caption(
            "O **CSV exportado** inclui as colunas da tabela principal e **campos adicionais** "
            "(Resultado %, pedido, frete detalhado, base imposto, dados fiscais da NF quando existirem"
            + (", **SKU_Normalizado**" if "SKU_Normalizado" in filt.columns else "")
            + ")."
        )
    elif "SKU_Normalizado" in filt.columns:
        st.caption(
            "O **CSV exportado** inclui a coluna **SKU_Normalizado** (chave de join), útil para cruzar com a planilha de custo."
        )
    st.download_button(
        "Exportar CSV (filtrado)",
        _export.to_csv(index=False).encode("utf-8-sig"),
        file_name="faturamento_filtrado.csv",
        mime="text/csv",
        key=f"fat_dl_csv_{_oid}",
    )


def _render_kpi_card(
    label: str,
    value: str,
    icon: str,
    css_class: str,
    *,
    frete_variant: bool = False,
) -> None:
    _ = css_class, frete_variant  # compat. com assinatura antiga — UI só com componentes nativos
    with st.container(border=True):
        st.metric(f"{icon} {label}", value)


def _frete_meta_for_render(load_info: dict[str, object]) -> dict[str, object]:
    keys = (
        "vendas_arquivo",
        "frete_arquivo",
        "frete_tabular",
        "debug_logs",
        "avisos",
        "linhas",
    )
    return {k: load_info[k] for k in keys if k in load_info}


def _frete_org_widget_suffix(org_id: str) -> str:
    return "".join(c if c.isalnum() else "_" for c in org_id)[:80]


def _render_frete_operacional_ui(
    org_id: str,
    df_frete: pd.DataFrame,
    meta_frete: dict[str, object],
    ts_proc: str,
    load_info: dict[str, object],
) -> None:
    """Filtros + contexto + KPIs + tabela + export (alinhado ao painel de repasse)."""
    _is_admin = _is_admin_mode()
    _sig = _frete_org_widget_suffix(org_id)

    try:
            _fdl_cp_inject_panel_styles()
            _fdl_frete_inject_panel_styles()
            _frete_stage_trace(2, "Filtros", "início")
            work = df_frete.copy()
            today = datetime.now(_BR_TZ).date()
            default_ini = today - timedelta(days=29)
            default_fim = today
        
            if "data_venda" in work.columns or "_data_venda_dt" in work.columns:
                dts = frete_series_normalize_sale_dt(frete_series_for_date_filter(work))
                d_min_data, d_max_data, have_dt = _series_datetime_bounds_dates(dts)
                if not have_dt:
                    d_min_data = d_max_data = today
            else:
                d_min_data = d_max_data = today
        
            picker_min = min(d_min_data, default_ini)
            picker_max = max(d_max_data, default_fim, today)
            if picker_max < picker_min:
                picker_min, picker_max = picker_max, picker_min

            # Janela «últimos 30 dias» ancorada em hoje pode não cruzar as datas reais do CSV
            # (ex.: export só 2025 com app em 2026 → recorte filtrado zerava KPIs e detalhamento).
            d_ini_try = max(d_min_data, default_ini)
            d_fim_try = min(d_max_data, default_fim)
            if d_ini_try > d_fim_try:
                d_ini_val, d_fim_val = d_min_data, d_max_data
            else:
                d_ini_val, d_fim_val = d_ini_try, d_fim_try
        
            estados: list[str] = []
            if "Estado" in work.columns:
                estados = sorted(
                    {str(x).strip() for x in work["Estado"].dropna().unique().tolist() if str(x).strip()}
                )
            situacao_opts: list[str] = list(FRETE_SITUACAO_FRETE_VALORES_FILTRO)
        
            with st.container(border=True):
                st.markdown('<p class="fdl-frete-filtros-h">Filtros</p>', unsafe_allow_html=True)
                st.markdown(
                    '<p class="fdl-frete-caption">Recorte por <strong>data da venda</strong> (dia civil) · estado, '
                    "situação do frete e busca. Por omissão: últimos 30 dias até hoje.</p>",
                    unsafe_allow_html=True,
                )
                r2 = st.columns((1, 1))
                with r2[0]:
                    st.caption("Início")
                    data_ini = st.date_input(
                        "Data da venda — início",
                        value=d_ini_val,
                        min_value=picker_min,
                        max_value=picker_max,
                        format="DD/MM/YYYY",
                        key=f"op_frete_d_ini_{_sig}",
                        label_visibility="collapsed",
                    )
                with r2[1]:
                    st.caption("Fim")
                    data_fim = st.date_input(
                        "Data da venda — fim",
                        value=d_fim_val,
                        min_value=picker_min,
                        max_value=picker_max,
                        format="DD/MM/YYYY",
                        key=f"op_frete_d_fim_{_sig}",
                        label_visibility="collapsed",
                    )
                r1 = st.columns((1, 1, 1.35))
                with r1[0]:
                    sel_est = _multiselect_stable(
                        f"op_frete_ms_est_{_sig}", "Estado da venda", estados, compact_label=True
                    )
                with r1[1]:
                    sel_sit = _multiselect_stable(
                        f"op_frete_ms_situacao_{_sig}",
                        FRETE_UI_SITUACAO_FRETE,
                        situacao_opts,
                        compact_label=True,
                    )
                with r1[2]:
                    busca = st.text_input(
                        "Busca (venda ou # anúncio)",
                        "",
                        placeholder="Venda, anúncio…",
                        key=f"op_frete_busca_{_sig}",
                    )
                    busca = busca.strip().lower()
                _clr_l, _clr_r = st.columns((1.55, 1))
                with _clr_r:
                    if st.button(
                        "Limpar filtros desta vista",
                        key=f"op_frete_clear_{_sig}",
                        use_container_width=True,
                        help="Repor datas aos limites úteis da base, limpar estado/situação e a busca.",
                    ):
                        for _k in (
                            f"op_frete_d_ini_{_sig}",
                            f"op_frete_d_fim_{_sig}",
                            f"op_frete_ms_est_{_sig}",
                            f"op_frete_ms_situacao_{_sig}",
                            f"op_frete_busca_{_sig}",
                        ):
                            st.session_state.pop(_k, None)
                        st.rerun()

            data_ini = _safe_streamlit_date(data_ini, d_ini_val)
            data_fim = _safe_streamlit_date(data_fim, d_fim_val)

            if data_fim < data_ini:
                st.warning("A data final não pode ser anterior à data inicial. Ajuste o período.")
                data_fim = data_ini
        
            tbl = work
            if sel_est and "Estado" in tbl.columns:
                tbl = tbl[tbl["Estado"].isin(sel_est)]
            if sel_sit:
                sit_tbl = compute_frete_situacao_frete_column(tbl)
                tbl = tbl.loc[sit_tbl.isin(sel_sit)]
            if busca:
                m = (
                    tbl[FRETE_UI_N_VENDA].fillna("").astype(str).str.lower().str.contains(busca, regex=False)
                    if FRETE_UI_N_VENDA in tbl.columns
                    else pd.Series(False, index=tbl.index)
                )
                if FRETE_UI_ANUNCIO in tbl.columns:
                    m = m | tbl[FRETE_UI_ANUNCIO].fillna("").astype(str).str.lower().str.contains(
                        busca, regex=False
                    )
                tbl = tbl.loc[m]
        
            if "data_venda" in tbl.columns or "_data_venda_dt" in tbl.columns:
                dd = frete_series_normalize_sale_dt(frete_series_for_date_filter(tbl))
                if dd.notna().any():
                    ini = pd.Timestamp(data_ini)
                    fim = pd.Timestamp(data_fim) + pd.Timedelta(days=1)
                    tbl = tbl.loc[dd.notna() & (dd >= ini) & (dd < fim)]
        
            tbl_show = tbl[[c for c in tbl.columns if not str(c).startswith("_")]].copy()
            if "data_venda" not in tbl_show.columns and "_data_venda_dt" in tbl.columns:
                tbl_show["data_venda"] = tbl["_data_venda_dt"]
            if FRETE_UI_CLASSIFICACAO in tbl_show.columns:
                tbl_show = tbl_show.drop(columns=[FRETE_UI_CLASSIFICACAO])
        
            _miss_req = [c for c in (FRETE_UI_N_VENDA, FRETE_UI_DIFERENCA) if c not in tbl_show.columns]
            if _miss_req:
                st.error(
                    "Colunas obrigatórias ausentes para a Conciliação de Frete: "
                    + ", ".join(repr(c) for c in _miss_req)
                    + ". Verifique o CSV materializado (dataset_frete_app.csv) ou o export ML."
                )
                if _frete_debug_ui_enabled():
                    st.json({"colunas_presentes": list(tbl_show.columns)[:120]})
                return
        
            if _frete_debug_ui_enabled():
                st.caption("Debug Frete: filtros aplicados — a seguir KPIs e tabelas.")
        
            _va = str(meta_frete.get("vendas_arquivo", "—"))
            _ts_esc = str(ts_proc)
            _pl = "Todas"
            if estados and sel_est and len(sel_est) < len(estados):
                _pl = ", ".join(sel_est[:2]) + ("..." if len(sel_est) > 2 else "")
            elif estados and not sel_est:
                _pl = "Todas"
            elif not estados:
                _pl = "—"
            _fdl_ui_gap_section_lg()
            st.caption(
                f"Estado (filtro): **{_pl}** · Atualizado: **{_ts_esc}** · "
                f"Venda: **{data_ini.strftime('%d/%m/%Y')}**–**{data_fim.strftime('%d/%m/%Y')}** · "
                f"Fonte: **{_va}**"
            )

            _rec_key = f"op_frete_recebido_{_sig}"
            if _rec_key not in st.session_state:
                st.session_state[_rec_key] = {}
            rec_map: dict[str, bool] = st.session_state[_rec_key]
        
            nv_s = tbl_show[FRETE_UI_N_VENDA].map(lambda x: str(x).strip() if pd.notna(x) else "")
            recebido_series = nv_s.map(lambda x: FRETE_VAL_RECEBIDO_SIM if rec_map.get(x) else FRETE_VAL_RECEBIDO_NAO)
            recebido_series.index = tbl_show.index
            _frete_stage_trace(2, "Filtros", f"concluída — {len(tbl_show)} linhas após filtros")
    except Exception as exc:
        _frete_stage_error(2, "Filtros e preparação da tabela filtrada", exc)
        return

    try:
        _frete_stage_trace(3, "KPIs e anúncios", "início")
        kpi_ex = frete_kpis_executivos(tbl_show)
        tbl_cob_maior = frete_tabela_anuncios_cobrado_maior(tbl_show)
        tbl_repasse = frete_tabela_anuncios_repasse_frete(tbl_show, recebido_series)

        _fdl_ui_gap_section_lg()
        st.markdown(
            '<p class="fdl-frete-section-title">Indicadores</p>'
            '<p class="fdl-frete-section-note">Volume de vendas e montantes |Δ| no recorte (mesma regra de sempre).</p>',
            unsafe_allow_html=True,
        )
        _fdl_ui_gap_section()
        _render_frete_indicadores_kpis(kpi_ex, n_linhas=len(tbl_show))

        if _is_admin and FRETE_UI_STATUS_CONC in tbl_show.columns:
            st.caption(
                "Modo técnico: existe **Status conciliação** nos dados; a priorização segue **Situação do Frete**."
            )

        _sem_anuncio = FRETE_UI_ANUNCIO not in tbl_show.columns
        st.divider()
        _fdl_ui_gap_section()
        st.markdown(
            '<p class="fdl-frete-section-title">Cobrado a maior (por anúncio)</p>'
            '<p class="fdl-frete-section-note">Anúncios onde o frete cobrado excede o esperado — prioridade para recuperação.</p>',
            unsafe_allow_html=True,
        )
        if _sem_anuncio:
            st.info("Inclua o **# do anúncio** no export de vendas para agregar por anúncio.")
        elif tbl_cob_maior.empty:
            st.info("Sem anúncios com cobrança a maior neste recorte. Ajuste período ou critérios.")
        else:
            _h1 = min(420, 120 + 36 * max(len(tbl_cob_maior), 1))
            st.dataframe(
                _format_frete_anuncio_tabela_display(tbl_cob_maior),
                use_container_width=True,
                hide_index=True,
                height=_h1,
            )

        st.divider()
        _fdl_ui_gap_section()
        st.markdown(
            '<p class="fdl-frete-section-title">Repasse de frete (por anúncio)</p>'
            '<p class="fdl-frete-section-note">Anúncios com repasse a validar; «Recebido?» no detalhamento abaixo.</p>',
            unsafe_allow_html=True,
        )
        if _sem_anuncio:
            pass
        elif tbl_repasse.empty:
            st.info("Sem repasse de frete a validar neste recorte. Ajuste período ou critérios.")
        else:
            _h2 = min(420, 120 + 36 * max(len(tbl_repasse), 1))
            st.dataframe(
                _format_frete_anuncio_tabela_display(tbl_repasse),
                use_container_width=True,
                hide_index=True,
                height=_h2,
            )
    
        _frete_stage_trace(3, "KPIs e anúncios", "concluída")
    except Exception as exc:
        _frete_stage_error(3, "Indicadores executivos e tabelas por anúncio", exc)
        return
    try:
        _frete_stage_trace(4, "Exportação e detalhe", "início")
        for w in meta_frete.get("avisos") or []:
            st.info(w)

        st.divider()
        _fdl_ui_gap_section()
        st.markdown(
            '<p class="fdl-frete-section-title">Detalhamento das vendas</p>'
            '<p class="fdl-frete-section-note">Grelha executiva do recorte · exporte CSV/Excel ou ajuste «Recebido?» quando disponível.</p>',
            unsafe_allow_html=True,
        )
        _fdl_ui_gap_section()

        t_export_view = dataframe_frete_conciliacao_principal(tbl_show, layout="executivo")
        csv_bytes = t_export_view.to_csv(index=False).encode("utf-8-sig")
        t_excel = dataframe_frete_conciliacao_principal(tbl_show, layout="executivo")
        excel_buf = BytesIO()
        with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
            t_excel.to_excel(writer, index=False, sheet_name="Frete")
            ws = writer.sheets["Frete"]
            header_row = [cell.value for cell in ws[1]]
            for c_data in ("Data da venda",):
                if c_data in header_row:
                    col_idx = header_row.index(c_data) + 1
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            cell.number_format = oxl_number_formats.FORMAT_DATE_DDMMYY
        excel_buf.seek(0)
        with st.container(border=True):
            st.caption("Exportar recorte filtrado")
            btn1, btn2 = st.columns([1, 1])
            with btn1:
                st.download_button(
                    "Exportar CSV",
                    data=csv_bytes,
                    file_name="conciliacao_frete_filtrada.csv",
                    mime="text/csv",
                    use_container_width=True,
                    key=f"op_frete_dl_csv_{_sig}",
                )
            with btn2:
                st.download_button(
                    "Exportar Excel",
                    data=excel_buf.getvalue(),
                    file_name="conciliacao_frete_filtrada.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"op_frete_dl_xlsx_{_sig}",
                )

        st.write("")
        st.write("")
        st.write("")

        t_grid = _dataframe_frete_grid(tbl_show, _fmt_brl_ptbr_celula, _col_referencia_como_texto)
        t_main = dataframe_frete_conciliacao_principal(t_grid, layout="executivo")
        t_main = _frete_conciliacao_grid_com_icones(t_main)
        t_display = frete_executivo_display_styled(t_main)
        _h_df = min(560, 140 + max(18 * min(len(t_main), 80), 120))

        if t_main.empty:
            st.info(
                "**Nenhuma venda** com os filtros atuais. Alargue o período de datas ou limpe a busca / multiselects."
            )
        else:
            st.dataframe(
                t_display,
                column_config=_column_config_frete(t_main),
                use_container_width=True,
                hide_index=True,
                height=_h_df,
            )
            st.caption(
                "Use o ícone **olho** na barra da tabela para mostrar ou ocultar colunas. "
                "Em **Situação do Frete**, os ícones indicam o estado (ex.: ✅ OK, ⬆️ cobrado a maior, 🚚 repasse). "
                f"**{len(t_main)}** linhas no filtro atual."
            )
    
        if _is_admin and load_info.get("frete_consume") in ("live", "live_fallback"):
            st.caption(
                "Modo técnico: **tempo real** — dados calculados diretamente das fontes; ficheiro consolidado indisponível ou falhou."
            )
    
    
        _frete_stage_trace(4, "Exportação e detalhe", "concluída")
    except Exception as exc:
        _frete_stage_error(4, "Exportação, detalhe de vendas, tabela principal e editor", exc)
        return


def _devolucoes_pool_excluded_redundant(d: pd.DataFrame) -> set[str]:
    """Colunas que não devem aparecer em lado nenhum (redundância com ``status_ml_texto``)."""
    out: set[str] = set()
    if "status_ml_texto" in d.columns and "Descrição do status" in d.columns:
        out.add("Descrição do status")
    return out


def _devolucoes_is_dup_suffix_col(name: object) -> bool:
    import re

    return bool(re.search(r"\.\d+$", str(name or "")))


def _devolucoes_audit_column_names() -> tuple[str, ...]:
    return (
        "vinculo_tipo",
        "vinculo_confianca",
        "vinculo_detalhe",
        "reembolso_valor_inferido",
        "lib_n_eventos",
        "lib_ultima_data_pagamento",
        "lib_descricoes_amostra",
        "lib_soma_net_debito",
        "jaci_cep_15155038",
        "jaci_endereco_score",
        "jaci_endereco_normalizado",
        "arquivo_origem_venda",
        "candidato_motivo",
        "empresa",
        "cliente_id",
        "empresa_id",
        "cnpj",
        "org_id",
    )


# Colunas de origem usadas só na tabela principal enxuta (o resto vai para a camada técnica).
_DEVOLUCOES_SLIM_SOURCE_COLS: tuple[str, ...] = (
    "N° de venda",
    "Data da venda",
    "status_interno",
    "acao_sugerida",
    "status_ml_texto",
    "classificacao_reembolso",
)


def _devolucoes_build_fila_operacional_view(d: pd.DataFrame) -> pd.DataFrame:
    """DataFrame só para exibição: 7 colunas com rótulos de negócio."""
    idx = d.index
    out: dict[str, object] = {}
    out["N.º da venda"] = d["N° de venda"] if "N° de venda" in d.columns else pd.Series("", index=idx)
    out["Data"] = d["Data da venda"] if "Data da venda" in d.columns else pd.Series("", index=idx)
    out["Situação"] = d["status_interno"] if "status_interno" in d.columns else pd.Series("", index=idx)
    out["O que fazer"] = d["acao_sugerida"] if "acao_sugerida" in d.columns else pd.Series("", index=idx)
    out["Mensagem do Mercado Livre"] = (
        d["status_ml_texto"] if "status_ml_texto" in d.columns else pd.Series("", index=idx)
    )
    out["Reembolso"] = (
        d["classificacao_reembolso"] if "classificacao_reembolso" in d.columns else pd.Series("", index=idx)
    )
    if "lib_n_eventos" in d.columns:
        _ln = pd.to_numeric(d["lib_n_eventos"], errors="coerce").fillna(0)
        out["Liberações"] = _ln.map(lambda x: "Com registro" if int(x) > 0 else "Sem registro")
    else:
        out["Liberações"] = pd.Series("Sem registro", index=idx)
    return pd.DataFrame(out, index=idx)


def _devolucoes_cols_tabela_detalhe(d: pd.DataFrame, principais: list[str]) -> list[str]:
    """
    Restante do export + auditoria + colunas ``.1`` — expander fechado por defeito.
    Ordem: colunas do DataFrame (meio útil), bloco auditoria, sufixos duplicados no fim.
    """
    pool_ex = _devolucoes_pool_excluded_redundant(d)
    ps = set(principais)
    g3 = _devolucoes_audit_column_names()
    g3_set = set(g3)
    out: list[str] = []
    seen: set[str] = set()

    for c in d.columns:
        if c in ps or c in pool_ex or c in seen:
            continue
        if c in g3_set or _devolucoes_is_dup_suffix_col(c):
            continue
        seen.add(c)
        out.append(c)
    for c in g3:
        if c in d.columns and c not in ps and c not in pool_ex and c not in seen:
            seen.add(c)
            out.append(c)
    for c in d.columns:
        if c in ps or c in pool_ex or c in seen:
            continue
        if _devolucoes_is_dup_suffix_col(c):
            seen.add(c)
            out.append(c)
    return out


def _painel_devolucoes_operacional(
    df: pd.DataFrame, load_info: dict[str, object], ts_proc: str
) -> None:
    """Painel só leitura sobre o materializado (candidatas a devolução)."""
    _is_admin = _is_admin_mode()
    n = int(len(df))
    st.markdown("### Fila de devoluções — Mercado Livre")
    st.markdown(
        "Vendas com devolução, retorno físico, reembolso ligado ao caso ou revisão operacional. "
        "Cancelamentos comerciais isolados já foram excluídos da fila."
    )
    st.caption(
        f"Use os indicadores e os filtros para priorizar o que precisa de ação agora. "
        f"**{_fmt_int_ptbr(n)}** vendas neste ficheiro · Atualizado **{ts_proc}**"
    )
    if df.empty:
        st.info(
            "Não há vendas nesta fila. Quando houver devoluções ou casos correlatos materializados, elas aparecerão aqui."
        )
        return

    d = df.copy()
    if "empresa" not in d.columns:
        d["empresa"] = _dataset_empresa_label()

    si_series = d["status_interno"].fillna("").astype(str).str.strip() if "status_interno" in d.columns else pd.Series("", index=d.index)
    n_conf = int((si_series == "Conferência física pendente").sum())
    n_rev = int((si_series == "Revisar financeiramente").sum())
    n_cob = int((si_series == "Cobrar plataforma").sum())
    n_pag = int((si_series == "Pagamento liberado ao vendedor").sum())
    n_reemb = int((si_series == "Reembolso ao comprador").sum())
    _lib_n = (
        pd.to_numeric(d["lib_n_eventos"], errors="coerce").fillna(0)
        if "lib_n_eventos" in d.columns
        else pd.Series(0, index=d.index)
    )
    n_vinc = int((_lib_n > 0).sum())

    st.markdown(
        '<p class="fdl-frete-section-title">Indicadores</p>'
        '<p class="fdl-frete-section-note">Contagens por tipo de situação nesta fila.</p>',
        unsafe_allow_html=True,
    )
    with st.container(border=True):
        k1, k2, k3, k4, k5 = st.columns(5)
        with k1:
            st.metric(
                "Conferência física pendente",
                _fmt_int_ptbr(n_conf),
                help="Devolução ou retorno ao vendedor: conferir recebimento ou situação física do pacote.",
            )
        with k2:
            st.metric(
                "Revisar no financeiro",
                _fmt_int_ptbr(n_rev),
                help="Há movimento de valores para conferir no Mercado Pago ou nas liberações.",
            )
        with k3:
            st.metric(
                "Cobrar Mercado Livre",
                _fmt_int_ptbr(n_cob),
                help="Problema de envio ou falha atribuível ao fluxo do Mercado Livre.",
            )
        with k4:
            st.metric(
                "Dinheiro liberado para você",
                _fmt_int_ptbr(n_pag),
                help="Caso encerrado ou valor liberado a favor da loja.",
            )
        with k5:
            st.metric(
                "Reembolso ao comprador",
                _fmt_int_ptbr(n_reemb),
                help="Estorno ou devolução de dinheiro ao comprador.",
            )
        k6, k7, _sp = st.columns([1, 1, 2])
        with k6:
            st.metric("Total na fila", _fmt_int_ptbr(n), help="Total de linhas neste ficheiro materializado.")
        with k7:
            st.metric(
                "Com registro em liberações",
                _fmt_int_ptbr(n_vinc),
                help="Pelo menos um lançamento nas liberações do Mercado Livre ligado a esta venda.",
            )

    st.markdown(
        '<p class="fdl-frete-section-title">Filtros</p>'
        '<p class="fdl-frete-section-note">A empresa é a selecionada no painel (igual repasse e frete).</p>',
        unsafe_allow_html=True,
    )

    _prio3 = (
        "Conferência física pendente",
        "Revisar financeiramente",
        "Cobrar plataforma",
    )
    _recorte_devolucoes: list[tuple[str, str]] = [
        ("Todos", "__all__"),
        ("Só os 3 prioritários (conferência / financeiro / ML)", "__prio3__"),
        ("Conferência física pendente", "Conferência física pendente"),
        ("Revisar no financeiro", "Revisar financeiramente"),
        ("Cobrar Mercado Livre", "Cobrar plataforma"),
        ("Dinheiro liberado para você", "Pagamento liberado ao vendedor"),
        ("Reembolso ao comprador", "Reembolso ao comprador"),
        ("Personalizado (escolha em Situação)", "__custom__"),
    ]
    _recorte_labels = [t[0] for t in _recorte_devolucoes]

    recorte_i = st.selectbox(
        "Recorte rápido",
        options=range(len(_recorte_labels)),
        format_func=lambda i: _recorte_labels[i],
        index=0,
        key="fdl_devolucoes_recorte_rapido",
    )
    recorte_val = _recorte_devolucoes[int(recorte_i)][1]

    busca = st.text_input(
        "Busca por n.º da venda (contém)",
        "",
        key="fdl_devolucoes_busca_venda",
        placeholder="Ex.: 20000158…",
    )

    opts_si_full = (
        sorted({str(x) for x in d["status_interno"].dropna().unique()})
        if "status_interno" in d.columns
        else []
    )
    opts_si_full = [x for x in opts_si_full if str(x).strip()]

    opts_ac = (
        sorted({str(x) for x in d["acao_sugerida"].dropna().unique()})
        if "acao_sugerida" in d.columns
        else []
    )
    opts_ac = [x for x in opts_ac if str(x).strip()]

    opts_mot = (
        sorted({str(x) for x in d["candidato_motivo"].dropna().unique()})
        if "candidato_motivo" in d.columns
        else []
    )
    opts_mot = [x for x in opts_mot if str(x).strip()]

    f1, f2 = st.columns(2)
    with f1:
        si = st.multiselect(
            "Situação",
            options=opts_si_full,
            default=[],
            key="fdl_devolucoes_status_interno",
            help="Vazio = não filtra. Use com «Recorte rápido» = Personalizado.",
        )
    with f2:
        ac = st.multiselect(
            "O que fazer",
            options=opts_ac,
            default=[],
            key="fdl_devolucoes_acao_sugerida",
            help="Filtra pela ação sugerida para a operação.",
        )

    filt_lib = st.selectbox(
        "Só com registro em liberações",
        options=["Todos", "Com registro", "Sem registro"],
        index=0,
        key="fdl_devolucoes_filtro_lib",
        help="«Com registro»: pelo menos um evento nas liberações ML para esta venda.",
    )

    mot: list[str] = []
    if len(opts_mot) > 1:
        with st.expander("Filtros avançados", expanded=False):
            mot = st.multiselect(
                "Motivo da fila",
                options=opts_mot,
                default=[],
                key="fdl_devolucoes_motivo",
                help="Por que a venda entrou na fila materializada (materializado).",
            )

    if busca.strip() and "N° de venda" in d.columns:
        d = d[d["N° de venda"].astype(str).str.contains(busca.strip(), case=False, na=False)]

    if recorte_val == "__prio3__" and "status_interno" in d.columns:
        d = d[d["status_interno"].astype(str).str.strip().isin(_prio3)]
    elif recorte_val not in ("__all__", "__custom__") and "status_interno" in d.columns:
        d = d[d["status_interno"].astype(str).str.strip() == recorte_val]

    if "status_interno" in d.columns and recorte_val in ("__all__", "__custom__") and si:
        d = d[d["status_interno"].astype(str).isin(si)]

    if ac and "acao_sugerida" in d.columns:
        d = d[d["acao_sugerida"].astype(str).isin(ac)]

    if filt_lib == "Com registro" and "lib_n_eventos" in d.columns:
        d = d[pd.to_numeric(d["lib_n_eventos"], errors="coerce").fillna(0) > 0]
    elif filt_lib == "Sem registro" and "lib_n_eventos" in d.columns:
        d = d[pd.to_numeric(d["lib_n_eventos"], errors="coerce").fillna(0) <= 0]

    if mot and "candidato_motivo" in d.columns:
        d = d[d["candidato_motivo"].astype(str).isin(mot)]

    d = d.copy()
    _cols_det = _devolucoes_cols_tabela_detalhe(d, list(_DEVOLUCOES_SLIM_SOURCE_COLS))
    d_fila = _devolucoes_build_fila_operacional_view(d)

    _cfg_fila: dict[str, object] = {
        "N.º da venda": st.column_config.TextColumn("N.º da venda", width="small"),
        "Data": st.column_config.TextColumn("Data", width="small"),
        "Situação": st.column_config.TextColumn("Situação", width="medium"),
        "O que fazer": st.column_config.TextColumn(
            "O que fazer",
            width="medium",
            help="Próximo passo sugerido para a loja.",
        ),
        "Mensagem do Mercado Livre": st.column_config.TextColumn(
            "Mensagem do Mercado Livre",
            width="large",
            help="Texto do export de vendas (não duplicamos «Descrição do status» quando é o mesmo conteúdo).",
        ),
        "Reembolso": st.column_config.TextColumn("Reembolso", width="small"),
        "Liberações": st.column_config.TextColumn("Liberações", width="small"),
    }

    _h_main = min(620, 140 + min(len(d_fila), 20) * 34)
    st.markdown(
        '<p class="fdl-frete-section-title">Fila</p>'
        '<p class="fdl-frete-section-note">Uma linha por venda: situação, ação e mensagem do ML. Detalhes técnicos ficam abaixo, fechados por defeito.</p>',
        unsafe_allow_html=True,
    )
    st.dataframe(
        d_fila,
        use_container_width=True,
        hide_index=True,
        height=_h_main,
        column_config=_cfg_fila,
    )

    if _cols_det:
        _h_det = min(420, 120 + min(len(d), 10) * 26)
        with st.expander("Detalhes técnicos e export completo", expanded=False):
            st.caption(
                f"**{_fmt_int_ptbr(len(_cols_det))}** colunas: valores, vínculo fino com liberações, Jaci, ficheiro de origem, restante do export ML e colunas duplicadas («.1»). "
                "Para auditoria e conferência profunda — não necessário no dia a dia."
            )

            def _devolucoes_column_config_tecnico(cols: list[str]) -> dict[str, object] | None:
                cfg: dict[str, object] = {}
                if "reembolso_valor_inferido" in cols:
                    cfg["reembolso_valor_inferido"] = st.column_config.NumberColumn(
                        "Reembolso inferido (BRL)",
                        format="%.2f",
                    )
                if "Total (BRL)" in cols:
                    cfg["Total (BRL)"] = st.column_config.NumberColumn("Total (BRL)", format="%.2f")
                if "Cancelamentos e reembolsos (BRL)" in cols:
                    cfg["Cancelamentos e reembolsos (BRL)"] = st.column_config.NumberColumn(
                        "Cancelam. / reemb. (BRL)",
                        format="%.2f",
                    )
                if "lib_n_eventos" in cols:
                    cfg["lib_n_eventos"] = st.column_config.NumberColumn("Eventos liberações", format="%d")
                if "lib_ultima_data_pagamento" in cols:
                    cfg["lib_ultima_data_pagamento"] = st.column_config.TextColumn("Últ. pag. liberações", width="small")
                if "jaci_cep_15155038" in cols:
                    cfg["jaci_cep_15155038"] = st.column_config.CheckboxColumn("CEP Jaci")
                if "candidato_motivo" in cols:
                    cfg["candidato_motivo"] = st.column_config.TextColumn("Motivo fila", width="small")
                if "empresa" in cols:
                    cfg["empresa"] = st.column_config.TextColumn("Empresa", width="small")
                return cfg or None

            st.dataframe(
                d[_cols_det],
                use_container_width=True,
                hide_index=True,
                height=_h_det,
                column_config=_devolucoes_column_config_tecnico([c for c in _cols_det if c in d.columns]),
            )

    st.caption(
        f"**{_fmt_int_ptbr(len(d_fila))}** linhas após filtros · Vista principal só com o essencial para operar."
    )

    if _is_admin:
        with st.expander("Admin — caminho do materializado", expanded=False):
            st.caption(str(load_info.get("devolucoes_path_resolved", "")))


def _painel_frete_emergencial(
    org_id: str, df_frete: pd.DataFrame, load_info: dict[str, object], ts_proc: str
) -> None:
    """
    Apresentação do painel Frete. O carregamento é feito em _load_frete_data (ponto único).
    """
    _is_admin = _is_admin_mode()

    if load_info.get("frete_fontes_error"):
        st.error("Erro ao localizar fontes de Frete / vendas ML.")
        if _is_admin:
            st.code(str(load_info.get("frete_fontes_error")), language="text")
        return

    if load_info.get("frete_no_vendas_source"):
        if _is_admin:
            st.warning(
                "Sem fonte de vendas ML para Frete. Defina **FDL_FRETE_VENDAS_URL** nos Secrets (Cloud) "
                "ou coloque ficheiros .xlsx/.csv em **Vendas - Mercado Livre** sob **FDL_BASE_DIR**."
            )
            st.caption(str(BASE_DIR))
        else:
            st.warning(
                "Não foi possível localizar o ficheiro de vendas do Mercado Livre para esta conciliação. "
                "Contacte o administrador."
            )
        return

    if load_info.get("frete_placeholder_vendas_url"):
        st.error(
            "O URL de **FDL_FRETE_VENDAS_URL** parece um **placeholder**. "
            "Nos Secrets, substitua por o **link completo** do Excel (Partilhar → copiar ligação do ficheiro)."
        )
        return

    if load_info.get("frete_ml_validation_failed") or load_info.get("frete_loader_error"):
        st.error("O ficheiro não parece ser o **export de vendas ML** (detalhe envios).")
        st.warning(
            "Confirme que **FDL_FRETE_VENDAS_URL** aponta para o mesmo tipo de ficheiro que está em "
            "**Vendas - Mercado Livre** no OneDrive (não use a planilha de **Repasse**)."
        )
        if _is_admin:
            st.code(str(load_info.get("frete_loader_error", "")), language="text")
        return

    if load_info.get("frete_vendas_from_url"):
        if _is_admin:
            st.success("**FDL_FRETE_VENDAS_URL** detetado — a base foi descarregada a partir do SharePoint.")
        else:
            st.success("Dados de vendas carregados a partir da nuvem.")
    elif load_info.get("frete_fonte_local_path") and _is_admin:
        st.caption(f"Fonte local: `{load_info['frete_fonte_local_path']}`")

    if _is_admin and load_info.get("frete_mat_from_repasse_sibling"):
        st.caption(
            "Frete: **dataset_frete_app.csv** ao lado do repasse (`.../repasse/current/` → "
            "`.../frete/current/`), via **FDL_REPASSE_MATERIALIZED_PATH** ou **FDL_PRECOMPUTED_PATH**."
        )

    if load_info.get("frete_consume") == "materialized" and _is_admin:
        for _line in (load_info.get("debug_logs") or []):
            if _line:
                st.caption(f"[frete-debug] {_line}")

    meta = _frete_meta_for_render(load_info)
    if _frete_debug_ui_enabled() and _is_admin:
        try:
            with st.expander("Diagnóstico Frete (opt-in: FDL_DEBUG_FRETE_UI=1)", expanded=True):
                st.write("### Etapa 1/4 — Dataset (carregado antes desta página)")
                st.write("**Estado:**", "carregado" if not df_frete.empty else "DataFrame vazio")
                st.write("**Linhas × colunas:**", df_frete.shape)
                _cols = list(df_frete.columns)
                st.write("**Nº de colunas:**", len(_cols))
                st.write("**Colunas (início):**", _cols[:35])
                if len(_cols) > 35:
                    st.caption(f"… e mais {len(_cols) - 35} colunas.")
                _req_ui = {
                    FRETE_UI_N_VENDA: "N.º venda (filtros / Recebido?)",
                    FRETE_ML_COL: "Frete cobrado (ML)",
                    "Estado": "Estado da venda",
                    FRETE_UI_DIFERENCA: "KPIs e situação",
                }
                st.write("**Colunas esperadas pela UI:**")
                for _k, _desc in _req_ui.items():
                    st.write(f"- `{_k}`: {'OK' if _k in _cols else 'AUSENTE'} — {_desc}")
                st.write("**Fonte (consume):**", load_info.get("frete_consume"))
                st.write("**frete_arquivo:**", load_info.get("frete_arquivo"))
                st.write("**Alvo materializado (path/URL resolvido):**", load_info.get("frete_materialized_target"))
                st.write("**Rótulo vendas (UI):**", meta.get("vendas_arquivo", load_info.get("vendas_arquivo")))
                st.write(
                    "**Próximo passo:**",
                    "etapas 2–4 em _render_frete_operacional_ui (filtros → KPIs → export/detalle)",
                )
        except Exception as exc:
            st.error("Diagnóstico Frete (expander) falhou — o restante da página tenta continuar.")
            st.caption(str(exc))
    if df_frete.empty:
        if load_info.get("frete_consume") == "materialized":
            if _is_admin:
                st.info(
                    "Não há linhas no **frete materializado** (ficheiro com 0 linhas). "
                    "Regere `dataset_frete_app.csv` com `processing/materialize_cliente_5.ps1` a partir das pastas "
                    "Esquilo/Wood e publique os CSV atualizados no repositório (ou aloje o ficheiro e use URL nos Secrets)."
                )
            else:
                st.info("Sem registos de frete para este recorte. Contacte o suporte se precisar de ajuda.")
        elif _is_admin:
            st.info(
                "Não há linhas de frete para exibir. Verifique o export de vendas ML (modo live) "
                "ou o ficheiro consolidado (modo materializado)."
            )
        else:
            st.info("Sem registos de frete para exibir. Tente mais tarde ou contacte o suporte.")
        return
    try:
        _render_frete_operacional_ui(org_id, df_frete, meta, ts_proc, load_info)
    except Exception as exc:
        st.error("Erro ao renderizar a Conciliação de Frete (detalhe abaixo).")
        st.exception(exc)


def _build_pdf_bytes(df: pd.DataFrame) -> bytes:
    buff = BytesIO()
    c = canvas.Canvas(buff, pagesize=A4)
    w, h = A4
    y = h - 36
    c.setFont("Helvetica-Bold", 11)
    c.drawString(30, y, "Conciliação Operacional - Exportação")
    y -= 20
    c.setFont("Helvetica", 8)
    cols = list(df.columns)
    c.drawString(30, y, " | ".join(cols))
    y -= 12
    for row in df.astype(str).itertuples(index=False):
        line = " | ".join(list(row))
        if len(line) > 170:
            line = line[:167] + "..."
        c.drawString(30, y, line)
        y -= 11
        if y < 40:
            c.showPage()
            y = h - 36
            c.setFont("Helvetica", 8)
    c.save()
    return buff.getvalue()


def _pick_col_by_tokens(columns: list[str], tokens: list[str]) -> str:
    for c in columns:
        n = unicodedata.normalize("NFKD", str(c)).encode("ascii", "ignore").decode().lower()
        if all(t in n for t in tokens):
            return c
    return ""


def _resolve_col_data_emissao(columns: list[str]) -> str:
    """Prioriza o nome literal da tabela final; fallback só se encoding divergir."""
    if "Data de emissão" in columns:
        return "Data de emissão"
    return _pick_col_by_tokens(columns, ["data", "emiss"])


def _parse_data_emissao_final(series: pd.Series) -> pd.Series:
    """
    Tabela final persiste emissão como YYYY-MM-DD (integracao_notas + etapa4b).
    Parse fixo — evita ambiguidade de dayfirst na UI.
    """
    s = series.fillna("").astype(str).str.strip()
    s = s.str.replace("NaT", "", regex=False).str.replace("None", "", regex=False)
    s = s.mask(s.str.lower().isin({"none", "nan", "nat", "<na>", "null"}), "")
    return pd.to_datetime(s, format="%Y-%m-%d", errors="coerce")


def _parse_data_pagamento_final(series: pd.Series) -> pd.Series:
    """Tabela final: parse robusto (ISO/mixed, com/sem timezone)."""
    s = series.fillna("").astype(str).str.strip()
    s = s.str.replace("NaT", "", regex=False).str.replace("None", "", regex=False)
    s = s.mask(s.str.lower().isin({"none", "nan", "nat", "<na>", "null"}), "")
    t = pd.to_datetime(s, errors="coerce", format="mixed", utc=True)
    try:
        t = t.dt.tz_convert(_BR_TZ).dt.tz_localize(None)
    except Exception:  # noqa: BLE001
        t = pd.to_datetime(s, errors="coerce", format="mixed")
    return t


def _first_series(df: pd.DataFrame, col: str) -> pd.Series:
    """
    Retorna a primeira série quando há colunas duplicadas com o mesmo nome.
    """
    obj = df[col]
    if isinstance(obj, pd.DataFrame):
        return obj.iloc[:, 0]
    return obj


def _drop_duplicate_columns_keep_first(df: pd.DataFrame) -> pd.DataFrame:
    """Remove colunas com nome duplicado, preservando a primeira ocorrência."""
    if df.empty:
        return df
    return df.loc[:, ~df.columns.duplicated()].copy()


def _effective_periodo_repasse_series(df: pd.DataFrame) -> tuple[pd.Series, str]:
    """
    Datas para filtro e ordenação do repasse: usa ``Data período repasse`` quando existe no dataset
    materializado; caso contrário replica o critério da etapa4b (pagamento quando há data; senão emissão)
    apenas a partir das colunas já carregadas — sem ler fontes brutas nem refazer o pipeline.
    """
    if "Data período repasse" in df.columns:
        s = _parse_data_pagamento_final(_first_series(df, "Data período repasse"))
        return s, "Data período repasse"
    pay = _parse_data_pagamento_final(_first_series(df, "Data de pagamento"))
    col_em = _resolve_col_data_emissao(list(df.columns))
    emi = _parse_data_emissao_final(df[col_em]) if col_em else pd.Series(pd.NaT, index=df.index)
    pay_dt = pd.to_datetime(pay, errors="coerce")
    emi_dt = pd.to_datetime(emi, errors="coerce")
    out = pay_dt.copy()
    need_emi = pay_dt.isna() & emi_dt.notna()
    out = out.where(~need_emi, emi_dt)
    return (
        out,
        "Data de pagamento / emissão (regenere o repasse para gravar **Data período repasse** no Parquet)",
    )


def _repasse_ui_periodo_series(df: pd.DataFrame) -> tuple[pd.Series, str]:
    """PR5: Parquet → só ``Data período repasse``; legado → ``_effective_periodo_repasse_series``."""
    if _repasse_use_parquet():
        return repasse_ui_periodo_series_parquet(
            df,
            parse_data_periodo_repasse_column=lambda d: _parse_data_pagamento_final(
                _first_series(d, COL_DATA_PERIODO_REPASSE)
            ),
        )
    return _effective_periodo_repasse_series(df)


def _painel_conciliacao_fragment(base: pd.DataFrame, ts_proc: str) -> None:
    """
    Filtros + validação de ações + fila/tabela de repasse.

    Não usar @st.fragment aqui: ao mudar para «Frete», o fragment deixava de ser invocado e o Streamlit
    podia mostrar ecrã em branco (desincronização da árvore de widgets entre vistas).
    """
    if base.empty or "Data de pagamento" not in base.columns:
        st.warning("Sem dados de repasse para esta vista. Contacte o suporte se o problema continuar.")
        return

    _rpq = _repasse_use_parquet()
    acao_col = repasse_ui_acao_column(use_parquet=_rpq)

    _fdl_cp_inject_panel_styles()
    _fdl_repasse_inject_panel_styles()
    _fdl_ui_gap_tight()

    dp_series_full, _periodo_src_label = _repasse_ui_periodo_series(base)

    # Chaves de widget por org: sem isto, ao mudar de empresa o estado do Streamlit podia manter
    # limites/valores de outra org e parecer que o calendário «começa» na data errada.
    _rep_wk = _frete_org_widget_suffix(_active_org.org_id)

    with st.container(border=True):
        st.markdown('<p class="fdl-repasse-filtros-h">Filtros</p>', unsafe_allow_html=True)
        if _rpq:
            st.markdown(
                '<p class="fdl-repasse-caption">Recorte pelo <strong>período de repasse</strong> usando apenas a coluna '
                "materializada <strong>Data período repasse</strong> (valores vazios permanecem fora do intervalo de datas). "
                "Plataforma, ação, situação e busca.</p>",
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                '<p class="fdl-repasse-caption">Recorte pelo <strong>período de repasse</strong>: coluna materializada '
                "<strong>Data período repasse</strong> quando existe no Parquet; em bases antigas usa pagamento e, "
                "se vazio, <strong>data de emissão</strong> (mesmo critério da etapa4b, só sobre linhas já carregadas). "
                "Plataforma, ação, situação e busca.</p>",
                unsafe_allow_html=True,
            )
        _d_min, _d_max, has_dp_base = _series_datetime_bounds_dates(dp_series_full)
        today_rep = datetime.now(_BR_TZ).date()
        picker_min = min(_d_min, today_rep - timedelta(days=3 * 365))
        picker_max = max(_d_max, today_rep)
        if picker_max < picker_min:
            picker_min, picker_max = picker_max, picker_min
        # O Streamlit mantém data_input no session_state: após rematerializar com datas mais recentes,
        # o «Fim» podia ficar preso (ex.: 23/03) embora a base já vá até 31/03. Repor início/fim quando
        # os limites reais da coluna mudarem.
        _bounds_sig_key = f"op_repasse_dp_bounds_sig_{_rep_wk}"
        _bounds_sig = (_d_min.isoformat(), _d_max.isoformat())
        if st.session_state.get(_bounds_sig_key) != _bounds_sig:
            st.session_state[_bounds_sig_key] = _bounds_sig
            st.session_state[f"op_repasse_d_pag_ini_{_rep_wk}"] = _d_min
            st.session_state[f"op_repasse_d_pag_fim_{_rep_wk}"] = _d_max
        plats = (
            nf_grain_plataforma_ui_options(base["Plataforma"])
            if "Plataforma" in base.columns
            else []
        )
        r2 = st.columns((1, 1))
        with r2[0]:
            data_pag_ini = st.date_input(
                "Pagamento — início",
                value=_d_min,
                min_value=picker_min,
                max_value=picker_max,
                format="DD/MM/YYYY",
                key=f"op_repasse_d_pag_ini_{_rep_wk}",
            )
        with r2[1]:
            data_pag_fim = st.date_input(
                "Pagamento — fim",
                value=_d_max,
                min_value=picker_min,
                max_value=picker_max,
                format="DD/MM/YYYY",
                key=f"op_repasse_d_pag_fim_{_rep_wk}",
            )
        data_pag_ini = _safe_streamlit_date(data_pag_ini, _d_min)
        data_pag_fim = _safe_streamlit_date(data_pag_fim, _d_max)
        r1 = st.columns((1, 1, 1))
        with r1[0]:
            sel_plat = _multiselect_stable(
                f"op_ms_plat_{_rep_wk}", "Plataforma", plats, compact_label=True
            )
        # Recalcula opções dependentes da plataforma selecionada para evitar filtros «presos»
        # de outra plataforma (ex.: seleção anterior de ML zerando Shopee).
        base_opts = base.copy()
        if "Plataforma" in base_opts.columns and sel_plat:
            base_opts = base_opts[base_opts["Plataforma"].isin(sel_plat)].copy()
        acoes = sorted(
            [
                x
                for x in base_opts[acao_col].dropna().unique().tolist()
                if str(x).strip()
            ]
        )
        sit = sorted(
            [x for x in base_opts["Situação"].dropna().unique().tolist() if str(x).strip()]
        )
        with r1[1]:
            sel_acao = _multiselect_stable(
                f"op_ms_acao_{_rep_wk}", "Ação sugerida", acoes, compact_label=True
            )
        with r1[2]:
            sel_sit = _multiselect_stable(f"op_ms_sit_{_rep_wk}", "Situação", sit, compact_label=True)
        busca = st.text_input(
            "Buscar venda, pedido ou NF",
            placeholder="Venda, pedido, NF…",
            key=f"op_repasse_busca_txt_{_rep_wk}",
        ).strip().lower()
        if not has_dp_base:
            st.info(
                "Sem datas na coluna de período da tabela: o filtro por datas não aplica (todas as linhas aparecem). "
                "Regenere o repasse com o pipeline atual para obter **Data período repasse**."
            )
        _clr_l, _clr_r = st.columns((1.55, 1))
        with _clr_l:
            pass
        with _clr_r:
            if st.button(
                "Limpar filtros desta vista",
                key=f"op_repasse_clear_{_rep_wk}",
                use_container_width=True,
                help="Repor período à base, limpar plataforma/ação/situação e a busca.",
            ):
                for _k in (
                    _bounds_sig_key,
                    f"op_repasse_d_pag_ini_{_rep_wk}",
                    f"op_repasse_d_pag_fim_{_rep_wk}",
                    f"op_ms_plat_{_rep_wk}",
                    f"op_ms_acao_{_rep_wk}",
                    f"op_ms_sit_{_rep_wk}",
                    f"op_repasse_busca_txt_{_rep_wk}",
                ):
                    st.session_state.pop(_k, None)
                st.rerun()

    _fdl_ui_gap_tight()

    if data_pag_fim < data_pag_ini:
        st.warning("A data final não pode ser anterior à data inicial. Ajuste o período.")
        data_pag_fim = data_pag_ini
    
    tabela = base.copy()
    if "Plataforma" in tabela.columns and sel_plat:
        tabela = tabela[tabela["Plataforma"].isin(sel_plat)]
    if sel_acao:
        tabela = tabela[tabela[acao_col].isin(sel_acao)]
    if sel_sit:
        tabela = tabela[tabela["Situação"].isin(sel_sit)]
    if busca:
        m_busca = (
            tabela["N° de venda"].fillna("").astype(str).str.lower().str.contains(busca, regex=False)
            | tabela["ID do pedido"].fillna("").astype(str).str.lower().str.contains(busca, regex=False)
            | tabela["Número da nota"].fillna("").astype(str).str.lower().str.contains(busca, regex=False)
        )
        tabela = tabela[m_busca]
    
    _dp_filt = _repasse_ui_periodo_series(tabela)[0]
    if _dp_filt.notna().any():
        _dd = _dp_filt.dt.normalize()
        _ini_ts = pd.Timestamp(data_pag_ini)
        _fim_ts = pd.Timestamp(data_pag_fim) + pd.Timedelta(days=1)
        m_data = _dp_filt.notna() & (_dd >= _ini_ts) & (_dd < _fim_ts)
        tabela = tabela.loc[m_data].copy()
    if repasse_ui_apply_pipeline_exclusao_na_ui(use_parquet=_rpq):
        tabela = _excluir_linhas_fora_conciliacao(tabela)
    if repasse_ui_apply_filtro_somente_linhas_com_data_pagamento(use_parquet=_rpq):
        if "Data de pagamento" in tabela.columns:
            _dp_somente_tabela = _parse_data_pagamento_final(_first_series(tabela, "Data de pagamento"))
            tabela = tabela.loc[_dp_somente_tabela.notna()].copy()

    if "Plataforma" in base.columns:
        n_plat = len(plats)
        # Multiselect vazio = não filtra por plataforma (mostra todas) — não confundir com «nenhuma linha».
        if not n_plat:
            plataforma_label = "—"
        elif len(sel_plat) == 0 or (n_plat and len(sel_plat) == n_plat):
            plataforma_label = "Todas"
        else:
            plataforma_label = ", ".join(sel_plat[:2]) + ("..." if len(sel_plat) > 2 else "")
    else:
        plataforma_label = "Mercado Livre"
    
    _pag_caption = (
        f"Período: **{data_pag_ini.strftime('%d/%m/%Y')}** a **{data_pag_fim.strftime('%d/%m/%Y')}** "
        f"({_periodo_src_label})"
    )
    if not has_dp_base:
        _pag_caption += " — **filtro por data inativo** (sem datas na coluna de período)"
    _caption_dp_extra = (
        ""
        if _rpq
        else " · Tabela e resumo: só linhas com **data de pagamento**."
    )
    st.caption(
        f"Plataforma **{plataforma_label}** · Atualizado **{ts_proc}** · {_pag_caption}{_caption_dp_extra}"
    )

    _fdl_ui_gap_tight()

    # Tipos numéricos para a base já filtrada (tabela e totais nas colunas)
    tabela["Valor da nota"] = pd.to_numeric(tabela["Valor da nota"], errors="coerce").fillna(0.0)
    tabela["Total BRL"] = pd.to_numeric(tabela.get("Total BRL"), errors="coerce")
    tabela["Valor a receber"] = pd.to_numeric(tabela.get("Valor a receber"), errors="coerce")
    tabela["Valor pago"] = pd.to_numeric(tabela.get("Valor pago"), errors="coerce")
    tabela["Diferença"] = pd.to_numeric(tabela.get("Diferença"), errors="coerce")

    st.markdown(
        '<p class="fdl-repasse-section-title">Resumo por ação</p>'
        + (
            '<p class="fdl-repasse-section-note">Recorte atual (dataset carregado + filtros de sessão). '
            "Tamanho da fila e distribuição por ação sugerida.</p>"
            if _rpq
            else '<p class="fdl-repasse-section-note">Apenas linhas com data de pagamento. '
            "Tamanho da fila e distribuição por ação sugerida no recorte atual.</p>"
        ),
        unsafe_allow_html=True,
    )
    if _repasse_vendas_liberacoes_only():
        acoes_validacao = ["Baixado", "Analisar diferença", "Verificar recebimento"]
    else:
        acoes_validacao = ["Ok", "Baixado" if _repasse_sem_bling() else "Baixar no Bling", "Analisar diferença"]
    contagens_acao = {a: int(tabela[acao_col].eq(a).sum()) for a in acoes_validacao}
    contagens_acao["Zerado"] = int(tabela[acao_col].eq("Revisar venda zerada").sum())
    _render_repasse_resumo_por_acao_kpis(contagens_acao, n_linhas=len(tabela))

    _fdl_ui_gap_section()
    st.divider()

    # Tabela operacional — Data de emissão: mesma coluna da tabela final, parse ISO (sem dayfirst).
    col_data_emissao = _resolve_col_data_emissao(list(tabela.columns))
    if _repasse_vendas_liberacoes_only():
        exibir_cols = [
            "N° de venda",
            "Total BRL",
            "Valor a receber",
            "Valor pago",
            "Diferença",
            acao_col,
            "Plataforma",
        ]
    else:
        exibir_cols = [
            "N° de venda",
            "ID do pedido",
            "Total BRL",
            "Número da nota",
            "Valor da nota",
            "Valor a receber",
            "Diferença",
            "Situação",
            acao_col,
        ]
    if "Data de pagamento" in tabela.columns:
        exibir_cols.append("Data de pagamento")
    if "Valor pago" in tabela.columns:
        exibir_cols.append("Valor pago")
    
    exibir_cols = [c for c in exibir_cols if c in tabela.columns]
    if not exibir_cols:
        st.warning("Não foi possível apresentar a tabela com o recorte atual.")
        tabela_exibir = pd.DataFrame()
    else:
        tabela_exibir = tabela[exibir_cols].copy()
        tabela_exibir = _drop_duplicate_columns_keep_first(tabela_exibir)
        if "Valor da nota" in tabela_exibir.columns:
            tabela_exibir["Valor da nota"] = pd.to_numeric(
                _first_series(tabela_exibir, "Valor da nota"), errors="coerce"
            )
        else:
            tabela_exibir["Valor da nota"] = 0.0
        if "Valor a receber" in tabela_exibir.columns:
            tabela_exibir["Valor a receber"] = pd.to_numeric(
                _first_series(tabela_exibir, "Valor a receber"), errors="coerce"
            )
        else:
            tabela_exibir["Valor a receber"] = 0.0
        if "Valor pago" in tabela_exibir.columns:
            tabela_exibir["Valor pago"] = pd.to_numeric(
                _first_series(tabela_exibir, "Valor pago"), errors="coerce"
            )
        else:
            tabela_exibir["Valor pago"] = 0.0
        if "Diferença" in tabela_exibir.columns:
            tabela_exibir["Diferença"] = pd.to_numeric(
                _first_series(tabela_exibir, "Diferença"), errors="coerce"
            )
        else:
            tabela_exibir["Diferença"] = 0.0
        if col_data_emissao:
            tabela_exibir["Data de emissão"] = _parse_data_emissao_final(
                tabela.loc[tabela_exibir.index, col_data_emissao]
            )
        else:
            tabela_exibir["Data de emissão"] = pd.NaT
        tabela_exibir["Data de pagamento"] = _parse_data_pagamento_final(
            tabela.loc[tabela_exibir.index, "Data de pagamento"]
            if "Data de pagamento" in tabela.columns
            else pd.Series("", index=tabela_exibir.index)
        )
        _periodo_tab, _ = _repasse_ui_periodo_series(tabela)
        tabela_exibir["Período repasse"] = _periodo_tab.loc[tabela_exibir.index]
        tabela_exibir["Valor da nota"] = tabela_exibir["Valor da nota"].fillna(0.0)
        tabela_exibir["Valor a receber"] = tabela_exibir["Valor a receber"].fillna(0.0)
        tabela_exibir["Valor pago"] = tabela_exibir["Valor pago"].fillna(0.0)
        tabela_exibir["Diferença"] = tabela_exibir["Diferença"].fillna(0.0)
        _rename_map: dict[str, str] = {
            "N° de venda": "Número da venda",
            "ID do pedido": "Número do pedido",
        }
        if COL_ACAO_LEGACY_UI in tabela_exibir.columns:
            _rename_map[COL_ACAO_LEGACY_UI] = "Ação sugerida"
        tabela_exibir = tabela_exibir.rename(columns=_rename_map)
        tabela_exibir = _drop_duplicate_columns_keep_first(tabela_exibir)
        tabela_exibir = tabela_exibir.drop(columns=["Total BRL"], errors="ignore")
        _ordem_final = [
            "Número da venda",
            "Número do pedido",
            "Número da nota",
            "Data de emissão",
            "Período repasse",
            "Data de pagamento",
            "Valor da nota",
            "Valor a receber",
            "Valor pago",
            "Diferença",
            "Situação",
            "Ação sugerida",
        ]
        tabela_exibir = tabela_exibir[[c for c in _ordem_final if c in tabela_exibir.columns]]
        # Ordenação alinhada ao filtro: «Período repasse» evita esconder Shopee 2026 só com emissão
        # quando «Data de pagamento» ainda é 2025 ou vazia.
        if not tabela_exibir.empty:
            _sort_key = (
                "Período repasse"
                if "Período repasse" in tabela_exibir.columns
                else (
                    "Data de emissão"
                    if "Data de emissão" in tabela_exibir.columns
                    else "Data de pagamento"
                )
            )
            if _sort_key in tabela_exibir.columns:
                tabela_exibir = tabela_exibir.sort_values(
                    by=_sort_key, ascending=False, na_position="last"
                ).reset_index(drop=True)
        for _dc in ("Data de emissão", "Período repasse", "Data de pagamento"):
            if _dc in tabela_exibir.columns:
                tabela_exibir[_dc] = pd.to_datetime(tabela_exibir[_dc], errors="coerce")
        for _ref in ("Número da venda", "Número do pedido", "Número da nota"):
            if _ref in tabela_exibir.columns:
                # object + str puro: o Glide Data Grid formata int/float com separadores de milhar
                _txt = _col_referencia_como_texto(tabela_exibir[_ref])
                _obj: list[str | None] = []
                for x in _txt:
                    if pd.isna(x) or x == "":
                        _obj.append(None)
                    else:
                        _obj.append(str(x))
                tabela_exibir[_ref] = pd.Series(_obj, dtype=object, index=tabela_exibir.index)
    
    tabela_exibir_grid, _repasse_n_recorte_total, _repasse_grid_truncated = repasse_ui_grid_display_slice(
        tabela_exibir, cap=REPASSE_UI_GRID_ROW_CAP
    )
    if _repasse_grid_truncated:
        st.info(
            f"Mostrando apenas as primeiras **{REPASSE_UI_GRID_ROW_CAP:,}** linhas de "
            f"**{_repasse_n_recorte_total:,}** neste recorte. **Afine os filtros** (período, plataforma, "
            "ação, situação ou busca) para trabalhar com menos linhas na grelha."
        )

    if _fdl_safe_mode():
        st.warning(
            "**Modo seguro (FDL_SAFE_MODE=1)** — sem fila HTML, sem exports Excel/PDF, "
            "sem `column_config` na tabela (apenas `st.dataframe` simples)."
        )
        st.metric("Linhas (filtro atual)", _repasse_n_recorte_total)
        _grid_safe = _dataframe_conciliacao_somente_grid(tabela_exibir_grid)
        st.dataframe(
            _grid_safe,
            use_container_width=True,
            height=min(560, 140 + max(18 * min(len(tabela_exibir_grid), 80), 120)),
        )
        st.caption("Desative FDL_SAFE_MODE para voltar à UI completa.")
        return

    st.markdown(
        '<p class="fdl-repasse-section-title">Fila operacional</p>'
        '<p class="fdl-repasse-section-note">Tratativa no recorte · exporte Excel/PDF quando o volume permitir.</p>',
        unsafe_allow_html=True,
    )
    _fdl_ui_gap_section()

    # Guardrail de estabilidade: evita trabalho pesado em cada troca de filtro
    # para não arriscar ecrã em branco por timeout/memória no Streamlit Cloud.
    _max_rows_heavy_export = 3000
    excel_bytes: bytes | None = None
    pdf_bytes: bytes | None = None
    heavy_exports_enabled = len(tabela_exibir) <= _max_rows_heavy_export
    if heavy_exports_enabled:
        try:
            tabela_excel = tabela_exibir.copy()
            excel_buf = BytesIO()
            with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
                tabela_excel.to_excel(writer, index=False, sheet_name="Conciliação")
                ws = writer.sheets["Conciliação"]
                header_row = [cell.value for cell in ws[1]]
                for c_data in ("Data de emissão", "Período repasse", "Data de pagamento"):
                    if c_data in header_row:
                        col_idx = header_row.index(c_data) + 1
                        for row_idx in range(2, ws.max_row + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            if cell.value is not None:
                                cell.number_format = oxl_number_formats.FORMAT_DATE_DDMMYY
            excel_buf.seek(0)
            excel_bytes = excel_buf.getvalue()
            pdf_bytes = _build_pdf_bytes(tabela_exibir)
        except Exception:  # noqa: BLE001
            heavy_exports_enabled = False

    with st.container(border=True):
        st.caption("Exportar recorte filtrado")
        btn2, btn3 = st.columns([1, 1])
        with btn2:
            if heavy_exports_enabled and excel_bytes is not None:
                st.download_button(
                    "Exportar Excel",
                    data=excel_bytes,
                    file_name="conciliacao_operacional_filtrada.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            else:
                st.button("Exportar Excel", disabled=True, use_container_width=True)
        with btn3:
            if heavy_exports_enabled and pdf_bytes is not None:
                st.download_button(
                    "Exportar PDF",
                    data=pdf_bytes,
                    file_name="conciliacao_operacional_filtrada.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            else:
                st.button("Exportar PDF", disabled=True, use_container_width=True)
        if not heavy_exports_enabled:
            st.caption(
                f"Excel/PDF desativados para recortes acima de {_max_rows_heavy_export:,} linhas (estabilidade)."
            )

    _fdl_ui_gap_section_lg()

    tabela_grid = _dataframe_conciliacao_somente_grid(tabela_exibir_grid)
    _cfg_grid = (
        _column_config_conciliacao(tabela_grid, moeda_como_texto=True)
        if not tabela_grid.empty
        else None
    )
    _n_grid_total = _repasse_n_recorte_total
    _n_grid_vis = len(tabela_exibir_grid)
    _h_grid = min(520, 44 + min(_n_grid_vis, 16) * 34) if _n_grid_vis else 160
    _use_styler = repasse_ui_apply_grid_styler(grid_truncated=_repasse_grid_truncated)
    _disp_grid: object = (
        _repasse_fila_operacional_styler(tabela_grid)
        if (not tabela_grid.empty and _use_styler)
        else tabela_grid
    )

    if tabela_exibir.empty:
        st.info(
            "Nenhum registo corresponde aos filtros. Alargue o período, limpe a busca ou ajuste os critérios."
        )
        st.dataframe(
            _disp_grid,
            use_container_width=True,
            height=_h_grid,
            hide_index=True,
            column_config=_cfg_grid,
        )
    else:
        st.dataframe(
            _disp_grid,
            use_container_width=True,
            height=_h_grid,
            hide_index=True,
            column_config=_cfg_grid,
        )
    if _repasse_grid_truncated:
        st.caption(
            f"**{_fmt_int_ptbr(_n_grid_total)} linhas** no recorte · na grelha: primeiras "
            f"**{_fmt_int_ptbr(_n_grid_vis)}** (limite de estabilidade **{REPASSE_UI_GRID_ROW_CAP:,}**)."
        )
    else:
        st.caption(f"{_fmt_int_ptbr(_n_grid_total)} linhas no recorte.")

_admin_mode = _is_admin_mode()

_fv = st.session_state["op_financeiro_view"]
_fdl_product_area = str(st.session_state.get(SESSION_FDL_PRODUCT_AREA_KEY, FDL_PRODUCT_AREA_FINANCEIRO))
if _fdl_product_area not in (
    FDL_PRODUCT_AREA_FINANCEIRO,
    FDL_PRODUCT_AREA_FATURAMENTO_DRE,
    FDL_PRODUCT_AREA_COMERCIAL_PEDIDOS,
    FDL_PRODUCT_AREA_APURACAO_FISCAL,
):
    _fdl_product_area = FDL_PRODUCT_AREA_FINANCEIRO
    st.session_state[SESSION_FDL_PRODUCT_AREA_KEY] = _fdl_product_area
_fdl_global_trace(f"rerun: area={_fdl_product_area} vista={_fv}")
frete_df = pd.DataFrame()
frete_info: dict[str, object] = {}
devolucoes_df = pd.DataFrame()
devolucoes_info: dict[str, object] = {}
faturamento_df = pd.DataFrame()
faturamento_info: dict[str, object] = {}
if _fv == "frete":
    # Ponto único de carga Frete (materializado → live); não carrega repasse/precomputed.
    try:
        _fdl_global_trace("frete: a carregar _load_frete_data")
        with st.spinner("A carregar dados de Frete…"):
            frete_df, frete_info, ts_proc = _load_frete_data(_active_org.org_id)
        if _admin_mode:
            if frete_info.get("frete_consume") == "live_fallback":
                st.warning(
                    "Frete: ficheiro consolidado indisponível ou com erro — em uso **fonte em tempo real** (fallback)."
                )
                st.caption(f"Path/URL tentado: `{frete_info.get('frete_materialized_target', '')}`")
                st.caption(f"Erro: {frete_info.get('frete_materialized_error', '')}")
            elif frete_info.get("frete_consume") == "materialized":
                _t_disp = str(frete_info.get("frete_materialized_target", ""))[:500]
                st.caption(f"Frete: ficheiro consolidado (`{_t_disp}`).")
            elif frete_info.get("frete_mat_note"):
                st.info(str(frete_info["frete_mat_note"]))
    except Exception as exc:
        if _strict_materialized() and isinstance(exc, ValueError):
            st.error(str(exc))
            st.stop()
        err_text = str(exc).strip() or exc.__class__.__name__
        if _expose_load_errors():
            st.error("Erro ao carregar os dados de Frete.")
            st.exception(exc)
        elif _admin_mode:
            st.warning("Dados de Frete indisponíveis no momento.")
            st.caption(f"Detalhe técnico: {exc}")
        else:
            st.warning("Dados indisponíveis no momento. Tente novamente em instantes.")
            with st.expander("Detalhes para suporte", expanded=False):
                st.code(err_text, language="text")
        st.stop()

    tabela_geral = pd.DataFrame()
    info = frete_info
    _fdl_global_trace("frete: dados carregados")
elif _fv == "devolucoes" and "devolucoes" in _enabled_modules:
    try:
        _fdl_global_trace("devolucoes: a carregar _load_devolucoes_data")
        with st.spinner("A carregar fila de Devoluções…"):
            devolucoes_df, devolucoes_info, ts_proc = _load_devolucoes_data(_active_org.org_id)
        if _admin_mode:
            _t_d = str(devolucoes_info.get("devolucoes_materialized_target", ""))[:500]
            _p_d = str(devolucoes_info.get("devolucoes_path_resolved", ""))[:500]
            st.caption(f"Devoluções: materializado · alvo=`{_t_d}` · lido=`{_p_d}`")
    except Exception as exc:
        if _strict_materialized() and isinstance(exc, ValueError):
            st.error(str(exc))
            st.stop()
        err_text = str(exc).strip() or exc.__class__.__name__
        if _expose_load_errors():
            st.error("Erro ao carregar os dados de **Devoluções**.")
            st.exception(exc)
        elif _admin_mode:
            st.warning("Dados de Devoluções indisponíveis no momento.")
            st.caption(f"Detalhe técnico: {exc}")
        else:
            st.warning("Dados indisponíveis no momento. Tente novamente em instantes.")
            with st.expander("Detalhes para suporte", expanded=False):
                st.code(err_text, language="text")
        st.stop()

    tabela_geral = pd.DataFrame()
    info = devolucoes_info
    _fdl_global_trace("devolucoes: dados carregados")
elif (
    _fdl_product_area
    in (
        FDL_PRODUCT_AREA_FATURAMENTO_DRE,
        FDL_PRODUCT_AREA_COMERCIAL_PEDIDOS,
        FDL_PRODUCT_AREA_APURACAO_FISCAL,
    )
    and "faturamento" in _enabled_modules
    and not _user_perfil_acesso_operacional_only()
):
    _allowed_org_key = ",".join(sorted(o.org_id for o in _app_ctx.organizations))
    _fdl_global_trace("faturamento_dre: a carregar _load_faturamento_dataframe_cached")
    with st.spinner("A carregar dados de Resultado Gerencial…"):
        faturamento_df, faturamento_info, ts_proc = _load_faturamento_dataframe_cached(
            _faturamento_load_cache_signature(
                FAT_DRE_CACHE_ACTIVE_ORG_PLACEHOLDER,
                consolidado=True,
                allowed_org_ids_key=_allowed_org_key,
            ),
            FAT_DRE_CACHE_ACTIVE_ORG_PLACEHOLDER,
            True,
            _allowed_org_key,
        )
    fc = str(faturamento_info.get("faturamento_consume", "")).strip()
    if fc == "materialized" and _admin_mode:
        _t_disp = str(faturamento_info.get("faturamento_materialized_target", ""))[:500]
        _lay = str(faturamento_info.get("faturamento_data_layout", ""))
        _src = str(faturamento_info.get("faturamento_resolution_source", ""))
        _pf = str(faturamento_info.get("faturamento_path_final_resolved", ""))[:500]
        _pf_full = str(faturamento_info.get("faturamento_path_final_resolved", "")).strip()
        _n_loaded = faturamento_info.get("faturamento_row_count_loaded")
        st.caption(
            f"Faturamento: layout **{_lay}** · origem `{_src}` · alvo=`{_t_disp}` · path=`{_pf}`"
        )
        with st.expander("Admin — ficheiro lido e validação", expanded=False):
            st.markdown(
                "| Campo | Valor |\n| --- | --- |\n"
                f"| **faturamento_resolution_source** | `{_src}` |\n"
                f"| **faturamento_row_count_loaded** | `{_n_loaded}` |\n"
            )
            st.code(_pf_full or "(vazio)", language="text")
            st.caption("**faturamento_path_final_resolved** (completo, acima).")
            _slug = str(_materialized_cliente_slug()).strip()
            _expect_marker = f"data_products/{_slug}/faturamento/current/dataset_faturamento_app.csv".casefold()
            _norm = _pf_full.replace("\\", "/").casefold()
            _expect_pq = f"data_products/{_slug}/faturamento/current/dataset.parquet".casefold()
            if _slug and (_expect_marker in _norm or _expect_pq in _norm):
                st.success(
                    f"Path alinha ao V2 canónico em `data_products/{_slug}/faturamento/current/` "
                    "(CSV ou Parquet, conforme existir no disco)."
                )
            elif _pf_full:
                st.warning(
                    "O path **não** contém o sufixo canónico esperado para o slug atual — confirme **FDL_MATERIALIZED_CLIENTE_SLUG** e **FDL_FATURAMENTO_MATERIALIZED_PATH**."
                )
            if _n_loaded is not None:
                st.markdown(_faturamento_admin_metadata_rowcount_message(_pf_full, int(_n_loaded)))
        _cnt = faturamento_info.get("faturamento_status_custo_counts")
        if isinstance(_cnt, dict) and _cnt:
            st.caption(
                f"**Status_Custo** após escopo (**{faturamento_info.get('faturamento_escopo', '—')}**): `{_cnt}` · "
                f"linhas carregadas: **{faturamento_info.get('faturamento_row_count_loaded', '—')}** → "
                f"**{faturamento_info.get('linhas', '—')}** após escopo."
            )
        if faturamento_info.get("faturamento_escopo") == FAT_DRE_ESCOPO_CONSOLIDADO:
            st.caption(
                "**Consolidado:** conjunto = todas as organizações permitidas ao utilizador (intersecção com `org_id` no ficheiro). "
                "A **Empresa** na barra lateral não restringe este conjunto; o recorte por marca é o multiselect **Empresa** no módulo."
            )
        st.caption(
            "**V2 canónico:** `data_products/<FDL_MATERIALIZED_CLIENTE_SLUG>/faturamento/current/` — com **CSV e Parquet**, "
            "o app lê o **CSV** primeiro. Sem path explícito, este ficheiro tem **prioridade** sobre o CSV «irmão do repasse» "
            "(join fiscal: Nota_Data_Emissao, etc.). O **slug** tem de bater com a materialização (ex. **cliente_5**)."
        )
        if faturamento_info.get("faturamento_scope_note"):
            st.caption(str(faturamento_info["faturamento_scope_note"]))
        if faturamento_info.get("faturamento_note_v2_canonical_available"):
            st.info(str(faturamento_info["faturamento_note_v2_canonical_available"]))
        with st.expander("Debug — `faturamento_info` (materializado)", expanded=False):
            try:
                st.json(json.loads(json.dumps(faturamento_info, default=str)))
            except (TypeError, ValueError):
                st.write({k: str(v)[:2000] for k, v in faturamento_info.items()})
    elif fc == "missing_config":
        if _admin_mode:
            st.warning(
                str(
                    faturamento_info.get(
                        "faturamento_note",
                        "Dados de faturamento não configurados ou não encontrados.",
                    )
                )
            )
        else:
            st.warning("Dados de faturamento não disponíveis. Contacte o administrador.")
    elif fc == "error":
        st.warning("Não foi possível carregar os dados de **Resultado Gerencial**.")
        if _admin_mode:
            st.caption(str(faturamento_info.get("faturamento_materialized_error", "")))
            st.caption(f"Alvo: `{faturamento_info.get('faturamento_materialized_target', '')}`")
    elif fc == "unsupported":
        if _admin_mode:
            st.warning(str(faturamento_info.get("faturamento_note", "Modo de consumo não suportado.")))
        else:
            st.warning("Esta vista não está disponível na configuração atual. Contacte o administrador.")

    if not faturamento_df.empty and "empresa" not in faturamento_df.columns:
        if str(faturamento_info.get("faturamento_data_layout", "")).strip() != "v2":
            faturamento_df = faturamento_df.copy()
            faturamento_df["empresa"] = _dataset_empresa_label()

    if not faturamento_df.empty:
        faturamento_df = _filtrar_df_col_empresa_por_contexto(faturamento_df)

    _nf_ctx = faturamento_info.get("faturamento_nf_df")
    if isinstance(_nf_ctx, pd.DataFrame):
        faturamento_info = {
            **faturamento_info,
            "faturamento_nf_df": _filtrar_df_col_empresa_por_contexto(_nf_ctx),
        }
    _nf_panel_ctx = faturamento_info.get("faturamento_nf_panel_df")
    if isinstance(_nf_panel_ctx, pd.DataFrame):
        faturamento_info = {
            **faturamento_info,
            "faturamento_nf_panel_df": _filtrar_df_col_empresa_por_contexto(_nf_panel_ctx),
        }

    faturamento_info = {**faturamento_info, "linhas": int(len(faturamento_df))}
    if not str(faturamento_info.get("faturamento_escopo") or "").strip():
        faturamento_info = {**faturamento_info, "faturamento_escopo": FAT_DRE_ESCOPO_CONSOLIDADO}
    tabela_geral = pd.DataFrame()
    info = faturamento_info
    _fdl_global_trace(f"faturamento_dre: após filtro empresa ({len(faturamento_df)} linhas)")
else:
    try:
        _fdl_global_trace("repasse: a carregar _load_data (cache por org/config)")
        with st.spinner("A carregar dados (a ir buscar o ficheiro à nuvem, se aplicável)…"):
            tabela_geral, info, ts_proc = _load_repasse_dataframe_cached(
                _repasse_load_cache_signature(_active_org.org_id)
            )
            if _admin_mode:
                if info.get("repasse_consume") == "live_fallback":
                    st.warning(
                        "Repasse: ficheiro consolidado indisponível ou com erro — em uso **fonte em tempo real** (fallback)."
                    )
                    st.caption(f"Path/URL tentado: `{info.get('repasse_materialized_target', '')}`")
                    st.caption(f"Erro: {info.get('repasse_materialized_error', '')}")
                elif info.get("repasse_materialized_note"):
                    st.info(str(info["repasse_materialized_note"]))
                elif info.get("repasse_consume") == "materialized":
                    st.caption(
                        f"Repasse: ficheiro consolidado (`{info.get('repasse_materialized_target', '')}`)."
                    )
    except Exception as exc:
        if _strict_materialized() and isinstance(exc, ValueError):
            st.error(str(exc))
            st.stop()
        err_text = str(exc).strip() or exc.__class__.__name__
        if _expose_load_errors():
            st.error("Erro ao carregar os dados. Ajuste Secrets/URL ou use o detalhe abaixo.")
            st.exception(exc)
        elif _admin_mode:
            st.warning("Dados indisponíveis no momento.")
            st.caption(f"Detalhe técnico: {exc}")
        else:
            st.warning("Dados indisponíveis no momento. Tente novamente em instantes.")
            with st.expander("Detalhes para suporte", expanded=False):
                st.code(err_text, language="text")
        od_url = _onedrive_public_url()
        if (
            _admin_mode
            and _data_source_mode() in {"onedrive", "filesystem"}
            and od_url
            and ":f:/" in od_url.lower()
        ):
            st.info(
                "A configuração atual usa **link de pasta** do SharePoint (`:f:/`), que depende de "
                "acesso à API Microsoft e costuma falhar na Streamlit Cloud. "
                "Use **FDL_DATA_SOURCE = \"precomputed\"** com `FDL_PRECOMPUTED_URL` (link direto ao "
                "`.csv` ou `.xlsx`) ou `FDL_PRECOMPUTED_PATH` no servidor."
            )
        st.stop()

    # Cache antigo do Streamlit ou pickle sem a coluna — alinhar ao pipeline atual.
    if "empresa" not in tabela_geral.columns:
        tabela_geral = tabela_geral.copy()
        tabela_geral["empresa"] = _dataset_empresa_label()

    tabela_geral = _filtrar_df_col_empresa_por_contexto(
        tabela_geral,
        repasse_org_scoped_fallback=True,
    )
    info = {**info, "linhas": int(len(tabela_geral))}
    _fdl_global_trace(f"repasse: após filtro empresa ({len(tabela_geral)} linhas)")

_ts_raw = str(ts_proc).strip() if ts_proc is not None else ""
_dt_sb = _fdl_parse_ts_proc(_ts_raw)
_sb_ts_display = (
    _dt_sb.strftime("%d/%m/%Y %H:%M") if _dt_sb is not None else (_ts_raw if _ts_raw else "—")
)

_fdl_global_trace("04: antes da sidebar (dados carregados)")
_inject_fdl_professional_theme()
if _bootstrap_debug_enabled() and _admin_mode:
    with st.expander("Diagnóstico bootstrap (FDL_DEBUG_BOOTSTRAP=1)", expanded=True):
        st.write("**Última etapa:**", st.session_state.get("_fdl_bootstrap_stage", "—"))
        st.write("**Área:**", _fdl_product_area, "· **Vista financeiro:**", _fv)
        st.write("**Modo seguro (FDL_SAFE_MODE):**", _fdl_safe_mode())
        st.write("**Layout mínimo (FDL_MINIMAL_LAYOUT, omisso=on):**", _fdl_minimal_layout())
        _n_linhas_dbg = (
            len(tabela_geral)
            if _fv == "repasse"
            else (
                len(faturamento_df)
                if _fdl_product_area
                in (FDL_PRODUCT_AREA_FATURAMENTO_DRE, FDL_PRODUCT_AREA_COMERCIAL_PEDIDOS)
                else (
                    "— (apuracao fiscal)"
                    if _fdl_product_area == FDL_PRODUCT_AREA_APURACAO_FISCAL
                    else "— (vista frete)"
                )
            )
        )
        st.write("**Linhas tabela_geral (repasse) / faturamento_df:**", _n_linhas_dbg)
        _lg = st.session_state.get("_fdl_bootstrap_log")
        if isinstance(_lg, list) and _lg:
            st.write("**Log de etapas (esta execução):**")
            for _i, _line in enumerate(_lg, 1):
                st.caption(f"{_i}. {_line}")

with st.sidebar:
    _fdl_sidebar_inject_layout_css()
    _sb_view = st.session_state.get("op_financeiro_view", "repasse")
    _sb_area = st.session_state.get(SESSION_FDL_PRODUCT_AREA_KEY, FDL_PRODUCT_AREA_FINANCEIRO)

    st.write("")

    _cli_raw = str(st.session_state.get("cliente", "") or _app_ctx.display_name or "").strip()
    if not _cli_raw:
        _u = str(st.session_state.get("usuario", "") or "").strip()
        _cli_raw = _u.split("@", 1)[0] if "@" in _u else (_u or "Conta")
    _cli_nome = html.escape(_cli_raw)
    _sb_tagline = "Da operação ao insight"
    _brand_inner = (
        '<div class="fdl-sb-wordmark" role="img" aria-label="FDL Analytics">'
        '<span class="fdl-sb-wordmark-fdl">FDL</span>'
        '<span class="fdl-sb-wordmark-analytics"> Analytics</span>'
        "</div>"
        f'<div class="fdl-sb-tagline fdl-sb-tagline--after-logo">{html.escape(_sb_tagline)}</div>'
        '<div class="fdl-sb-client-row"><div class="fdl-sb-client-block">'
        '<span class="fdl-sb-client-icon" aria-hidden="true">👤</span>'
        f'<span class="fdl-sb-client-name">{_cli_nome}</span>'
        "</div></div>"
    )
    st.markdown(
        '<div class="fdl-sb-brand-shell">'
        f'<div class="fdl-sb-brand">{_brand_inner}</div>'
        "</div>",
        unsafe_allow_html=True,
    )
    st.divider()

    _empresas_usuario = list(st.session_state["empresas_permitidas"])
    _nomes_nav = nomes_permitidos_com_registro(_empresas_usuario)

    _has_gerencial = "faturamento" in _enabled_modules and not _user_perfil_acesso_operacional_only()
    _has_operacional = (
        "repasse" in _enabled_modules
        or "frete" in _enabled_modules
        or "devolucoes" in _enabled_modules
    )
    _first_nav_section = True

    _lbl_repasse = "Conciliação de Repasse"
    _lbl_frete = "Conciliação de Frete"
    _lbl_devolucoes = "Controle de Devoluções"
    _lbl_fat_dre = "Resultado Gerencial"

    if _has_gerencial:
        _sec_cls = (
            "fdl-sb-section-label fdl-sb-section-label--first"
            if _first_nav_section
            else "fdl-sb-section-label"
        )
        st.markdown(f'<p class="{_sec_cls}">GERENCIAL</p>', unsafe_allow_html=True)
        _first_nav_section = False
        st.button(
            f"📊 {_lbl_fat_dre}",
            key="fdl_mod_faturamento_dre",
            use_container_width=True,
            type="primary" if _sb_area == FDL_PRODUCT_AREA_FATURAMENTO_DRE else "secondary",
            on_click=_sb_nav_set_faturamento_dre,
            help=(
                "Visão de negócio. Carga consolidada das organizações permitidas; escolha uma ou mais "
                "marcas (ou todas) no filtro **Empresa** dentro do painel."
            ),
        )
        st.markdown(
            '<p class="fdl-sb-nav-item-hint">DRE, margem e desempenho</p>',
            unsafe_allow_html=True,
        )
        st.button(
            "🛒 Comercial & pedidos",
            key="fdl_mod_comercial_pedidos",
            use_container_width=True,
            type="primary" if _sb_area == FDL_PRODUCT_AREA_COMERCIAL_PEDIDOS else "secondary",
            on_click=_sb_nav_set_comercial_pedidos,
            help=(
                "Análise comercial sobre pedidos atendidos; receita por linha = **Vl_Venda** da tabela materializada "
                "(fallback lista×qtd). Sem NF. Filtros no painel; base consolidada como Resultado Gerencial."
            ),
        )
        st.markdown(
            '<p class="fdl-sb-nav-item-hint">Pedidos e status operacional</p>',
            unsafe_allow_html=True,
        )

    _has_fiscal_apuracao = (
        "faturamento" in _enabled_modules and not _user_perfil_acesso_operacional_only()
    )
    if _has_fiscal_apuracao:
        _sec_cls = (
            "fdl-sb-section-label fdl-sb-section-label--first"
            if _first_nav_section
            else "fdl-sb-section-label"
        )
        st.markdown(f'<p class="{_sec_cls}">FISCAL</p>', unsafe_allow_html=True)
        _first_nav_section = False
        st.button(
            "🧾 Apuração Fiscal",
            key="fdl_mod_apuracao_fiscal",
            use_container_width=True,
            type="primary" if _sb_area == FDL_PRODUCT_AREA_APURACAO_FISCAL else "secondary",
            on_click=_sb_nav_set_apuracao_fiscal,
            help="Notas emitidas, base fiscal líquida e imposto no período — visão fiscal consolidada.",
        )
        st.markdown(
            '<p class="fdl-sb-nav-item-hint">Notas, base tributável e imposto</p>',
            unsafe_allow_html=True,
        )

    if _has_operacional:
        _sec_cls = (
            "fdl-sb-section-label fdl-sb-section-label--first"
            if _first_nav_section
            else "fdl-sb-section-label"
        )
        st.markdown(f'<p class="{_sec_cls}">OPERACIONAL</p>', unsafe_allow_html=True)
        _first_nav_section = False

        if _sb_area == FDL_PRODUCT_AREA_FINANCEIRO and _nomes_nav:
            _op_hints = []
            if "repasse" in _enabled_modules:
                _op_hints.append("Repasse")
            if "frete" in _enabled_modules:
                _op_hints.append("Frete")
            if "devolucoes" in _enabled_modules:
                _op_hints.append("Devoluções")
            st.markdown(
                f'<p class="fdl-sb-org-hint">{" · ".join(_op_hints) or "Operacional"}</p>',
                unsafe_allow_html=True,
            )
            _org_idx = 0
            for i, n in enumerate(_nomes_nav):
                _o = organizacao_por_nome_cadastrado(n)
                if _o and _o.org_id == _app_ctx.active_org_id:
                    _org_idx = i
                    break
            _sel_nome = st.selectbox(
                "Empresa ativa",
                options=_nomes_nav,
                index=_org_idx,
                key="operacional_empresa_ativa_select",
                label_visibility="visible",
                help=(
                    "Define qual organização carregar para Repasse, Frete e Devoluções. "
                    "Em Resultado Gerencial e Comercial, o recorte por marca fica no filtro Empresa do painel."
                ),
            )
            _chosen_org = organizacao_por_nome_cadastrado(_sel_nome)
            if _chosen_org and _chosen_org.org_id != _app_ctx.active_org_id:
                st.session_state[SESSION_ACTIVE_ORG_KEY] = _chosen_org.org_id
                st.rerun()

        if "repasse" in _enabled_modules:
            st.button(
                f"💸 {_lbl_repasse}",
                key="fdl_mod_repasse",
                use_container_width=True,
                type="primary"
                if _sb_area == FDL_PRODUCT_AREA_FINANCEIRO and _sb_view == "repasse"
                else "secondary",
                on_click=_sb_nav_set_repasse,
            )
        if "frete" in _enabled_modules:
            st.button(
                f"🚚 {_lbl_frete}",
                key="fdl_mod_frete",
                use_container_width=True,
                type="primary"
                if _sb_area == FDL_PRODUCT_AREA_FINANCEIRO and _sb_view == "frete"
                else "secondary",
                on_click=_sb_nav_set_frete,
            )
        if "devolucoes" in _enabled_modules:
            st.button(
                f"↩️ {_lbl_devolucoes}",
                key="fdl_mod_devolucoes",
                use_container_width=True,
                type="primary"
                if _sb_area == FDL_PRODUCT_AREA_FINANCEIRO and _sb_view == "devolucoes"
                else "secondary",
                on_click=_sb_nav_set_devolucoes,
                help="Fila operacional: só vendas candidatas a devolução/reembolso/mediação ou com eventos correlatos nas liberações.",
            )

    if _admin_mode:
        st.markdown(
            f'<p class="fdl-sb-footer-admin">{html.escape(_sidebar_version_display())}</p>',
            unsafe_allow_html=True,
        )

    if _admin_mode and _data_source_mode() == "upload_zip":
        _render_cloud_data_loader()

    st.markdown('<div class="fdl-sb-footer-spacer"></div>', unsafe_allow_html=True)
    st.divider()

    if _admin_mode and st.button(
        "Atualizar dados",
        use_container_width=True,
        help="Limpa caches e recarrega os dados a partir da fonte configurada.",
        key="fdl_sb_admin_refresh",
        type="secondary",
    ):
        st.cache_data.clear()
        for _k in list(st.session_state.keys()):
            if str(_k).startswith("_frete_cache_") or str(_k).startswith("_devolucoes_cache_"):
                st.session_state.pop(_k, None)
        st.rerun()

    st.button(
        "Sair",
        icon="🚪",
        use_container_width=True,
        help="Encerra a sessão neste navegador.",
        type="tertiary",
        key="fdl_sb_logout",
        on_click=_sb_logout_click,
    )

    _data_esc = html.escape(str(_sb_ts_display).strip() or "—")
    st.markdown(
        f"""
<div style="
    position: absolute;
    bottom: 16px;
    left: 0;
    right: 0;
    text-align: center;
    font-size: 0.72rem;
    color: #94a3b8;
    padding: 0 16px;
    border-top: 1px solid #e2e8f0;
    padding-top: 12px;
">
    🕐 Dados de <strong>{_data_esc}</strong>
</div>
""",
        unsafe_allow_html=True,
    )

_fdl_global_trace("05: após sidebar — antes do hero / painel principal")

if _fv == "repasse" and _fdl_product_area == FDL_PRODUCT_AREA_FINANCEIRO:
    try:
        if _repasse_use_parquet():
            _fdl_global_trace(
                "repasse: base Parquet (sem map_acao, sem Ação sugerida operacional, sem exclusão/filtro N° venda na UI)"
            )
            tabela = tabela_geral.copy()
            tabela["Valor pago"] = pd.to_numeric(tabela.get("Valor pago"), errors="coerce")
            tabela_operacional_base = tabela.copy()
            _fdl_global_trace(f"repasse: base pronta ({len(tabela_operacional_base)} linhas)")
        else:
            _fdl_global_trace("repasse: a preparar base (map_acao / filtros negócio)")
            _acao_baixa = "Baixado" if _repasse_sem_bling() else "Baixar no Bling"
            map_acao = {
                "Ok": "Ok",
                "Baixar no Bling": _acao_baixa,
                "Baixado": _acao_baixa,
                "Analisar manualmente": "Analisar diferença",
                "Verificar título no Bling": "Verificar recebimento",
                "Revisar venda zerada": "Revisar venda zerada",
                "Verificar faturamento": "Verificar faturamento",
            }
            tabela_geral["Ação sugerida operacional"] = (
                tabela_geral["Ação sugerida"].map(map_acao).fillna(tabela_geral["Ação sugerida"])
            )
            tabela = tabela_geral.copy()

            # Mantém também linhas sem pagamento para não ocultar plataformas
            # em cenários onde o extrato ainda não foi consolidado.
            tabela["Valor pago"] = pd.to_numeric(tabela.get("Valor pago"), errors="coerce")

            # Exibição operacional focada em vendas:
            # mantém somente linhas com N° de venda preenchido.
            tabela["N° de venda"] = tabela["N° de venda"].fillna("").astype(str).str.strip()
            tabela = tabela[tabela["N° de venda"].ne("")].copy()

            # Base operacional antes dos filtros da UI (mesma regra de negócio de sempre).
            tabela_operacional_base = tabela.copy()
            _fdl_global_trace(f"repasse: base pronta ({len(tabela_operacional_base)} linhas)")
    except Exception as exc:
        _fdl_global_trace(f"repasse: ERRO na preparação da base — {exc.__class__.__name__}")
        st.error("Erro ao preparar a base de **Conciliação de Repasse** (colunas ou dados incompatíveis).")
        st.exception(exc)
        tabela_operacional_base = pd.DataFrame()
else:
    tabela_operacional_base = pd.DataFrame()

if _fdl_product_area == FDL_PRODUCT_AREA_COMERCIAL_PEDIDOS and "faturamento" in _enabled_modules:
    _render_financeiro_header(
        segment="Comercial",
        title="Comercial & pedidos",
        subtitle="ABC receita e giro, tendência 3 meses e sugestão de compra · só pedidos atendidos.",
        kicker_area="Comercial & pedidos",
        compact_spacing=True,
    )
elif _fdl_product_area == FDL_PRODUCT_AREA_FINANCEIRO and _fv == "devolucoes":
    _render_financeiro_header(
        segment="Operacional",
        title="Controle de Devoluções",
        subtitle="Fila de candidatas · devolução, reembolso, mediação, reclamação e eventos nas liberações (não é o export completo de vendas).",
        compact_spacing=True,
    )
elif _fdl_product_area == FDL_PRODUCT_AREA_FINANCEIRO and _fv == "repasse":
    _render_financeiro_header(
        segment="Repasse",
        title="Conciliação de Repasse",
        subtitle="Fila de tratativa · repasses, NF e divergências no mesmo recorte.",
        compact_spacing=True,
    )
elif _fdl_product_area == FDL_PRODUCT_AREA_FINANCEIRO and _fv == "frete":
    _render_financeiro_header(
        segment="Frete",
        title="Conciliação de Frete",
        subtitle="Frete cobrado vs esperado · anúncios e vendas no mesmo recorte.",
        compact_spacing=True,
    )
elif _fdl_product_area == FDL_PRODUCT_AREA_FINANCEIRO:
    _render_financeiro_header(
        segment="Financeiro",
        title="Financeiro operacional",
        subtitle="Escolha Repasse, Frete ou Devoluções na barra lateral.",
        compact_spacing=True,
    )

if _fv == "repasse" and _fdl_product_area == FDL_PRODUCT_AREA_FINANCEIRO:
    try:
        _fdl_global_trace("repasse: a renderizar _painel_conciliacao_fragment (filtros UI)")
        _painel_conciliacao_fragment(tabela_operacional_base, ts_proc)
        _fdl_global_trace("repasse: painel concluído")
    except Exception as exc:
        _fdl_global_trace(f"repasse: ERRO no painel — {exc.__class__.__name__}")
        st.error("Erro ao renderizar a **Conciliação de Repasse** (filtros ou tabela).")
        st.exception(exc)
elif _fdl_product_area == FDL_PRODUCT_AREA_APURACAO_FISCAL and "faturamento" in _enabled_modules:
    try:
        _fdl_global_trace("apuracao_fiscal: painel fiscal")
        from app.pages.apuracao_fiscal import render_apuracao_fiscal_page

        render_apuracao_fiscal_page(
            faturamento_df,
            faturamento_info,
            ts_proc,
            org_id=_active_org.org_id,
            org_display_name=_active_org.display_name,
        )
        _fdl_global_trace("apuracao_fiscal: painel concluído")
    except Exception as exc:
        _fdl_global_trace(f"apuracao_fiscal: ERRO — {exc.__class__.__name__}")
        st.error("Erro ao renderizar **Apuração Fiscal**.")
        st.exception(exc)
elif _fdl_product_area == FDL_PRODUCT_AREA_FATURAMENTO_DRE and "faturamento" in _enabled_modules:
    try:
        _fdl_global_trace("faturamento_dre: vista mínima Etapa 1")
        _render_faturamento_dre_minimal(
            faturamento_df,
            faturamento_info,
            ts_proc,
            org_id=_active_org.org_id,
            org_display_name=_active_org.display_name,
        )
        _fdl_global_trace("faturamento_dre: painel concluído")
    except Exception as exc:
        _fdl_global_trace(f"faturamento_dre: ERRO no painel — {exc.__class__.__name__}")
        st.error("Erro ao renderizar **Resultado Gerencial**.")
        st.exception(exc)
elif _fdl_product_area == FDL_PRODUCT_AREA_COMERCIAL_PEDIDOS and "faturamento" in _enabled_modules:
    try:
        _fdl_global_trace("comercial_pedidos: painel")
        _render_comercial_pedidos_analise(faturamento_df, faturamento_info, ts_proc)
        _fdl_global_trace("comercial_pedidos: painel concluído")
    except Exception as exc:
        _fdl_global_trace(f"comercial_pedidos: ERRO — {exc.__class__.__name__}")
        st.error("Erro ao renderizar **Comercial & pedidos**.")
        st.exception(exc)
elif _fv == "devolucoes" and "devolucoes" in _enabled_modules:
    try:
        _fdl_global_trace("devolucoes: painel")
        _painel_devolucoes_operacional(devolucoes_df, devolucoes_info, ts_proc)
        _fdl_global_trace("devolucoes: painel concluído")
    except Exception as exc:
        _fdl_global_trace(f"devolucoes: ERRO — {exc.__class__.__name__}")
        st.error("Erro ao renderizar **Controle de Devoluções**.")
        st.exception(exc)
elif _fv == "frete":
    try:
        _fdl_global_trace("frete: a renderizar _painel_frete_emergencial")
        _painel_frete_emergencial(_active_org.org_id, frete_df, frete_info, ts_proc)
        _fdl_global_trace("frete: painel concluído")
    except Exception as exc:
        _fdl_global_trace(f"frete: ERRO no painel — {exc.__class__.__name__}")
        st.error("Não foi possível carregar o painel de Frete.")
        st.exception(exc)

_fdl_global_trace("99: fim do script app_operacional")

