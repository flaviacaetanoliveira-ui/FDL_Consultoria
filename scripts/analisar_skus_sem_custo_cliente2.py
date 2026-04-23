"""
Analisa SKUs com Status_Custo = SKU_SEM_CORRESPONDENCIA no dataset cliente_2,
cruza com Custos.xlsx (match exato na chave de join, match no código bruto, parciais).

Saídas em data_products/cliente_2/faturamento/current/:
  - SKUs_Sem_Custo_ATUALIZADO.xlsx
  - SKUs_Sem_Custo_ATUALIZADO_analise.csv (detalhe + colunas de match)
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from processing.faturamento.normalize import normalize_sku_join_key_scalar

STATUS_SEM = "SKU_SEM_CORRESPONDENCIA"
DEFAULT_PARQUET = ROOT / "data_products/cliente_2/faturamento/current/dataset.parquet"
DEFAULT_CUSTO = Path(
    r"C:\Users\diieg\OneDrive - FDL Consultoria\Cursor\Pedro\Cliente_2\Custos.xlsx"
)
DEFAULT_OUT_DIR = ROOT / "data_products/cliente_2/faturamento/current"


def _resumo_sem_custo(sem: pd.DataFrame) -> pd.DataFrame:
    if sem.empty:
        return pd.DataFrame(
            columns=[
                "org_id",
                "SKU_Normalizado",
                "Código_Pedido",
                "Quantidade",
                "Receita",
                "NFs",
            ]
        )
    g = sem.groupby(["org_id", "SKU_Normalizado"], sort=False, dropna=False)

    rows: list[dict[str, object]] = []
    for (oid, sku), idx in g.groups.items():
        sub = sem.loc[idx]
        c0 = sub["Código"].dropna().astype(str).str.strip() if "Código" in sub.columns else pd.Series(dtype=str)
        cod_ex = str(c0.iloc[0]) if len(c0) else ""
        n_nf = 0
        if "Nota_Numero_Normalizado" in sub.columns:
            n_nf = int(
                sub["Nota_Numero_Normalizado"]
                .fillna("")
                .astype(str)
                .str.strip()
                .replace("", pd.NA)
                .nunique(dropna=True)
            )
        rows.append(
            {
                "org_id": oid,
                "SKU_Normalizado": sku,
                "Código_Pedido": cod_ex,
                "Quantidade": float(pd.to_numeric(sub.get("Quantidade", 0), errors="coerce").fillna(0).sum()),
                "Receita": float(pd.to_numeric(sub.get("Vl_Venda", 0), errors="coerce").fillna(0).sum()),
                "NFs": n_nf,
            }
        )
    out = pd.DataFrame(rows)
    return out.sort_values("Receita", ascending=False).reset_index(drop=True)


def _load_custo_codes(path: Path) -> tuple[str, list[str], dict[str, str]]:
    df = pd.read_excel(path, header=1, engine="openpyxl")
    # Mesma convenção do metadata cliente_2: coluna SKU é a segunda (índice 1).
    sku_col = df.columns[1]
    raw: list[str] = []
    for x in df[sku_col].dropna().tolist():
        s = str(x).strip()
        if s and s.lower() not in ("nan", "none"):
            raw.append(s)
    norm_set: set[str] = set()
    norm_to_raw: dict[str, str] = {}
    for r in raw:
        k = normalize_sku_join_key_scalar(r)
        if not k:
            continue
        norm_set.add(k)
        norm_to_raw.setdefault(k, r)
    return str(sku_col), raw, norm_to_raw


def _partial_matches(sku: str, codigos_lower: list[tuple[str, str]], max_n: int = 8) -> list[str]:
    """codigos_lower: (lower_raw, raw) para cada código da planilha."""
    sku_l = str(sku).strip().casefold()
    if len(sku_l) < 4:
        return []
    hits: list[tuple[int, str]] = []
    seen: set[str] = set()
    for low, raw in codigos_lower:
        if len(low) < 3 or sku_l == low:
            continue
        if sku_l in low or low in sku_l:
            if raw in seen:
                continue
            seen.add(raw)
            hits.append((abs(len(sku_l) - len(low)), raw))
    hits.sort(key=lambda t: (t[0], t[1]))
    return [raw for _, raw in hits[:max_n]]


def _numeric_prefix_matches(sku: str, norm_keys: set[str], max_n: int = 6) -> list[str]:
    if not sku.isdigit() or len(sku) < 4:
        return []
    out: list[str] = []
    p = sku[:4]
    for k in norm_keys:
        if k.isdigit() and len(k) >= 4 and k[:4] == p and k != sku:
            out.append(k)
            if len(out) >= max_n:
                break
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--parquet", type=Path, default=DEFAULT_PARQUET)
    ap.add_argument("--custo", type=Path, default=DEFAULT_CUSTO)
    ap.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    args = ap.parse_args()

    df = pd.read_parquet(args.parquet)
    sem = df[df["Status_Custo"].astype(str).str.strip() == STATUS_SEM].copy()
    resumo = _resumo_sem_custo(sem)

    print(f"=== {len(resumo)} SKUs únicos SEM CUSTO (agregado) ===\n")
    print(f"Linhas de pedido sem custo: {len(sem)}")
    print(f"Receita total (Vl_Venda) afetada: R$ {resumo['Receita'].sum():,.2f}\n")

    if not args.custo.is_file():
        print(f"ERRO: planilha não encontrada: {args.custo}", file=sys.stderr)
        return 1

    sku_col_name, raw_list, norm_to_raw = _load_custo_codes(args.custo)
    norm_keys = set(norm_to_raw.keys())
    codigos_lower = [(r.casefold(), r) for r in raw_list]

    analise_rows: list[dict[str, object]] = []
    for _, row in resumo.iterrows():
        sku = str(row["SKU_Normalizado"]).strip().casefold()
        cod_ped = str(row["Código_Pedido"]).strip() if pd.notna(row["Código_Pedido"]) else ""
        cod_ped_n = normalize_sku_join_key_scalar(cod_ped) if cod_ped else ""

        notes: list[str] = []
        match_exato = sku in norm_keys
        if match_exato:
            notes.append(f"EXATO_JOIN:{norm_to_raw.get(sku, sku)}")
        if cod_ped_n and cod_ped_n != sku and cod_ped_n in norm_keys:
            notes.append(f"EXATO_COD_PEDIDO:{norm_to_raw.get(cod_ped_n, cod_ped)}")
        cod_lower = cod_ped.casefold()
        if cod_lower and cod_lower != sku and cod_lower in {r.casefold() for r in raw_list}:
            # match bruto lowercase na planilha (sem mesma normalização)
            pass

        partial = _partial_matches(sku, codigos_lower, max_n=6)
        if partial and not match_exato:
            notes.append("PARCIAL:" + "; ".join(partial[:5]))

        num_pref = _numeric_prefix_matches(sku, norm_keys, max_n=5)
        if num_pref and not match_exato:
            notes.append("PREFIXO_NUM:" + "; ".join(num_pref))

        if not notes:
            notes.append("SEM_SUGESTAO")

        analise_rows.append(
            {
                **row.to_dict(),
                "Match_Notas": " | ".join(notes),
                "Coluna_SKU_planilha": sku_col_name,
            }
        )

    analise = pd.DataFrame(analise_rows)

    # Categorização
    cat_counts: dict[str, tuple[int, float]] = {}
    for _, r in analise.iterrows():
        sku = str(r["SKU_Normalizado"]).casefold()
        rec = float(r["Receita"])
        m = str(r["Match_Notas"])
        if m.startswith("EXATO"):
            c = "Match exato (join)"
        elif m.startswith("PARCIAL") or "PARCIAL:" in m:
            c = "Só match parcial / substring"
        elif "PREFIXO_NUM" in m:
            c = "Numérico — possível prefixo na planilha"
        elif sku.isdigit():
            c = "Numérico — sem sugestão"
        else:
            c = "Alfanumérico — sem sugestão"
        n, s = cat_counts.get(c, (0, 0.0))
        cat_counts[c] = (n + 1, s + rec)

    print("=== RESUMO POR CATEGORIA (heurística) ===\n")
    for c, (n, s) in sorted(cat_counts.items(), key=lambda x: -x[1][1]):
        print(f"{c}: {n} SKUs, R$ {s:,.2f}")
    print()

    print("=== Top 25 por receita ===\n")
    top = analise.head(25)[
        ["org_id", "SKU_Normalizado", "Código_Pedido", "Receita", "Match_Notas"]
    ]
    for _, r in top.iterrows():
        print(
            f"{str(r['SKU_Normalizado']):<22} | {str(r['org_id']):<12} | R$ {r['Receita']:>10,.2f} | {r['Match_Notas'][:100]}"
        )

    args.out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = args.out_dir / "SKUs_Sem_Custo_ATUALIZADO.xlsx"
    csv_path = args.out_dir / "SKUs_Sem_Custo_ATUALIZADO_analise.csv"

    analise.to_csv(csv_path, index=False, encoding="utf-8-sig")

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "SKUs_Sem_Custo"
    headers = list(analise.columns)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    fill_pedro = PatternFill("solid", fgColor="FFFF00")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for ri, row in enumerate(analise.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if headers[ci - 1] == "Receita":
                c.number_format = "#,##0.00"
            if headers[ci - 1] == "Match_Notas" and str(val).startswith("SEM_SUGESTAO"):
                c.fill = fill_pedro

    widths = [12, 22, 22, 12, 14, 8, 70, 18]
    for i, w in enumerate(widths, 1):
        if i <= len(headers):
            ws.column_dimensions[get_column_letter(i)].width = min(w, 50)

    w2 = wb.create_sheet("Resumo_Categorias")
    w2.append(["Categoria", "Qtd_SKUs", "Receita"])
    for c, (n, s) in sorted(cat_counts.items(), key=lambda x: -x[1][1]):
        w2.append([c, n, s])
    w2.append(["TOTAL", len(analise), float(analise["Receita"].sum())])
    for row in w2.iter_rows(min_row=1, max_row=w2.max_row, min_col=1, max_col=3):
        for cell in row:
            if cell.row == 1:
                cell.font = Font(bold=True)
            if cell.column == 3 and cell.row > 1 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

    wb.save(xlsx_path)
    print(f"\nCSV: {csv_path}")
    print(f"Excel: {xlsx_path}")
    print(f"Planilha custo: {args.custo} (coluna SKU: {sku_col_name!r}, {len(raw_list)} códigos brutos)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
