"""
Exporta SKUs sem custo (Status_Custo = SKU_SEM_CORRESPONDENCIA) para revisão na Custos.xlsx.

Saídas (por defeito em data_products/cliente_2/faturamento/current/):
  - SKUs_Sem_Custo_Para_Pedro.xlsx (detalhe + resumo por org)
  - SKUs_Sem_Custo_Para_Pedro.csv

Uso:
  python scripts/export_skus_sem_custo_pedro.py
  python scripts/export_skus_sem_custo_pedro.py --parquet caminho/dataset.parquet --out-dir pasta
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
STATUS_SEM = "SKU_SEM_CORRESPONDENCIA"


def _build_relatorio(sem: pd.DataFrame) -> pd.DataFrame:
    if sem.empty:
        return pd.DataFrame(
            columns=[
                "org_id",
                "Empresa",
                "SKU_Normalizado",
                "Código_pedido",
                "Qtd_Total",
                "Preco_Lista_Medio",
                "NFs_distintas",
                "Primeira_emissao_NF",
                "Ultima_emissao_NF",
                "Receita_Vl_Venda",
            ]
        )
    pl_col = "Preço de lista" if "Preço de lista" in sem.columns else ""
    nf_col = "Nota_Numero_Normalizado" if "Nota_Numero_Normalizado" in sem.columns else ""
    emi_col = "Nota_Data_Emissao" if "Nota_Data_Emissao" in sem.columns else ""

    rows: list[dict[str, object]] = []
    for (oid, sku), group in sem.groupby(["org_id", "SKU_Normalizado"], sort=False, dropna=False):
        q = pd.to_numeric(group.get("Quantidade", pd.Series(0, index=group.index)), errors="coerce").fillna(0.0)
        vl = pd.to_numeric(group.get("Vl_Venda", pd.Series(0, index=group.index)), errors="coerce").fillna(0.0)
        pl = (
            pd.to_numeric(group[pl_col], errors="coerce").fillna(0.0)
            if pl_col
            else pd.Series(0.0, index=group.index)
        )
        cod = group["Código"].dropna().astype(str).str.strip() if "Código" in group.columns else pd.Series(dtype=str)
        cod_ex = str(cod.iloc[0]) if len(cod) else ""
        n_nf = 0
        if nf_col:
            n_nf = int(group[nf_col].fillna("").astype(str).str.strip().replace("", pd.NA).nunique(dropna=True))
        emi_min = emi_max = pd.NaT
        if emi_col:
            dt = pd.to_datetime(group[emi_col], errors="coerce", dayfirst=False)
            if dt.notna().any():
                emi_min = dt.min()
                emi_max = dt.max()
        emp = ""
        if "empresa" in group.columns:
            e = group["empresa"].dropna().astype(str).str.strip()
            emp = str(e.iloc[0]) if len(e) else ""
        rows.append(
            {
                "org_id": oid,
                "SKU_Normalizado": sku,
                "Empresa": emp,
                "Código_pedido": cod_ex,
                "Qtd_Total": float(q.sum()),
                "Preco_Lista_Medio": float(pl.mean()) if len(pl) else 0.0,
                "NFs_distintas": n_nf,
                "Primeira_emissao_NF": emi_min,
                "Ultima_emissao_NF": emi_max,
                "Receita_Vl_Venda": float(vl.sum()),
            }
        )

    out = pd.DataFrame(rows)
    return out.sort_values("Receita_Vl_Venda", ascending=False).reset_index(drop=True)


def _print_resumo_console(df: pd.DataFrame, sem: pd.DataFrame) -> None:
    total = len(df)
    com = (df["Status_Custo"].astype(str).str.strip() == "CUSTO_OK").sum() if "Status_Custo" in df.columns else 0
    print("=== RESUMO (dataset atual) ===\n")
    print(f"Total de linhas: {total:,}")
    print(f"Com CUSTO_OK: {com:,} ({100 * com / total:.1f}%)")
    print(f"Sem custo (SKU_SEM_CORRESPONDENCIA): {len(sem):,} ({100 * len(sem) / total:.1f}%)")
    print("\n=== POR org_id ===\n")
    for org in sorted(df["org_id"].dropna().unique(), key=str):
        o = df[df["org_id"] == org]
        s = o[o["Status_Custo"].astype(str).str.strip() == STATUS_SEM]
        rv = pd.to_numeric(s.get("Vl_Venda", pd.Series(0.0, index=s.index)), errors="coerce").fillna(0.0).sum()
        print(f"{org}:")
        print(f"  Linhas sem custo: {len(s)}")
        print(f"  SKUs únicos: {s['SKU_Normalizado'].nunique()}")
        print(f"  Soma Vl_Venda (afetado): R$ {float(rv):,.2f}")
        print()


def _write_xlsx(path: Path, rel: pd.DataFrame) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "SKUs_Sem_Custo"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_pedro = PatternFill("solid", fgColor="FFFF00")

    headers = [
        "org_id",
        "Empresa",
        "SKU_Normalizado",
        "Código_pedido",
        "Qtd_Total",
        "Preço_lista_médio",
        "NFs_distintas",
        "Primeira_emissao_NF",
        "Ultima_emissao_NF",
        "Receita_Vl_Venda",
        "CUSTO_PREENCHER",
    ]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border

    disp = rel.copy()
    if not disp.empty and "Primeira_emissao_NF" in disp.columns:
        for c in ("Primeira_emissao_NF", "Ultima_emissao_NF"):
            if c in disp.columns:
                disp[c] = pd.to_datetime(disp[c], errors="coerce").dt.strftime("%d/%m/%Y")
                disp[c] = disp[c].fillna("")

    for r_idx in range(len(disp)):
        row = disp.iloc[r_idx]
        vals = [
            row.get("org_id", ""),
            row.get("Empresa", ""),
            row.get("SKU_Normalizado", ""),
            row.get("Código_pedido", ""),
            float(row.get("Qtd_Total", 0) or 0),
            float(row.get("Preco_Lista_Medio", 0) or 0),
            int(row.get("NFs_distintas", 0) or 0),
            row.get("Primeira_emissao_NF", ""),
            row.get("Ultima_emissao_NF", ""),
            float(row.get("Receita_Vl_Venda", 0) or 0),
            "",
        ]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx + 2, column=c_idx, value=val)
            cell.border = border
            if c_idx in (5, 6, 10):
                cell.number_format = "#,##0.00"
            if c_idx == 11:
                cell.fill = fill_pedro

    widths = (14, 22, 22, 18, 12, 16, 14, 14, 14, 18, 22)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("Resumo_por_org")
    if rel.empty:
        ws2.cell(row=1, column=1, value="Sem linhas SKU_SEM_CORRESPONDENCIA")
    else:
        resumo = (
            rel.groupby("org_id", sort=False)
            .agg(SKUs_sem_custo=("SKU_Normalizado", "count"), Receita_afetada=("Receita_Vl_Venda", "sum"))
            .reset_index()
        )
        for col, h in enumerate(["org_id", "SKUs_sem_custo", "Receita_afetada"], 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = border
        for r_idx, row in enumerate(resumo.itertuples(index=False), 2):
            ws2.cell(row=r_idx, column=1, value=row[0]).border = border
            ws2.cell(row=r_idx, column=2, value=int(row[1])).border = border
            c3 = ws2.cell(row=r_idx, column=3, value=float(row[2]))
            c3.number_format = "R$ #,##0.00"
            c3.border = border
        ws2.column_dimensions["A"].width = 16
        ws2.column_dimensions["B"].width = 18
        ws2.column_dimensions["C"].width = 20

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    p = argparse.ArgumentParser(description="Exporta SKUs sem custo para revisão (Excel + CSV).")
    p.add_argument(
        "--parquet",
        type=Path,
        default=ROOT / "data_products/cliente_2/faturamento/current/dataset.parquet",
        help="Caminho para dataset.parquet",
    )
    p.add_argument(
        "--out-dir",
        type=Path,
        default=ROOT / "data_products/cliente_2/faturamento/current",
        help="Pasta de saída",
    )
    args = p.parse_args()
    pq = args.parquet.expanduser().resolve()
    out_dir = args.out_dir.expanduser().resolve()
    if not pq.is_file():
        print(f"Parquet não encontrado: {pq}", file=sys.stderr)
        return 1

    df = pd.read_parquet(pq)
    if "Status_Custo" not in df.columns or "SKU_Normalizado" not in df.columns:
        print("Colunas mínimas ausentes (Status_Custo / SKU_Normalizado).", file=sys.stderr)
        return 1

    sem = df[df["Status_Custo"].astype(str).str.strip() == STATUS_SEM].copy()
    _print_resumo_console(df, sem)

    rel = _build_relatorio(sem)
    meta_path = out_dir / "metadata.json"
    if meta_path.is_file():
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
        print("metadata generated_at:", meta.get("generated_at"), "\n")

    xlsx_path = out_dir / "SKUs_Sem_Custo_Para_Pedro.xlsx"
    csv_path = out_dir / "SKUs_Sem_Custo_Para_Pedro.csv"

    out_dir.mkdir(parents=True, exist_ok=True)
    rel.to_csv(csv_path, index=False, encoding="utf-8-sig")
    _write_xlsx(xlsx_path, rel)

    print("=== TOP 30 (por Receita_Vl_Venda) ===\n")
    if rel.empty:
        print("(nenhum)")
    else:
        print(rel.head(30).to_string(index=False))
    print(f"\nSKUs únicos sem custo: {len(rel)}")
    if not rel.empty:
        print(f"Soma Receita_Vl_Venda: R$ {float(rel['Receita_Vl_Venda'].sum()):,.2f}")
    print(f"\nCSV:  {csv_path}")
    print(f"XLSX: {xlsx_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
