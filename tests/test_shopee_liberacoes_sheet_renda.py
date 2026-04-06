"""Liberações Shopee: dados na aba «Renda» (4ª aba no export típico)."""

from __future__ import annotations

from pathlib import Path

import pytest

from etapa2_liberacoes import read_shopee_liberacoes_input_file, resolve_shopee_liberacoes_sheet


@pytest.fixture
def xlsx_quatro_abas_renda(tmp_path: Path) -> Path:
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    p = tmp_path / "lib_shopee.xlsx"
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Resumo"
    ws0.append(["col errada", "x"])
    wb.create_sheet("Tab2")
    wb.create_sheet("Tab3")
    ws_renda = wb.create_sheet("Renda")
    ws_renda.append(
        [
            "ID do pedido",
            "Data de conclusão do pagamento",
            "Quantia total lançada (R$)",
        ]
    )
    ws_renda.append(["251234567", "01/02/2026 10:00", "100,50"])
    wb.save(p)
    return p


@pytest.fixture
def xlsx_sem_nome_renda_quarta_com_dados(tmp_path: Path) -> Path:
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    p = tmp_path / "lib_sem_nome.xlsx"
    wb = Workbook()
    wb.active.title = "A"
    wb.create_sheet("B")
    wb.create_sheet("C")
    ws4 = wb.create_sheet("D")
    ws4.append(
        [
            "ID do pedido",
            "Data de conclusão do pagamento",
            "Quantia total lançada (R$)",
        ]
    )
    ws4.append(["999", "15/03/2026", "42,00"])
    wb.save(p)
    return p


def test_resolve_prefere_aba_renda(xlsx_quatro_abas_renda: Path) -> None:
    assert resolve_shopee_liberacoes_sheet(xlsx_quatro_abas_renda) == "Renda"


def test_resolve_quarta_aba_sem_nome_renda(xlsx_sem_nome_renda_quarta_com_dados: Path) -> None:
    assert resolve_shopee_liberacoes_sheet(xlsx_sem_nome_renda_quarta_com_dados) == 3


def test_read_shopee_liberacoes_le_aba_renda(
    xlsx_quatro_abas_renda: Path,
) -> None:
    df = read_shopee_liberacoes_input_file(xlsx_quatro_abas_renda)
    assert "ID do pedido" in df.columns
    assert len(df) >= 1
    assert df["ID do pedido"].astype(str).str.contains("251234567", na=False).any()


def test_read_shopee_liberacoes_quarta_aba_por_indice(
    xlsx_sem_nome_renda_quarta_com_dados: Path,
) -> None:
    df = read_shopee_liberacoes_input_file(xlsx_sem_nome_renda_quarta_com_dados)
    assert "ID do pedido" in df.columns
    assert (df["ID do pedido"].astype(str).str.strip() == "999").any()
